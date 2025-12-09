# reportes_licitaciones.py
import os
from tkinter import filedialog, messagebox

# Excel
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

# PDF
try:
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except Exception:
    REPORTLAB_AVAILABLE = False


class ReportesLicitaciones:
    """
    Generador de reportes (Excel y PDF) para la vista 'Competidores y Resultados'.
    Marca en VERDE a todos los GANADORES (competidores y nuestra empresa)
    y, dentro del “pack” expandible, resalta los lotes ganados.
    """

    # ========= Helpers =========
    def _ask_path(self, defext, title_sufijo):
        default_name = f"Resultados_Detallados_{title_sufijo}"
        return filedialog.asksaveasfilename(
            initialfile=default_name,
            defaultextension=defext,
            filetypes=[("Excel", "*.xlsx")] if defext == ".xlsx" else [("PDF", "*.pdf")]
        )

    def _map_ganadores_por_lote(self, lic):
        """
        Devuelve:
        - winners_by_lot: { '1': {'ganador': 'NOMBRE', 'nuestra_empresa': 'Z... o None', 'es_nuestro': bool}, ... }
        - winners_global: set con nombres de participantes que ganaron al menos un lote
        - winners_nuestros_lotes: set con números de lote ganados por nosotros
        """
        winners_by_lot = {}
        winners_global = set()
        winners_nuestros_lotes = set()

        # Los atributos existen en tu modelo Lote: ganador_nombre, ganado_por_nosotros, empresa_nuestra
        # (estos se cargan desde la tabla licitacion_ganadores_lote) :contentReference[oaicite:0]{index=0} :contentReference[oaicite:1]{index=1}
        for lote in getattr(lic, "lotes", []):
            num = str(getattr(lote, "numero", ""))
            ganador = getattr(lote, "ganador_nombre", "") or ""
            es_nuestro = bool(getattr(lote, "ganado_por_nosotros", False))
            emp_ntra = getattr(lote, "empresa_nuestra", None)
            if ganador:
                winners_by_lot[num] = {
                    "ganador": ganador,
                    "nuestra_empresa": emp_ntra,
                    "es_nuestro": es_nuestro
                }
                winners_global.add(ganador)
                if es_nuestro:
                    winners_nuestros_lotes.add(num)

        return winners_by_lot, winners_global, winners_nuestros_lotes

    def _es_fila_ganadora_participante(self, nombre_participante, winners_global, lic):
        """
        True si ese participante ganó algún lote o si es nuestra empresa adjudicada.
        """
        if nombre_participante in winners_global:
            return True

        # También consideramos nuestra empresa (si la licitación quedó adjudicada a una de las nuestras)
        if getattr(lic, "estado", "") == "Adjudicada":
            ntras = {str(e) for e in getattr(lic, "empresas_nuestras", [])}
            if getattr(lic, "adjudicada_a", "") in ntras:
                # En tu vista agregas un “participante” tipo “➡️ Nuestras Empresas …”.
                # A ese le damos verde cuando realmente ganamos.
                if "Nuestra" in nombre_participante or "Nuestras" in nombre_participante:
                    return True
        return False

    # ========= Excel =========
    def exportar_resultados_excel(self, licitacion):
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("Librería faltante", "Instala openpyxl:  pip install openpyxl")
            return
        ruta = self._ask_path(".xlsx", "Excel.xlsx")
        if not ruta:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resultados Detallados"

        # Encabezados (con nueva columna GANADOR)
        headers = [
            "Participante / Lote Ofertado",
            "Monto Ofertado",
            "Monto Habilitado (Fase A)",
            "Estado Fase A",
            "Monto Base Lote",
            "% Diferencia",
            "Ganador"
        ]
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")

        # Estilos
        fill_ganador = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")  # verde suave
        font_rojo = Font(color="FF0000")

        winners_by_lot, winners_global, winners_nuestros_lotes = self._map_ganadores_por_lote(licitacion)

        # Construimos la misma estructura que ves en pantalla:
        # 1) “competidores” (incluye nuestra fila “➡️ … (Nuestra Oferta)” que tú ya generas en la vista)
        participantes = [o.__dict__ for o in getattr(licitacion, "oferentes_participantes", [])]

        # Agregar nuestra “fila” de oferta (igual que en tu UI)
        nuestras = ", ".join(str(e) for e in getattr(licitacion, "empresas_nuestras", [])) or "Nuestras Empresas"
        nuestras_ofertas = [
            {'lote_numero': l.numero, 'monto': l.monto_ofertado, 'paso_fase_A': l.fase_A_superada}
            for l in getattr(licitacion, "lotes", [])
            if l.participamos
        ]
        participantes.append({
            "nombre": f"➡️ {nuestras} (Nuestra Oferta)",
            "es_nuestra": True,
            "ofertas_por_lote": nuestras_ofertas
        })

        # 2) Orden como en tu tabla: monto habilitado total (los que no tienen = al final)
        def _monto_hab_total(p):
            ofertas = p.get("ofertas_por_lote", [])
            # Si Fase A ya evaluada, se suman solo los que pasaron; si no, 0 (tu misma lógica)
            # Para Excel asumimos que la vista ya está en “Resultados”, así que usamos paso_fase_A
            return sum(o.get("monto", 0) for o in ofertas if o.get("paso_fase_A", False))

        participantes_ordenados = sorted(
            participantes,
            key=lambda p: _monto_hab_total(p) if _monto_hab_total(p) > 0 else float("inf")
        )

        # 3) Escribir filas
        for p in participantes_ordenados:
            nombre = p.get("nombre", "")
            es_ganador_fila = self._es_fila_ganadora_participante(nombre, winners_global, licitacion)

            # Fila “padre” (participante)
            row = [nombre, "", f"RD$ {_monto_hab_total(p):,.2f}" if _monto_hab_total(p) > 0 else "N/D",
                   "", "", "", "Sí" if es_ganador_fila else "No"]
            ws.append(row)
            if es_ganador_fila:
                for c in ws[ws.max_row]:
                    c.fill = fill_ganador

            # Hijas: una por cada lote ofertado por ese participante
            for oferta in sorted(p.get("ofertas_por_lote", []), key=lambda o: str(o.get("lote_numero", ""))):
                num = str(oferta.get("lote_numero", ""))
                lot_obj = next((l for l in getattr(licitacion, "lotes", []) if str(l.numero) == num), None)
                nombre_lote = getattr(lot_obj, "nombre", "N/E")
                monto_base = float(getattr(lot_obj, "monto_base", 0) or 0)
                monto_of = float(oferta.get("monto", 0) or 0)

                dif = ""
                if monto_base > 0 and monto_of > 0:
                    dif = f"{((monto_of - monto_base)/monto_base)*100:,.2f}%"

                paso_a = oferta.get("paso_fase_A", False)
                estado_a = "✅" if paso_a else "❌"

                # ¿Ganó este lote?
                info_g = winners_by_lot.get(num, {})
                gano_este_lote = False
                # a) si el ganador es este participante
                if info_g.get("ganador") and info_g["ganador"] in nombre:
                    gano_este_lote = True
                # b) si el lote es “nuestro”
                if info_g.get("es_nuestro") and p.get("es_nuestra"):
                    gano_este_lote = True

                fila_lote = [
                    f"   ↳ Lote {num}: {nombre_lote}",
                    f"RD$ {monto_of:,.2f}" if monto_of > 0 else "RD$ 0.00",
                    "",                       # Monto Habilitado ya resumido en la fila padre
                    estado_a,
                    f"RD$ {monto_base:,.2f}" if monto_base > 0 else "N/D",
                    dif or "N/D",
                    "Sí" if gano_este_lote else "No"
                ]
                ws.append(fila_lote)

                # Estilos por lote
                if not paso_a:
                    # si no pasó Fase A, pinto en rojo esa fila
                    for c in ws[ws.max_row]:
                        c.font = font_rojo
                if gano_este_lote:
                    for c in ws[ws.max_row]:
                        c.fill = fill_ganador

        # Ajuste de ancho
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                v = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(v))
            ws.column_dimensions[col_letter].width = max_len + 2

        try:
            wb.save(ruta)
            messagebox.showinfo("Éxito", f"Exportado a:\n{ruta}")
        except PermissionError:
            messagebox.showerror("Error", "Cierra el archivo si está abierto e inténtalo de nuevo.")

    # ========= PDF =========
    def exportar_resultados_pdf(self, licitacion):
        if not REPORTLAB_AVAILABLE:
            messagebox.showerror("Librería faltante", "Instala reportlab:  pip install reportlab")
            return
        ruta = self._ask_path(".pdf", "PDF.pdf")
        if not ruta:
            return

        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name="right", alignment=TA_RIGHT))
        styles.add(ParagraphStyle(name="center", alignment=TA_CENTER))
        styles.add(ParagraphStyle(name="small", fontSize=9, leading=11))
        doc = SimpleDocTemplate(ruta, pagesize=letter)

        elems = []
        elems.append(Paragraph("Resultados Detallados", styles["h1"]))
        elems.append(Paragraph(f"{licitacion.nombre_proceso}", styles["h2"]))
        elems.append(Spacer(1, 0.2 * inch))

        winners_by_lot, winners_global, winners_nuestros_lotes = self._map_ganadores_por_lote(licitacion)

        # Tabla (con columna "Ganador")
        head = ["Participante / Lote", "Monto Ofertado", "Habilitado (A)", "Estado A", "Base Lote", "% Dif.", "Ganador"]
        data = [head]
        tstyle = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('GRID', (0,0), (-1,-1), 0.6, colors.black),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ])

        # Igual que en Excel:
        participantes = [o.__dict__ for o in getattr(licitacion, "oferentes_participantes", [])]
        nuestras = ", ".join(str(e) for e in getattr(licitacion, "empresas_nuestras", [])) or "Nuestras Empresas"
        nuestras_ofertas = [
            {'lote_numero': l.numero, 'monto': l.monto_ofertado, 'paso_fase_A': l.fase_A_superada}
            for l in getattr(licitacion, "lotes", [])
            if l.participamos
        ]
        participantes.append({"nombre": f"➡️ {nuestras} (Nuestra Oferta)", "es_nuestra": True, "ofertas_por_lote": nuestras_ofertas})

        def _monto_hab_total(p):
            return sum(o.get("monto", 0) for o in p.get("ofertas_por_lote", []) if o.get("paso_fase_A", False))

        participantes_ordenados = sorted(
            participantes,
            key=lambda p: _monto_hab_total(p) if _monto_hab_total(p) > 0 else float("inf")
        )

        current_row = 1
        for p in participantes_ordenados:
            nombre = p.get("nombre", "")
            es_ganador_fila = self._es_fila_ganadora_participante(nombre, winners_global, licitacion)

            padre = [
                Paragraph(f"<b>{nombre}</b>", styles["Normal"]),
                "", Paragraph(f"RD$ {_monto_hab_total(p):,.2f}", styles["right"]),
                "", "", "", "Sí" if es_ganador_fila else "No"
            ]
            data.append(padre)
            if es_ganador_fila:
                tstyle.add('BACKGROUND', (0, current_row), (-1, current_row), colors.lightgreen)
            current_row += 1

            for oferta in sorted(p.get("ofertas_por_lote", []), key=lambda o: str(o.get("lote_numero", ""))):
                num = str(oferta.get("lote_numero", ""))
                lot_obj = next((l for l in getattr(licitacion, "lotes", []) if str(l.numero) == num), None)
                nombre_lote = getattr(lot_obj, "nombre", "N/E")
                base = float(getattr(lot_obj, "monto_base", 0) or 0)
                monto = float(oferta.get("monto", 0) or 0)

                dif = ""
                if base > 0 and monto > 0:
                    dif = f"{((monto - base)/base)*100:.2f}%"

                paso_a = oferta.get("paso_fase_A", False)
                estado_a = "✅" if paso_a else "❌"

                info_g = winners_by_lot.get(num, {})
                gano_este_lote = False
                if info_g.get("ganador") and info_g["ganador"] in nombre:
                    gano_este_lote = True
                if info_g.get("es_nuestro") and p.get("es_nuestra"):
                    gano_este_lote = True

                fila = [
                    Paragraph(f"&nbsp;&nbsp;&nbsp;↳ Lote {num}: {nombre_lote}", styles["small"]),
                    Paragraph(f"RD$ {monto:,.2f}", styles["right"]),
                    "", Paragraph(estado_a, styles["center"]),
                    Paragraph(f"RD$ {base:,.2f}" if base > 0 else "N/D", styles["right"]),
                    Paragraph(dif or "N/D", styles["right"]),
                    "Sí" if gano_este_lote else "No"
                ]
                data.append(fila)

                if not paso_a:
                    # texto en rojo
                    tstyle.add('TEXTCOLOR', (0, current_row), (-1, current_row), colors.red)
                if gano_este_lote:
                    tstyle.add('BACKGROUND', (0, current_row), (-1, current_row), colors.lightgreen)

                current_row += 1

        table = Table(
            data,
            colWidths=[3.5*inch, 1.3*inch, 1.3*inch, 0.9*inch, 1.3*inch, 0.9*inch, 0.9*inch],
            repeatRows=1
        )
        table.setStyle(tstyle)
        elems.append(table)

        try:
            doc.build(elems)
            messagebox.showinfo("Éxito", f"Exportado a:\n{ruta}")
        except PermissionError:
            messagebox.showerror("Error", "Cierra el PDF si está abierto e inténtalo de nuevo.")
