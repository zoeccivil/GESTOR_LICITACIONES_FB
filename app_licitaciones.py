import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
import datetime, json, os, sys, subprocess, logging, shutil
from tkcalendar import DateEntry
from ttkthemes import ThemedTk
from logic_licitaciones import DatabaseManager
import sys
import subprocess
from ttkthemes import ThemedTk
import logging
import shutil
# Soporte opcional para exportar a iCalendar
try:
    from ics import Calendar, Event
    ICS_AVAILABLE = True
except Exception:
    ICS_AVAILABLE = False
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import inch
from reportlab.lib import colors





# Dependencias de Reportes y Gráficos (Opcionales)
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
try:
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
try:
    from matplotlib.figure import Figure
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    import matplotlib.pyplot as plt
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False
try:
    from ics import Calendar, Event
    ICS_AVAILABLE = True
except ImportError:
    ICS_AVAILABLE = False

import traceback

import io
import platform
from zipfile import ZipFile, ZIP_DEFLATED

# PDF/QR
try:
    import qrcode
    from PyPDF2 import PdfMerger, PdfReader
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    QR_AVAILABLE = True
except ImportError:
    QR_AVAILABLE = False


# Pega esta NUEVA clase en tu archivo, junto a los otros diálogos

class DialogoResultadosIA(simpledialog.Dialog):
    """Muestra los documentos extraídos por la IA y permite al usuario seleccionarlos."""
    def __init__(self, parent, documentos_extraidos):
        self.documentos = documentos_extraidos
        super().__init__(parent, "Resultados del Análisis de IA")

    def body(self, master):
        self.geometry("700x450")
        label = ttk.Label(master, text="La IA ha extraído los siguientes documentos del pliego. Selecciona los que deseas importar:")
        label.pack(padx=10, pady=10)

        tree_frame = ttk.Frame(master)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        self.tree = ttk.Treeview(tree_frame, columns=('nombre', 'categoria'), show='tree headings')
        self.tree.column("#0", width=40, anchor=tk.CENTER)
        self.tree.heading("#0", text="Sel.")
        self.tree.heading('nombre', text='Nombre del Documento')
        self.tree.heading('categoria', text='Categoría Sugerida')
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.selection_status = {}
        for i, doc in enumerate(self.documentos):
            nombre = doc.get('nombre', 'N/A')
            categoria = doc.get('categoria', 'Técnica')
            self.selection_status[i] = tk.BooleanVar(value=True) # Por defecto, todos seleccionados
            check_text = '☑' if self.selection_status[i].get() else '☐'
            self.tree.insert('', tk.END, text=check_text, values=(nombre, categoria), iid=i)
            
        self.tree.bind("<Button-1>", self._toggle_selection)
        
        return self.tree

    def _toggle_selection(self, event):
        row_id_str = self.tree.identify_row(event.y)
        if not row_id_str: return
        
        row_id = int(row_id_str)
        self.selection_status[row_id].set(not self.selection_status[row_id].get())
        check_text = '☑' if self.selection_status[row_id].get() else '☐'
        self.tree.item(row_id, text=check_text)
        
    def apply(self):
        self.result = [doc for i, doc in enumerate(self.documentos) if self.selection_status[i].get()]

# Pega esta NUEVA función en tu archivo (fuera de cualquier clase)

def _extraer_texto_de_pdf(ruta_archivo):
    """Extrae el texto de un archivo PDF."""
    try:
        from PyPDF2 import PdfReader
        texto_completo = ""
        with open(ruta_archivo, 'rb') as f:
            reader = PdfReader(f)
            for page in reader.pages:
                texto_completo += page.extract_text() + "\n"
        return texto_completo
    except Exception as e:
        messagebox.showerror("Error de Lectura", f"No se pudo leer el archivo PDF:\n{e}")
        return None

class LoggingList(list):
    """Una subclase de list que registra sus modificaciones en el visor de diagnóstico."""
    def __init__(self, name, app_instance, *args):
        super(LoggingList, self).__init__(*args)
        self.name = name
        self.app = app_instance # Guardamos la referencia a la app principal

    def _log_action(self, action_name, item_details=""):
        # Extraemos el contexto de la llamada para saber quién modificó la lista
        contexto = []
        for line in traceback.format_stack(limit=5)[:-2]: # Ignoramos las llamadas internas
            contexto.append(line.strip())
        
        payload = {
            "lista_modificada": self.name,
            "accion": action_name,
            "tamaño_anterior": len(self),
            "detalles": str(item_details),
            "contexto_llamada": "\n".join(contexto)
        }
        # Usamos la función de log de la app principal
        self.app.debug_log("MODIFICACIÓN DE LISTA", payload)

    def append(self, item):
        self._log_action("append", item)
        super(LoggingList, self).append(item)

    def remove(self, item):
        self._log_action("remove", item)
        super(LoggingList, self).remove(item)

    def extend(self, iterable):
        self._log_action("extend")
        super(LoggingList, self).extend(iterable)

    def pop(self, *args):
        self._log_action("pop")
        return super(LoggingList, self).pop(*args)

    def clear(self):
        self._log_action("clear")
        super(LoggingList, self).clear()

    def __setitem__(self, key, value):
        if isinstance(key, slice):
            self._log_action("__setitem__[:]")
        else:
            self._log_action(f"__setitem__[{key}]", value)
        super(LoggingList, self).__setitem__(key, value)

    def __delitem__(self, key):
        self._log_action(f"__delitem__[{key}]")
        super(LoggingList, self).__delitem__(key)
# =================================================================================
# 0.0 CLASE DE AYUDA PARA TOOLTIPS
# =================================================================================
class Tooltip:
    """
    Crea un tooltip (ventana emergente) para un widget determinado.
    """
    def __init__(self, widget, text_func):
        self.widget = widget
        self.text_func = text_func
        self.tooltip_window = None
        widget.bind("<Enter>", self.enter)
        widget.bind("<Leave>", self.leave)
        widget.bind("<ButtonPress>", self.leave)

    def enter(self, event=None):
        self.schedule()

    def leave(self, event=None):
        self.unschedule()
        self.hidetip()

    def schedule(self):
        self.unschedule()
        self.id = self.widget.after(500, self.showtip) # Espera 500ms para mostrar

    def unschedule(self):
        id = getattr(self, 'id', None)
        if id:
            self.widget.after_cancel(id)

    def showtip(self, event=None):
        text = self.text_func()
        if not text:
            return

        self.hidetip() 
        
        # Obtenemos la posición del cursor directamente para mayor fiabilidad
        x, y = self.widget.winfo_pointerxy()
        x += 20
        y += 15
        
        self.tooltip_window = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        
        label = ttk.Label(tw, text=text, justify=tk.LEFT,
                          background="#ffffe0", relief=tk.SOLID, borderwidth=1,
                          wraplength=500, 
                          font=("tahoma", "8", "normal"))
        label.pack(ipadx=1)

    def hidetip(self):
        tw = self.tooltip_window
        self.tooltip_window = None
        if tw:
            tw.destroy()


# En gestor_licitaciones_db_2.py
# Pega esta nueva clase completa cerca del inicio del archivo

class ScrollableFrame(ttk.Frame):
    """Un frame con una barra de desplazamiento vertical y scroll de mouse."""
    def __init__(self, container, *args, **kwargs):
        super().__init__(container, *args, **kwargs)

        self.canvas = tk.Canvas(self, borderwidth=0, highlightthickness=0)
        self.scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)
        self.canvas_frame = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")

        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        self.scrollable_frame.bind("<Configure>", self.on_frame_configure)
        self.canvas.bind("<Configure>", self.on_canvas_configure)

        # Enlaza el scroll solo cuando el mouse está SOBRE el canvas
        self.canvas.bind('<Enter>', self._bind_mouse_wheel)
        self.canvas.bind('<Leave>', self._unbind_mouse_wheel)

        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")

    def _bind_mouse_wheel(self, event):
        self.canvas.bind_all("<MouseWheel>", self._on_mouse_wheel)

    def _unbind_mouse_wheel(self, event):
        self.canvas.unbind_all("<MouseWheel>")

    def _on_mouse_wheel(self, event):
        """Maneja el evento de la rueda del mouse para hacer scroll."""
        # En Windows, delta es +/-120. En otros SO puede variar.
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def on_frame_configure(self, event):
        """Actualiza la región de scroll del canvas."""
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def on_canvas_configure(self, event):
        """Asegura que el frame interior ocupe todo el ancho del canvas."""
        self.canvas.itemconfig(self.canvas_frame, width=event.width)




def setup_logging():
    """Configura un sistema de registro de errores en un archivo."""
    logging.basicConfig(
        level=logging.ERROR,
        filename='error_log.txt',
        filemode='a',
        format='%(asctime)s - %(levelname)s - %(message)s\n\n'
    )

def handle_exception(exc_type, exc_value, exc_traceback):
    """Maneja cualquier excepción no controlada y la escribe en el log."""
    if issubclass(exc_type, KeyboardInterrupt):
        sys.__excepthook__(exc_type, exc_value, exc_traceback)
        return
    
    logging.error("Excepción no controlada:", exc_info=(exc_type, exc_value, exc_traceback))
    
    messagebox.showerror(
        "Error Inesperado",
        "La aplicación ha encontrado un error inesperado.\n\n"
        "Se ha guardado un registro del error en 'error_log.txt'.\n"
        "Por favor, reinicie la aplicación."
    )

sys.excepthook = handle_exception

# =================================================================================
# 0.1 CLASE DE GENERACIÓN DE REPORTES
# =================================================================================

# --- IMPORTS/FLAGS (déjalos arriba junto a tus otros imports) ---
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import inch
from reportlab.lib import colors
from tkinter import messagebox

OPENPYXL_AVAILABLE = True
REPORTLAB_AVAILABLE = True


class ReportGenerator:
    # --------------------- API PÚBLICA ---------------------
    def generate_bid_results_report(self, licitacion, file_path):
        if file_path.endswith('.xlsx'):
            if not OPENPYXL_AVAILABLE:
                messagebox.showerror("Librería Faltante", "La librería 'openpyxl' es necesaria para exportar a Excel. Instala: pip install openpyxl")
                return
            self._generate_bid_excel(licitacion, file_path)
        elif file_path.endswith('.pdf'):
            if not REPORTLAB_AVAILABLE:
                messagebox.showerror("Librería Faltante", "La librería 'reportlab' es necesaria para exportar a PDF. Instala: pip install reportlab")
                return
            self._generate_bid_pdf(licitacion=licitacion, file_path=file_path)

    def generate_institution_history_report(self, all_bids, file_path):
        """all_bids puede ser:
           - list[Licitacion]
           - dict[str, list[Licitacion]]  (clave = institución)
        """
        if file_path.endswith('.xlsx'):
            if not OPENPYXL_AVAILABLE:
                messagebox.showerror("Librería Faltante", "Se necesita 'openpyxl' para exportar a Excel.")
                return
            self._generate_institution_excel(all_bids, file_path)
        elif file_path.endswith('.pdf'):
            if not REPORTLAB_AVAILABLE:
                messagebox.showerror("Librería Faltante", "Se necesita 'reportlab' para exportar a PDF.")
                return
            self._generate_institution_pdf(all_bids, file_path)

    # --------------------- HELPERS INTERNOS ---------------------
    @staticmethod
    def _norm(s):
        s = (s or "").strip()
        s = s.replace("➡️", "").replace("(Nuestra Oferta)", "")
        while "  " in s:
            s = s.replace("  ", " ")
        return s.upper()

    def _map_ganadores_por_lote(self, lic):
        """Devuelve mapa { 'num': {'ganador': str, 'es_nuestro': bool} }."""
        winners = {}
        for lote in getattr(lic, "lotes", []):
            winners[str(getattr(lote, "numero", ""))] = {
                "ganador": (getattr(lote, "ganador_nombre", "") or "").strip(),
                "es_nuestro": bool(getattr(lote, "ganado_por_nosotros", False)),
                "empresa_nuestra": getattr(lote, "empresa_nuestra", None)
            }
        return winners

    # --------------------- EXCEL RESULTADOS ---------------------
    def _generate_bid_excel(self, licitacion, file_path):
        wb = openpyxl.Workbook()
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"

        # Estilos
        header_font = Font(bold=True, size=14)
        title_font = Font(bold=True)
        bold_font = Font(bold=True)
        winner_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")   # verde ganador
        winner_ours_fill = PatternFill(start_color="B7E1A1", end_color="B7E1A1", fill_type="solid")
        red_italic = Font(italic=True, color="FF0000")

        # ===== RESUMEN =====
        ws_resumen['A1'] = f"Reporte de Licitación: {licitacion.nombre_proceso}"
        ws_resumen['A1'].font = header_font
        resumen = [
            ("Código Proceso", licitacion.numero_proceso),
            ("Institución", licitacion.institucion),
            ("Nuestras Empresas", ", ".join(str(e) for e in licitacion.empresas_nuestras)),
            ("Estado Actual", licitacion.estado),
            ("Monto Base Total", f"RD$ {licitacion.get_monto_base_total():,.2f}"),
            ("Monto Ofertado Total", f"RD$ {licitacion.get_oferta_total():,.2f}"),
            ("Diferencia (%)", f"{licitacion.get_diferencia_porcentual():.2f}%"),
            ("Progreso Docs", f"{licitacion.get_porcentaje_completado():.1f}%"),
        ]
        r = 3
        for k, v in resumen:
            ws_resumen.cell(row=r, column=1, value=k).font = title_font
            ws_resumen.cell(row=r, column=2, value=v)
            r += 1

        # ===== RESULTADOS COMPETIDORES =====
        ws = wb.create_sheet("Resultados Competidores")
        ws.append(["Participante / Lote", "Monto Ofertado", "Monto Habilitado (A)", "Estado A", "Base Lote", "% Dif.", "Ganador", "Empresa Nuestra"])
        for c in ws[1]:
            c.font = bold_font
            c.alignment = Alignment(horizontal='center')

        winners_by_lot = self._map_ganadores_por_lote(licitacion)
        nuestras_empresas = {self._norm(str(e)) for e in getattr(licitacion, "empresas_nuestras", [])}

        # Construir lista de participantes (competidores + nuestra fila agregada)
        participantes = [o.__dict__ for o in getattr(licitacion, "oferentes_participantes", [])]
        # Fila "nuestra oferta"
        nuestras = ", ".join(str(e) for e in getattr(licitacion, "empresas_nuestras", [])) or "Nuestras Empresas"
        nuestras_ofertas = [
            {'lote_numero': l.numero, 'monto': l.monto_ofertado, 'paso_fase_A': l.fase_A_superada}
            for l in getattr(licitacion, "lotes", [])
            if getattr(l, "participamos", False)
        ]
        participantes.append({"nombre": f"➡️ {nuestras} (Nuestra Oferta)", "es_nuestra": True, "ofertas_por_lote": nuestras_ofertas})

        def _monto_hab_total(p):
            return sum(o.get('monto', 0) for o in p.get("ofertas_por_lote", []) if o.get('paso_fase_A', False))

        participantes_orden = sorted(
            participantes,
            key=lambda it: _monto_hab_total(it) if _monto_hab_total(it) > 0 else float('inf')
        )

        for p in participantes_orden:
            nombre = p.get("nombre", "")
            # Fila padre
            fila_padre = [nombre, "", f"RD$ {_monto_hab_total(p):,.2f}" if _monto_hab_total(p) > 0 else "N/D", "", "", "", "", ""]
            ws.append(fila_padre)
            row_padre = ws.max_row
            ws.cell(row=row_padre, column=1).font = bold_font

            gano_alguno = 0

            # Hijas por lote
            for oferta in sorted(p.get("ofertas_por_lote", []), key=lambda o: str(o.get('lote_numero', ''))):
                num = str(oferta.get('lote_numero', ''))
                lot = next((l for l in getattr(licitacion, "lotes", []) if str(l.numero) == num), None)
                nombre_lote = getattr(lot, "nombre", "N/E")
                base = float(getattr(lot, "monto_base", 0) or 0)
                monto = float(oferta.get('monto', 0) or 0)
                pasoA = bool(oferta.get('paso_fase_A', False))

                dif = ""
                if base > 0 and monto > 0:
                    dif = f"{((monto - base)/base)*100:.2f}%"

                info_g = winners_by_lot.get(num, {})
                ganador_real = self._norm(info_g.get("ganador", ""))
                ganador_txt = info_g.get("ganador", "")
                es_nuestro_ganador = bool(info_g.get("es_nuestro", False))

                # ¿esta fila (participante) ganó este lote?
                nombres_en_padre = {x.strip() for x in self._norm(nombre).split(",") if x.strip()}
                es_ganador_esta_fila = False
                if ganador_real:
                    if p.get("es_nuestra") and (ganador_real in nuestras_empresas):
                        es_ganador_esta_fila = True
                    elif ganador_real in nombres_en_padre:
                        es_ganador_esta_fila = True
                    elif self._norm(nombre).startswith(ganador_real):
                        es_ganador_esta_fila = True

                fila = [
                    f"   ↳ Lote {num}: {nombre_lote}",
                    f"RD$ {monto:,.2f}" if monto > 0 else "RD$ 0.00",
                    "",                      # habilitado (resumen en padre)
                    "✅" if pasoA else "❌",
                    f"RD$ {base:,.2f}" if base > 0 else "N/D",
                    dif or "N/D",
                    "Sí" if es_ganador_esta_fila else "No",
                    info_g.get("empresa_nuestra") if es_nuestro_ganador else ""
                ]
                ws.append(fila)
                row = ws.max_row

                if not pasoA:
                    for c in ws[row]:
                        c.font = red_italic

                if es_ganador_esta_fila:
                    for c in ws[row]:
                        c.fill = winner_ours_fill if (p.get("es_nuestra") and es_nuestro_ganador) else winner_fill
                    gano_alguno += 1

            # Si ganó al menos un lote -> pintar padre y setear “Ganador”
            if gano_alguno > 0:
                for c in ws[row_padre]:
                    c.fill = winner_ours_fill if p.get("es_nuestra") else winner_fill
                ws.cell(row=row_padre, column=7, value=f"Sí ({gano_alguno})")

        # Ajuste de columnas
        for col in ws.columns:
            max_len = 0
            letter = col[0].column_letter
            for cell in col:
                v = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(v))
            ws.column_dimensions[letter].width = max_len + 2

        wb.save(file_path)

    # --------------------- PDF RESULTADOS ---------------------
    def _generate_bid_pdf(self, licitacion, file_path):


        lic = licitacion

        # 1) Documento en horizontal (Carta apaisado) + márgenes estándar
        doc = SimpleDocTemplate(
            file_path,
            pagesize=landscape(letter),
            leftMargin=0.5*inch, rightMargin=0.5*inch,
            topMargin=0.5*inch, bottomMargin=0.5*inch
        )

        # 2) Estilos con WRAP
        styles = getSampleStyleSheet()
        # texto pequeño y con wrapping agresivo (CJK fuerza corte aunque no haya espacios)
        styles.add(ParagraphStyle(name="small", fontSize=8, leading=10, wordWrap='CJK', alignment=TA_LEFT))
        styles.add(ParagraphStyle(name="small_right", fontSize=8, leading=10, wordWrap='CJK', alignment=TA_RIGHT))
        styles.add(ParagraphStyle(name="small_center", fontSize=8, leading=10, wordWrap='CJK', alignment=TA_CENTER))
        styles.add(ParagraphStyle(name="hsmall", fontSize=9, leading=11, wordWrap='CJK', alignment=TA_LEFT))

        elems = []
        elems.append(Paragraph("Resultados Detallados", styles["h1"]))
        elems.append(Paragraph(lic.nombre_proceso, styles["h2"]))
        elems.append(Spacer(1, 0.15*inch))

        # 3) Cabecera de la tabla
        head = [
            "Participante / Lote",
            "Monto Ofertado",
            "Habilitado (A)",
            "Estado A",
            "Base Lote",
            "% Dif.",
            "Ganador",
            "Empresa Nuestra"
        ]

        # 4) Datos y estilo base de la tabla
        data = [[Paragraph(h, styles["small_center"]) for h in head]]
        tstyle = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#DDDDDD")),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 0.6, colors.black),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('ALIGN', (1,1), (1,-1), 'RIGHT'),  # Monto Ofertado
            ('ALIGN', (2,1), (2,-1), 'RIGHT'),  # Habilitado (A)
            ('ALIGN', (3,1), (3,-1), 'CENTER'), # Estado A
            ('ALIGN', (4,1), (4,-1), 'RIGHT'),  # Base Lote
            ('ALIGN', (5,1), (5,-1), 'RIGHT'),  # % Dif.
            ('ALIGN', (6,1), (6,-1), 'CENTER'), # Ganador
            ('ALIGN', (7,1), (7,-1), 'LEFT'),   # Empresa Nuestra
        ])

        winners_by_lot = self._map_ganadores_por_lote(lic)
        nuestras_empresas = {self._norm(str(e)) for e in getattr(lic, "empresas_nuestras", [])}

        # 5) Construcción de participantes (incluimos nuestra fila)
        participantes = [o.__dict__ for o in getattr(lic, "oferentes_participantes", [])]
        nuestras = ", ".join(str(e) for e in getattr(lic, "empresas_nuestras", [])) or "Nuestras Empresas"
        nuestras_ofertas = [
            {'lote_numero': l.numero, 'monto': l.monto_ofertado, 'paso_fase_A': l.fase_A_superada}
            for l in getattr(lic, "lotes", [])
            if getattr(l, "participamos", False)
        ]
        participantes.append({"nombre": f"➡️ {nuestras} (Nuestra Oferta)", "es_nuestra": True, "ofertas_por_lote": nuestras_ofertas})

        def _monto_hab_total(p):
            return sum(o.get('monto', 0) for o in p.get("ofertas_por_lote", []) if o.get('paso_fase_A', False))

        participantes_orden = sorted(
            participantes,
            key=lambda it: _monto_hab_total(it) if _monto_hab_total(it) > 0 else float('inf')
        )

        current = 1
        for p in participantes_orden:
            nombre = p.get("nombre", "")
            padre = [
                Paragraph(f"<b>{nombre}</b>", styles["hsmall"]),
                Paragraph("", styles["small_right"]),
                Paragraph(f"RD$ {_monto_hab_total(p):,.2f}" if _monto_hab_total(p) > 0 else "N/D", styles["small_right"]),
                Paragraph("", styles["small_center"]),
                Paragraph("", styles["small_right"]),
                Paragraph("", styles["small_right"]),
                Paragraph("", styles["small_center"]),
                Paragraph("", styles["small"])
            ]
            data.append(padre)
            row_padre = current
            current += 1

            gano_alguno = 0
            for oferta in sorted(p.get("ofertas_por_lote", []), key=lambda o: str(o.get('lote_numero', ''))):
                num = str(oferta.get('lote_numero', ''))
                lot = next((l for l in getattr(lic, "lotes", []) if str(l.numero) == num), None)
                nombre_lote = getattr(lot, "nombre", "N/E")
                base = float(getattr(lot, "monto_base", 0) or 0)
                monto = float(oferta.get("monto", 0) or 0)
                pasoA = bool(oferta.get("paso_fase_A", False))

                dif = ""
                if base > 0 and monto > 0:
                    dif = f"{((monto - base)/base)*100:.2f}%"

                info_g = winners_by_lot.get(num, {})
                ganador_real_norm = self._norm(info_g.get("ganador", ""))
                ganador_txt = info_g.get("ganador", "")
                es_nuestro_ganador = bool(info_g.get("es_nuestro", False))

                nombres_en_padre = {x.strip() for x in self._norm(nombre).split(",") if x.strip()}
                es_ganador_esta_fila = False
                if ganador_real_norm:
                    if p.get("es_nuestra") and (ganador_real_norm in nuestras_empresas):
                        es_ganador_esta_fila = True
                    elif ganador_real_norm in nombres_en_padre:
                        es_ganador_esta_fila = True
                    elif self._norm(nombre).startswith(ganador_real_norm):
                        es_ganador_esta_fila = True

                fila = [
                    Paragraph(f"↳ Lote {num}: {nombre_lote}", styles["small"]),
                    Paragraph(f"RD$ {monto:,.2f}", styles["small_right"]),
                    Paragraph("", styles["small_right"]),  # habilitado (se resume arriba)
                    Paragraph("✅" if pasoA else "❌", styles["small_center"]),
                    Paragraph(f"RD$ {base:,.2f}" if base > 0 else "N/D", styles["small_right"]),
                    Paragraph(dif or "N/D", styles["small_right"]),
                    Paragraph("Sí" if es_ganador_esta_fila else "No", styles["small_center"]),
                    Paragraph(info_g.get("empresa_nuestra") if es_nuestro_ganador else "", styles["small"])
                ]
                data.append(fila)

                if not pasoA:
                    tstyle.add('TEXTCOLOR', (0, current), (-1, current), colors.red)
                if es_ganador_esta_fila:
                    tstyle.add('BACKGROUND', (0, current), (-1, current), colors.lightgreen)
                    gano_alguno += 1

                current += 1

            if gano_alguno > 0:
                tstyle.add('BACKGROUND', (0, row_padre), (-1, row_padre), colors.lightgreen)
                # poner "Sí (N)" en columna Ganador del padre
                data[row_padre][6] = Paragraph(f"Sí ({gano_alguno})", styles["small_center"])

        # 6) Columnas: usamos TODO el ancho útil (doc.width) con proporciones
        # Ajustadas para horizontal: suman 1.00
        ratios = [0.30, 0.13, 0.12, 0.08, 0.12, 0.10, 0.08, 0.07]
        col_widths = [doc.width * r for r in ratios]

        table = Table(
            data,
            colWidths=col_widths,
            repeatRows=1,
            splitByRow=True   # <- permite partir la tabla en saltos de página sin romper
        )
        table.setStyle(tstyle)
        elems.append(table)

        doc.build(elems)
    # --------------------- EXCEL/PDF HISTÓRICO POR INSTITUCIÓN ---------------------
    def _generate_institution_excel(self, all_bids, file_path):
        """Acepta list[Licitacion] o dict[str, list[Licitacion]]."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Histórico por Institución"

        headers = ["Institución", "Proceso", "Nuestras Empresas", "Monto Ofertado Total", "Estado", "Fase A Habilitada", "Comentarios", "Empresa Nuestra Adjudicada"]
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True)

        def _append_row(lic):
            empresas_str = ", ".join(str(e) for e in lic.empresas_nuestras)
            habilitado_str = "Sí" if getattr(lic, "fase_A_superada", False) else "No"
            adjudicada_ntra = ""
            if getattr(lic, "estado", "") == "Adjudicada":
                if getattr(lic, "adjudicada_a", "") in {str(e) for e in lic.empresas_nuestras}:
                    adjudicada_ntra = lic.adjudicada_a
            ws.append([getattr(lic, "institucion", ""), lic.nombre_proceso, empresas_str,
                       lic.get_oferta_total(), lic.estado, habilitado_str,
                       getattr(lic, "motivo_descalificacion", ""), adjudicada_ntra])

        if isinstance(all_bids, dict):
            for _, lst in all_bids.items():
                for lic in lst:
                    _append_row(lic)
        else:
            for lic in all_bids:
                _append_row(lic)

        # Ajuste de columnas
        for col in ws.columns:
            max_len = 0
            letter = col[0].column_letter
            for cell in col:
                v = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(v))
            ws.column_dimensions[letter].width = max_len + 2

        wb.save(file_path)

    def _generate_institution_pdf(self, all_bids, file_path):
        """Acepta list[Licitacion] o dict[str, list[Licitacion]]."""
        doc = SimpleDocTemplate(file_path, pagesize=letter)
        styles = getSampleStyleSheet()
        elems = []

        elems.append(Paragraph("Histórico de Licitaciones por Institución", styles['h1']))
        elems.append(Spacer(1, 0.25*inch))

        headers = ["Institución", "Proceso", "Nuestras Empresas", "Monto Ofertado Total", "Estado", "Fase A", "Comentarios", "Empresa Nuestra Adjudicada"]
        data = [headers]
        tstyle = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('GRID', (0,0), (-1,-1), 0.6, colors.black),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ])

        def _add(lic):
            empresas = ", ".join(str(e) for e in lic.empresas_nuestras)
            habilitado = "Sí" if getattr(lic, "fase_A_superada", False) else "No"
            adjudicada_ntra = ""
            if getattr(lic, "estado", "") == "Adjudicada":
                if getattr(lic, "adjudicada_a", "") in {str(e) for e in lic.empresas_nuestras}:
                    adjudicada_ntra = lic.adjudicada_a
            data.append([
                getattr(lic, "institucion", ""),
                Paragraph(lic.nombre_proceso, styles['Normal']),
                empresas,
                f"RD$ {lic.get_oferta_total():,.2f}",
                lic.estado,
                habilitado,
                Paragraph(getattr(lic, "motivo_descalificacion", "") or "", styles['Normal']),
                adjudicada_ntra
            ])

        if isinstance(all_bids, dict):
            for _, lst in all_bids.items():
                for lic in lst:
                    _add(lic)
        else:
            for lic in all_bids:
                _add(lic)

        table = Table(data, colWidths=[1.6*inch, 2.3*inch, 2.0*inch, 1.3*inch, 1.0*inch, 0.8*inch, 1.6*inch, 1.7*inch], repeatRows=1)
        table.setStyle(tstyle)
        elems.append(table)

        doc.build(elems)


# =================================================================================
# 1. CLASES DE DATOS
# =================================================================================

# REEMPLAZA tu clase Lote entera con esta versión final:

class Lote:
    def __init__(self, **kwargs):
        self.numero = kwargs.get("numero", "")
        self.nombre = kwargs.get("nombre", "")
        self.monto_base = float(kwargs.get("monto_base", 0.0) or 0.0)
        self.monto_base_personal = float(kwargs.get("monto_base_personal", 0.0) or 0.0)
        self.monto_ofertado = float(kwargs.get("monto_ofertado", 0.0) or 0.0)
        self.participamos = kwargs.get("participamos", True) # Por defecto participamos si se crea un lote
        self.fase_A_superada = kwargs.get("fase_A_superada", True) # Por defecto se asume superada
        self.ganador_nombre = kwargs.get("ganador_nombre", "")
        self.ganado_por_nosotros = kwargs.get("ganado_por_nosotros", False)
        self.empresa_nuestra = kwargs.get("empresa_nuestra", None)

    def to_dict(self):
        return {
            "numero": self.numero,
            "nombre": self.nombre,
            "monto_base": self.monto_base,
            "monto_base_personal": self.monto_base_personal,
            "monto_ofertado": self.monto_ofertado,
            "participamos": self.participamos,
            "fase_A_superada": self.fase_A_superada,
            "empresa_nuestra": self.empresa_nuestra # Se asegura de incluir la empresa
        }

class Oferente:
    def __init__(self, **kwargs):
        self.nombre = kwargs.get("nombre", "")
        self.comentario = kwargs.get("comentario", "")
        self.ofertas_por_lote = kwargs.get("ofertas_por_lote", [])

    def to_dict(self):
        return {
            "nombre": self.nombre, "comentario": self.comentario,
            "ofertas_por_lote": self.ofertas_por_lote,
        }

    def get_monto_total_ofertado(self, solo_habilitados=False):
        ofertas_a_sumar = self.ofertas_por_lote
        if solo_habilitados:
            ofertas_a_sumar = [o for o in self.ofertas_por_lote if o.get('paso_fase_A', True)]
        return sum(oferta.get('monto', 0) for oferta in ofertas_a_sumar)

# En la clase Documento, dentro de gestor_licitaciones_db_2.py

class Documento:
# En la clase Documento, REEMPLAZA el método __init__ con este:

# En la clase Documento, REEMPLAZA el método __init__ con este:

    def __init__(self, **kwargs):
        self.id = kwargs.get("id")
        self.codigo = kwargs.get("codigo")
        self.nombre = kwargs.get("nombre")
        self.categoria = kwargs.get("categoria")
        self.comentario = kwargs.get("comentario", "")
        self.presentado = kwargs.get("presentado", False)
        # --- CAMBIO AQUÍ ---
        # El valor por defecto ahora es "Subsanable" en lugar de "No Definido"
        self.subsanable = kwargs.get("subsanable", "Subsanable") 
        # --- FIN DEL CAMBIO ---
        self.ruta_archivo = kwargs.get("ruta_archivo", "")
        self.empresa_nombre = kwargs.get("empresa_nombre", None)
        self.responsable = kwargs.get("responsable", "Sin Asignar")
        self.revisado = kwargs.get("revisado", False)
        self.obligatorio = kwargs.get("obligatorio", False)
        self.orden_pliego = kwargs.get("orden_pliego", None)
        
    def to_dict(self):
        return {
            "id": self.id, # <-- AÑADE ESTA LÍNEA
            "codigo": self.codigo, "nombre": self.nombre, "categoria": self.categoria, 
            "comentario": self.comentario, "presentado": self.presentado, 
            "subsanable": self.subsanable, "ruta_archivo": self.ruta_archivo,
            "empresa_nombre": self.empresa_nombre,
            "responsable": self.responsable,
            "revisado": self.revisado,
            "obligatorio": self.obligatorio
        }
    def __str__(self):
        estado = "✅" if self.presentado else "❌"
        adjunto = "📎" if self.ruta_archivo else ""
        revisado_str = "👁️" if self.revisado else "" # <-- LÍNEA NUEVA
        comentario_str = f"({self.comentario})" if self.comentario else ""
        sub_str = {'Subsanable': '(S)', 'No Subsanable': '(NS)'}.get(self.subsanable, '')
        return f"{estado} {revisado_str} {adjunto} [{self.codigo}] {self.nombre} {sub_str} {comentario_str}".strip()

class Empresa:
    def __init__(self, nombre):
        self.nombre = nombre
    def to_dict(self):
        return {"nombre": self.nombre}
    def __str__(self):
        return self.nombre

# En gestor_licitaciones_db.py

class Licitacion:
# En la clase Licitacion

# En la clase Licitacion

    def __init__(self, **kwargs):
        self.id = kwargs.get("id")
        self.nombre_proceso = kwargs.get("nombre_proceso", "")
        self.numero_proceso = kwargs.get("numero_proceso", "")
        self.institucion = kwargs.get("institucion", "")
        empresas = kwargs.get("empresas_nuestras", [])
        print(f"DEBUG: Cargando licitación '{self.numero_proceso}'. Datos de 'empresas_nuestras': {empresas}")
        self.empresas_nuestras = [Empresa(e["nombre"]) for e in empresas]
        self.estado = kwargs.get("estado", "Iniciada")
        self.fase_A_superada = kwargs.get("fase_A_superada", False)
        self.fase_B_superada = kwargs.get("fase_B_superada", False)
        self.adjudicada = kwargs.get("adjudicada", False)
        self.adjudicada_a = kwargs.get("adjudicada_a", "")
        self.motivo_descalificacion = kwargs.get("motivo_descalificacion", "")
        self.docs_completos_manual = kwargs.get("docs_completos_manual", False)
        self.last_modified = kwargs.get("last_modified")
        self.riesgos = [Riesgo(**r) for r in kwargs.get("riesgos", [])]

        try:
            self.fecha_creacion = datetime.datetime.strptime(
                kwargs.get("fecha_creacion", str(datetime.date.today())), 
                '%Y-%m-%d'
            ).date()
        except (ValueError, TypeError):
            self.fecha_creacion = datetime.date.today()

        self.lotes = [Lote(**data) for data in kwargs.get("lotes", [])]
        self.oferentes_participantes = [Oferente(**data) for data in kwargs.get("oferentes_participantes", [])]
        
        documentos_data = kwargs.get("documentos_solicitados", [])
        for doc_data in documentos_data:
            if doc_data.get("categoria") == "Tecnica":
                doc_data["categoria"] = "Técnica"
        self.documentos_solicitados = [Documento(**doc) for doc in documentos_data]

        # ▼▼▼ AÑADE ESTA LÍNEA ▼▼▼
        self.fallas_fase_a = kwargs.get("fallas_fase_a", [])
        # ▲▲▲ FIN DE LA LÍNEA ▲▲▲

        cronograma_cargado = kwargs.get('cronograma', {})
        self.cronograma = {}
        eventos_posibles = [
            "Presentacion de Ofertas", "Apertura de Ofertas", "Informe de Evaluacion Tecnica",
            "Notificaciones de Subsanables", "Notificacion de Habilitacion Sobre B",
            "Apertura de Oferta Economica", "Adjudicacion"
        ]
        for evento in eventos_posibles:
            datos_evento = cronograma_cargado.get(evento)
            if isinstance(datos_evento, dict):
                self.cronograma[evento] = datos_evento
            elif isinstance(datos_evento, str):
                self.cronograma[evento] = {"fecha_limite": datos_evento, "estado": "Pendiente"}
            else:
                self.cronograma[evento] = {"fecha_limite": None, "estado": "Pendiente"}
    
    def to_dict(self):
        return {
            "id": self.id,
            "nombre_proceso": self.nombre_proceso,
            "numero_proceso": self.numero_proceso,
            "institucion": self.institucion,
            "empresas_nuestras": [e.to_dict() for e in self.empresas_nuestras],
            "estado": self.estado,
            "fase_A_superada": self.fase_A_superada,
            "fase_B_superada": self.fase_B_superada,
            "adjudicada": self.adjudicada,
            "adjudicada_a": self.adjudicada_a,
            "motivo_descalificacion": self.motivo_descalificacion,
            "docs_completos_manual": self.docs_completos_manual,
            "last_modified": self.last_modified,
            "fecha_creacion": str(self.fecha_creacion),
            "lotes": [l.to_dict() for l in self.lotes],
            "oferentes_participantes": [o.to_dict() for o in self.oferentes_participantes],
            "documentos_solicitados": [d.to_dict() for d in self.documentos_solicitados],
            "cronograma": self.cronograma,
            "riesgos": [r.to_dict() for r in self.riesgos],
            # ▼▼▼ AÑADE ESTA LÍNEA ▼▼▼
            "fallas_fase_a": self.fallas_fase_a
            # ▲▲▲ FIN DE LA LÍNEA ▲▲▲
        }

    def get_monto_base_personal_total(self):
        """Calcula la suma de los montos base personales de los lotes en los que participamos."""
        return sum(lote.monto_base_personal for lote in self.lotes if lote.participamos)

    def get_diferencia_bases_porcentual(self):
        """
        Calcula la diferencia porcentual entre el presupuesto base personal total
        y el presupuesto base de la licitación total.
        Un resultado negativo significa que tu presupuesto es menor que el de la licitación.
        """
        monto_base_licitacion = self.get_monto_base_total()
        monto_base_personal = self.get_monto_base_personal_total()

        if monto_base_licitacion > 0:
            return ((monto_base_personal - monto_base_licitacion) / monto_base_licitacion) * 100
        return 0.0

    # En gestor_licitaciones_db_2.py, dentro de la clase Licitacion

    def to_summary_dict(self):
        """Devuelve un diccionario con un resumen de la licitación, para el log."""
        return {
            "numero_proceso": self.numero_proceso,
            "nombre_proceso": self.nombre_proceso,
            "institucion": self.institucion,
            "empresa_nuestra": str(", ".join(str(e) for e in self.empresas_nuestras)
),
            "estado": self.estado,
            "monto_ofertado_total": self.get_oferta_total(),
            "cantidad_lotes": len(self.lotes),
            "cantidad_documentos": len(self.documentos_solicitados)
        }

    def get_matriz_ofertas(self):
        """
        Crea una matriz (diccionario) con todas las ofertas válidas (Fase A superada)
        organizadas por Lote y luego por Oferente.
        
        Devuelve: {lote_numero: {oferente_nombre: {'monto': X, 'plazo': Y}, ...}, ...}
        """
        matriz = {str(lote.numero): {} for lote in self.lotes}
        
        for oferente in self.oferentes_participantes:
            for oferta in oferente.ofertas_por_lote:
                # Solo consideramos ofertas que pasaron la Fase A
                if oferta.get('paso_fase_A', False):
                    lote_num_str = str(oferta.get('lote_numero'))
                    if lote_num_str in matriz:
                        matriz[lote_num_str][oferente.nombre] = {
                            'monto': oferta.get('monto', 0),
                            'plazo': oferta.get('plazo_entrega', 0)
                        }
        return matriz


    def get_riesgo_total_score(self):
        """Calcula la puntuación total de riesgo sumando (impacto * probabilidad) de cada riesgo."""
        if not self.riesgos:
            return 0
        return sum(r.impacto * r.probabilidad for r in self.riesgos)



    def calcular_mejor_paquete_individual(self):
        """
        Calcula el costo total si se adjudicara cada lote al oferente más barato
        para ese lote en específico.

        Devuelve: {'monto_total': X, 'detalle': {lote_numero: oferente_nombre, ...}}
        """
        matriz = self.get_matriz_ofertas()
        monto_total = 0.0
        detalle_adjudicacion = {}

        for lote_num, ofertas_lote in matriz.items():
            if not ofertas_lote:  # Si no hay ofertas para este lote, se ignora
                continue

            # Encontrar el oferente con el monto más bajo para este lote
            mejor_oferente = min(ofertas_lote, key=lambda oferente: ofertas_lote[oferente]['monto'])
            monto_minimo = ofertas_lote[mejor_oferente]['monto']
            
            monto_total += monto_minimo
            detalle_adjudicacion[lote_num] = mejor_oferente
            
        return {'monto_total': monto_total, 'detalle': detalle_adjudicacion}

    def calcular_mejor_paquete_por_oferente(self):
        """
        Calcula el costo total por cada oferente que participa en todos los lotes
        requeridos y devuelve el mejor paquete de un único oferente.

        Devuelve: {'monto_total': X, 'oferente': oferente_nombre} o None si nadie ofertó a todo.
        """
        paquetes_completos = {}
        # Obtenemos solo los lotes en los que nosotros participamos
        lotes_participantes = {str(l.numero) for l in self.lotes if l.participamos}
        
        if not lotes_participantes:
            return None

        for oferente in self.oferentes_participantes:
            ofertas_validas = [o for o in oferente.ofertas_por_lote if o.get('paso_fase_A', False)]
            lotes_ofertados_por_competidor = {str(o['lote_numero']) for o in ofertas_validas}

            # Verificamos si el oferente ha ofertado en TODOS los lotes en los que participamos
            if lotes_participantes.issubset(lotes_ofertados_por_competidor):
                # Sumamos solo los montos de los lotes relevantes
                monto_paquete = sum(o['monto'] for o in ofertas_validas if str(o['lote_numero']) in lotes_participantes)
                paquetes_completos[oferente.nombre] = monto_paquete
        
        if not paquetes_completos:
            return None # Ningún oferente ofertó a todos los lotes necesarios

        mejor_oferente = min(paquetes_completos, key=paquetes_completos.get)
        monto_mas_bajo = paquetes_completos[mejor_oferente]

        return {'monto_total': monto_mas_bajo, 'oferente': mejor_oferente}

# En la clase Licitacion, reemplaza este método

    def get_monto_base_total(self, solo_participados: bool = False) -> float:
        """
        Suma de montos base. Prioriza el presupuesto personal si es > 0.
        - si solo_participados=True: solo lotes elegibles.
        """
        total = 0.0
        lotes_a_evaluar = self._lotes_elegibles_para_porcentaje() if solo_participados else getattr(self, "lotes", [])
        
        for lote in lotes_a_evaluar:
            # Prioridad 1: Monto Base Personal, pero solo si es un valor positivo.
            base = float(getattr(lote, "monto_base_personal", 0.0) or 0.0)
            
            # Prioridad 2: Si el personal es 0, usar el Monto Base de la licitación.
            if base <= 0:
                base = float(getattr(lote, "monto_base", 0.0) or 0.0)
            
            total += base
        return total    

    def get_oferta_total(self, solo_participados: bool = False) -> float:
        """
        Suma de ofertas.
        - si solo_participados=True: solo lotes elegibles (participamos o con oferta > 0)
        """
        total = 0.0
        lotes = self._lotes_elegibles_para_porcentaje() if solo_participados else getattr(self, "lotes", [])
        for lote in lotes:
            total += float(getattr(lote, "monto_ofertado", 0) or 0.0)
        return total

# En la clase Licitacion, REEMPLAZA este método:

# En la clase Licitacion, REEMPLAZA este método por completo:

    def get_diferencia_porcentual(self, solo_participados=False, usar_base_personal=True):
        """
        Calcula la diferencia porcentual entre la oferta y el monto base.
        - solo_participados: Si es True, solo considera lotes en los que participamos o que tienen oferta.
        - usar_base_personal: Si es True, prioriza el monto base personal. Si es False, usa solo el de la licitación.
        """
        lotes_a_considerar = self.lotes
        if solo_participados:
            lotes_a_considerar = [
                l for l in self.lotes
                if bool(getattr(l, 'participamos', False)) or (float(getattr(l, 'monto_ofertado', 0) or 0) > 0)
            ]

        base_total = 0.0
        oferta_total = 0.0

        for lote in lotes_a_considerar:
            oferta = float(getattr(lote, 'monto_ofertado', 0) or 0)
            
            # --- LÓGICA MODIFICADA ---
            if usar_base_personal:
                base = float(getattr(lote, 'monto_base_personal', 0.0) or 0.0)
                if base <= 0:
                    base = float(getattr(lote, 'monto_base', 0.0) or 0.0)
            else:
                base = float(getattr(lote, 'monto_base', 0.0) or 0.0)
            
            base_total += base
            
            # --- CORRECCIÓN CLAVE ---
            # Si el lote está en la lista de 'lotes_a_considerar',
            # su oferta siempre debe ser sumada, sin doble verificación.
            oferta_total += oferta
            # --- FIN DE LA CORRECCIÓN ---

        if base_total == 0:
            return 0.0

        return ((oferta_total - base_total) / base_total) * 100.0

    
    def get_porcentaje_completado(self):
        total_docs = len(self.documentos_solicitados)
        if total_docs == 0:
            return 100.0 if self.docs_completos_manual else 0.0
        docs_completados = sum(1 for doc in self.documentos_solicitados if doc.presentado)
        return (docs_completados / total_docs) * 100
    
    def _lotes_elegibles_para_porcentaje(self):
        """Lotes que cuentan para %Dif: participamos o ya tienen oferta (>0)."""
        elegibles = []
        for lote in getattr(self, "lotes", []):
            participa = bool(getattr(lote, "participamos", False))
            oferta = float(getattr(lote, "monto_ofertado", 0) or 0.0)
            if participa or oferta > 0:
                elegibles.append(lote)
        return elegibles

# En la clase Licitacion, reemplaza este método:
    def get_dias_restantes(self):
        hoy = datetime.date.today()
        eventos_futuros_pendientes = []
        
        # Define un orden de prioridad para los eventos
        hitos_prioridad = [
            "Presentacion de Ofertas", "Apertura de Ofertas", "Informe de Evaluacion Tecnica",
            "Notificaciones de Subsanables", "Notificacion de Habilitacion Sobre B",
            "Apertura de Oferta Economica", "Adjudicacion"
        ]

        for nombre_evento in hitos_prioridad:
            datos_evento = self.cronograma.get(nombre_evento)
            if datos_evento and datos_evento.get("estado") == "Pendiente" and datos_evento.get("fecha_limite"):
                try:
                    fecha_limite = datetime.datetime.strptime(datos_evento["fecha_limite"], '%Y-%m-%d').date()
                    if fecha_limite >= hoy:
                        # Al encontrar el primer evento futuro y pendiente según la prioridad, lo usamos y salimos del bucle.
                        eventos_futuros_pendientes.append((fecha_limite, nombre_evento))
                        break
                except (ValueError, TypeError):
                    continue

        if eventos_futuros_pendientes:
            proxima_fecha, proximo_evento = eventos_futuros_pendientes[0]
            diferencia = (proxima_fecha - hoy).days
            
            # --- LÓGICA CORREGIDA ---
            # Ahora, todos los casos incluyen el nombre del evento.
            if diferencia == 0:
                return f"Hoy: {proximo_evento}"
            elif diferencia == 1:
                return f"Mañana: {proximo_evento}"
            else:
                return f"Faltan {diferencia} días para: {proximo_evento}"
        
        # El resto de la lógica para estados finalizados permanece igual.
        if self.estado == "Adjudicada (Ganada)": return "✅ Adjudicada"
        if self.estado in ["Adjudicada (Perdida)", "Descalificado Fase A", "Descalificado Fase B", "Desierta", "Cancelada"]: return "❌ Finalizada"
        
        return "Fases cumplidas"
    
    def clone(self):
        return Licitacion(**self.to_dict())

# =================================================================================
# 2. VENTANAS SECUNDARIAS
# =================================================================================

class VentanaGanadoresPorLote(tk.Toplevel):
    """
    Ventana para seleccionar el ganador de cada lote de una licitación.
    Permite elegir con qué empresa nuestra fue adjudicado cada lote.
    """
    def __init__(self, parent, licitacion, db):
        super().__init__(parent)
        self.parent_app = parent
        self.licitacion = licitacion
        self.db = db

        self.title(f"Seleccionar Ganadores - {licitacion.nombre_proceso}")
        self.geometry("800x400")
        self.grab_set()

        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # --- Tabla de lotes ---
        cols = ("lote", "ganador", "nuestra_empresa")
        self.tree = ttk.Treeview(main_frame, columns=cols, show="headings", height=10)
        self.tree.heading("lote", text="Lote")
        self.tree.heading("ganador", text="Ganador")
        self.tree.heading("nuestra_empresa", text="Empresa Nuestra (si aplica)")
        self.tree.column("lote", width=120, anchor=tk.CENTER)
        self.tree.column("ganador", width=220, anchor=tk.W)
        self.tree.column("nuestra_empresa", width=250, anchor=tk.W)

        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # --- Poblar tabla ---
        for lote in self.licitacion.lotes:
            self.tree.insert("", tk.END, iid=str(lote.numero), 
                             values=(f"Lote {lote.numero} - {lote.nombre}", "", ""))

        # --- Frame inferior ---
        bottom_frame = ttk.Frame(self, padding="5")
        bottom_frame.pack(fill=tk.X)

        ttk.Button(bottom_frame, text="Asignar Ganador", command=self.asignar_ganador).pack(side=tk.LEFT, padx=5)
        ttk.Button(bottom_frame, text="Cerrar", command=self.destroy).pack(side=tk.RIGHT, padx=5)

    def asignar_ganador(self):
        """Abrir un diálogo para asignar el ganador de un lote seleccionado"""
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Seleccionar Lote", "Debes seleccionar un lote en la tabla.")
            return

        lote_id = selected[0]
        lote_obj = next((l for l in self.licitacion.lotes if str(l.numero) == lote_id), None)
        if not lote_obj:
            return

        # Empresas disponibles
        competidores = [oferente.nombre for oferente in self.licitacion.oferentes_participantes]
        nuestras_empresas = [str(e) for e in self.licitacion.empresas_nuestras]

        # --- Ventana emergente ---
        win = tk.Toplevel(self)
        win.title(f"Asignar Ganador - Lote {lote_obj.numero}")
        win.geometry("400x250")
        win.transient(self)
        win.grab_set()

        ttk.Label(win, text=f"Lote {lote_obj.numero}: {lote_obj.nombre}", font=("Arial", 11, "bold")).pack(pady=10)

        ttk.Label(win, text="Seleccionar Ganador:").pack(pady=5)
        combo_ganador = ttk.Combobox(win, values=competidores, state="readonly", width=40)
        combo_ganador.pack(pady=5)

        ttk.Label(win, text="Si es nuestra empresa, selecciona cuál:").pack(pady=5)
        combo_empresa = ttk.Combobox(win, values=nuestras_empresas, state="readonly", width=40)
        combo_empresa.pack(pady=5)

        def guardar_ganador():
            ganador_nombre = combo_ganador.get()
            empresa_nuestra = combo_empresa.get() if combo_empresa.get() else None

            if not ganador_nombre:
                messagebox.showerror("Error", "Debes seleccionar un ganador.")
                return

            # Guardar en base de datos
            ok = self.db.marcar_ganador_lote(self.licitacion.id, lote_obj.numero, ganador_nombre, empresa_nuestra)
            if ok:
                self.tree.item(lote_id, values=(f"Lote {lote_obj.numero} - {lote_obj.nombre}", ganador_nombre, empresa_nuestra or ""))
                messagebox.showinfo("Éxito", f"Ganador asignado: {ganador_nombre}")
                win.destroy()
            else:
                messagebox.showerror("Error", "No se pudo guardar el ganador en la base de datos.")

        ttk.Button(win, text="Guardar", command=guardar_ganador).pack(pady=15)



class VentanaVisorDebug(tk.Toplevel):
    """Una ventana no modal para mostrar logs de diagnóstico en tiempo real."""
    def __init__(self, parent):
        super().__init__(parent)
        self.parent_app = parent
        self.title("Visor de Diagnóstico")
        self.geometry("700x400")

        self.text_widget = tk.Text(self, wrap=tk.WORD, font=("Consolas", 10), state="disabled")
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.text_widget.yview)
        self.text_widget.configure(yscrollcommand=scrollbar.set)
        
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Al cerrar, solo ocultamos la ventana, no la destruimos para no perder el log
        self.protocol("WM_DELETE_WINDOW", self.on_closing)

    def add_log_entry(self, message):
        """Añade un mensaje al visor y se desplaza hasta el final."""
        if not self.winfo_exists(): return
        self.text_widget.config(state="normal")
        self.text_widget.insert(tk.END, message + "\n")
        self.text_widget.see(tk.END)
        self.text_widget.config(state="disabled")

    def on_closing(self):
        """Maneja el cierre de la ventana a través del botón [X]."""
        # Notifica a la app principal que la ventana se ha cerrado
        self.parent_app.on_debug_viewer_closed()
        self.destroy()



class VentanaVistaLotes(tk.Toplevel):
    """
    Una ventana de solo lectura para mostrar los detalles de los lotes
    de una licitación específica.
    """
    def __init__(self, parent, licitacion):
        super().__init__(parent)
        self.licitacion = licitacion
        self.title(f"Detalle de Lotes: {licitacion.numero_proceso}")
        self.geometry("1200x450")  # Hacemos la ventana un poco más ancha
        self.grab_set()

        main_frame = ttk.Frame(self, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # --- Treeview para los lotes (con la nueva columna) ---
        tree_frame = ttk.LabelFrame(main_frame, text=f"Lotes para '{licitacion.nombre_proceso}'", padding=10)
        tree_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        cols = ("participar","fase_a","numero","nombre","monto_base",
                "monto_personal","dif_bases","monto_ofertado","dif_lic","dif_pers")

        self.tree_lotes = ttk.Treeview(tree_frame, columns=cols, show="headings")
        
        headings = {
            "participar": "Participa", "fase_a": "Fase A OK", "numero": "N°", "nombre": "Nombre del Lote",
            "monto_base": "Base Licitación", "monto_personal": "Base Personal",
            "dif_bases": "% Dif. Bases",
            "monto_ofertado": "Nuestra Oferta", "dif_lic": "% Oferta vs Licit.", "dif_pers": "% Oferta vs Pers."
        }
        for col, text in headings.items():
            self.tree_lotes.heading(col, text=text)
        
        # Ajuste de anchos de columna
        for col in ["participar", "fase_a", "numero", "dif_lic", "dif_pers", "dif_bases"]:
            self.tree_lotes.column(col, width=90, anchor=tk.CENTER)
        self.tree_lotes.column("nombre", width=250)
        for col in ["monto_base", "monto_personal", "monto_ofertado"]:
            self.tree_lotes.column(col, anchor=tk.E, width=120)

        self.tree_lotes.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree_lotes.yview)
        self.tree_lotes.configure(yscroll=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.tree_lotes.tag_configure('descalificado', foreground='red')
        self.tree_lotes.tag_configure('no_participa', foreground='grey')

        self.actualizar_tree_lotes()

        # --- Frame de Resumen Financiero ---
        summary_frame = ttk.LabelFrame(main_frame, text="Resumen Financiero (Solo lotes donde participamos)", padding=10)
        summary_frame.pack(fill=tk.X, pady=(10, 0))
        
        monto_base_total = self.licitacion.get_monto_base_total()
        monto_personal_total = self.licitacion.get_monto_base_personal_total()
        monto_ofertado_total = self.licitacion.get_oferta_total()
        diferencia_bases = self.licitacion.get_diferencia_bases_porcentual()

        ttk.Label(summary_frame, text="Monto Base Licitación Total:", font=("Helvetica", 10, "bold")).grid(row=0, column=0, sticky=tk.W)
        ttk.Label(summary_frame, text=f"RD$ {monto_base_total:,.2f}").grid(row=0, column=1, sticky=tk.E, padx=10)
        
        ttk.Label(summary_frame, text="Monto Base Personal Total:", font=("Helvetica", 10, "bold")).grid(row=1, column=0, sticky=tk.W)
        ttk.Label(summary_frame, text=f"RD$ {monto_personal_total:,.2f} ({diferencia_bases:.2f}%)").grid(row=1, column=1, sticky=tk.E, padx=10)

        ttk.Label(summary_frame, text="Monto Ofertado Total:", font=("Helvetica", 10, "bold")).grid(row=2, column=0, sticky=tk.W)
        ttk.Label(summary_frame, text=f"RD$ {monto_ofertado_total:,.2f}").grid(row=2, column=1, sticky=tk.E, padx=10)
        
        summary_frame.columnconfigure(1, weight=1)

        ttk.Button(main_frame, text="Cerrar", command=self.destroy).pack(pady=15, ipady=4)

    def actualizar_tree_lotes(self):
        """
        Rellena el Treeview de lotes con:
        - Participa
        - Fase A
        - Número
        - Nombre
        - Monto base
        - Monto base personal
        - % Dif. Bases (personal vs base)
        - Monto ofertado (si participamos)
        - % Dif. Lic. (ofertado vs base)
        - % Dif. Pers. (ofertado vs personal)
        Además, configura el iid del item = índice del lote, para que otros métodos
        que convierten selección->índice funcionen correctamente.
        """
        # 1) Seguridad: si no hay tree, salimos silenciosamente
        if not hasattr(self, "tree_lotes") or self.tree_lotes is None:
            return

        # 2) Limpiar filas actuales
        for item in self.tree_lotes.get_children():
            self.tree_lotes.delete(item)

        # 3) Asegurar lista de lotes
        lotes = getattr(self.licitacion, "lotes", []) or []

        # 4) Insertar filas
        for idx, lote in enumerate(sorted(lotes, key=lambda l: getattr(l, "numero", 0))):
            # Tags visuales
            tags = []
            if not getattr(lote, "participamos", True):
                tags.append("no_participa")
            if getattr(lote, "participamos", True) and not getattr(lote, "fase_A_superada", False):
                tags.append("descalificado")

            # Valores seguros
            base = float(getattr(lote, "monto_base", 0.0) or 0.0)
            base_pers = float(getattr(lote, "monto_base_personal", 0.0) or 0.0)
            ofertado = float(getattr(lote, "monto_ofertado", 0.0) or 0.0)
            participa = bool(getattr(lote, "participamos", True))
            fase_a = bool(getattr(lote, "fase_A_superada", False))

            # % diferencias con protección /0
            dif_bases_pct = ((base_pers - base) / base * 100.0) if base > 0 else 0.0
            dif_lic_pct   = ((ofertado  - base) / base * 100.0) if (base > 0 and participa) else 0.0
            dif_pers_pct  = ((ofertado  - base_pers) / base_pers * 100.0) if (base_pers > 0 and participa) else 0.0

            values = (
                "Sí" if participa else "No",
                "Sí" if fase_a else "No",
                getattr(lote, "numero", ""),
                getattr(lote, "nombre", ""),
                f"RD$ {base:,.2f}",
                f"RD$ {base_pers:,.2f}",
                f"{dif_bases_pct:.2f}%",
                (f"RD$ {ofertado:,.2f}" if participa else "N/A"),
                (f"{dif_lic_pct:.2f}%"   if participa else "N/A"),
                (f"{dif_pers_pct:.2f}%"  if participa else "N/A"),
            )

            # Insertamos con iid = índice, para que otros métodos puedan hacer int(iid)
            self.tree_lotes.insert("", "end", iid=str(idx), values=values, tags=tuple(tags))

        # 5) (Opcional) estilos visuales
        try:
            self.tree_lotes.tag_configure("no_participa", foreground="#888888")
            self.tree_lotes.tag_configure("descalificado", foreground="#B00000")
        except Exception:
            pass





# En gestor_licitaciones_db_2.py
# Reemplaza esta clase por completo

class DialogoGestionarOfertaLote(simpledialog.Dialog):
    def __init__(self, parent, title, lotes_disponibles, initial_data=None):
        self.lotes_disponibles = lotes_disponibles
        self.initial_data = initial_data or {}
        super().__init__(parent, title)

    def body(self, master):
        # ... (código para Lote y Monto sin cambios)
        ttk.Label(master, text="Seleccionar Lote:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        self.lote_var = tk.StringVar()
        lote_nombres = [f"{lote.numero} - {lote.nombre}" for lote in self.lotes_disponibles]
        self.lote_combo = ttk.Combobox(master, textvariable=self.lote_var, values=lote_nombres, state="readonly", width=38)
        self.lote_combo.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(master, text="Monto Ofertado:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        self.monto_entry = ttk.Entry(master, width=40)
        self.monto_entry.grid(row=1, column=1, padx=5, pady=5)

        ttk.Label(master, text="Plazo de Entrega (días):").grid(row=2, column=0, sticky=tk.W, padx=5, pady=5)
        self.plazo_entry = ttk.Entry(master, width=40)
        self.plazo_entry.grid(row=2, column=1, padx=5, pady=5)

        # --- INICIO CAMPO NUEVO ---
        ttk.Label(master, text="Garantía (meses):").grid(row=3, column=0, sticky=tk.W, padx=5, pady=5)
        self.garantia_entry = ttk.Entry(master, width=40)
        self.garantia_entry.grid(row=3, column=1, padx=5, pady=5)
        # --- FIN CAMPO NUEVO ---

        self.paso_fase_A_var = tk.BooleanVar()
        ttk.Checkbutton(master, text="Oferta habilitada (Pasó Fase A)", variable=self.paso_fase_A_var).grid(row=4, column=0, columnspan=2, sticky=tk.W, padx=5, pady=5)

        if self.initial_data:
            # ... (código para llenar los campos existentes sin cambios)
            self.plazo_entry.insert(0, self.initial_data.get('plazo_entrega', 0))
            self.garantia_entry.insert(0, self.initial_data.get('garantia_meses', 0)) # <-- LÍNEA NUEVA
            self.paso_fase_A_var.set(self.initial_data.get('paso_fase_A', True))
        else:
            # ...
            self.plazo_entry.insert(0, '0')
            self.garantia_entry.insert(0, '0') # <-- LÍNEA NUEVA
        return self.monto_entry

    def apply(self):
        try:
            monto = float(self.monto_entry.get() or 0)
            plazo = int(self.plazo_entry.get() or 0)
            garantia = int(self.garantia_entry.get() or 0) # <-- LÍNEA NUEVA
            lote_seleccionado_str = self.lote_var.get()

            if not lote_seleccionado_str: 
                self.result = None; return

            lote_numero = lote_seleccionado_str.split(" - ", 1)[0]
            self.result = {
                "lote_numero": lote_numero, "monto": monto,
                "paso_fase_A": self.paso_fase_A_var.get(),
                "plazo_entrega": plazo,
                "garantia_meses": garantia # <-- LÍNEA NUEVA
            }
        except (ValueError, TypeError):
            messagebox.showerror("Error de Datos", "Monto debe ser un número. Plazo y Garantía deben ser números enteros.", parent=self)
            self.result = None


class DialogoGestionarEntidad(simpledialog.Dialog):
    def __init__(self, parent, title, entity_type, initial_data=None):
        self.entity_type = entity_type
        self.initial_data = initial_data or {}
        super().__init__(parent, title)

    def body(self, master):
        self.entries = {}
        
        # Define los campos según el tipo de entidad
        if self.entity_type == 'competidor':
            fields = [("Nombre", "nombre"), ("RNC", "rnc"), 
                      ("No. RPE", "rpe"), ("Representante", "representante")]
        else: # Para 'empresa' o 'institucion'
            fields = [("Nombre", "nombre"), ("RNC", "rnc"), ("Teléfono", "telefono"), 
                      ("Correo", "correo"), ("Dirección", "direccion")]

        for i, (label_text, key) in enumerate(fields):
            ttk.Label(master, text=f"{label_text}:").grid(row=i, column=0, sticky=tk.W, padx=5, pady=3)
            entry = ttk.Entry(master, width=40)
            entry.grid(row=i, column=1, padx=5, pady=3)
            entry.insert(0, self.initial_data.get(key) or "")
            self.entries[key] = entry
        
        return self.entries["nombre"] # Foco inicial

    def apply(self):
        self.result = {key: entry.get().strip() for key, entry in self.entries.items()}
        if not self.result["nombre"]:
            messagebox.showerror("Error", "El nombre no puede estar vacío.", parent=self)
            self.result = None

class DialogoGestionarRiesgo(simpledialog.Dialog):
    """Un formulario para crear o editar un riesgo con todos sus detalles."""
    def __init__(self, parent, title="Gestionar Riesgo", initial_data=None):
        self.initial_data = initial_data if initial_data else {}
        self.categorias = ["Técnico", "Legal", "Financiero", "Plazos", "Reputacional", "Otro"]
        super().__init__(parent, title)

    def body(self, master):
        # --- Descripción ---
        ttk.Label(master, text="Descripción del Riesgo:").grid(row=0, column=0, columnspan=2, sticky=tk.W, padx=5, pady=2)
        self.desc_entry = ttk.Entry(master, width=50)
        self.desc_entry.grid(row=1, column=0, columnspan=2, sticky=tk.EW, padx=5, pady=2)
        self.desc_entry.insert(0, self.initial_data.get('descripcion', ''))

        # --- Categoría ---
        ttk.Label(master, text="Categoría:").grid(row=2, column=0, sticky=tk.W, padx=5, pady=5)
        self.categoria_var = tk.StringVar(value=self.initial_data.get('categoria', self.categorias[0]))
        ttk.Combobox(master, textvariable=self.categoria_var, values=self.categorias, state="readonly").grid(row=2, column=1, sticky=tk.EW, padx=5, pady=5)

        # --- Impacto y Probabilidad ---
        self.impacto_var = tk.IntVar(value=self.initial_data.get('impacto', 1))
        self.prob_var = tk.IntVar(value=self.initial_data.get('probabilidad', 1))

        ttk.Label(master, text="Impacto (1=Bajo, 5=Muy Alto):").grid(row=3, column=0, sticky=tk.W, padx=5, pady=2)
        ttk.Scale(master, from_=1, to=5, variable=self.impacto_var, orient=tk.HORIZONTAL).grid(row=3, column=1, sticky=tk.EW, padx=5, pady=2)

        ttk.Label(master, text="Probabilidad (1=Baja, 5=Muy Alta):").grid(row=4, column=0, sticky=tk.W, padx=5, pady=2)
        ttk.Scale(master, from_=1, to=5, variable=self.prob_var, orient=tk.HORIZONTAL).grid(row=4, column=1, sticky=tk.EW, padx=5, pady=2)

        # --- Mitigación ---
        ttk.Label(master, text="Plan de Mitigación:").grid(row=5, column=0, columnspan=2, sticky=tk.W, padx=5, pady=2)
        self.mitigacion_text = tk.Text(master, width=50, height=4)
        self.mitigacion_text.grid(row=6, column=0, columnspan=2, sticky=tk.EW, padx=5, pady=2)
        self.mitigacion_text.insert("1.0", self.initial_data.get('mitigacion', ''))
        
        return self.desc_entry

    def apply(self):
        descripcion = self.desc_entry.get().strip()
        if not descripcion:
            messagebox.showwarning("Dato Requerido", "La descripción del riesgo no puede estar vacía.", parent=self)
            self.result = None
            return

        self.result = {
            "descripcion": descripcion,
            "categoria": self.categoria_var.get(),
            "impacto": self.impacto_var.get(),
            "probabilidad": self.prob_var.get(),
            "mitigacion": self.mitigacion_text.get("1.0", tk.END).strip()
        }




# REEMPLAZA esta clase por completo

class DialogoGestionarLote(simpledialog.Dialog):
    def __init__(self, parent, title="Gestionar Lote", initial_data=None, participating_companies=None):
        self.initial_data = initial_data
        self.participating_companies = participating_companies or []
        super().__init__(parent, title)

# En la clase DialogoGestionarLote, REEMPLAZA este método:
    def body(self, master):
        master.columnconfigure(1, weight=1)
        
        # --- Campos existentes (sin cambios) ---
        ttk.Label(master, text="Número de Lote:").grid(row=0, sticky=tk.W, padx=5, pady=3)
        self.numero_entry = ttk.Entry(master, width=40)
        self.numero_entry.grid(row=0, column=1, sticky=tk.EW)
        
        ttk.Label(master, text="Nombre del Lote:").grid(row=1, sticky=tk.W, padx=5, pady=3)
        self.nombre_entry = ttk.Entry(master, width=40)
        self.nombre_entry.grid(row=1, column=1, sticky=tk.EW)
        
        ttk.Label(master, text="Monto Base (Licitación):").grid(row=2, sticky=tk.W, padx=5, pady=3)
        self.monto_base_entry = ttk.Entry(master, width=40)
        self.monto_base_entry.grid(row=2, column=1, sticky=tk.EW)
        
        ttk.Label(master, text="Monto Base (Presupuesto Personal):").grid(row=3, sticky=tk.W, padx=5, pady=3)
        self.monto_personal_entry = ttk.Entry(master, width=40)
        self.monto_personal_entry.grid(row=3, column=1, sticky=tk.EW)
        
        ttk.Label(master, text="Nuestra Oferta para el Lote:").grid(row=4, sticky=tk.W, padx=5, pady=3)
        self.monto_ofertado_entry = ttk.Entry(master, width=40)
        self.monto_ofertado_entry.grid(row=4, column=1, sticky=tk.EW)

        # --- INICIO DE LA MODIFICACIÓN ---
        ttk.Label(master, text="Asignar a Empresa:").grid(row=5, sticky=tk.W, padx=5, pady=3)
        self.empresa_var = tk.StringVar()
        
        # Siempre añadimos una opción para no asignar empresa
        opciones_empresa = ["(Sin Asignar)"] + (self.participating_companies or [])
        
        self.empresa_combo = ttk.Combobox(master, textvariable=self.empresa_var, values=opciones_empresa, state="readonly")
        self.empresa_combo.grid(row=5, column=1, sticky=tk.EW)
        
        # Eliminamos el bloqueo condicional. El combobox siempre estará activo.
        if self.initial_data and getattr(self.initial_data, 'empresa_nuestra', None):
            self.empresa_combo.set(self.initial_data.empresa_nuestra)
        else:
            # Por defecto, seleccionamos "Sin Asignar"
            self.empresa_combo.set("(Sin Asignar)")
        # --- FIN DE LA MODIFICACIÓN ---

        # Poblar datos si estamos editando
        if self.initial_data:
            self.numero_entry.insert(0, self.initial_data.numero)
            self.nombre_entry.insert(0, self.initial_data.nombre)
            self.monto_base_entry.insert(0, self.initial_data.monto_base)
            self.monto_personal_entry.insert(0, getattr(self.initial_data, 'monto_base_personal', 0.0))
            self.monto_ofertado_entry.insert(0, self.initial_data.monto_ofertado)
            
        return self.numero_entry

    # Ahora, en la misma clase DialogoGestionarLote, REEMPLAZA el método apply:
    def apply(self):
        try:
            # Leemos el valor seleccionado
            empresa_seleccionada = self.empresa_var.get()
            # Si es "(Sin Asignar)" o está vacío, lo guardamos como None
            empresa_final = None if empresa_seleccionada == "(Sin Asignar)" or not empresa_seleccionada else empresa_seleccionada
            
            # Creamos el objeto Lote con el valor correcto
            self.result = Lote(
                numero=self.numero_entry.get(),
                nombre=self.nombre_entry.get(),
                monto_base=float(self.monto_base_entry.get() or 0),
                monto_base_personal=float(self.monto_personal_entry.get() or 0),
                monto_ofertado=float(self.monto_ofertado_entry.get() or 0),
                empresa_nuestra=empresa_final # Guardar la empresa o None
            )
        except (ValueError, TypeError):
            messagebox.showerror("Error de Datos", "Los montos deben ser números válidos.", parent=self)
            self.result = None

    def apply(self):
        try:
            # Creamos el objeto Lote con el nuevo campo 'empresa_nuestra'
            self.result = Lote(
                numero=self.numero_entry.get(),
                nombre=self.nombre_entry.get(),
                monto_base=float(self.monto_base_entry.get() or 0),
                monto_base_personal=float(self.monto_personal_entry.get() or 0),
                monto_ofertado=float(self.monto_ofertado_entry.get() or 0),
                empresa_nuestra=self.empresa_var.get() or None # Guardar la empresa seleccionada
            )
        except (ValueError, TypeError):
            messagebox.showerror("Error de Datos", "Los montos deben ser números válidos.", parent=self)
            self.result = None

class DialogoGestionarCriterioBNB(simpledialog.Dialog):
    """Un diálogo para agregar o editar un criterio Bid/No-Bid con todos sus campos."""
    def __init__(self, parent, title="Gestionar Criterio", initial_data=None):
        self.initial_data = initial_data or {}
        super().__init__(parent, title)

    def body(self, master):
        ttk.Label(master, text="Nombre del Criterio:").grid(row=0, sticky=tk.W, padx=5, pady=5)
        self.nombre_entry = ttk.Entry(master, width=40)
        self.nombre_entry.grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Label(master, text="Peso (ej: 0.2 para 20%):").grid(row=1, sticky=tk.W, padx=5, pady=5)
        self.peso_entry = ttk.Entry(master, width=40)
        self.peso_entry.grid(row=1, column=1, padx=5, pady=5)

        # Si estamos editando, poblamos los campos con los datos existentes
        if self.initial_data:
            self.nombre_entry.insert(0, self.initial_data.get('nombre', ''))
            self.peso_entry.insert(0, self.initial_data.get('peso', ''))
        
        return self.nombre_entry # Foco inicial

    def apply(self):
        nombre = self.nombre_entry.get().strip()
        if not nombre:
            messagebox.showerror("Dato Requerido", "El nombre del criterio no puede estar vacío.", parent=self)
            self.result = None
            return

        try:
            peso = float(self.peso_entry.get())
            if not (0 < peso <= 1.0):
                messagebox.showerror("Valor Inválido", "El peso debe ser un número mayor que 0 y menor o igual que 1.0.", parent=self)
                self.result = None
                return
        except (ValueError, TypeError):
            messagebox.showerror("Error de Formato", "El peso debe ser un número válido (ej: 0.25).", parent=self)
            self.result = None
            return
            
        self.result = {"nombre": nombre, "peso": peso}



class DialogoGestionarOferente(simpledialog.Dialog):
    def __init__(self, parent, title="Gestionar Competidor", initial_data=None):
        self.initial_data = initial_data
        super().__init__(parent, title)
    def body(self, master):
        ttk.Label(master, text="Nombre del Competidor:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=3)
        self.nombre_entry = ttk.Entry(master, width=40)
        self.nombre_entry.grid(row=0, column=1, padx=5, pady=3)
        
        ttk.Label(master, text="Comentario Adicional:").grid(row=1, column=0, columnspan=2, sticky=tk.W, padx=5, pady=3)
        self.comentario_text = tk.Text(master, width=50, height=4)
        self.comentario_text.grid(row=2, column=0, columnspan=2, padx=5, pady=3)

        if self.initial_data:
            self.nombre_entry.insert(0, self.initial_data.nombre)
            self.comentario_text.insert("1.0", self.initial_data.comentario)
        return self.nombre_entry
    def apply(self):
        nombre = self.nombre_entry.get().strip()
        if not nombre:
            messagebox.showerror("Error", "El nombre del competidor no puede estar vacío.", parent=self)
            self.result = None; return
        self.result = {"nombre": nombre, "comentario": self.comentario_text.get("1.0", tk.END).strip()}

class DialogoSeleccionarEmpresa(simpledialog.Dialog):
    def __init__(self, parent, title, lista_empresas_nombres):
        self.lista_empresas = lista_empresas_nombres
        self.result = None
        super().__init__(parent, title)

    def body(self, master):
        ttk.Label(master, text="Selecciona la empresa para la licitación duplicada:").pack(padx=10, pady=5)
        self.listbox = tk.Listbox(master, width=50, height=10)
        self.listbox.pack(padx=10, pady=10)
        
        for empresa_nombre in sorted(self.lista_empresas):
            self.listbox.insert(tk.END, empresa_nombre)
            
        return self.listbox

    def apply(self):
        try:
            self.result = self.listbox.get(self.listbox.curselection()[0])
        except IndexError:
            self.result = None


# En gestor_licitaciones_db_2.py
# Pega esta nueva clase de datos

class Riesgo:
    """Representa un riesgo individual asociado a una licitación."""
    def __init__(self, **kwargs):
        self.id = kwargs.get('id', None)
        self.licitacion_id = kwargs.get('licitacion_id', None)
        self.descripcion = kwargs.get('descripcion', '')
        self.categoria = kwargs.get('categoria', 'Técnico')
        self.impacto = kwargs.get('impacto', 1) # Escala de 1 a 5
        self.probabilidad = kwargs.get('probabilidad', 1) # Escala de 1 a 5
        self.mitigacion = kwargs.get('mitigacion', '')

    def to_dict(self):
        return {
            'id': self.id,
            'licitacion_id': self.licitacion_id,
            'descripcion': self.descripcion,
            'categoria': self.categoria,
            'impacto': self.impacto,
            'probabilidad': self.probabilidad,
            'mitigacion': self.mitigacion
        }

from collections import defaultdict
import numpy as np

class VentanaDashboardGlobal(tk.Toplevel):

    def __init__(self, parent, all_bids):
        super().__init__(parent)
        self.parent_app = parent
        self.all_bids = all_bids
                # --- INICIO DE LÍNEAS NUEVAS ---
        self.search_competidor_var = tk.StringVar()
        self.datos_competidores_completos = []
        self.tree_competidores = None
        # --- FIN DE LÍNEAS NUEVAS ---

        self.title("Dashboard Global de Licitaciones")
        self.geometry("1400x900")
        self.grab_set()

        if not MATPLOTLIB_AVAILABLE:
            ttk.Label(self, text="Error: La librería 'matplotlib' no está instalada.", font=("Helvetica", 12)).pack(pady=50)
            return

        main_frame = ttk.Frame(self, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)

        self._crear_estilos_modernos()
        self.crear_widgets_filtros(main_frame)

        self.notebook = ttk.Notebook(main_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True, pady=10)

        self.tab_resumen = ScrollableFrame(self.notebook)
        self.tab_competencia = ttk.Frame(self.notebook, padding="10")
        self.tab_fallas_a = ttk.Frame(self.notebook, padding="10")

        self.notebook.add(self.tab_resumen, text="📊 Resumen General")
        self.notebook.add(self.tab_competencia, text="🤺 Análisis de Competencia")
        self.notebook.add(self.tab_fallas_a, text="🔍 Fallas Fase A")

        self._aplicar_filtros()

    def _nuestras_empresas_de(self, lic):
        empresas = set()
        for lote in getattr(lic, "lotes", []):
            nombre_empresa_lote = getattr(lote, "empresa_nuestra", None)
            if nombre_empresa_lote and isinstance(nombre_empresa_lote, str) and nombre_empresa_lote.strip():
                empresas.add(nombre_empresa_lote.strip())
        
        if not empresas:
            for item in getattr(lic, "empresas_nuestras", []):
                nombre = None
                if hasattr(item, 'nombre'):
                    nombre = item.nombre
                elif isinstance(item, dict) and 'nombre' in item:
                    nombre = item['nombre']
                
                if nombre and isinstance(nombre, str) and nombre.strip():
                    empresas.add(nombre.strip())
        return empresas


# Pega este bloque completo DENTRO de la clase VentanaDashboardGlobal

# DENTRO DE LA CLASE VentanaDashboardGlobal:
# REEMPLAZA los métodos de análisis de fallas con este nuevo bloque completo.

# En la clase VentanaDashboardGlobal, REEMPLAZA este método:

    def generar_analisis_fallas(self, bids):
        """
        Puebla la pestaña de 'Fallas Fase A' con un diseño de dos columnas:
        - Izquierda: Tabla de impacto y tabla de detalle.
        - Derecha: Gráfico Top 10 de toda la altura.
        """
        for widget in self.tab_fallas_a.winfo_children():
            widget.destroy()

        self.datos_fallas_completos = self.parent_app.db.obtener_todas_las_fallas()
        if not self.datos_fallas_completos:
            ttk.Label(self.tab_fallas_a, text="No hay datos de fallas en Fase A para analizar.").pack(pady=50)
            return
            
        instituciones = sorted(list(set(f[0] for f in self.datos_fallas_completos)))

        filtro_frame = ttk.Frame(self.tab_fallas_a)
        filtro_frame.pack(fill=tk.X, padx=10, pady=(5,10))
        ttk.Label(filtro_frame, text="Filtrar por Institución:").pack(side=tk.LEFT, padx=(0,5))
        self.fallas_inst_var = tk.StringVar(value="Todas")
        inst_combo = ttk.Combobox(filtro_frame, textvariable=self.fallas_inst_var, values=["Todas"] + instituciones, state="readonly", width=50)
        inst_combo.pack(side=tk.LEFT, fill=tk.X, expand=True)
        inst_combo.bind("<<ComboboxSelected>>", self._actualizar_vista_fallas)

        # --- INICIO DEL REDISEÑO DEL LAYOUT ---
        
        # 1. El panel principal ahora es HORIZONTAL, para crear la división izquierda-derecha.
        main_pane = ttk.PanedWindow(self.tab_fallas_a, orient=tk.HORIZONTAL)
        main_pane.pack(fill=tk.BOTH, expand=True)

        # 2. Creamos un nuevo panel VERTICAL para la columna izquierda que contendrá las dos tablas.
        left_column_pane = ttk.PanedWindow(main_pane, orient=tk.VERTICAL)
        # Asignamos el mismo peso a ambas columnas para que inicien del mismo ancho.
        main_pane.add(left_column_pane, weight=1)

        # 3. Creamos la tabla de análisis de impacto y la añadimos a la columna izquierda.
        self.tabla_fallas_pct_frame = ttk.LabelFrame(left_column_pane, text="Análisis de Impacto por Documento", padding=5)
        self.tree_fallas_pct = ttk.Treeview(self.tabla_fallas_pct_frame, columns=('doc', 'fallas', 'pct'), show='headings')
        self.tree_fallas_pct.heading('doc', text='Documento'); self.tree_fallas_pct.heading('fallas', text='N° Fallas'); self.tree_fallas_pct.heading('pct', text='% del Total')
        self.tree_fallas_pct.column('fallas', width=60, anchor='center', stretch=False); self.tree_fallas_pct.column('pct', width=65, anchor='e', stretch=False)
        self.tree_fallas_pct.pack(fill=tk.BOTH, expand=True)
        self.tree_fallas_pct.bind("<<TreeviewSelect>>", self._actualizar_detalle_fallas_empresa)
        left_column_pane.add(self.tabla_fallas_pct_frame, weight=1) # Ocupa la mitad superior izquierda

        # 4. Creamos la tabla de detalle y la añadimos a la columna izquierda.
        self.detalle_fallas_frame = ttk.LabelFrame(left_column_pane, text="Detalle de Fallas por Empresa (Seleccione un documento)", padding=5)
        self.tree_detalle_fallas = ttk.Treeview(self.detalle_fallas_frame, columns=('empresa', 'rnc', 'institucion', 'tipo'), show='headings')
        self.tree_detalle_fallas.heading('empresa', text='Empresa'); self.tree_detalle_fallas.heading('rnc', text='RNC'); self.tree_detalle_fallas.heading('institucion', text='Institución'); self.tree_detalle_fallas.heading('tipo', text='Tipo')
        self.tree_detalle_fallas.column('rnc', width=120, anchor='center'); self.tree_detalle_fallas.column('institucion', width=250); self.tree_detalle_fallas.column('tipo', width=100, anchor='center')
        self.tree_detalle_fallas.pack(fill=tk.BOTH, expand=True)
        left_column_pane.add(self.detalle_fallas_frame, weight=1) # Ocupa la mitad inferior izquierda

        # 5. Creamos el panel del gráfico y lo añadimos directamente al panel principal (columna derecha).
        self.grafico_fallas_frame = ttk.LabelFrame(main_pane, text="Top 10 Documentos con Más Fallas", padding=5)
        self.fig_fallas = Figure(figsize=(5, 4), dpi=100, facecolor="white")
        self.ax_fallas = self.fig_fallas.add_subplot(111)
        self.canvas_fallas = FigureCanvasTkAgg(self.fig_fallas, master=self.grafico_fallas_frame)
        self.canvas_fallas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        main_pane.add(self.grafico_fallas_frame, weight=1)
        
        # --- FIN DEL REDISEÑO DEL LAYOUT ---
        
        self._actualizar_vista_fallas()

    def _actualizar_vista_fallas(self, event=None):
        from collections import Counter
        institucion_sel = self.fallas_inst_var.get()
        
        if institucion_sel == "Todas":
            datos_filtrados = self.datos_fallas_completos
        else:
            datos_filtrados = [f for f in self.datos_fallas_completos if f[0] == institucion_sel]
        
        contador_fallas = Counter(item[2] for item in datos_filtrados)
        
        # CORRECCIÓN: Calculamos el total de fallas para el nuevo porcentaje
        total_fallas_filtradas = len(datos_filtrados)
            
        self._poblar_tabla_y_grafico_fallas(contador_fallas, total_fallas_filtradas)
        
        self.tree_detalle_fallas.delete(*self.tree_detalle_fallas.get_children())
        self.detalle_fallas_frame.config(text="Detalle de Fallas por Empresa (Seleccione un documento en la tabla de arriba)")

    def _poblar_tabla_y_grafico_fallas(self, contador_fallas, total_fallas_filtradas):
        import matplotlib.pyplot as plt
        self.tree_fallas_pct.delete(*self.tree_fallas_pct.get_children())
        self.ax_fallas.clear()
        
        if not contador_fallas:
            self.ax_fallas.text(0.5, 0.5, "Sin datos para esta selección", ha='center', va='center', fontsize=12)
            self.ax_fallas.set_xticks([]); self.ax_fallas.set_yticks([])
            self.canvas_fallas.draw()
            return

        datos_tabla = []
        for doc_nombre, num_fallas in contador_fallas.items():
            # CORRECCIÓN: Nuevo cálculo del porcentaje de impacto
            porcentaje = (num_fallas / total_fallas_filtradas * 100) if total_fallas_filtradas > 0 else 0
            datos_tabla.append((doc_nombre, num_fallas, porcentaje))
        
        for doc, fallas, pct in sorted(datos_tabla, key=lambda x: x[1], reverse=True):
            self.tree_fallas_pct.insert('', 'end', values=(doc, fallas, f"{pct:.1f}%"), iid=doc)
            
        top_items = contador_fallas.most_common(10)
        labels = [item[0] for item in top_items][::-1]
        counts = [item[1] for item in top_items][::-1]
        colors = plt.get_cmap('viridis', len(counts))(range(len(counts)))
        bars = self.ax_fallas.barh(labels, counts, color=colors)
        self.ax_fallas.bar_label(bars, padding=3, fontsize=8, color='black', fmt='%d')
        self.ax_fallas.set_xlabel("Cantidad de Fallas Registradas", fontsize=9)
        self.ax_fallas.tick_params(axis='y', labelsize=8)
        self.ax_fallas.set_yticklabels([])
        if counts:
            self.ax_fallas.set_xlim(right=max(counts) * 1.15)
        self.ax_fallas.legend(handles=bars, labels=labels, title="Documentos", fontsize=8, loc='upper center', bbox_to_anchor=(0.5, -0.15), ncol=2)
        self.fig_fallas.subplots_adjust(bottom=0.3)
        self.canvas_fallas.draw()

    def _actualizar_detalle_fallas_empresa(self, event=None):
        """Rediseño completo para mostrar RNC, Institución y ordenar por tipo."""
        self.tree_detalle_fallas.delete(*self.tree_detalle_fallas.get_children())
        
        if not self.tree_fallas_pct.selection(): return
            
        doc_seleccionado = self.tree_fallas_pct.selection()[0]
        institucion_sel = self.fallas_inst_var.get()
        
        self.detalle_fallas_frame.config(text=f"Empresas que fallaron en: '{doc_seleccionado}'")

        # 1. Creamos un mapa de RNCs para una búsqueda rápida
        rnc_map = {e['nombre']: e.get('rnc', 'N/D') for e in self.parent_app.empresas_registradas}
        rnc_map.update({c['nombre']: c.get('rnc', 'N/D') for c in self.parent_app.competidores_maestros})

        # 2. Filtramos los datos y separamos por tipo (nuestra vs competidor)
        nuestras_fallaron = []
        competidores_fallaron = []
        
        for inst, participante, doc_nombre, es_nuestro, _ in self.datos_fallas_completos:
            if doc_nombre == doc_seleccionado and (institucion_sel == "Todas" or inst == institucion_sel):
                rnc = rnc_map.get(participante, 'N/D')
                if es_nuestro:
                    nuestras_fallaron.append((participante, rnc, inst))
                else:
                    competidores_fallaron.append((participante, rnc, inst))
        
        # 3. Insertamos en la tabla, primero las nuestras y luego los competidores
        # Usamos set() para eliminar duplicados y luego sorted() para ordenar alfabéticamente
        if nuestras_fallaron:
            for empresa, rnc, inst in sorted(list(set(nuestras_fallaron))):
                self.tree_detalle_fallas.insert('', 'end', values=(empresa, rnc, inst, "Nuestra"))

        if competidores_fallaron:
            for empresa, rnc, inst in sorted(list(set(competidores_fallaron))):
                self.tree_detalle_fallas.insert('', 'end', values=(empresa, rnc, inst, "Competidor"))




    def _display_empresas_de(self, lic):
        """Devuelve un string amigable para mostrar nuestras empresas de una licitación."""
        emps = sorted(self._nuestras_empresas_de(lic))
        return ", ".join(emps) if emps else "(sin empresa)"

    def _es_ganada_por_nosotros(self, lic):
        """Una licitación finalizada 'Adjudicada' es GANADA si algún lote tiene ganado_por_nosotros=True."""
        if getattr(lic, "estado", "") != "Adjudicada":
            return False
        try:
            return any(getattr(l, "ganado_por_nosotros", False) for l in lic.lotes)
        except Exception:
            return False

    def _contar_ganadas_perdidas(self, licitaciones_filtradas):
        """Cuenta ganadas y perdidas usando los lotes."""
        ganadas = 0
        perdidas = 0
        estados_perdida_directa = {"Descalificado Fase A", "Descalificado Fase B", "Desierta", "Cancelada"}

        for lic in licitaciones_filtradas:
            estado = getattr(lic, "estado", "")
            if estado == "Adjudicada":
                if self._es_ganada_por_nosotros(lic):
                    ganadas += 1
                else:
                    perdidas += 1
            elif estado in estados_perdida_directa:
                perdidas += 1
        return ganadas, perdidas


    def _filtrar_y_repoblar_competidores(self, *args):
        """Filtra la lista completa de competidores y actualiza la tabla."""
        termino = self.search_competidor_var.get().lower()
        
        if not termino:
            datos_filtrados = self.datos_competidores_completos
        else:
            datos_filtrados = [
                comp for comp in self.datos_competidores_completos
                if termino in comp.get('nombre', '').lower() or \
                termino in comp.get('rnc', '').lower()
            ]
        
        self._poblar_tabla_competidores(datos_filtrados)

    def _poblar_tabla_competidores(self, datos):
        """Limpia y rellena la tabla de competidores con los datos proporcionados."""
        if self.tree_competidores:
            self.tree_competidores.delete(*self.tree_competidores.get_children())
            for item in datos:
                self.tree_competidores.insert(
                    "",
                    tk.END,
                    values=(
                        item['nombre'],
                        item['rnc'],
                        item['participaciones'],
                        f"{item['pct_promedio']:.2f}%"
                    )
                )



    def _analizar_competidores_pct(self, bids):
        """
        Calcula, por competidor, el PORCENTAJE DE DIFERENCIA PROMEDIO y ahora
        también incluye el RNC desde el catálogo maestro.
        """
        stats = {}  # nombre -> {'sum_pct': float, 'count': int}
        
        # Creamos un mapa de Nombre -> RNC para una búsqueda rápida
        rnc_map = {comp.get('nombre', ''): comp.get('rnc', '') for comp in self.parent_app.competidores_maestros}

        for lic in bids:
            base_por_lote = {}
            for lote in getattr(lic, 'lotes', []):
                base = getattr(lote, 'monto_base_personal', 0) or getattr(lote, 'monto_base', 0) or 0
                base_por_lote[str(getattr(lote, 'numero', ''))] = float(base) if base else 0.0

            for comp in getattr(lic, 'oferentes_participantes', []):
                nombre = getattr(comp, 'nombre', '').strip() or '—'
                for o in getattr(comp, 'ofertas_por_lote', []):
                    lote_num = str(o.get('lote_numero'))
                    oferta = float(o.get('monto', 0) or 0)
                    base = float(base_por_lote.get(lote_num, 0) or 0)
                    if base > 0 and oferta > 0:
                        pct = (oferta - base) / base * 100.0
                        if nombre not in stats:
                            stats[nombre] = {'sum_pct': 0.0, 'count': 0}
                        stats[nombre]['sum_pct'] += pct
                        stats[nombre]['count'] += 1

        salida = []
        for nombre, agg in stats.items():
            count = agg['count']
            pct_prom = (agg['sum_pct'] / count) if count else 0.0
            salida.append({
                'nombre': nombre,
                'rnc': rnc_map.get(nombre, ''), # <-- Añadimos el RNC aquí
                'participaciones': count,
                'pct_promedio': pct_prom
            })
        salida.sort(key=lambda x: (-x['participaciones'], x['pct_promedio']))
        return salida

    def _aplicar_filtros(self):
        # --- CAMBIO IMPORTANTE ---
        # Ahora limpiamos el contenido DENTRO del scroll, no el scroll en sí
        for widget in self.tab_resumen.scrollable_frame.winfo_children():
            widget.destroy()

        # Las otras pestañas se limpian como antes
        for widget in self.tab_competencia.winfo_children():
            widget.destroy()

        # El resto de la función sigue igual...
        inst_filter = self.inst_filter_var.get()
        code_filter = self.code_filter_var.get().lower()
        start_date = self.start_date_entry.get_date() if self.start_date_entry.get() else None
        end_date = self.end_date_entry.get_date() if self.end_date_entry.get() else None

        filtered_bids = self.all_bids[:]
        if inst_filter != "Todas":
            filtered_bids = [b for b in filtered_bids if b.institucion == inst_filter]
        if code_filter:
            filtered_bids = [b for b in filtered_bids if code_filter in b.numero_proceso.lower()]
        if start_date:
            filtered_bids = [b for b in filtered_bids if b.fecha_creacion >= start_date]
        if end_date:
            filtered_bids = [b for b in filtered_bids if b.fecha_creacion <= end_date]

        if not filtered_bids:
            ttk.Label(self.tab_resumen.scrollable_frame, text="No se encontraron licitaciones con los filtros aplicados.").pack(pady=50)
            ttk.Label(self.tab_competencia, text="No se encontraron datos de competencia con los filtros aplicados.").pack(pady=50)
            ttk.Label(self.tab_causas, text="No se encontraron datos con los filtros aplicados.").pack(pady=50)
            return

        self.generar_resumen_general(filtered_bids)
        self.generar_analisis_competencia(filtered_bids)
        self.generar_analisis_fallas(filtered_bids)


    # En gestor_licitaciones_db_2.py
    # Pega este nuevo método dentro de la clase VentanaDashboardGlobal

    def _calcular_tiempo_preparacion(self, bids):
        """
        Calcula el tiempo promedio en días entre la fecha de creación y la fecha
        de presentación de ofertas para las licitaciones finalizadas.
        """
        import datetime
        diferencias_dias = []
        
        estados_finalizados = ["Adjudicada (Ganada)", "Adjudicada (Perdida)", "Descalificado Fase A", "Descalificado Fase B"]

        for b in bids:
            # Solo consideramos licitaciones en las que se llegó a presentar oferta
            if b.estado in estados_finalizados:
                try:
                    fecha_presentacion_str = b.cronograma.get("Presentacion de Ofertas", {}).get("fecha_limite")
                    if fecha_presentacion_str:
                        fecha_creacion = b.fecha_creacion
                        fecha_presentacion = datetime.datetime.strptime(fecha_presentacion_str, '%Y-%m-%d').date()
                        
                        # Calculamos la diferencia y la añadimos a la lista
                        diferencia = (fecha_presentacion - fecha_creacion).days
                        if diferencia >= 0: # Solo consideramos diferencias positivas
                            diferencias_dias.append(diferencia)
                except (ValueError, TypeError):
                    # Ignoramos si las fechas no son válidas
                    continue
        
        if not diferencias_dias:
            return 0
            
        # Devolvemos el promedio de días
        return sum(diferencias_dias) / len(diferencias_dias)




    def _calcular_top_instituciones(self, bids, top_n=5):
        """Cuenta las licitaciones por institución y devuelve las N principales."""
        from collections import Counter

        # Contamos la frecuencia de cada institución en la lista de licitaciones
        contador_instituciones = Counter(b.institucion for b in bids)

        # Devolvemos las 'top_n' más comunes como una lista de tuplas (institucion, cantidad)
        return contador_instituciones.most_common(top_n)

    def crear_grafico_top_instituciones(self, parent, bids):
        """Crea y devuelve un frame con un gráfico de barras para el top de instituciones."""
        frame = ttk.LabelFrame(parent, text=f"Top 5 Instituciones (por # de Licitaciones)")

        top_data = self._calcular_top_instituciones(bids, top_n=5)
        if not top_data:
            ttk.Label(frame, text="No hay suficientes datos.").pack(pady=20)
            return frame

        # Preparamos los datos para el gráfico
        # El [::-1] es para que la barra más grande quede arriba en el gráfico horizontal
        instituciones = [item[0] for item in top_data][::-1]
        cantidades = [item[1] for item in top_data][::-1]

        fig = Figure(figsize=(4, 3), dpi=100)
        ax = fig.add_subplot(111)

        # Creamos un gráfico de barras horizontales
        bars = ax.barh(instituciones, cantidades, color='#6666c2')
        ax.set_xlabel("Cantidad de Licitaciones")

        # Añadimos etiquetas con los números al final de cada barra
        ax.bar_label(bars, padding=3)

        # Ajustamos el layout para que no se corten los nombres
        fig.tight_layout()

        canvas = FigureCanvasTkAgg(fig, master=frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        return frame

# Pega este NUEVO método dentro de la clase VentanaDashboardGlobal

    def _crear_panel_card(self, parent, titulo):
        """
        Crea un panel personalizado con borde y título que es compatible con todos los temas.
        Devuelve el frame principal (la tarjeta) y un frame interior para el contenido.
        """
        import tkinter as tk # Nos aseguramos que tk esté importado
        
        # 1. Frame exterior que actúa como el borde de la tarjeta.
        card_frame = tk.Frame(parent, relief="solid", borderwidth=1, background="white")
        
        # 2. Frame interior con padding para el contenido.
        inner_frame = tk.Frame(card_frame, background="white", padx=10, pady=10)
        inner_frame.pack(fill=tk.BOTH, expand=True)
        
        # 3. Etiqueta para el título en la parte superior.
        title_label = ttk.Label(inner_frame, text=titulo, 
                                font=("Segoe UI", 12, "bold"), 
                                background="white", 
                                foreground="#333333")
        title_label.pack(anchor="nw", pady=(0, 10))
        
        # 4. Devolvemos el frame principal y el frame donde irá el contenido (el gráfico).
        return card_frame, inner_frame

    def _crear_estilos_modernos(self):
        style = ttk.Style(self)
        import tkinter as tk # Aseguramos que tk esté disponible

        # --- Estilo para los indicadores KPI (este funciona bien) ---
        style.configure("KPI.TLabelframe", padding=10)
        style.configure("KPI.TLabelframe.Label",
                        font=("Segoe UI", 11, "bold"),
                        foreground="#333333")

        # --- INICIO DE LA SOLUCIÓN DEFINITIVA PARA EL ESTILO "CARD" ---
        
        # PASO 1: Copiar la ESTRUCTURA (Layout) de un LabelFrame estándar.
        # Esta es la línea crucial que el tema 'arc' necesita para evitar el error "Layout not found".
        try:
            style.layout("Card.TLabelframe", style.layout("TLabelFrame"))
        except tk.TclError:
            # Esto es una medida de seguridad por si algún tema no soporta el comando.
            pass
        
        # PASO 2: Configurar las PROPIEDADES (colores, bordes) de esa estructura.
        # Usamos 'relief' porque es la forma más compatible de crear el borde.
        style.configure("Card.TLabelFrame", 
                        padding=10, 
                        background="white", 
                        borderwidth=1, 
                        relief="solid")
    
        # PASO 3: Asegurarse de que el TÍTULO del panel también tenga fondo blanco.
        style.configure("Card.TLabelFrame.Label",
                        font=("Segoe UI", 12, "bold"),
                        background="white", 
                        foreground="#333333")
        # --- FIN DE LA SOLUCIÓN DEFINITIVA ---
        
    def generar_resumen_general(self, bids):
        frame_principal = self.tab_resumen.scrollable_frame
        for widget in frame_principal.winfo_children():
            widget.destroy()

        stats_exito = self._calcular_tasa_exito(bids)
        total_lotes_ganados = sum(sum(1 for l in lic.lotes if l.ganado_por_nosotros) for lic in bids if lic.estado == "Adjudicada")

        frame_principal.rowconfigure(1, weight=1)
        frame_principal.rowconfigure(2, weight=1)
        frame_principal.columnconfigure(0, weight=3)
        frame_principal.columnconfigure(1, weight=1)

        top_panel = ttk.Frame(frame_principal)
        top_panel.grid(row=0, column=0, columnspan=2, sticky="nsew", pady=(0, 15))
        for i in range(3): top_panel.columnconfigure(i, weight=1)

        # NOTA: Los KPIs usan un estilo diferente que SÍ funciona, los dejamos como están.
        kpi_rendimiento = ttk.LabelFrame(top_panel, text="Rendimiento (Finalizadas)", style="KPI.TLabelframe")
        kpi_rendimiento.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        self.crear_kpi_exito(kpi_rendimiento, stats_exito)

        kpi_lotes = ttk.LabelFrame(top_panel, text="Lotes Ganados (Finalizadas)", style="KPI.TLabelframe")
        kpi_lotes.grid(row=0, column=1, sticky="nsew", padx=10)
        ttk.Label(kpi_lotes, text="Total de Lotes Adjudicados:", font=("Segoe UI", 10)).grid(sticky="ew")
        ttk.Label(kpi_lotes, text=f"{total_lotes_ganados}", font=("Segoe UI", 18, "bold"), foreground="#1E7D32").grid(sticky="ew")

        kpi_financiero = ttk.LabelFrame(top_panel, text="Análisis Financiero", style="KPI.TLabelframe")
        kpi_financiero.grid(row=0, column=2, sticky="nsew", padx=(10, 0))
        self.crear_kpis_financieros(kpi_financiero, bids)

        # Aquí usamos nuestro nuevo constructor para los gráficos principales
        frame_rendimiento_empresa = self.crear_grafico_rendimiento_por_empresa(frame_principal, bids)
        frame_rendimiento_empresa.grid(row=1, column=0, sticky="nsew", pady=(0, 10))

        frame_dist_estados = self.crear_grafico_distribucion_estados(frame_principal, bids)
        frame_dist_estados.grid(row=2, column=0, sticky="nsew")

        frame_lic_empresa_tabla = self.crear_tabla_licitaciones_por_empresa(frame_principal, bids)
        frame_lic_empresa_tabla.grid(row=1, rowspan=2, column=1, sticky="nsew", padx=(10, 0))
        

    def _calcular_tasa_exito(self, bids):
        stats = {'total': {'ganadas': 0, 'perdidas': 0}, 'por_empresa': {}}
        estados_perdida_directa = {"Descalificado Fase A", "Descalificado Fase B", "Desierta", "Cancelada"}
        for lic in bids:
            empresas_participantes = self._nuestras_empresas_de(lic) or {"(sin empresa)"}
            es_ganada = lic.estado == "Adjudicada" and any(l.ganado_por_nosotros for l in lic.lotes)
            es_perdida = (lic.estado == "Adjudicada" and not es_ganada) or lic.estado in estados_perdida_directa
            
            if es_ganada:
                stats['total']['ganadas'] += 1
                for empresa in empresas_participantes:
                    stats['por_empresa'].setdefault(empresa, {'ganadas': 0, 'perdidas': 0})['ganadas'] += 1
            elif es_perdida:
                stats['total']['perdidas'] += 1
                for empresa in empresas_participantes:
                    stats['por_empresa'].setdefault(empresa, {'ganadas': 0, 'perdidas': 0})['perdidas'] += 1
        
        tot_fin = stats['total']['ganadas'] + stats['total']['perdidas']
        stats['total']['tasa_exito'] = (stats['total']['ganadas'] / tot_fin * 100.0) if tot_fin > 0 else 0.0
        if "(sin empresa)" in stats['por_empresa'] and (stats['por_empresa']['(sin empresa)']['ganadas'] + stats['por_empresa']['(sin empresa)']['perdidas']) == 0:
            del stats['por_empresa']['(sin empresa)']
        for data in stats['por_empresa'].values():
            fin_emp = data['ganadas'] + data['perdidas']
            data['tasa_exito'] = (data['ganadas'] / fin_emp * 100.0) if fin_emp > 0 else 0.0
        return stats

    def _calcular_montos_adjudicados(self, bids):
        total_adjudicado = 0
        for lic in bids:
            if lic.estado == "Adjudicada":
                for lote in lic.lotes:
                    if lote.ganado_por_nosotros:
                        total_adjudicado += lote.monto_ofertado
        return total_adjudicado



    def _calcular_stats_precios(self, bids):
        diferencias = {'ganadas': [], 'perdidas': []}
        estados_ganada = ["Adjudicada (Ganada)"]
        estados_perdida = ["Adjudicada (Perdida)", "Descalificado Fase A", "Descalificado Fase B"]
        for lic in bids:
            if lic.get_oferta_total() > 0:
                dif = lic.get_diferencia_porcentual(solo_participados=True)
                if lic.estado in estados_ganada: diferencias['ganadas'].append(dif)
                elif lic.estado in estados_perdida: diferencias['perdidas'].append(dif)
        avg_ganadas = sum(diferencias['ganadas']) / len(diferencias['ganadas']) if diferencias['ganadas'] else 0
        avg_perdidas = sum(diferencias['perdidas']) / len(diferencias['perdidas']) if diferencias['perdidas'] else 0
        return {'avg_ganadas': avg_ganadas, 'avg_perdidas': avg_perdidas}

    def crear_widgets_filtros(self, parent_frame):
        filter_frame = ttk.LabelFrame(parent_frame, text="Filtros del Dashboard", padding=10)
        filter_frame.pack(fill=tk.X, pady=(0, 10))
        ttk.Label(filter_frame, text="Institución:").grid(row=0, column=0, padx=(0,5))
        instituciones = sorted(list(set(b.institucion for b in self.all_bids)))
        self.inst_filter_var = tk.StringVar(value="Todas")
        ttk.Combobox(filter_frame, textvariable=self.inst_filter_var, values=["Todas"] + instituciones, state="readonly", width=30).grid(row=0, column=1, padx=5)
        ttk.Label(filter_frame, text="Código Proceso:").grid(row=0, column=2, padx=(10,5))
        self.code_filter_var = tk.StringVar()
        ttk.Entry(filter_frame, textvariable=self.code_filter_var, width=20).grid(row=0, column=3, padx=5)
        ttk.Label(filter_frame, text="Desde:").grid(row=0, column=4, padx=(10,5))
        self.start_date_entry = DateEntry(filter_frame, width=12, locale='es_ES', date_pattern='y-mm-dd')
        self.start_date_entry.grid(row=0, column=5); self.start_date_entry.delete(0, "end")
        ttk.Label(filter_frame, text="Hasta:").grid(row=0, column=6, padx=(10,5))
        self.end_date_entry = DateEntry(filter_frame, width=12, locale='es_ES', date_pattern='y-mm-dd')
        self.end_date_entry.grid(row=0, column=7); self.end_date_entry.delete(0, "end")
        ttk.Button(filter_frame, text="🔍 Aplicar Filtros", command=self._aplicar_filtros).grid(row=0, column=8, padx=10)
        ttk.Button(filter_frame, text="🧹 Limpiar", command=self._limpiar_filtros).grid(row=0, column=9)

    def _limpiar_filtros(self):
        self.inst_filter_var.set("Todas")
        self.code_filter_var.set("")
        self.start_date_entry.delete(0, "end")
        self.end_date_entry.delete(0, "end")
        self._aplicar_filtros()


# En la clase VentanaDashboardGlobal, REEMPLAZA este método:

    def generar_analisis_competencia(self, bids):
        # Limpiar la pestaña por si se reaplican los filtros
        for w in self.tab_competencia.winfo_children():
            w.destroy()

        # 1. Calcular todos los datos y guardarlos
        self.datos_competidores_completos = self._analizar_competidores_pct(bids)

        if not self.datos_competidores_completos:
            ttk.Label(self.tab_competencia,
                    text="No hay datos de ofertas de competidores en las licitaciones filtradas."
                    ).pack(pady=50)
            return

        # 2. Crear el Frame para el buscador
        search_frame = ttk.Frame(self.tab_competencia, padding=(0, 0, 0, 10))
        search_frame.pack(fill=tk.X)
        ttk.Label(search_frame, text="🔍 Buscar Competidor (por Nombre o RNC):").pack(side=tk.LEFT, padx=(0, 5))
        search_entry = ttk.Entry(search_frame, textvariable=self.search_competidor_var)
        search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.search_competidor_var.trace_add("write", self._filtrar_y_repoblar_competidores)

        # 3. Crear la tabla (Treeview) con la nueva columna RNC
        self.tree_competidores = ttk.Treeview(
            self.tab_competencia,
            columns=("nombre", "rnc", "participaciones", "pct_promedio"),
            show="headings"
        )
        self.tree_competidores.heading("nombre", text="Nombre del Competidor")
        self.tree_competidores.heading("rnc", text="RNC")
        self.tree_competidores.heading("participaciones", text="# de Lotes Ofertados")
        self.tree_competidores.heading("pct_promedio", text="% Diferencia Promedio (oferta vs base)")

        self.tree_competidores.column("nombre", width=320, anchor=tk.W)
        self.tree_competidores.column("rnc", width=120, anchor=tk.W)
        self.tree_competidores.column("participaciones", width=160, anchor=tk.CENTER)
        self.tree_competidores.column("pct_promedio", width=220, anchor=tk.E)

        self.tree_competidores.pack(fill=tk.BOTH, expand=True)

        # --- LÍNEA AÑADIDA: Conectamos el evento de doble clic ---
        self.tree_competidores.bind("<Double-1>", self._abrir_perfil_competidor_desde_dashboard)
        # --- FIN DE LA LÍNEA AÑADIDA ---

        # 4. Poblar la tabla por primera vez con todos los datos
        self._poblar_tabla_competidores(self.datos_competidores_completos)

# En la clase VentanaDashboardGlobal, PEGA este nuevo método:

    def _abrir_perfil_competidor_desde_dashboard(self, event=None):
        """Maneja el evento de doble clic en la tabla de competidores."""
        # Asegurarse de que el treeview y la selección existan
        if not self.tree_competidores or not self.tree_competidores.selection():
            return

        try:
            # Obtener el item seleccionado
            item_seleccionado = self.tree_competidores.selection()[0]
            
            # Obtener los valores de la fila; el nombre del competidor es el primer valor (índice 0)
            nombre_competidor = self.tree_competidores.item(item_seleccionado, 'values')[0]

            # Abrir la ventana del perfil del competidor, pasándole los datos necesarios
            VentanaPerfilCompetidor(self, nombre_competidor, self.all_bids)
        except IndexError:
            # Esto puede ocurrir si los valores de la fila están vacíos, es una medida de seguridad
            pass


    def crear_kpi_exito(self, parent_frame, stats):
        # CORRECCIÓN: Ya no creamos un LabelFrame aquí.
        # Usamos directamente el 'parent_frame' que nos pasan.
        parent_frame.columnconfigure(tuple(range(6)), weight=1)
        
        ttk.Label(parent_frame, text="Tasa de Éxito:", font=("Helvetica", 12, "bold")).grid(row=0, column=0, sticky="e")
        ttk.Label(parent_frame, text=f"{stats['total']['tasa_exito']:.1f}%", font=("Helvetica", 16, "bold"), foreground="#007bff").grid(row=0, column=1, sticky="w")
        
        ttk.Separator(parent_frame, orient='vertical').grid(row=0, column=2, sticky='ns', padx=20)
        
        ttk.Label(parent_frame, text="Ganadas:", font=("Helvetica", 10)).grid(row=0, column=3, sticky="e")
        ttk.Label(parent_frame, text=f"{stats['total']['ganadas']}", font=("Helvetica", 12, "bold"), foreground="green").grid(row=0, column=4, sticky="w", padx=5)
        
        ttk.Label(parent_frame, text="Perdidas:", font=("Helvetica", 10)).grid(row=0, column=5, sticky="e")
        ttk.Label(parent_frame, text=f"{stats['total']['perdidas']}", font=("Helvetica", 12, "bold"), foreground="red").grid(row=0, column=6, sticky="w", padx=5)

    def crear_kpis_financieros(self, parent_frame, bids):
        total_monto_base = sum(b.get_monto_base_total(solo_participados=True) for b in bids)
        total_monto_ofertado = sum(b.get_oferta_total(solo_participados=True) for b in bids)
        total_adjudicado = self._calcular_montos_adjudicados(bids)
        parent_frame.columnconfigure(tuple(range(2)), weight=1)
        
        ttk.Label(parent_frame, text="Monto Base Total:", font=("Helvetica", 10, "bold")).grid(row=0, column=0, sticky=tk.W)
        ttk.Label(parent_frame, text=f"RD$ {total_monto_base:,.2f}").grid(row=0, column=1, sticky=tk.E, padx=10)
        
        ttk.Label(parent_frame, text="Monto Ofertado Total:", font=("Helvetica", 10, "bold")).grid(row=1, column=0, sticky=tk.W)
        ttk.Label(parent_frame, text=f"RD$ {total_monto_ofertado:,.2f}").grid(row=1, column=1, sticky=tk.E, padx=10)

        ttk.Label(parent_frame, text="Monto Adjudicado (Nosotros):", font=("Helvetica", 10, "bold")).grid(row=2, column=0, sticky=tk.W)
        ttk.Label(parent_frame, text=f"RD$ {total_adjudicado:,.2f}", font=("Helvetica", 10, "bold"), foreground="green").grid(row=2, column=1, sticky=tk.E, padx=10)


    def crear_grafico_distribucion_estados(self, parent, bids):
        card, content = self._crear_panel_card(parent, "Distribución de Estados")

        stats = {"Ganada": 0, "Perdida": 0, "En Proceso": 0}
        for lic in bids:
            if lic.estado == "Adjudicada":
                stats["Ganada" if any(l.ganado_por_nosotros for l in lic.lotes) else "Perdida"] += 1
            elif lic.estado in ["Descalificado Fase A", "Descalificado Fase B", "Desierta", "Cancelada"]:
                stats["Perdida"] += 1
            else:
                stats["En Proceso"] += 1

        if sum(stats.values()) == 0:
            ttk.Label(content, text="Sin datos para mostrar.", background="white").pack(pady=20)
            return card

        sorted_stats = sorted(stats.items(), key=lambda item: item[1], reverse=True)
        labels = [f"{k} ({v})" for k, v in sorted_stats]
        sizes = [v for k, v in sorted_stats]
        colors = ['#2E7D32' if 'Ganada' in L else '#C62828' if 'Perdida' in L else '#FFAB00' for L in labels]

        fig = Figure(figsize=(4, 2.5), dpi=100, facecolor="white")
        ax = fig.add_subplot(111)
        ax.set_facecolor("white")
        bars = ax.barh(labels, sizes, color=colors)
        ax.bar_label(bars, padding=3)
        ax.invert_yaxis()
        ax.set_xlabel("Cantidad de Licitaciones")
        fig.tight_layout()
        canvas = FigureCanvasTkAgg(fig, master=content)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        return card

    def crear_tabla_licitaciones_por_empresa(self, parent, bids):
        card, content = self._crear_panel_card(parent, "Resumen por Empresa")

        stats_por_empresa = defaultdict(lambda: {'participaciones': 0, 'ganadas': 0, 'monto_adjudicado': 0.0})
        for lic in bids:
            empresas_participantes = self._nuestras_empresas_de(lic) or ["(sin empresa)"]
            es_ganada = lic.estado == "Adjudicada" and any(l.ganado_por_nosotros for l in lic.lotes)
            for empresa in empresas_participantes:
                stats_por_empresa[empresa]['participaciones'] += 1
                if es_ganada:
                    stats_por_empresa[empresa]['ganadas'] += 1
                    for lote in lic.lotes:
                        if lote.ganado_por_nosotros and lote.empresa_nuestra == empresa:
                            stats_por_empresa[empresa]['monto_adjudicado'] += lote.monto_ofertado

        if "(sin empresa)" in stats_por_empresa and stats_por_empresa["(sin empresa)"]['participaciones'] == 0:
            del stats_por_empresa["(sin empresa)"]

        tree = ttk.Treeview(content, columns=("empresa", "participaciones", "ganadas", "monto_adjudicado"), show="headings")
        tree.heading("empresa", text="Empresa")
        tree.heading("participaciones", text="Participa")
        tree.heading("ganadas", text="Ganadas")
        tree.heading("monto_adjudicado", text="Monto Adjudicado")
        tree.column("empresa", anchor=tk.W, width=180)
        tree.column("participaciones", anchor=tk.CENTER, width=60)
        tree.column("ganadas", anchor=tk.CENTER, width=60)
        tree.column("monto_adjudicado", anchor=tk.E, width=120)

        sorted_stats = sorted(stats_por_empresa.items(), key=lambda item: item[1]['participaciones'], reverse=True)
        for emp, data in sorted_stats:
            monto_str = f"RD$ {data['monto_adjudicado']:,.2f}"
            tree.insert("", tk.END, values=(emp, data['participaciones'], data['ganadas'], monto_str))

        tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        return card

    def crear_grafico_rendimiento_por_empresa(self, parent, bids):
        card, content = self._crear_panel_card(parent, "Rendimiento por Empresa")

        stats_por_empresa = defaultdict(lambda: {'participaciones': 0, 'ganadas': 0})
        for lic in bids:
            empresas_participantes = self._nuestras_empresas_de(lic)
            if not empresas_participantes: continue
            es_ganada = lic.estado == "Adjudicada" and any(l.ganado_por_nosotros for l in lic.lotes)
            for empresa in empresas_participantes:
                stats_por_empresa[empresa]['participaciones'] += 1
                if es_ganada: stats_por_empresa[empresa]['ganadas'] += 1

        sorted_data = sorted(stats_por_empresa.items(), key=lambda item: item[1]['participaciones'], reverse=True)
        labels = [item[0] for item in sorted_data]
        participaciones = [item[1]['participaciones'] for item in sorted_data]
        ganadas = [item[1]['ganadas'] for item in sorted_data]

        if not labels:
            ttk.Label(content, text="Sin datos para mostrar.", background="white").pack(pady=20)
            return card

        y = np.arange(len(labels))
        height = 0.35
        fig = Figure(figsize=(8, 4), dpi=100, facecolor="white")
        ax = fig.add_subplot(111)
        ax.set_facecolor("white")
        rects1 = ax.barh(y + height/2, participaciones, height, label='Participaciones', color='#42A5F5')
        rects2 = ax.barh(y - height/2, ganadas, height, label='Ganadas', color='#2E7D32')
        ax.set_ylabel('Empresa')
        ax.set_xlabel('Cantidad de Licitaciones')
        ax.set_yticks(y, labels)
        ax.invert_yaxis()
        ax.legend()
        ax.bar_label(rects1, padding=3)
        ax.bar_label(rects2, padding=3)
        fig.tight_layout()
        canvas = FigureCanvasTkAgg(fig, master=content)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        return card


    def crear_grafico_estados(self, parent, bids):
        frame = ttk.LabelFrame(parent, text="Distribución de Estados")

        # Contadores
        stats = {"Ganada": 0, "Perdida": 0, "En Proceso": 0}
        estados_perdida_directa = {"Descalificado Fase A", "Descalificado Fase B", "Desierta", "Cancelada"}

        for lic in bids:
            estado = getattr(lic, "estado", "")
            if estado == "Adjudicada":
                if any(getattr(l, "ganado_por_nosotros", False) for l in lic.lotes):
                    stats["Ganada"] += 1
                else:
                    stats["Perdida"] += 1
            elif estado in estados_perdida_directa:
                stats["Perdida"] += 1
            else:
                stats["En Proceso"] += 1

        # Preparar datos del gráfico
        labels, sizes, colors_pie = [], [], []
        colors_map = {'Ganada': '#2E7D32', 'Perdida': '#C62828', 'En Proceso': '#FFAB00'}

        for estado, cantidad in stats.items():
            if cantidad > 0:
                labels.append(f"{estado} ({cantidad})")
                sizes.append(cantidad)
                colors_pie.append(colors_map[estado])

        if not sizes:
            ttk.Label(frame, text="Sin datos para mostrar.").pack(pady=10)
            return frame

        fig = Figure(figsize=(4, 3), dpi=100)
        ax = fig.add_subplot(111)
        wedges, _, autotexts = ax.pie(
            sizes, autopct='%1.1f%%', startangle=90, colors=colors_pie,
            pctdistance=0.85, wedgeprops=dict(width=0.4, edgecolor='w')
        )
        ax.text(0, 0, f'Total\n{sum(sizes)}', ha='center', va='center', fontsize=12)
        import matplotlib.pyplot as plt
        plt.setp(autotexts, size=8, weight="bold", color="white")
        ax.legend(wedges, labels, title="Estados", loc="center left", bbox_to_anchor=(0.95, 0.5))
        fig.tight_layout(pad=1.5)

        FigureCanvasTkAgg(fig, master=frame).get_tk_widget().pack(fill=tk.BOTH, expand=True)
        return frame


    def crear_grafico_exito_empresa(self, parent, stats):
        frame = ttk.LabelFrame(parent, text="Tasa de Éxito por Empresa")
        datos_grafico = {e: d['tasa_exito'] for e, d in stats['por_empresa'].items()
                        if (d['ganadas'] + d['perdidas']) > 0}
        if not datos_grafico:
            ttk.Label(frame, text="Sin datos suficientes.").pack(pady=10)
            return frame

        sorted_data = sorted(datos_grafico.items(), key=lambda item: item[1], reverse=True)
        empresas = [item[0] for item in sorted_data]
        tasas = [item[1] for item in sorted_data]

        fig = Figure(figsize=(4, 3), dpi=100); ax = fig.add_subplot(111)
        bars = ax.barh(empresas, tasas, color='#007bff')
        ax.set_xlabel("Tasa de Éxito (%)"); ax.set_xlim(0, 100)
        ax.bar_label(bars, fmt='%.1f%%', padding=3)
        fig.tight_layout()
        FigureCanvasTkAgg(fig, master=frame).get_tk_widget().pack(fill=tk.BOTH, expand=True)
        return frame


    def crear_grafico_licitaciones_por_empresa(self, parent, bids):
        from collections import Counter

        frame = ttk.LabelFrame(parent, text="Licitaciones por Empresa")
        # La línea "frame.pack(...)" que estaba aquí ha sido ELIMINADA.
        
        conteo = Counter()
        for b in bids:
            emps = self._nuestras_empresas_de(b)
            if emps:
                conteo.update(emps)
            else:
                conteo.update(["(sin empresa)"])

        tree = ttk.Treeview(frame, columns=("empresa", "cantidad"), show="headings", height=6)
        tree.heading("empresa", text="Empresa")
        tree.heading("cantidad", text="Cantidad")
        tree.column("empresa", anchor=tk.W, width=220)
        tree.column("cantidad", anchor=tk.CENTER, width=80)

        for emp, n in sorted(conteo.items(), key=lambda x: (-x[1], x[0])):
            tree.insert("", tk.END, values=(emp, n))

        tree.pack(fill=tk.BOTH, expand=True) # Esto es correcto, posiciona el Treeview DENTRO del frame
        return frame

    def crear_tabla_por_mes(self, parent, bids):
        frame = ttk.LabelFrame(parent, text="Licitaciones por Mes")

        # (mes, empresa) -> cantidad
        counts = {}
        for b in bids:
            # Si no existe fecha, evita romper
            try:
                month_key = b.fecha_creacion.strftime("%Y-%m")
            except Exception:
                month_key = "N/D"

            empresas = self._nuestras_empresas_de(b)
            if not empresas:
                empresas = {"(sin empresa)"}

            for emp in empresas:
                counts[(month_key, emp)] = counts.get((month_key, emp), 0) + 1

        tree = ttk.Treeview(frame, columns=("mes", "empresa", "cantidad"), show="headings", height=5)
        tree.heading("mes", text="Mes")
        tree.heading("empresa", text="Empresa")
        tree.heading("cantidad", text="Cantidad")

        tree.column("mes", anchor=tk.W, width=100)
        tree.column("empresa", anchor=tk.W, width=150)
        tree.column("cantidad", anchor=tk.CENTER, width=80)

        for (month, company), count in sorted(counts.items(), key=lambda x: (x[0][0], x[0][1])):
            tree.insert("", tk.END, values=(month, company, count))

        tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        return frame

        
    def crear_analisis_precios(self, parent, stats):
        frame = ttk.LabelFrame(parent, text="Análisis de Estrategia de Precios (% Diferencia vs Monto Base)")
        if stats['avg_ganadas'] == 0 and stats['avg_perdidas'] == 0:
            ttk.Label(frame, text="No hay datos suficientes.").pack(); return frame
        
        fig = Figure(figsize=(5, 2.5), dpi=100); ax = fig.add_subplot(111)
        labels = ['% Dif. Promedio (Ganadas)', '% Dif. Promedio (Perdidas)']
        values = [stats['avg_ganadas'], stats['avg_perdidas']]; colors = ['#2E7D32', '#C62828']
        bars = ax.bar(labels, values, color=colors)
        ax.set_ylabel('% de Descuento'); ax.bar_label(bars, fmt='%.1f%%', padding=3)
        ax.axhline(y=stats['avg_ganadas'], color='blue', linestyle='--', linewidth=1)
        ax.text(1.5, stats['avg_ganadas'], f' Estrategia Ganadora: {stats["avg_ganadas"]:.1f}%', va='center', color='blue')
        fig.tight_layout()
        FigureCanvasTkAgg(fig, master=frame).get_tk_widget().pack(fill=tk.BOTH, expand=True)
        return frame
        

class DialogoAgregarDocumento(simpledialog.Dialog):
    def __init__(self, parent, title=None, initial_data=None, categorias=None, empresa_actual=None):
        self.initial_data = initial_data
        self.categorias = categorias or ["Legal", "Financiera", "Técnica", "Sobre B", "Otros"]
        self.empresa_actual = empresa_actual
        super().__init__(parent, title)
        
    def body(self, master):
        # Campos estándar del documento
        ttk.Label(master, text="Categoría:").grid(row=0, sticky=tk.W, padx=5, pady=2)
        self.categoria_var = tk.StringVar()
        self.categoria_combo = ttk.Combobox(master, textvariable=self.categoria_var, values=self.categorias, state="readonly", width=38)
        self.categoria_combo.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(master, text="Código:").grid(row=1, sticky=tk.W, padx=5, pady=2)
        self.codigo_entry = ttk.Entry(master, width=40)
        self.codigo_entry.grid(row=1, column=1, padx=5, pady=5)
        
        ttk.Label(master, text="Nombre:").grid(row=2, sticky=tk.W, padx=5, pady=2)
        self.nombre_entry = ttk.Entry(master, width=40)
        self.nombre_entry.grid(row=2, column=1, padx=5, pady=5)
        
        ttk.Label(master, text="Comentario:").grid(row=3, sticky=tk.W, padx=5, pady=2)
        self.comentario_entry = ttk.Entry(master, width=40)
        self.comentario_entry.grid(row=3, column=1, padx=5, pady=5)

        # --- NUEVO WIDGET: CHECKBOX PARA GUARDAR EN PLANTILLA ---
        ttk.Separator(master, orient='horizontal').grid(row=4, column=0, columnspan=2, sticky='ew', pady=10)
        self.guardar_plantilla_var = tk.BooleanVar(value=False)
        self.check_plantilla = ttk.Checkbutton(
            master, 
            text=f"Guardar como plantilla para '{self.empresa_actual}'",
            variable=self.guardar_plantilla_var
        )
        self.check_plantilla.grid(row=5, column=0, columnspan=2, sticky=tk.W, padx=5)

        # Si no hay empresa, se deshabilita la opción
        if not self.empresa_actual:
            self.check_plantilla.config(state="disabled")

        # Poblar datos si estamos editando
        if self.initial_data:
            self.categoria_var.set(self.initial_data.categoria)
            self.codigo_entry.insert(0, self.initial_data.codigo)
            self.nombre_entry.insert(0, self.initial_data.nombre)
            self.comentario_entry.insert(0, self.initial_data.comentario)
        elif self.categorias:
            self.categoria_combo.current(0)
            
        return self.codigo_entry

    def apply(self):
        # El resultado ahora es un diccionario que incluye el estado del checkbox
        if self.codigo_entry.get() and self.nombre_entry.get() and self.categoria_var.get():
            self.result = {
                "codigo": self.codigo_entry.get(),
                "nombre": self.nombre_entry.get(),
                "categoria": self.categoria_var.get(),
                "comentario": self.comentario_entry.get(),
                "guardar_plantilla": self.guardar_plantilla_var.get()
            }

class VentanaAgregarLicitacion(tk.Toplevel):
    # ... (sin cambios)
    def __init__(self, parent, lista_empresas, lista_instituciones, callback_guardar):
        super().__init__(parent)
        self.parent = parent
        self.lista_empresas = lista_empresas
        self.lista_instituciones = lista_instituciones
        self.callback_guardar = callback_guardar
        self.institucion_seleccionada = None
        self.empresa_seleccionada = None
        self.lotes_temp = []

        self.title("Agregar Nueva Licitación"); self.geometry("950x700"); self.grab_set()

        main_frame = ttk.Frame(self, padding="10"); main_frame.pack(fill=tk.BOTH, expand=True)

        left_frame = ttk.Frame(main_frame); left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)
        right_frame = ttk.Frame(main_frame); right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=5)
        
# En VentanaAgregarLicitacion, dentro de __init__
        # Reemplaza el bloque inst_frame

# En el __init__ de VentanaAgregarLicitacion, REEMPLAZA el bloque de inst_frame:

        inst_frame = ttk.LabelFrame(left_frame, text="A. Seleccione la Institución", padding=10)
        inst_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # Frame para la lista y el scrollbar
        list_container = ttk.Frame(inst_frame)
        list_container.pack(fill=tk.BOTH, expand=True)

        self.inst_listbox = tk.Listbox(list_container, exportselection=False)
        scrollbar = ttk.Scrollbar(list_container, orient="vertical", command=self.inst_listbox.yview)
        self.inst_listbox.configure(yscrollcommand=scrollbar.set)
        
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.inst_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.inst_listbox.bind('<<ListboxSelect>>', self.confirmar_seleccion_institucion)
        
        # Frame para el botón y la etiqueta de selección
        bottom_inst_frame = ttk.Frame(inst_frame)
        bottom_inst_frame.pack(fill=tk.X, pady=(5,0))

        self.label_inst_sel = ttk.Label(bottom_inst_frame, text="Actual: NINGUNA", font=("Helvetica", 9, "italic"))
        self.label_inst_sel.pack(side=tk.LEFT, padx=5)

        ttk.Button(bottom_inst_frame, text="➕ Agregar", command=self._agregar_nueva_institucion, style="Small.TButton").pack(side=tk.RIGHT)        # --- CAMBIO: El botón "Seleccionar" se elimina porque ya no es necesario ---
        self.label_inst_sel.pack(pady=5)

# En VentanaAgregarLicitacion, dentro de __init__
        # Reemplaza el bloque emp_frame

# En el __init__ de VentanaAgregarLicitacion, REEMPLAZA el bloque de emp_frame

        emp_frame = ttk.LabelFrame(left_frame, text="B. Seleccione su(s) Empresa(s)", padding=10)
        emp_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        self.empresas_seleccionadas = [] # Atributo para guardar los nombres

        self.label_emp_sel = ttk.Label(emp_frame, text="Ninguna seleccionada", foreground="red", wraplength=350)
        self.label_emp_sel.pack(fill=tk.X, expand=True, pady=5, padx=5)

        ttk.Button(emp_frame, text="Seleccionar Empresas...", command=self._abrir_selector_empresas_para_agregar).pack(pady=5)
        details_frame = ttk.LabelFrame(right_frame, text="C. Complete los Detalles", padding=10)
        details_frame.pack(fill=tk.X, pady=(0, 10))
        self.nombre_var = tk.StringVar()
        self.codigo_var = tk.StringVar() # Asegúrate que esta línea esté presente

        # Campo para el Nombre
        ttk.Label(details_frame, text="Nombre de la Licitación:").grid(row=0, column=0, sticky=tk.W)
        ttk.Entry(details_frame, textvariable=self.nombre_var, width=50).grid(row=0, column=1, pady=5)
        
        # --- LÍNEAS RESTAURADAS ---
        # Este es el campo que faltaba
        ttk.Label(details_frame, text="Código del Proceso:").grid(row=1, column=0, sticky=tk.W)
        ttk.Entry(details_frame, textvariable=self.codigo_var, width=50).grid(row=1, column=1, pady=5)
        # --- FIN DE LÍNEAS RESTAURADAS ---

        # Campo para el Kit
        ttk.Label(details_frame, text="Aplicar Kit de Requisitos:").grid(row=2, column=0, sticky=tk.W)
        self.kit_var = tk.StringVar()
        self.kit_combo = ttk.Combobox(details_frame, textvariable=self.kit_var, state="disabled", width=48)
        self.kit_combo.grid(row=2, column=1, pady=5)
        self.kits_disponibles = []        # --- FIN DEL CÓDIGO NUEVO ---

        self.crear_gestor_lotes(right_frame)
        # ... (resto del método __init__)        
        ttk.Button(self, text="💾 Guardar Licitación", command=self.guardar_licitacion).pack(pady=15, ipady=5)
        self.actualizar_listas()


# Pega estos dos nuevos métodos DENTRO de la clase VentanaAgregarLicitacion

# Pega esta nueva función DENTRO de la clase VentanaAgregarLicitacion

    def _agregar_nueva_institucion(self):
        """Abre un diálogo para agregar una nueva institución al catálogo."""
        dialogo = DialogoGestionarEntidad(self, "Agregar Nueva Institución", "institucion")
        if dialogo.result and dialogo.result.get('nombre'):
            nueva_institucion = dialogo.result
            nombre_nuevo = nueva_institucion['nombre']

            # 1. Verificar si ya existe
            if any(i['nombre'].lower() == nombre_nuevo.lower() for i in self.lista_instituciones):
                messagebox.showerror("Error", f"La institución '{nombre_nuevo}' ya existe.", parent=self)
                return

            # 2. Guardar en la base de datos
            self.parent.db.save_single_institucion(nueva_institucion)
            
            # 3. Actualizar las listas en memoria (la de esta ventana y la de la app principal)
            self.lista_instituciones.append(nueva_institucion)
            self.parent.instituciones_registradas.append(nueva_institucion)
            
            # 4. Refrescar la Listbox
            self.actualizar_listas()
            
            # 5. Seleccionar automáticamente la nueva institución
            for i, item in enumerate(self.inst_listbox.get(0, tk.END)):
                if item == nombre_nuevo:
                    self.inst_listbox.selection_set(i)
                    self.inst_listbox.see(i)
                    self.confirmar_seleccion_institucion() # Simular clic para actualizar todo
                    break
            
            messagebox.showinfo("Éxito", f"Institución '{nombre_nuevo}' agregada correctamente.", parent=self)



    def _abrir_selector_empresas_para_agregar(self):
        """Abre el diálogo para que el usuario elija las empresas."""
        # Creamos objetos Empresa temporales para pasarlos al diálogo
        empresas_obj_actuales = [Empresa(nombre) for nombre in self.empresas_seleccionadas]

        dialogo = DialogoSeleccionarNuestrasEmpresas(
            self,
            self.lista_empresas, # lista de dicts {'nombre':...}
            empresas_obj_actuales
        )
        if dialogo.result is not None:
            self.empresas_seleccionadas = dialogo.result # Guardar la lista de nombres
            self._actualizar_display_empresas()

    def _actualizar_display_empresas(self):
        """Actualiza la etiqueta para mostrar la selección."""
        if not self.empresas_seleccionadas:
            self.label_emp_sel.config(text="Ninguna seleccionada", foreground="red")
        else:
            texto = ", ".join(sorted(self.empresas_seleccionadas))
            self.label_emp_sel.config(text=texto, foreground="black")

    def crear_gestor_lotes(self, parent_frame):
        lotes_frame = ttk.LabelFrame(parent_frame, text="D. Lotes del Proceso", padding=10)
        lotes_frame.pack(fill=tk.BOTH, expand=True)
        tree_frame = ttk.Frame(lotes_frame); tree_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        self.tree_lotes = ttk.Treeview(tree_frame, columns=("numero", "nombre", "monto_base", "monto_ofertado"), show="headings", height=5)
        self.tree_lotes.heading("numero", text="N°"); self.tree_lotes.heading("nombre", text="Nombre Lote")
        self.tree_lotes.heading("monto_base", text="Monto Base"); self.tree_lotes.heading("monto_ofertado", text="Nuestra Oferta")
        self.tree_lotes.column("numero", width=40, anchor=tk.CENTER); self.tree_lotes.column("nombre", width=200)
        self.tree_lotes.column("monto_base", anchor=tk.E); self.tree_lotes.column("monto_ofertado", anchor=tk.E)
        
        self.tree_lotes.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree_lotes.yview)
        self.tree_lotes.configure(yscroll=scrollbar.set); scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        btn_frame = ttk.Frame(lotes_frame); btn_frame.pack(fill=tk.X, pady=5)
        ttk.Button(btn_frame, text="➕ Agregar Lote", command=self.agregar_lote).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="✏️ Editar Lote", command=self.editar_lote).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="🗑️ Eliminar Lote", command=self.eliminar_lote).pack(side=tk.LEFT, padx=5)

    def actualizar_tree_lotes(self):
        for i in self.tree_lotes.get_children(): self.tree_lotes.delete(i)
        for lote in self.lotes_temp:
            self.tree_lotes.insert("", tk.END, values=(lote.numero, lote.nombre, f"RD$ {lote.monto_base:,.2f}", f"RD$ {lote.monto_ofertado:,.2f}"))

# En la clase VentanaAgregarLicitacion

    def agregar_lote(self):
        # Pasamos la lista de empresas seleccionadas al diálogo
        dialogo = DialogoGestionarLote(self, participating_companies=self.empresas_seleccionadas)
        if dialogo.result:
            self.lotes_temp.append(dialogo.result)
            self.actualizar_tree_lotes()
    
    def editar_lote(self):
        if not (selected_item := self.tree_lotes.focus()):
            messagebox.showwarning("Sin Selección", "Selecciona un lote.", parent=self)
            return
        idx = self.tree_lotes.index(selected_item)
        lote_a_editar = self.lotes_temp[idx]
        # Pasamos la lista de empresas seleccionadas al diálogo
        dialogo = DialogoGestionarLote(self, initial_data=lote_a_editar, participating_companies=self.empresas_seleccionadas)
        if dialogo.result:
            self.lotes_temp[idx] = dialogo.result
            self.actualizar_tree_lotes()
    def eliminar_lote(self):
        if not (selected_item := self.tree_lotes.focus()): messagebox.showwarning("Sin Selección", "Selecciona un lote.", parent=self); return
        if messagebox.askyesno("Confirmar", "¿Eliminar el lote seleccionado?", parent=self):
            self.lotes_temp.pop(self.tree_lotes.index(selected_item)); self.actualizar_tree_lotes()

# En la clase VentanaAgregarLicitacion, reemplaza este método

# En la clase VentanaAgregarLicitacion, REEMPLAZA el método guardar_licitacion

    def guardar_licitacion(self):
        if not self.institucion_seleccionada:
            messagebox.showerror("Campo Requerido", "Debe seleccionar una institución.", parent=self)
            return
        if not self.empresas_seleccionadas:
            messagebox.showerror("Campo Requerido", "Debe seleccionar al menos una empresa participante.", parent=self)
            return
        if not self.nombre_var.get().strip() or not self.codigo_var.get().strip():
            messagebox.showerror("Campo Requerido", "Nombre y Código no pueden estar vacíos.", parent=self)
            return
        if not self.lotes_temp:
            messagebox.showerror("Campo Requerido", "Debe agregar al menos un lote.", parent=self)
            return

        # Formatear las empresas para el constructor de Licitacion
        empresas_data = [{'nombre': nombre} for nombre in self.empresas_seleccionadas]

        datos = {
            "nombre_proceso": self.nombre_var.get().strip(),
            "numero_proceso": self.codigo_var.get().strip(),
            "institucion": self.institucion_seleccionada,
            "empresas_nuestras": empresas_data, # <--- Cambio clave aquí
            "lotes": [l.to_dict() for l in self.lotes_temp],
            "documentos_solicitados": []
        }
        
        # ... (el resto de la lógica para aplicar kits sigue igual) ...
        kit_seleccionado_nombre = self.kit_var.get()
        if kit_seleccionado_nombre and kit_seleccionado_nombre != " (Ninguno) ":
            kit_id_seleccionado = next((kit[0] for kit in self.kits_disponibles if kit[1] == kit_seleccionado_nombre), None)
            if kit_id_seleccionado:
                cursor = self.parent.db.cursor
                cursor.execute("SELECT documento_maestro_id FROM kit_items WHERE kit_id = ?", (kit_id_seleccionado,))
                ids_docs_maestros = [row[0] for row in cursor.fetchall()]
                if ids_docs_maestros:
                    documentos_del_kit = [doc for doc in self.parent.documentos_maestros if doc.id in ids_docs_maestros]
                    for doc_maestro in documentos_del_kit:
                        nuevo_doc = Documento(
                            codigo=doc_maestro.codigo, nombre=doc_maestro.nombre,
                            categoria=doc_maestro.categoria, comentario=doc_maestro.comentario
                        )
                        datos["documentos_solicitados"].append(nuevo_doc.to_dict())

        self.callback_guardar(Licitacion(**datos))
        messagebox.showinfo("Éxito", "Licitación agregada correctamente.", parent=self)
        self.destroy()

    def confirmar_seleccion_institucion(self, event=None):
        print("DEBUG: Se llamó a 'confirmar_seleccion_institucion'.")
        try:
            # Obtenemos el índice del elemento seleccionado
            seleccion_indices = self.inst_listbox.curselection()
            if not seleccion_indices:
                print("DEBUG: No hay selección en la lista de instituciones.")
                return

            self.institucion_seleccionada = self.inst_listbox.get(seleccion_indices[0])
            self.label_inst_sel.config(text=f"Actual: {self.institucion_seleccionada}")
            print(f"DEBUG: 'self.institucion_seleccionada' AHORA ES ==> '{self.institucion_seleccionada}'")

            self.kit_combo.set('')
            self.kits_disponibles = []
            cursor = self.parent.db.cursor
            cursor.execute("SELECT id, nombre_kit FROM kits_de_requisitos WHERE institucion_nombre = ? ORDER BY nombre_kit",
                           (self.institucion_seleccionada,))
            self.kits_disponibles = cursor.fetchall()

            if self.kits_disponibles:
                nombres_kits = [kit[1] for kit in self.kits_disponibles]
                self.kit_combo['values'] = [" (Ninguno) "] + nombres_kits
                self.kit_combo.config(state="readonly")
                self.kit_combo.current(0)
            else:
                self.kit_combo['values'] = []
                self.kit_combo.config(state="disabled")

        except tk.TclError:
            self.institucion_seleccionada = None
            print("DEBUG: Ocurrió un TclError en la selección de institución.")
            self.kit_combo.set('')
            self.kit_combo.config(state="disabled")

    def confirmar_seleccion_empresa(self, event=None):
        print("DEBUG: Se llamó a 'confirmar_seleccion_empresa'.")
        try:
            # Obtenemos el índice del elemento seleccionado
            seleccion_indices = self.emp_listbox.curselection()
            if not seleccion_indices:
                print("DEBUG: No hay selección en la lista de empresas.")
                return

            self.empresa_seleccionada = self.emp_listbox.get(seleccion_indices[0])
            self.label_emp_sel.config(text=f"Actual: {self.empresa_seleccionada}")
            print(f"DEBUG: 'self.empresa_seleccionada' AHORA ES ==> '{self.empresa_seleccionada}'")
        except tk.TclError:
            print("DEBUG: Ocurrió un TclError en la selección de empresa.")
            pass
    
# En la clase VentanaAgregarLicitacion, REEMPLAZA este método:

    def actualizar_listas(self):
        # Limpiamos la lista de instituciones (la de empresas ya no existe)
        self.inst_listbox.delete(0, tk.END)
        
        # Llenamos la lista de instituciones
        for inst in sorted(self.lista_instituciones, key=lambda i: i.get('nombre', '')):
            self.inst_listbox.insert(tk.END, inst['nombre'])
class VentanaVisorDocumentos(tk.Toplevel):
    """
    Visor liviano por pestañas: Todos, Legal, Financiera, Técnica, Sobre B.
    Solo lectura. Permite abrir Gestión y refrescar.
    """
    def __init__(self, parent, licitacion, categorias, on_refresh=None):
        super().__init__(parent)
        self.parent = parent
        self.licitacion = licitacion
        self.categorias = list(categorias or ["Legal", "Financiera", "Técnica", "Sobre B"])
        self.on_refresh = on_refresh

        self.title(f"Checklist de Documentos — {licitacion.nombre_proceso}")
        self.geometry("950x600")
        self.grab_set()

        self.nb = ttk.Notebook(self)
        self.nb.pack(fill=tk.BOTH, expand=True, padx=10, pady=(10, 6))

        # Pestaña "Todos" al inicio
        self._tabs = {}  # nombre_tab -> (frame, tree)
        self._crear_tab("Todos")
        for cat in self.categorias:
            self._crear_tab(cat)

        # Barra de acciones abajo
        bottom = ttk.Frame(self)
        bottom.pack(fill=tk.X, padx=10, pady=(0, 10))
        ttk.Button(bottom, text="🔄 Refrescar", command=self._refrescar).pack(side=tk.LEFT)
        ttk.Button(bottom, text="🛠️ Abrir Gestión…", command=self._abrir_gestion).pack(side=tk.LEFT, padx=6)
        ttk.Button(bottom, text="Cerrar", command=self.destroy).pack(side=tk.RIGHT)

        self._refrescar()

    # ---------- UI helpers ----------
    def _crear_tab(self, nombre):
        frm = ttk.Frame(self.nb, padding=6)
        self.nb.add(frm, text=nombre)

        # Tree + scrollbar vertical
        cols = ("estado", "codigo", "nombre", "categoria", "cond", "rev", "adj", "orden")
        tree = ttk.Treeview(frm, columns=cols, show="headings", height=14)
        vsb = ttk.Scrollbar(frm, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)

        tree.heading("estado", text="✓")
        tree.heading("codigo", text="Código")
        tree.heading("nombre", text="Documento")
        tree.heading("categoria", text="Categoría")
        tree.heading("cond", text="Condición")
        tree.heading("rev", text="👁️")
        tree.heading("adj", text="📎")
        tree.heading("orden", text="Orden")

        tree.column("estado", width=40, anchor=tk.CENTER, stretch=False)
        tree.column("codigo", width=120, anchor=tk.W)
        tree.column("nombre", width=460, anchor=tk.W)
        tree.column("categoria", width=100, anchor=tk.CENTER)
        tree.column("cond", width=100, anchor=tk.CENTER)
        tree.column("rev", width=50, anchor=tk.CENTER, stretch=False)
        tree.column("adj", width=50, anchor=tk.CENTER, stretch=False)
        tree.column("orden", width=60, anchor=tk.E, stretch=False)

        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        # Doble click: si hay archivo, lo abre
        tree.bind("<Double-1>", lambda e, t=tree: self._abrir_archivo_desde_tree(t))

        self._tabs[nombre] = (frm, tree)

    def _ord_val(self, d):
        v = getattr(d, "orden_pliego", None)
        try:
            return 999999 if v in (None, "", "None") else int(v)
        except Exception:
            return 999999

    # ---------- Data render ----------
    def _refrescar(self):
        # Limpiar
        for _, tree in self._tabs.values():
            for iid in tree.get_children():
                tree.delete(iid)

        docs = list(getattr(self.licitacion, "documentos_solicitados", []) or [])
        # Orden estable: orden_pliego -> código -> nombre
        docs.sort(key=lambda d: (self._ord_val(d), getattr(d, "codigo", "") or "", getattr(d, "nombre", "") or ""))

        # Llenar "Todos"
        self._poblar_tree(self._tabs["Todos"][1], docs)

        # Llenar por categoría
        docs_by_cat = {}
        for d in docs:
            docs_by_cat.setdefault(getattr(d, "categoria", "") or "Legal", []).append(d)

        for cat in self.categorias:
            tree = self._tabs.get(cat, (None, None))[1]
            if tree is None:
                continue
            self._poblar_tree(tree, docs_by_cat.get(cat, []))

        # Refrescar Detalles si pasaron un callback
        try:
            if callable(self.on_refresh):
                self.on_refresh()
        except Exception:
            pass

    def _poblar_tree(self, tree, docs):
        for d in docs:
            estado = "✓" if getattr(d, "presentado", False) else "✖"
            rev = "✓" if getattr(d, "revisado", False) else "—"
            adj = "✓" if getattr(d, "ruta_archivo", "") else "—"
            cond = (getattr(d, "subsanable", "") or "No Definido")
            tree.insert(
                "", tk.END,
                values=(
                    estado,
                    getattr(d, "codigo", "") or "",
                    getattr(d, "nombre", "") or "",
                    getattr(d, "categoria", "") or "",
                    cond,
                    rev,
                    adj,
                    self._ord_val(d)
                ),
            )

    # ---------- Acciones ----------
    def _abrir_gestion(self):
        # Reutilizamos la ventana de gestión de tu app
        try:
            if hasattr(self.parent, "abrir_gestion_docs"):
                self.parent.abrir_gestion_docs()
            else:
                messagebox.showinfo("Info", "No se encontró la acción para abrir Gestión.", parent=self)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir Gestión:\n{e}", parent=self)

    def _abrir_archivo_desde_tree(self, tree):
        # Identifica el doc por código+nombre (suficiente para abrir adjunto si existe)
        sel = tree.selection()
        if not sel:
            return
        vals = tree.item(sel[0], "values")
        codigo, nombre = vals[1], vals[2]
        for d in getattr(self.licitacion, "documentos_solicitados", []) or []:
            if (getattr(d, "codigo", "") or "") == codigo and (getattr(d, "nombre", "") or "") == nombre:
                ruta = getattr(d, "ruta_archivo", "") or ""
                if ruta:
                    try:
                        os.startfile(ruta)
                    except Exception as e:
                        messagebox.showerror("Error", f"No se pudo abrir el archivo:\n{e}", parent=self)
                break





class VentanaDetalles(tk.Toplevel):
    # --- NUEVO: soporte "un ganador por cada LOTE" ---
    # Esta versión agrega un panel "Ganadores por Lote" con RadioButtons por lote.
    # Cada lote tiene su propia StringVar, así podrás marcar varios ganadores (uno por lote)
    # dentro de la misma licitación (incluyendo a tu empresa, si participa en ese lote).



    def _nuestras_empresas_de(self, lic):
        empresas = set()
        for lote in getattr(lic, "lotes", []):
            nombre_empresa_lote = getattr(lote, "empresa_nuestra", None)
            if nombre_empresa_lote and isinstance(nombre_empresa_lote, str) and nombre_empresa_lote.strip():
                empresas.add(nombre_empresa_lote.strip())
        
        if not empresas:
            for item in getattr(lic, "empresas_nuestras", []):
                nombre = None
                if hasattr(item, 'nombre'): nombre = item.nombre
                elif isinstance(item, dict) and 'nombre' in item: nombre = item.get('nombre')
                if nombre and isinstance(nombre, str) and nombre.strip() and nombre.lower() != 'none':
                    empresas.add(nombre.strip())
        return empresas
        
    def _extraer_nombres_empresas(self, seq):
        """Devuelve un set de nombres de empresa a partir de str/dict/objeto."""
        out = set()
        for it in (seq or []):
            if isinstance(it, str):
                nombre = it
            elif isinstance(it, dict):
                nombre = it.get("nombre", "")
            else:
                # objeto con atributo nombre/razon_social
                nombre = getattr(it, "nombre", "") or getattr(it, "razon_social", "")
            nombre = (nombre or "").strip()
            if nombre and nombre.lower() != "none":
                out.add(nombre)
        return out


    def _display_empresas_de(self, lic):
        """Devuelve un string amigable para mostrar nuestras empresas de una licitación."""
        emps = sorted(self._nuestras_empresas_de(lic))
        return ", ".join(emps) if emps else "(sin empresa)"


# REEMPLAZA el método __init__ de tu clase VentanaDetalles con este:

# En la clase VentanaDetalles, REEMPLAZA el método __init__ con este:

    def __init__(self, parent, licitacion, callback_actualizar, documentos_maestros, categorias_documentos, db_manager, lista_instituciones):
        super().__init__(parent)
        self.parent_app = parent
        self.licitacion = licitacion
        self.callback_actualizar = callback_actualizar
        self.documentos_maestros = documentos_maestros
        self.categorias_documentos = categorias_documentos
        self.db = db_manager
        # --- LÍNEA NUEVA: Guardamos la lista de instituciones ---
        self.lista_instituciones = lista_instituciones
        
        # --- Variables de estado ---
        self.docs_manual_var = tk.BooleanVar(value=self.licitacion.docs_completos_manual)
        self.var_codigo = tk.StringVar(value=self.licitacion.numero_proceso)
        self.var_nombre = tk.StringVar(value=self.licitacion.nombre_proceso)
        # --- LÍNEA NUEVA: Creamos la variable para el combobox de institución ---
        self.var_institucion = tk.StringVar(value=self.licitacion.institucion)
        self.var_ganador_por_lote = {}

        # --- Configuración de la ventana ---
        self.title(f"Detalles de: {self.licitacion.nombre_proceso}")
        self.geometry("950x700")
        self.grab_set()

        main_frame = ttk.Frame(self, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)
        main_frame.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)

        # --- Notebook (Sin la pestaña de Empresas) ---
        notebook = ttk.Notebook(main_frame)
        notebook.grid(row=0, column=0, sticky="nsew", pady=5)

        tabs = {
            "Detalles Generales": "crear_widgets_generales",
            "Lotes del Proceso": "crear_widgets_lotes",
            "Resultados de Competidores": "crear_widgets_oferentes",
        }
        for text, method_name in tabs.items():
            tab = ttk.Frame(notebook, padding="10")
            notebook.add(tab, text=text)
            getattr(self, method_name)(tab)

        # --- Barra de botones inferior ---
        bottom = ttk.Frame(main_frame)
        bottom.grid(row=1, column=0, sticky="ew", pady=(8, 0))
        bottom.columnconfigure(1, weight=1)

        ttk.Button(bottom, text="Cerrar sin Guardar", command=self.destroy).grid(row=0, column=0, padx=6, sticky="w")
        self.btn_guardar_continuar = ttk.Button(bottom, text="💾 Guardar y Continuar", command=self._guardar_sin_cerrar)
        self.btn_guardar_continuar.grid(row=0, column=1, padx=6, sticky="e")
        ttk.Button(bottom, text="✅ Guardar y Cerrar", command=self.guardar_y_cerrar).grid(row=0, column=2, padx=6, ipadx=8, ipady=4, sticky="e")


# Pega este nuevo método DENTRO de la clase VentanaDetalles

# En la clase VentanaDetalles

    def _abrir_analisis_fase_a(self):
        """
        Abre la ventana de análisis de fallas en Fase A, verificando primero
        que no haya documentos sin guardar.
        """
        # --- INICIO DE LA LÓGICA MEJORADA ---
        # 1. Comprobar si hay documentos sin un ID de base de datos.
        documentos_sin_guardar = [
            doc for doc in self.licitacion.documentos_solicitados if not getattr(doc, 'id', None)
        ]

        if documentos_sin_guardar:
            messagebox.showinfo(
                "Guardar Cambios",
                f"Se han detectado {len(documentos_sin_guardar)} documento(s) nuevo(s) que aún no se han guardado en la base de datos.\n\n"
                "Por favor, usa el botón '💾 Guardar y Continuar' en la ventana de Detalles para registrar los cambios antes de analizar las fallas.",
                parent=self
            )
            return # Detenemos la acción para que el usuario pueda guardar.
        # --- FIN DE LA LÓGICA MEJORADA ---

        # 2. Si todo está guardado, abrimos la ventana como siempre.
        VentanaAnalisisFaseA(self, self.licitacion, self.db)
    def _get_empresas_catalogo(self):
        """Catálogo completo de empresas maestras (lista de nombres)."""
        try:
            return [e['nombre'] for e in self.db.get_empresas_maestras()]
        except Exception:
            return [e.get('nombre', '') for e in getattr(self.parent_app, 'empresas_registradas', []) if e.get('nombre')]

    def _empresas_seleccionadas_panel(self):
        """Empresas seleccionadas en la pestaña Empresas Nuestras (lista de nombres)."""
        if hasattr(self, "list_empresas") and self.list_empresas.winfo_exists():
            return [self.list_empresas.get(i) for i in self.list_empresas.curselection()]
        return []

    def _aplicar_empresas_a_lotes(self, modo="todos"):
        """
        Aplica la selección del panel a los lotes:
        - modo="todos": pone la misma empresa (o distribución) en TODOS los lotes.
        - modo="sin_empresa": solo rellena los lotes que no tienen empresa_nuestra.
        - modo="quitar": limpia empresa_nuestra en todos los lotes.
        Reglas:
        * Si hay 1 empresa seleccionada -> se aplica esa a los lotes objetivo.
        * Si hay 2+ seleccionadas -> se reparte round-robin (1,2,3,1,2,3,...).
        * Si no hay seleccionadas -> se avisa.
        """
        seleccion = self._empresas_seleccionadas_panel()
        lotes = getattr(self.licitacion, "lotes", [])

        if modo == "quitar":
            for l in lotes:
                l.empresa_nuestra = None
            self.actualizar_tree_lotes()
            try:
                self._rebuild_ganadores_ui()
            except Exception:
                pass
            messagebox.showinfo("Empresas por lote", "Se quitó la empresa de todos los lotes.", parent=self)
            return

        if not seleccion:
            messagebox.showwarning("Empresas Nuestras", "Selecciona al menos una empresa en la lista.", parent=self)
            return

        # Objetivo: todos o solo los vacíos
        objetivos = lotes if modo == "todos" else [l for l in lotes if not getattr(l, "empresa_nuestra", None)]

        if not objetivos:
            messagebox.showinfo("Empresas por lote", "No hay lotes a los que aplicar (ya tienen empresa asignada).", parent=self)
            return

        if len(seleccion) == 1:
            emp = seleccion[0]
            for l in objetivos:
                l.empresa_nuestra = emp
        else:
            # Reparto secuencial (round-robin)
            idx = 0
            n = len(seleccion)
            for l in objetivos:
                l.empresa_nuestra = seleccion[idx]
                idx = (idx + 1) % n

        # Refrescar
        self.actualizar_tree_lotes()
        try:
            self._rebuild_ganadores_ui()
        except Exception:
            pass
        messagebox.showinfo("Empresas por lote", "Asignación aplicada correctamente.", parent=self)






    def _mostrar_menu_lotes(self, event):
        iid = self.tree_lotes.identify_row(event.y)
        if iid:
            self.tree_lotes.selection_set(iid)
            menu = tk.Menu(self, tearoff=0)
            menu.add_command(label="📊 Comparar Ofertas para este Lote", command=self._abrir_comparador_ofertas)
            menu.post(event.x_root, event.y_root)

    def _abrir_comparador_ofertas(self):
        if not self.tree_lotes.selection():
            return
        item_id = self.tree_lotes.selection()[0]
        # Toma el número de lote desde los 'values' (columna 3 = numero)
        values = self.tree_lotes.item(item_id, "values")
        if not values:
            return
        num_lote = values[2]  # 'numero' es la 3ra columna en tu definición
        VentanaComparadorOfertas(self, self.licitacion, num_lote)

    # =======================
    # COMPETIDORES: agregar desde catálogo (sin cambios sustantivos)
    # =======================
    def _agregar_desde_lista(self):
        print("\n--- Acción: Agregar Competidor desde Catálogo ---")
        print(f"[ANTES] Tamaño del catálogo maestro: {len(self.parent_app.competidores_maestros)}")
        
        dialogo = DialogoSeleccionarCompetidores(self, self.parent_app.competidores_maestros, self.licitacion.oferentes_participantes)
        if dialogo.result:
            for comp_data in dialogo.result:
                print(f"  -> Agregando '{comp_data['nombre']}' a la licitación actual.")
                nuevo_oferente = Oferente(nombre=comp_data['nombre'], comentario="")
                self.licitacion.oferentes_participantes.append(nuevo_oferente)
            self._actualizar_tree_competidores()
            self._rebuild_ganadores_ui()  # NUEVO: reconstruir radios
        
        print(f"[DESPUÉS] Tamaño del catálogo maestro: {len(self.parent_app.competidores_maestros)}")
        print("--- Fin de la Acción ---\n")

    # =======================
    # ANALIZAR PLIEGO (tu código, sin cambios)
    # =======================
# En la clase VentanaDetalles, REEMPLAZA este método:

# En la clase VentanaDetalles, REEMPLAZA este método:

# En la clase VentanaDetalles, REEMPLAZA este método:

    def _analizar_pliego_con_ia(self):
        try:
            import google.generativeai as genai
        except ImportError:
            messagebox.showerror("Librería Faltante", "La librería 'google-generativeai' no está instalada. Ejecuta: pip install google-generativeai")
            return
        
        api_key = self.parent_app.api_key

        if not api_key:
            api_key = simpledialog.askstring("Clave API de Google AI", 
                                             "Por primera vez, introduce tu clave API de Google AI:", 
                                             parent=self, show='*')
            if not api_key:
                return
            
            self.parent_app.api_key = api_key
            self.parent_app._guardar_configuracion()
        
        try:
            genai.configure(api_key=api_key)
        except Exception as e:
            messagebox.showerror("Error de API", f"La clave API no es válida o no se pudo configurar:\n{e}", parent=self)
            self.parent_app.api_key = None
            self.parent_app._guardar_configuracion()
            return

        ruta_pliego = filedialog.askopenfilename(
            parent=self, title="Selecciona el Pliego de Condiciones (PDF)", filetypes=[("PDF files", "*.pdf")]
        )
        if not ruta_pliego: return
            
        self.config(cursor="wait")
        texto_pliego = _extraer_texto_de_pdf(ruta_pliego)
        if not texto_pliego:
            self.config(cursor="")
            return

        # --- CAMBIO CLAVE AQUÍ ---
        # Usamos el nombre del modelo más reciente y eficiente
        modelo = genai.GenerativeModel('gemini-1.5-flash-latest')
        # --- FIN DEL CAMBIO ---
        
        prompt = f"""
        Analiza el siguiente texto de un pliego de condiciones de una licitación. Extrae únicamente la lista de documentos requeridos.
        Formatea la respuesta como una lista JSON válida. Cada objeto en la lista debe tener dos claves: "nombre" (el nombre completo del documento) y "categoria" (clasifícalo en 'Legal', 'Financiera' o 'Técnica').
        No incluyas nada más en tu respuesta, solo la lista JSON.

        Texto del pliego:
        ---
        {texto_pliego[:8000]}
        ---
        """
        
        try:
            respuesta = modelo.generate_content(prompt)
            json_text = respuesta.text.strip().replace("```json", "").replace("```", "").strip()
            documentos_ia = json.loads(json_text)

            if not isinstance(documentos_ia, list):
                raise ValueError("La IA no devolvió una lista.")

            dialogo = DialogoResultadosIA(self, documentos_ia)
            if dialogo.result:
                codigos_existentes = {d.codigo for d in self.licitacion.documentos_solicitados}
                nuevos_agregados = 0
                for doc_data in dialogo.result:
                    codigo_temporal = "".join(filter(str.isalnum, doc_data['nombre']))[:10].upper()
                    if codigo_temporal not in codigos_existentes:
                        nuevo_doc = Documento(codigo=codigo_temporal, nombre=doc_data['nombre'], categoria=doc_data['categoria'])
                        self.licitacion.documentos_solicitados.append(nuevo_doc)
                        codigos_existentes.add(codigo_temporal)
                        nuevos_agregados += 1
                
                if nuevos_agregados > 0:
                    self.parent_app.actualizar_tabla_gui() 
                    self.actualizar_info_docs()
                    messagebox.showinfo("Éxito", f"Se importaron {nuevos_agregados} documentos.", parent=self)
        
        except Exception as e:
            messagebox.showerror("Error de IA", f"No se pudo procesar la respuesta de la IA:\n{e}\n\nAsegúrate de que el pliego contenga una lista clara de documentos.", parent=self)
        finally:
            self.config(cursor="")


    def _on_toggle_docs_manual(self):
        self.licitacion.docs_completos_manual = self.docs_manual_var.get()
        if hasattr(self, "progreso_docs_var"):
            self.progreso_docs_var.set(f"{self.licitacion.get_porcentaje_completado():.1f}%")
        try:
            self.callback_actualizar()
        except Exception:
            pass


# Pega estos dos nuevos métodos DENTRO de la clase VentanaDetalles

    def _abrir_selector_empresas(self):
        """Abre el diálogo de selección múltiple de empresas."""
        dialogo = DialogoSeleccionarNuestrasEmpresas(
            self,
            self.parent_app.empresas_registradas,
            self.licitacion.empresas_nuestras
        )
        if dialogo.result is not None: # El resultado puede ser una lista vacía
            # Actualizar el objeto licitacion en memoria
            self.licitacion.empresas_nuestras = [Empresa(nombre) for nombre in dialogo.result]
            # Actualizar la vista
            self._actualizar_display_empresas()
            # Opcional: Reconstruir la UI de ganadores por si cambiaron las opciones
            self._rebuild_ganadores_ui()

    def _actualizar_display_empresas(self):
        """Actualiza la etiqueta que muestra las empresas seleccionadas."""
        nombres = [str(e) for e in self.licitacion.empresas_nuestras]
        if not nombres:
            texto = "Ninguna empresa seleccionada."
            color = "red"
        else:
            texto = ", ".join(sorted(nombres))
            color = "black"
        
        self.label_empresas_seleccionadas.config(text=texto, foreground=color)

    # =======================
    # PESTAÑA: DETALLES GENERALES (sin cambios relevantes)
    # =======================

    def crear_widgets_generales(self, parent_frame):
        # --- Frame para Identificación y Empresas ---
        top_container = ttk.Frame(parent_frame)
        top_container.pack(fill=tk.X, pady=5)

        frm_ident = ttk.LabelFrame(top_container, text="Identificación del Proceso", padding=8)
        frm_ident.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        frm_ident.columnconfigure(1, weight=1)
        
        ttk.Label(frm_ident, text="Código:").grid(row=0, column=0, sticky="w", padx=(0, 6), pady=2)
        ttk.Entry(frm_ident, textvariable=self.var_codigo).grid(row=0, column=1, sticky="we", pady=2)
        
        ttk.Label(frm_ident, text="Nombre:").grid(row=1, column=0, sticky="w", padx=(0, 6), pady=2)
        ttk.Entry(frm_ident, textvariable=self.var_nombre).grid(row=1, column=1, sticky="we", pady=2)

        # --- INICIO DEL CÓDIGO AÑADIDO ---
        ttk.Label(frm_ident, text="Institución:").grid(row=2, column=0, sticky="w", padx=(0, 6), pady=2)
        nombres_instituciones = sorted([i['nombre'] for i in self.lista_instituciones])
        combo_inst = ttk.Combobox(frm_ident, textvariable=self.var_institucion, values=nombres_instituciones, state="readonly")
        combo_inst.grid(row=2, column=1, sticky="we", pady=2)
        # --- FIN DEL CÓDIGO AÑADIDO ---

        frm_empresas = ttk.LabelFrame(top_container, text="Nuestras Empresas", padding=8)
        frm_empresas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.label_empresas_seleccionadas = ttk.Label(frm_empresas, text="Cargando...", wraplength=300, justify=tk.LEFT)
        self.label_empresas_seleccionadas.pack(fill=tk.BOTH, expand=True, padx=5, pady=(0,5))
        ttk.Button(frm_empresas, text="Seleccionar...", command=self._abrir_selector_empresas).pack(anchor="se", padx=5)
        self._actualizar_display_empresas()
        
        # El resto del método continúa sin cambios...
        # --- Frame para Cronograma e Info General ---
        middle_container = ttk.Frame(parent_frame)
        middle_container.pack(fill=tk.X, pady=5)
        info_frame = ttk.LabelFrame(middle_container, text="Información General y Estado", padding="10")
        info_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10), anchor=tk.N)
        crono_frame = ttk.LabelFrame(middle_container, text="Cronograma del Proceso", padding="10")
        crono_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, anchor=tk.N)
        # ... (el resto del código que ya tenías)
        self.estado_var = tk.StringVar(value=self.licitacion.estado)
        self.adjudicado_a_var = tk.StringVar(value=self.licitacion.adjudicada_a)
        self.fase_b_var = tk.BooleanVar(value=self.licitacion.fase_B_superada)
        self.progreso_docs_var = tk.StringVar(value=f"{self.licitacion.get_porcentaje_completado():.1f}%")
        estados_posibles = [
            "Iniciada", "En Proceso", "Sobre A Entregado", "Sobre B Entregado",
            "Descalificado Fase A", "Descalificado Fase B", "Adjudicada", "Desierta", "Cancelada"
        ]
        ttk.Label(info_frame, text="Estado:").grid(row=0, column=0, sticky=tk.W, pady=3)
        combo_estado = ttk.Combobox(info_frame, textvariable=self.estado_var,
                                    values=estados_posibles, state="readonly", width=25)
        combo_estado.grid(row=0, column=1, sticky=tk.EW, pady=3)
        combo_estado.bind("<<ComboboxSelected>>", self._on_estado_change)
        ttk.Label(info_frame, text="Adjudicada a:").grid(row=1, column=0, sticky=tk.W, pady=3)
        nuestras_empresas = [str(e) for e in getattr(self.licitacion, "empresas_nuestras", [])]
        otros_participantes = [
            o.nombre for o in getattr(self.licitacion, "oferentes_participantes", [])
            if o.nombre not in nuestras_empresas
        ]
        nombres_participantes = nuestras_empresas + otros_participantes
        self.combo_adjudicado_a = ttk.Combobox(
            info_frame, textvariable=self.adjudicado_a_var,
            values=sorted(list(set(nombres_participantes))), state="disabled", width=25
        )
        self.combo_adjudicado_a.grid(row=1, column=1, sticky=tk.EW, pady=3)
        ttk.Label(info_frame, text="Progreso Docs:").grid(row=2, column=0, sticky=tk.W, pady=3)
        ttk.Label(info_frame, textvariable=self.progreso_docs_var,
                font=("Helvetica", 10, "bold")).grid(row=2, column=1, sticky=tk.W, pady=3)
        ttk.Checkbutton(
            info_frame, text="Documentación completa (sin requisitos)",
            variable=self.docs_manual_var, command=self._on_toggle_docs_manual
        ).grid(row=3, column=0, columnspan=2, sticky=tk.W, pady=3)
        ttk.Checkbutton(
            info_frame, text="Fase B (Sobres Económicos) superada",
            variable=self.fase_b_var
        ).grid(row=4, column=0, columnspan=2, sticky=tk.W, pady=3)
        ttk.Label(info_frame, text="Motivo Descalificación / Comentarios:"
                ).grid(row=5, column=0, columnspan=2, sticky=tk.W, pady=(10, 3))
        self.motivo_text = tk.Text(info_frame, width=40, height=4)
        self.motivo_text.grid(row=6, column=0, columnspan=2, sticky=tk.W)
        self.motivo_text.insert("1.0", self.licitacion.motivo_descalificacion)
        self.cronograma_entries, self.cronograma_estados_vars = {}, {}
        estados_crono = ["Pendiente", "Cumplido", "Incumplido"]
        for i, evento in enumerate(self.licitacion.cronograma.keys()):
            datos_evento = self.licitacion.cronograma.get(evento, {"fecha_limite": None, "estado": "Pendiente"})
            ttk.Label(crono_frame, text=f"{evento}:").grid(row=i, column=0, sticky=tk.W, pady=2, padx=5)
            date_entry = DateEntry(crono_frame, width=15, locale='es_ES', date_pattern='y-mm-dd')
            date_entry.grid(row=i, column=1, sticky=tk.EW, pady=2, padx=5)
            estado_var = tk.StringVar(value=datos_evento.get("estado", "Pendiente"))
            ttk.Combobox(crono_frame, textvariable=estado_var, values=estados_crono,
                        state="readonly", width=12).grid(row=i, column=2, sticky=tk.EW, pady=2, padx=5)
            if datos_evento.get("fecha_limite"):
                try:
                    date_entry.set_date(datos_evento["fecha_limite"])
                except Exception:
                    date_entry.delete(0, "end")
            else:
                date_entry.delete(0, "end")
            self.cronograma_entries[evento] = date_entry
            self.cronograma_estados_vars[evento] = estado_var
        docs_lf = ttk.LabelFrame(parent_frame, text="Documentos del Proceso", padding="10")
        docs_lf.pack(fill=tk.BOTH, expand=False, pady=10)
        btns = ttk.Frame(docs_lf); btns.pack(fill=tk.X)
        row1 = ttk.Frame(btns); row1.pack(fill=tk.X, pady=(0, 6))
        row2 = ttk.Frame(btns); row2.pack(fill=tk.X)
        ttk.Button(row1, text="👀 Ver checklist (visor)…",
                command=self.abrir_visor_docs, style="Small.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(row1, text="🛠️ Gestionar Documentos…",
                command=self.abrir_gestion_docs, style="Small.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(row1, text="🤖 Analizar Pliego con IA…",
                command=self._analizar_pliego_con_ia, style="Small.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(row1, text="🔀 Ordenar Docs (guardar)",
                command=self._ui_ordenar_docs_guardar, style="Small.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(row2, text="📄 Generar Expediente (PDF)",
                command=self._ui_generar_expediente_pdf, style="Small.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(row2, text="🗃️ Generar ZIP por Categoría",
                command=self._ui_generar_expediente_zip, style="Small.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(row2, text="🗂️ Abrir carpeta",
                command=self._ui_abrir_carpeta_destino, style="Small.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(row2, text="🔎 Validar faltantes",
                command=self.validar_faltantes, style="Small.TButton").pack(side=tk.LEFT, padx=5)
        self.actualizar_info_docs()
        self._on_estado_change()

    # =======================
    # PESTAÑA: LOTES (con llamada a reconstruir radios cuando cambian)
    # =======================
    def crear_widgets_lotes(self, parent_frame):
        tree_frame = ttk.Frame(parent_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        # ➕ añadimos "empresa_nuestra"
        cols = ("participar", "fase_a", "numero", "nombre", "monto_base",
                "monto_personal", "monto_ofertado", "dif_lic", "dif_pers", "empresa_nuestra")
        self.tree_lotes = ttk.Treeview(tree_frame, columns=cols, show="headings")

        headings = {
            "participar": "Participar",
            "fase_a": "Fase A OK",
            "numero": "N°",
            "nombre": "Nombre Lote",
            "monto_base": "Base Licitación",
            "monto_personal": "Base Personal",
            "monto_ofertado": "Nuestra Oferta",
            "dif_lic": "% Dif. Licit.",
            "dif_pers": "% Dif. Pers.",
            "empresa_nuestra": "Nuestra Empresa",
        }
        for col, text in headings.items():
            self.tree_lotes.heading(col, text=text)

        for col in ["participar", "fase_a", "numero", "dif_lic", "dif_pers"]:
            self.tree_lotes.column(col, width=80, anchor=tk.CENTER)
        self.tree_lotes.column("nombre", width=250)
        for col in ["monto_base", "monto_personal", "monto_ofertado"]:
            self.tree_lotes.column(col, anchor=tk.E, width=120)
        # tamaño razonable para la nueva columna
        self.tree_lotes.column("empresa_nuestra", width=180, anchor=tk.W)

        self.tree_lotes.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree_lotes.yview)
        # ✅ corrección: yscrollcommand (no 'yscroll')
        self.tree_lotes.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # clicks existentes
        self.tree_lotes.bind("<Button-1>", self._on_lote_cell_click)
        self.tree_lotes.tag_configure('descalificado', foreground='red')
        self.tree_lotes.bind("<Button-3>", self._mostrar_menu_lotes)

        # === Editor inline para columna "empresa_nuestra" con doble clic ===
        def _editar_empresa_inline(event):
            # ¿en qué celda se hizo doble clic?
            region = self.tree_lotes.identify("region", event.x, event.y)
            if region != "cell":
                return
            row_id = self.tree_lotes.identify_row(event.y)
            col_id = self.tree_lotes.identify_column(event.x)  # e.g. '#10'
            if not row_id or not col_id:
                return

            # ¿Es la columna empresa_nuestra?
            col_index = int(col_id.replace("#", "")) - 1
            if self.tree_lotes["columns"][col_index] != "empresa_nuestra":
                return

            # bbox de la celda para ubicar el combobox encima
            bbox = self.tree_lotes.bbox(row_id, col_id)
            if not bbox:
                return
            x, y, w, h = bbox

            # --- valores disponibles:
            # 1) primero, las empresas seleccionadas en la pestaña "Empresas Nuestras"
            # 2) si no hay, usa las empresas asignadas a la licitación (empresas_nuestras)
            # 3) si tampoco hay, usa el catálogo completo
            seleccionadas_panel = []
            if hasattr(self, "list_empresas") and self.list_empresas.winfo_exists():
                seleccionadas_panel = [self.list_empresas.get(i) for i in self.list_empresas.curselection()]

            seleccionadas_licit = [str(e) for e in getattr(self.licitacion, "empresas_nuestras", [])]

            if seleccionadas_panel:
                empresas_disp = seleccionadas_panel
            elif seleccionadas_licit:
                empresas_disp = seleccionadas_licit
            else:
                try:
                    empresas_disp = [e['nombre'] for e in self.db.get_empresas_maestras()]
                except Exception:
                    empresas_disp = [e.get('nombre','') for e in getattr(self.parent_app, 'empresas_registradas', []) if e.get('nombre')]

            # valor actual del lote
            valor_actual = self.tree_lotes.set(row_id, "empresa_nuestra")

            # editor
            combo = ttk.Combobox(self.tree_lotes, values=empresas_disp, state="readonly")
            combo.place(x=x, y=y, width=w, height=h)
            combo.set(valor_actual or "")

            def _guardar_y_cerrar():
                nuevo = combo.get().strip() or ""
                try:
                    combo.destroy()
                except Exception:
                    pass

                # Actualizar modelo
                # buscamos el objeto Lote asociado a la fila (tomamos 'numero' desde los values)
                try:
                    values = self.tree_lotes.item(row_id, "values")
                    num_lote = str(values[2])  # columna "numero" es la 3ra en 'cols'
                except Exception:
                    num_lote = None

                if num_lote is not None:
                    for l in getattr(self.licitacion, "lotes", []):
                        if str(l.numero) == str(num_lote):
                            l.empresa_nuestra = nuevo or None
                            break

                # Actualizar vista
                self.tree_lotes.set(row_id, "empresa_nuestra", nuevo)
                # reconstruir radios de ganadores para reflejar la empresa por lote
                try:
                    self._rebuild_ganadores_ui()
                except Exception:
                    pass

            # Guardar tanto al seleccionar como al salir con Enter o perder foco
            combo.bind("<<ComboboxSelected>>", lambda e: _guardar_y_cerrar())
            combo.bind("<Return>", lambda e: _guardar_y_cerrar())
            combo.bind("<FocusOut>", lambda e: _guardar_y_cerrar())
            combo.focus_set()

        self.tree_lotes.bind("<Double-1>", _editar_empresa_inline)

        # Botonera
        btn_frame = ttk.Frame(parent_frame)
        btn_frame.pack(fill=tk.X, pady=10)
        ttk.Button(btn_frame, text="➕ Agregar Lote", command=self.agregar_lote).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="✏️ Editar Lote", command=self.editar_lote).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="🗑️ Eliminar Lote", command=self.eliminar_lote).pack(side=tk.LEFT, padx=5)

        self.actualizar_tree_lotes()



    def actualizar_tree_lotes(self):
        """Rellena la tabla de lotes con los datos actuales de la licitación."""
        self.tree_lotes.delete(*self.tree_lotes.get_children())

        for l in getattr(self.licitacion, "lotes", []):
            # cálculo diferencias
            dif_lic_str = "N/D"
            dif_pers_str = "N/D"

            try:
                if l.monto_base and l.monto_ofertado:
                    dif = ((l.monto_base - l.monto_ofertado) / l.monto_base) * 100
                    dif_lic_str = f"{dif:.2f}%"
            except Exception:
                pass

            try:
                if l.monto_base_personal and l.monto_ofertado:
                    dif = ((l.monto_base_personal - l.monto_ofertado) / l.monto_base_personal) * 100
                    dif_pers_str = f"{dif:.2f}%"
            except Exception:
                pass

            values = (
                "✅" if getattr(l, "participamos", False) else "—",
                "✅" if getattr(l, "fase_A_superada", False) else "—",
                l.numero,
                l.nombre,
                f"RD$ {getattr(l, 'monto_base', 0):,.2f}",
                f"RD$ {getattr(l, 'monto_base_personal', 0):,.2f}",
                f"RD$ {getattr(l, 'monto_ofertado', 0):,.2f}",
                dif_lic_str,
                dif_pers_str,
                getattr(l, "empresa_nuestra", "") or ""  # 👈 NUEVA COLUMNA
            )

            self.tree_lotes.insert("", tk.END, iid=str(l.numero), values=values)


            
    def _on_lote_cell_click(self, event):
        """Toggle de Participar / Fase A OK al hacer click en su celda."""
        region = self.tree_lotes.identify("region", event.x, event.y)
        if region != "cell":
            return
        row_id = self.tree_lotes.identify_row(event.y)
        col_id = self.tree_lotes.identify_column(event.x)  # '#1', '#2', ...
        if not row_id or not col_id:
            return

        col_index = int(col_id.replace("#", "")) - 1
        col_name = self.tree_lotes["columns"][col_index]

        # Solo actuamos en estas dos columnas
        if col_name not in ("participar", "fase_a"):
            return

        # Obtener número de lote para mapear al objeto
        values = self.tree_lotes.item(row_id, "values")
        num_lote = str(values[2])  # 'numero' es la 3ra col

        # Buscar objeto lote
        lote = None
        for l in getattr(self.licitacion, "lotes", []):
            if str(l.numero) == num_lote:
                lote = l
                break
        if lote is None:
            return

        if col_name == "participar":
            nuevo = not bool(getattr(lote, "participamos", False))
            lote.participamos = nuevo
            # si marcamos participar y no hay empresa, deja sin empresa (puedes forzar a elegir luego)
            self.tree_lotes.set(row_id, "participar", "✅" if nuevo else "—")

        elif col_name == "fase_a":
            nuevo = not bool(getattr(lote, "fase_A_superada", False))
            lote.fase_A_superada = nuevo
            self.tree_lotes.set(row_id, "fase_a", "✅" if nuevo else "—")


# En la clase VentanaDetalles

    def agregar_lote(self):
        # Creamos una lista de nombres de las empresas participantes
        nombres_empresas = [str(e) for e in self.licitacion.empresas_nuestras]
        dialogo = DialogoGestionarLote(self, participating_companies=nombres_empresas)
        if dialogo.result:
            self.licitacion.lotes.append(dialogo.result)
            self.actualizar_tree_lotes()

    def editar_lote(self):
        selected_item = self.tree_lotes.focus()
        if not selected_item:
            messagebox.showwarning("Sin Selección", "Selecciona un lote.", parent=self)
            return
        
        lote_a_editar = next((l for l in self.licitacion.lotes if str(l.numero) == selected_item), None)
        if not lote_a_editar: return

        nombres_empresas = [str(e) for e in self.licitacion.empresas_nuestras]
        dialogo = DialogoGestionarLote(self, initial_data=lote_a_editar, participating_companies=nombres_empresas)
        if dialogo.result:
            # Reemplazar el lote en la lista
            for i, l in enumerate(self.licitacion.lotes):
                if str(l.numero) == str(lote_a_editar.numero):
                    self.licitacion.lotes[i] = dialogo.result
                    break
            self.actualizar_tree_lotes()

    def eliminar_lote(self):
        selected_item = self.tree_lotes.focus()
        if not selected_item:
            messagebox.showwarning("Sin Selección", "Selecciona un lote.", parent=self); return

        if messagebox.askyesno("Confirmar", "¿Eliminar lote?", parent=self):
            num_sel = str(selected_item)
            self.licitacion.lotes = [l for l in self.licitacion.lotes if str(l.numero) != num_sel]
            self.actualizar_tree_lotes()

    # =======================
    # PESTAÑA: COMPETIDORES + Ofertas + (NUEVO) Ganadores por Lote
    # =======================
    def crear_widgets_oferentes(self, parent_frame):
        # Panel principal izquierda/derecha
        main_pane = ttk.PanedWindow(parent_frame, orient=tk.HORIZONTAL)
        main_pane.pack(fill=tk.BOTH, expand=True)

        # ==== IZQUIERDA: COMPETIDORES ====
        self.left_frame = ttk.LabelFrame(main_pane, text="Competidores", padding=10)
        main_pane.add(self.left_frame, weight=1)

        self.tree_competidores = ttk.Treeview(self.left_frame, columns=("nombre",), show="headings", height=10)
        self.tree_competidores.heading("nombre", text="Nombre")
        self.tree_competidores.pack(fill=tk.BOTH, expand=True)
        self.tree_competidores.bind("<<TreeviewSelect>>", self._on_competidor_select)

        btn_comp_frame = ttk.Frame(self.left_frame)
        btn_comp_frame.pack(fill=tk.X, pady=(5, 0))
        ttk.Button(btn_comp_frame, text="Agregar Manual", command=self._agregar_competidor).pack(side=tk.LEFT, expand=True, padx=2)
        ttk.Button(btn_comp_frame, text="Agregar desde Catálogo...", command=self._agregar_desde_lista).pack(side=tk.LEFT, expand=True, padx=2)
        ttk.Button(btn_comp_frame, text="📥 Importar...", command=self._importar_competidores).pack(side=tk.LEFT, expand=True, padx=2)
        self.btn_edit_comp = ttk.Button(btn_comp_frame, text="Editar", command=self._editar_competidor, state="disabled")
        self.btn_edit_comp.pack(side=tk.LEFT, expand=True, padx=2)
        self.btn_del_comp = ttk.Button(btn_comp_frame, text="Eliminar", command=self._eliminar_competidor, state="disabled")
        self.btn_del_comp.pack(side=tk.LEFT, expand=True, padx=2)
        ttk.Button(btn_comp_frame, text="📊 Analizar Paquetes...", command=self._abrir_analisis_paquetes).pack(side=tk.LEFT, expand=True, padx=2)
        ttk.Button(btn_comp_frame, text="🔍 Analizar Fallas Fase A...", command=self._abrir_analisis_fase_a).pack(side=tk.LEFT, expand=True, padx=2)

        # donde haces:
        right_pane = ttk.PanedWindow(main_pane, orient=tk.VERTICAL)
        main_pane.add(right_pane, weight=3)
        self.right_pane = right_pane  # <--- añade esto

        right_top = ttk.LabelFrame(right_pane, text="Ofertas por Lote", padding=10)
        right_pane.add(right_top, weight=1)
        self.right_top = right_top  # <--- y esto

# En crear_widgets_oferentes, reemplaza la definición de self.tree_ofertas y sus encabezados

        self.tree_ofertas = ttk.Treeview(
            right_top,
            columns=("lote", "nombre_lote", "monto", "adjudicada"),
            show="headings",
            height=10
        )
        self.tree_ofertas.heading("lote", text="Lote")
        self.tree_ofertas.heading("nombre_lote", text="Nombre de Lote")
        self.tree_ofertas.heading("monto", text="Monto Ofertado")
        self.tree_ofertas.heading("adjudicada", text="Adjudicada")

        self.tree_ofertas.column("lote", width=80, anchor=tk.CENTER)
        self.tree_ofertas.column("nombre_lote", width=250, anchor=tk.W)
        self.tree_ofertas.column("monto", width=140, anchor=tk.E)
        self.tree_ofertas.column("adjudicada", width=100, anchor=tk.CENTER)

        self.tree_ofertas.tag_configure('ganador', background='#d4edda', font=('Helvetica', 9, 'bold'))

        # grid para permitir que el tree se expanda con el frame
        right_top.rowconfigure(0, weight=1)
        right_top.columnconfigure(0, weight=1)

        self.tree_ofertas.grid(row=0, column=0, sticky="nsew")

        # scrollbars
        sy = ttk.Scrollbar(right_top, orient="vertical",   command=self.tree_ofertas.yview)
        sx = ttk.Scrollbar(right_top, orient="horizontal", command=self.tree_ofertas.xview)
        self.tree_ofertas.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)
        sy.grid(row=0, column=1, sticky="ns")
        sx.grid(row=1, column=0, sticky="ew")

        self.tree_ofertas.bind("<Double-1>", lambda e: self._editar_oferta())

        btn_oferta_frame = ttk.Frame(right_top)
        btn_oferta_frame.grid(row=2, column=0, columnspan=2, sticky="ew", pady=(5, 0))
        self.btn_add_oferta  = ttk.Button(btn_oferta_frame, text="Agregar Oferta", command=self._agregar_oferta, state="disabled")
        self.btn_edit_oferta = ttk.Button(btn_oferta_frame, text="Editar Oferta",  command=self._editar_oferta, state="disabled")
        self.btn_del_oferta  = ttk.Button(btn_oferta_frame, text="Eliminar Oferta", command=self._eliminar_oferta, state="disabled")
        for b in (self.btn_add_oferta, self.btn_edit_oferta, self.btn_del_oferta):
            b.pack(side=tk.LEFT, expand=True, padx=2)
                # Ganadores por lote
        right_bottom = ttk.LabelFrame(right_pane, text="Ganadores por Lote (un ganador por cada lote)", padding=10)
        right_pane.add(right_bottom, weight=1)

        # Contenedor donde _rebuild_ganadores_ui colocará las filas por lote
        self.frame_ganadores = ttk.Frame(right_bottom)
        self.frame_ganadores.pack(fill="x", padx=6, pady=6)

        # Construir UI de ganadores (muestra TODOS los lotes y opciones)
        self._rebuild_ganadores_ui()
        self._actualizar_tree_competidores()
        # mostrar ofertas del primer competidor si quieres
        try:
            first = next(iter(self.licitacion.oferentes_participantes), None)
            self._actualizar_tree_ofertas(first)
        except Exception:
            pass
        # Botón para guardar ganadores
    def _guardar_ganadores(self):
        try:
            if not getattr(self, "var_ganador_por_lote", None):
                return

            for key, var in self.var_ganador_por_lote.items():
                val = var.get()

                # sin ganador → borrar fila
                if not val or val == "__NINGUNO__":
                    if self.licitacion.id is not None and hasattr(self.db, "borrar_ganador_lote"):
                        self.db.borrar_ganador_lote(self.licitacion.id, key)
                    # espejo en memoria
                    for l in getattr(self.licitacion, "lotes", []):
                        if str(l.numero) == str(key):
                            l.ganador_nombre = ""
                            l.ganado_por_nosotros = False
                            break
                    continue

                # ganador seleccionado
                if val.startswith("__NUESTRA__::"):
                    ganador = val.split("::", 1)[1]
                    empresa_nuestra = ganador
                else:
                    ganador = val
                    empresa_nuestra = None

                if self.licitacion.id is not None and hasattr(self.db, "marcar_ganador_lote"):
                    self.db.marcar_ganador_lote(self.licitacion.id, key, ganador, empresa_nuestra)

                # espejo en memoria
                for l in getattr(self.licitacion, "lotes", []):
                    if str(l.numero) == str(key):
                        l.ganador_nombre = ganador
                        l.ganado_por_nosotros = (empresa_nuestra is not None and empresa_nuestra == ganador)
                        break

            messagebox.showinfo("Éxito", "Ganadores guardados.", parent=self)

            # refrescar resúmenes/tabla si dejaste un callback
            if callable(getattr(self, "callback_actualizar", None)):
                self.callback_actualizar()

        except Exception as e:
            messagebox.showerror("Error al Guardar", f"Error guardando ganadores por lote:\n{e}", parent=self)



    def _actualizar_tree_competidores(self):
        """Rellena la tabla de competidores (panel izquierdo)."""
        if not hasattr(self, "tree_competidores"):
            return
        self.tree_competidores.delete(*self.tree_competidores.get_children())
        oferentes = getattr(self.licitacion, "oferentes_participantes", []) or []

        def _nombre(o):
            return (o.get("nombre") if isinstance(o, dict) else getattr(o, "nombre", "")) or ""

        for o in sorted(oferentes, key=_nombre):
            nom = _nombre(o)
            if nom:
                # opcional: iid=nom para facilitar búsquedas, pero no es obligatorio
                self.tree_competidores.insert("", "end", iid=nom, values=(nom,))

# En la clase VentanaDetalles, reemplaza este método

    def _actualizar_tree_ofertas(self, competidor=None):
        """Rellena la tabla 'Ofertas por Lote' para el competidor seleccionado."""
        if not hasattr(self, "tree_ofertas"):
            return

        self.tree_ofertas.delete(*self.tree_ofertas.get_children())

        if not competidor:
            return

        ofertas = getattr(competidor, "ofertas_por_lote", []) or []

        for of in sorted(ofertas, key=lambda o: str(o.get("lote_numero", ""))):
            lote_num = str(of.get("lote_numero", ""))
            monto = of.get("monto", 0.0) or 0.0
            
            # Buscar el objeto Lote para obtener su nombre
            lote_obj = next((l for l in self.licitacion.lotes if str(l.numero) == lote_num), None)
            nombre_lote = lote_obj.nombre if lote_obj else "N/D"

            # Verificar si esta oferta fue la ganadora
            es_ganador = bool(of.get("ganador", False))
            adjudicada_str = "Sí" if es_ganador else "No"
            
            tags = ('ganador',) if es_ganador else ()

            self.tree_ofertas.insert(
                "", "end", iid=lote_num,
                values=(lote_num, nombre_lote, f"RD$ {monto:,.2f}", adjudicada_str),
                tags=tags
            )


    def exportar_a_calendario(self):
        """Exporta el cronograma de la licitación a un archivo .ics (iCalendar)."""
        # Intentar importar ICS aquí por si el módulo no está disponible al inicio
        try:
            from ics import Calendar, Event
        except Exception:
            messagebox.showerror(
                "Dependencia faltante",
                "No se encontró el paquete 'ics'. Instálalo con:\n\npip install ics",
                parent=self
            )
            return

        # Pedir ruta de guardado
        default_filename = f"Cronograma_{self.licitacion.numero_proceso}.ics"
        file_path = filedialog.asksaveasfilename(
            parent=self,
            title="Guardar Cronograma como iCalendar",
            initialfile=default_filename,
            filetypes=[("iCalendar files", "*.ics")],
            defaultextension=".ics"
        )
        if not file_path:
            return

        # Construir calendario
        cal = Calendar()
        eventos_agregados = 0
        for nombre_evento, datos in (self.licitacion.cronograma or {}).items():
            fecha_str = (datos or {}).get("fecha_limite")
            if not fecha_str:
                continue
            try:
                # Interpretar YYYY-MM-DD
                import datetime
                fecha = datetime.datetime.strptime(fecha_str, "%Y-%m-%d").date()
                ev = Event(name=f"{self.licitacion.nombre_proceso}: {nombre_evento}", begin=fecha)
                ev.make_all_day()
                cal.events.add(ev)
                eventos_agregados += 1
            except Exception:
                # Ignorar fechas inválidas
                continue

        if eventos_agregados == 0:
            messagebox.showwarning("Sin Fechas", "No se encontraron fechas válidas para exportar.", parent=self)
            return

        # Guardar archivo .ics
        try:
            with open(file_path, "w", encoding="utf-8") as f:
                f.write(str(cal))
            messagebox.showinfo(
                "Éxito",
                f"Se exportaron {eventos_agregados} evento(s) al calendario.",
                parent=self
            )
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar el archivo:\n{e}", parent=self)




    def _on_competidor_select(self, event):
        competidor = self._get_selected_competidor()
        self._actualizar_tree_ofertas(competidor)
        state = "normal" if competidor else "disabled"
        for btn in [self.btn_edit_comp, self.btn_del_comp, self.btn_add_oferta]: btn.config(state=state)
        self.tree_ofertas.bind("<<TreeviewSelect>>", self._on_oferta_select)
        self._on_oferta_select(None)

    def _on_oferta_select(self, event):
        state = "normal" if self.tree_ofertas.selection() else "disabled"
        for btn in [self.btn_edit_oferta, self.btn_del_oferta]: btn.config(state=state)

    def _get_selected_competidor(self):
        sel = self.tree_competidores.selection()
        if not sel:
            return None
        nombre = self.tree_competidores.item(sel[0], "values")[0]
        for o in getattr(self.licitacion, "oferentes_participantes", []) or []:
            if (isinstance(o, dict) and o.get("nombre") == nombre) or (getattr(o, "nombre", None) == nombre):
                return o
        return None

    def _agregar_competidor(self):
        dialogo = DialogoGestionarOferente(self, title="Agregar Competidor")
        if dialogo and dialogo.result:
            if any(o.nombre.lower() == dialogo.result['nombre'].lower() for o in self.licitacion.oferentes_participantes):
                messagebox.showerror("Error", "Ya existe un competidor con ese nombre.", parent=self); return
            self.licitacion.oferentes_participantes.append(Oferente(**dialogo.result))
            self._actualizar_tree_competidores()

    def _editar_competidor(self):
        competidor = self._get_selected_competidor()
        if not competidor: return
        dialogo = DialogoGestionarOferente(self, title="Editar Competidor", initial_data=competidor)
        if dialogo and dialogo.result:
            competidor.nombre, competidor.comentario = dialogo.result["nombre"], dialogo.result["comentario"]
            self._actualizar_tree_competidores()

    def _importar_competidores(self):
        dialogo = DialogoSeleccionarLicitacion(
            self,
            "Seleccionar Licitación de Origen para Importar Competidores",
            self.parent_app.gestor_licitaciones,
            self.licitacion.numero_proceso
        )
        if not dialogo.result:
            return
        lic_origen = next((l for l in self.parent_app.gestor_licitaciones if l.numero_proceso == dialogo.result), None)
        if not lic_origen:
            messagebox.showerror("Error", "No se pudo encontrar la licitación de origen.", parent=self)
            return
        nombres_existentes = {o.nombre.lower() for o in self.licitacion.oferentes_participantes}
        nuevos_agregados = 0
        for competidor_origen in lic_origen.oferentes_participantes:
            if competidor_origen.nombre.lower() not in nombres_existentes:
                nuevo_oferente = Oferente(
                    nombre=competidor_origen.nombre,
                    comentario=competidor_origen.comentario
                )
                self.licitacion.oferentes_participantes.append(nuevo_oferente)
                nombres_existentes.add(nuevo_oferente.nombre.lower())
                nuevos_agregados += 1
        if nuevos_agregados > 0:
            self._actualizar_tree_competidores()
            messagebox.showinfo("Éxito", f"Se importaron {nuevos_agregados} nuevos competidores.", parent=self)
        else:
            messagebox.showinfo("Información", "No se encontraron competidores nuevos para importar.", parent=self)

    def _eliminar_competidor(self):
        competidor = self._get_selected_competidor()
        if not competidor: return
        if messagebox.askyesno("Confirmar Eliminación", f"¿Está seguro de eliminar a '{competidor.nombre}' de ESTA licitación?\n\n(No se borra del catálogo maestro)", parent=self):
            self.licitacion.oferentes_participantes.remove(competidor)
            self._actualizar_tree_competidores()

    def _agregar_oferta(self):
        competidor = self._get_selected_competidor()
        if not competidor: return
        numeros_ofertados = {str(o['lote_numero']) for o in competidor.ofertas_por_lote}
        lotes_disponibles = [l for l in self.licitacion.lotes if str(l.numero) not in numeros_ofertados]
        if not lotes_disponibles: 
            messagebox.showinfo("Información", "No hay más lotes disponibles para este competidor.", parent=self); return
        dialogo = DialogoGestionarOfertaLote(self, "Agregar Oferta", lotes_disponibles)
        if dialogo and dialogo.result:
            competidor.ofertas_por_lote.append(dialogo.result)
            self._actualizar_tree_ofertas(competidor)

    def _editar_oferta(self):
        competidor = self._get_selected_competidor()
        if not competidor: return
        try:
            oferta_ref = next(o for o in competidor.ofertas_por_lote if str(o['lote_numero']) == str(self.tree_ofertas.selection()[0]))
            dialogo = DialogoGestionarOfertaLote(self, "Editar Oferta", self.licitacion.lotes, initial_data=oferta_ref)
            if dialogo and dialogo.result:
                oferta_ref.update(dialogo.result)
                self._actualizar_tree_ofertas(competidor)
        except (IndexError, StopIteration): return

    def _eliminar_oferta(self):
        competidor = self._get_selected_competidor()
        if not competidor: return
        try:
            sel_iid = self.tree_ofertas.selection()[0]
            if messagebox.askyesno("Confirmar", f"¿Eliminar la oferta para el lote {sel_iid}?", parent=self):
                competidor.ofertas_por_lote = [o for o in competidor.ofertas_por_lote if str(o['lote_numero']) != str(sel_iid)]
                self._actualizar_tree_ofertas(competidor)
        except IndexError: return

    # =======================
    # NUEVO: UI y lógica de "Ganadores por Lote"
    # =======================
    def _make_scrollframe(self, parent, height=280):
        cont = ttk.Frame(parent); cont.pack(fill="both", expand=True)
        canvas = tk.Canvas(cont, borderwidth=0, height=height)
        vs = ttk.Scrollbar(cont, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vs.set)
        canvas.pack(side="left", fill="both", expand=True)
        vs.pack(side="right", fill="y")

        body = ttk.Frame(canvas)
        body_id = canvas.create_window((0, 0), window=body, anchor="nw")

        def _on_conf(event=None):
            canvas.configure(scrollregion=canvas.bbox("all"))
            try:
                canvas.itemconfigure(body_id, width=canvas.winfo_width())
            except Exception:
                pass
        body.bind("<Configure>", _on_conf)

        def _bind_wheel():
            def _wheel(evt):
                try:
                    canvas.yview_scroll(int(-1 * (evt.delta / 120)), "units")
                except Exception:
                    pass
            canvas.bind_all("<MouseWheel>", _wheel)
            return lambda: canvas.unbind_all("<MouseWheel>")

        self._scroll_unbinder = _bind_wheel()
        self._last_scroll_container = cont
        return body, canvas

    def _destroy_last_scrollframe(self):
        try:
            if getattr(self, "_scroll_unbinder", None):
                self._scroll_unbinder()
                self._scroll_unbinder = None
            if getattr(self, "_last_scroll_container", None):
                self._last_scroll_container.destroy()
                self._last_scroll_container = None
        except Exception:
            pass
    


    def _rebuild_ganadores_ui(self):
        """Reconstruye el panel de ganadores (un combobox por lote)."""
        if not hasattr(self, "frame_ganadores"):
            return

        # 0) limpiar scroll anterior
        self._destroy_last_scrollframe()
        for w in self.frame_ganadores.winfo_children():
            w.destroy()

        # 1) crear scroll-frame
        body, _ = self._make_scrollframe(self.frame_ganadores, height=300)

        # 2) variables por lote
        self.var_ganador_por_lote = {}

        # 3) mapa: lote -> set(nombres oferentes que ofertaron ese lote)
        oferentes_por_lote = {}
        for ofer in getattr(self.licitacion, "oferentes_participantes", []) or []:
            nombre = (ofer.get("nombre") if isinstance(ofer, dict) else getattr(ofer, "nombre", "")) or ""
            ofertas = (ofer.get("ofertas_por_lote", [])
                    if isinstance(ofer, dict)
                    else getattr(ofer, "ofertas_por_lote", [])) or []
            for of in ofertas:
                key = str(of.get("lote_numero", "")).strip()
                if key:
                    oferentes_por_lote.setdefault(key, set()).add(nombre)

        # 4) lotes ordenados
        lotes = sorted(
            getattr(self.licitacion, "lotes", []) or [],
            key=lambda l: int(str(l.numero)) if str(l.numero).isdigit() else str(l.numero)
        )
        body.columnconfigure(0, weight=1)

        for idx, lote in enumerate(lotes):
            key = str(lote.numero)
            self.var_ganador_por_lote[key] = tk.StringVar(value="__NINGUNO__")

            fila = ttk.Frame(body); fila.grid(row=idx, column=0, sticky="ew", pady=2)
            ttk.Label(fila, text=f"Lote {lote.numero}", width=10).pack(side="left", padx=(4, 8))

            # Opciones visibles
            opciones = [("__NINGUNO__", "— Sin ganador —")]
            for comp_name in sorted(oferentes_por_lote.get(key, [])):
                opciones.append((comp_name, comp_name))

            emp_lote = (getattr(lote, "empresa_nuestra", None) or "").strip()
            if emp_lote:
                opciones.append((f"__NUESTRA__::{emp_lote}", f"{emp_lote} (Nosotros)"))

            internos = [v for v, _ in opciones]
            visibles  = [t for _, t in opciones]
            combo = ttk.Combobox(fila, values=visibles, state="readonly")
            combo.pack(side="left", fill="x", expand=True)

            # Preselección desde datos del lote
            pre = "__NINGUNO__"
            g = (getattr(lote, "ganador_nombre", "") or "").strip()
            if g:
                pre = f"__NUESTRA__::{emp_lote}" if (emp_lote and g == emp_lote) else g
            self.var_ganador_por_lote[key].set(pre)
            try:
                combo.current(internos.index(pre))
            except ValueError:
                combo.current(0)

            def _on_select(evt, _key=key, _internos=internos, _combo=combo):
                sel = _combo.get()
                try:
                    idx_sel = list(_combo["values"]).index(sel)
                except ValueError:
                    idx_sel = 0
                self.var_ganador_por_lote[_key].set(_internos[idx_sel])

            combo.bind("<<ComboboxSelected>>", _on_select)


    def _ajustar_altura_panel_ganadores(self):
        """Asegura que el paned vertical muestre el frame inferior (ganadores)."""
        try:
            # right_pane es el PanedWindow vertical que creamos en crear_widgets_oferentes
            if hasattr(self, "right_pane") and self.right_pane is not None:
                self.right_pane.update_idletasks()
                h = max(self.right_pane.winfo_height(), 200)
                # mueve el sash para que el panel de abajo tenga ~40% de altura
                self.right_pane.sashpos(0, int(h * 0.6))
                # además, fija un mínimo para que nunca se colapse del todo
                self.right_pane.paneconfigure(self.right_top, minsize=120)
                self.right_pane.paneconfigure(self.frame_ganadores, minsize=120)
        except Exception:
            pass



    def _preseleccionar_ganadores_desde_modelo(self):
        """Si en los datos ya tienes marcado ganador por lote, reflejarlo en los radios."""
        for lote in self.licitacion.lotes:
            key = str(lote.numero)
            # 1) Si tu modelo del lote ya guarda el ganador (por ejemplo, lote.ganador_nombre)
            ganador = getattr(lote, 'ganador_nombre', '') or ''
            # 2) O si alguno de los competidores marcó 'ganador': True en su oferta para ese lote
            if not ganador:
                for comp in self.licitacion.oferentes_participantes:
                    for o in comp.ofertas_por_lote:
                        if str(o.get('lote_numero')) == str(lote.numero) and o.get('ganador'):
                            ganador = comp.nombre
                            break
                    if ganador:
                        break
            # 3) O si nosotros tenemos una marca en el lote (por ejemplo, lote.ganado_por_nosotros)
            if not ganador and getattr(lote, 'ganado_por_nosotros', False):
                ganador = "__NOSOTROS__"

            if key in self.var_ganador_por_lote:
                self.var_ganador_por_lote[key].set(ganador if ganador else "")

    # =======================
    # ESTADO / GUARDAR
    # =======================
    def _on_estado_change(self, event=None):
        # Ya no existe "Adjudicada a:", así que no hay nada que habilitar/deshabilitar
        pass

# En la clase VentanaDetalles, reemplaza este método

    def guardar_y_cerrar(self):
        if self._guardar_cambios():
            self.destroy()

# Pega esta nueva función DENTRO de la clase VentanaDetalles


# Pega esta nueva función DENTRO de la clase VentanaDetalles

# En la clase VentanaDetalles, reemplaza este método

    def _guardar_sin_cerrar(self):
        if self._guardar_cambios():
            # Feedback visual de que se guardó
            self.btn_guardar_continuar.config(text="¡Guardado!", state="disabled")
            self.after(1500, lambda: self.btn_guardar_continuar.config(text="💾 Guardar y Continuar", state="normal"))

            # --- LÓGICA DE REFRESCO AUTOMÁTICO ---
            competidor_seleccionado = self._get_selected_competidor()
            # La función _actualizar_tree_ofertas redibujará la tabla con los datos actualizados en memoria.
            self._actualizar_tree_ofertas(competidor_seleccionado)
            # --- FIN DE LA LÓGICA DE REFRESCO ---


    def _ensure_estructuras(self):
        if not isinstance(getattr(self.licitacion, "lotes", None), list):
            self.licitacion.lotes = []
        if not isinstance(getattr(self.licitacion, "oferentes_participantes", None), list):
            self.licitacion.oferentes_participantes = []
        if not isinstance(getattr(self.licitacion, "cronograma", None), dict):
            self.licitacion.cronograma = {}




    def _guardar_cambios(self):
        """Lógica central de guardado, devuelve True si tuvo éxito."""
        try:
            # ===== 1) LEER PESTAÑA: DETALLES GENERALES =====
            self.licitacion.numero_proceso = (self.var_codigo.get() or "").strip()
            self.licitacion.nombre_proceso = (self.var_nombre.get() or "").strip()
            self.licitacion.institucion    = (self.var_institucion.get() or "").strip()

            # Estado / Adjudicación
            self.licitacion.estado       = (self.estado_var.get() or "").strip()
            self.licitacion.adjudicada   = (self.licitacion.estado == "Adjudicada")
            self.licitacion.adjudicada_a = (self.adjudicado_a_var.get() or "").strip() if self.licitacion.adjudicada else ""

            # Fase B + comentarios
            self.licitacion.fase_B_superada = bool(self.fase_b_var.get())
            self.licitacion.motivo_descalificacion = (self.motivo_text.get("1.0", tk.END) or "").strip()

            # Cronograma (asegura diccionario)
            if not isinstance(getattr(self.licitacion, "cronograma", None), dict):
                self.licitacion.cronograma = {}
            for evento, date_entry in (self.cronograma_entries or {}).items():
                self.licitacion.cronograma[evento] = {
                    "fecha_limite": (date_entry.get() or None),
                    "estado": self.cronograma_estados_vars[evento].get()
                }

            # ===== 2) APLICAR GANADORES A MODELO =====
            self._aplicar_ganadores_por_lote_al_modelo()

            # Si está Adjudicada y no hay "adjudicada_a", intenta inferir de lotes nuestros
            if self.licitacion.adjudicada and not self.licitacion.adjudicada_a:
                empresa_detectada = None
                for lote in getattr(self.licitacion, "lotes", []):
                    if getattr(lote, "ganado_por_nosotros", False) and getattr(lote, "empresa_nuestra", None):
                        empresa_detectada = lote.empresa_nuestra
                        break
                if not empresa_detectada and getattr(self.licitacion, "empresas_nuestras", None):
                    empresa_detectada = str(self.licitacion.empresas_nuestras[0])
                self.licitacion.adjudicada_a = (empresa_detectada or self.licitacion.adjudicada_a or "").strip()

            # ===== 3) GUARDAR LICITACIÓN (con reintento suave) =====
            try:
                self.db.save_licitacion(self.licitacion)
            except Exception as e:
                msg = str(e).lower()
                if ("concurrencia" in msg) or ("last_modified" in msg) or ("modificado por otro" in msg):
                    # refrescamos timestamp y reintentamos una vez
                    if hasattr(self.db, "get_last_modified") and getattr(self.licitacion, "id", None):
                        try:
                            self.licitacion.last_modified = self.db.get_last_modified(self.licitacion.id)
                        except Exception:
                            pass
                    self.db.save_licitacion(self.licitacion)
                else:
                    raise

            # ===== 4) GUARDAR GANADORES POR LOTE =====
            self._persistir_ganadores_por_lote_si_posible()

            # ===== 5) REFRESCAR PANTALLA / LISTADOS =====
            if callable(self.callback_actualizar):
                self.callback_actualizar()

            return True

        except Exception as e:
            messagebox.showerror(
                "Error al Guardar",
                f"Ocurrió un error guardando los cambios:\n{e}",
                parent=self
            )
            return False


    def _aplicar_ganadores_por_lote_al_modelo(self):
        """
        Marca en memoria el ganador por cada lote según self.var_ganador_por_lote.
        """
        # 0) Limpiar marcas previas
        for comp in getattr(self.licitacion, "oferentes_participantes", []):
            for o in getattr(comp, "ofertas_por_lote", []):
                o['ganador'] = False
        for lote in getattr(self.licitacion, "lotes", []):
            lote.ganador_nombre = ""
            lote.ganado_por_nosotros = False

        # 1) Aplicar selección actual
        for lote in getattr(self.licitacion, "lotes", []):
            key = str(lote.numero)
            var = self.var_ganador_por_lote.get(key)
            if not var:
                continue

            val = var.get()
            if not val or val == "__NINGUNO__":
                continue
            
            # Determinar si el ganador es nuestra empresa
            es_nuestro = val.startswith("__NUESTRA__::")
            
            if es_nuestro:
                ganador = val.split("::", 1)[1]
                lote.ganador_nombre = ganador
                lote.ganado_por_nosotros = True
            else:
                # El ganador es un competidor
                ganador = val
                lote.ganador_nombre = ganador
                lote.ganado_por_nosotros = False
                # Marcar la oferta específica del competidor como ganadora
                for comp in getattr(self.licitacion, "oferentes_participantes", []):
                    if comp.nombre == ganador:
                        for o in getattr(comp, "ofertas_por_lote", []):
                            if str(o.get("lote_numero")) == key:
                                o['ganador'] = True

    def _persistir_ganadores_por_lote_si_posible(self):
        """
        Persiste ganadores en BD, adaptando a las firmas existentes:
        - marcar_ganador_lote(licitacion_id, lote_num, ganador, empresa_nuestra:str|None)
        - save_ganadores_por_lote(licitacion_id, mapping) donde mapping puede ser:
            * [(lote_num, ganador, empresa_nuestra)]  ó
            * [(lote_num, ganador, es_nuestro_bool)]
        (intentamos una y si falla por TypeError, probamos la otra)
        """
        try:
            if not getattr(self.licitacion, "id", None):
                return

            # Construimos la lista de ganadores presentes en memoria
            # y la info de empresa_nuestra si aplica
            mapping_emp = []   # [(lote_num, ganador, empresa_nuestra)]
            mapping_bool = []  # [(lote_num, ganador, es_nuestro_bool)]

            for lote in getattr(self.licitacion, "lotes", []):
                nombre = (getattr(lote, 'ganador_nombre', '') or '').strip()
                if not nombre:
                    continue
                es_nuestro = bool(getattr(lote, 'ganado_por_nosotros', False))
                emp = (getattr(lote, 'empresa_nuestra', '') or '').strip()
                empresa_nuestra = emp if (es_nuestro and emp) else None

                mapping_emp.append((str(getattr(lote, 'numero', '')), nombre, empresa_nuestra))
                mapping_bool.append((str(getattr(lote, 'numero', '')), nombre, es_nuestro))

            # Si no hay ganadores, no borramos lo que haya en BD para evitar pérdidas
            if not mapping_emp:
                return

            # Prioridad: usar API por-lote si existe
            if hasattr(self.db, 'marcar_ganador_lote'):
                for lote_num, ganador, empresa_nuestra in mapping_emp:
                    # Firma esperada: (licitacion_id, lote_num, ganador, empresa_nuestra:str|None)
                    self.db.marcar_ganador_lote(self.licitacion.id, lote_num, ganador, empresa_nuestra)
                return

            # Si no hay API por-lote, probamos la masiva
            if hasattr(self.db, 'save_ganadores_por_lote'):
                # Intento #1: asumiendo que recibe empresa_nuestra
                try:
                    self.db.save_ganadores_por_lote(self.licitacion.id, mapping_emp)
                    return
                except TypeError:
                    pass
                # Intento #2: asumiendo que recibe es_nuestro_bool
                try:
                    self.db.save_ganadores_por_lote(self.licitacion.id, mapping_bool)
                    return
                except TypeError:
                    pass
            # Si no hay métodos, salimos silenciosamente
        except Exception as e:
            try:
                print("[WARN] No se pudo persistir ganadores por lote:", e)
            except Exception:
                pass



    # =======================
    # RESTO: visor/generar docs/riesgos (tu código original)
    # =======================
    def abrir_visor_docs(self):
        VentanaVisorDocumentos(
            self,
            self.licitacion,
            categorias=self.categorias_documentos,
            on_refresh=self.actualizar_info_docs
        )

    def abrir_gestion_docs(self):
        VentanaGestionDocumentos(
            self,
            self.licitacion,
            callback=self.actualizar_info_docs,
            documentos_maestros=self.documentos_maestros,
            categorias=self.categorias_documentos,
            todas_las_licitaciones=self.parent_app.gestor_licitaciones,
            lista_responsables=self.parent_app.responsables_maestros
        )

    def actualizar_info_docs(self):
        try:
            if hasattr(self, "progreso_docs_var"):
                self.progreso_docs_var.set(f"{self.licitacion.get_porcentaje_completado():.1f}%")
        except Exception:
            pass

    def _abrir_analisis_paquetes(self):
        VentanaAnalisisPaquetes(self, self.licitacion)

    def _ui_abrir_carpeta_destino(self):
        carpeta = os.path.join(os.getcwd(), "expedientes")
        os.makedirs(carpeta, exist_ok=True)
        try:
            if platform.system() == "Windows":
                os.startfile(carpeta)
            elif platform.system() == "Darwin":
                subprocess.call(["open", carpeta])
            else:
                subprocess.call(["xdg-open", carpeta])
        except Exception as e:
            messagebox.showwarning("Aviso", f"Abrir carpeta falló:\n{e}", parent=self)

    def _construir_items_por_defecto(self):
        """
        Crea la lista de items (orden, doc_version_id, titulo) para el expediente.
        Ordena por 'orden_pliego' si existe; si no, al final.
        """
        docs_obj = list(getattr(self.licitacion, "documentos_solicitados", []) or [])
        docs_obj.sort(key=lambda d: getattr(d, "orden_pliego", 999999))
        items = []
        for i, d in enumerate(docs_obj, start=1):
            doc_id = getattr(d, "id", None)
            titulo = f"[{getattr(d, 'codigo', '') or ''}] {getattr(d, 'nombre', '') or ''}".strip()
            if doc_id is not None:
                items.append({'orden': i, 'doc_version_id': doc_id, 'titulo': titulo})
        return items

    def _ui_ordenar_docs_guardar(self):
        docs_obj = list(getattr(self.licitacion, "documentos_solicitados", []) or [])
        if not docs_obj:
            messagebox.showwarning("Sin documentos", "Esta licitación no tiene documentos cargados.", parent=self)
            return

        dlg = DialogoOrdenExpediente(self, documentos_obj=docs_obj)
        self.wait_window(dlg)
        if dlg.result_incluir is None:
            return

        orden_por_cat = dlg.result_orden
        pares_docid_orden = []
        orden_global = 1
        for cat in CATS_ORDEN_EXPD:
            for d in orden_por_cat.get(cat, []):
                setattr(d, "orden_pliego", orden_global)
                doc_id = getattr(d, "id", None)
                if doc_id is not None:
                    pares_docid_orden.append((doc_id, orden_global))
                orden_global += 1

        try:
            ok = self.db.guardar_orden_documentos(self.licitacion.id, pares_docid_orden)
            if not ok:
                messagebox.showwarning("Aviso", "No se pudo guardar el orden en la base de datos.", parent=self)
        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error guardando el orden:\n{e}", parent=self)

        try:
            orden_map = {doc_id: ordn for (doc_id, ordn) in pares_docid_orden}
            self.licitacion.documentos_solicitados.sort(
                key=lambda d: orden_map.get(getattr(d, "id", None), getattr(d, "orden_pliego", 999999))
            )
        except Exception:
            pass

        try:
            self.actualizar_info_docs()
        except Exception:
            pass

        messagebox.showinfo("Listo", "Se guardó el nuevo orden de los documentos.", parent=self)

    def _ui_generar_expediente_pdf(self):
        docs_obj = list(getattr(self.licitacion, "documentos_solicitados", []) or [])
        if not docs_obj:
            messagebox.showwarning("Sin documentos", "Esta licitación no tiene documentos cargados.", parent=self)
            return

        dlg = DialogoOrdenExpediente(self, documentos_obj=docs_obj)
        self.wait_window(dlg)
        if dlg.result_incluir is None:
            return

        incluir = dlg.result_incluir
        orden_por_cat = dlg.result_orden

        pares_docid_orden = []
        orden_global = 1
        for cat in CATS_ORDEN_EXPD:
            if cat not in orden_por_cat:
                continue
            for d in orden_por_cat[cat]:
                setattr(d, "orden_pliego", orden_global)
                doc_id = getattr(d, "id", None)
                if doc_id is not None:
                    pares_docid_orden.append((doc_id, orden_global))
                orden_global += 1

        try:
            ok = self.db.guardar_orden_documentos(self.licitacion.id, pares_docid_orden)
            if not ok:
                print("[WARN] No se pudo guardar el orden en BD.")
        except Exception as e:
            print("[WARN] Excepción al guardar orden en BD:", e)

        try:
            self.actualizar_info_docs()
        except Exception:
            pass

        carpeta = filedialog.askdirectory(parent=self, title="Selecciona la carpeta de salida de los PDF")
        if not carpeta:
            return
        os.makedirs(carpeta, exist_ok=True)

        generados = []
        for cat in CATS_ORDEN_EXPD:
            if cat not in orden_por_cat:
                continue
            if not incluir.get(cat, False):
                continue
            docs_cat = [d for d in orden_por_cat[cat] if isinstance(d, object)]
            if not docs_cat:
                continue

            items = []
            for i, d in enumerate(docs_cat, start=1):
                doc_id = getattr(d, "id", None)
                if doc_id is None:
                    continue
                titulo = f"[{getattr(d, 'codigo', '') or ''}] {getattr(d, 'nombre', '') or ''}".strip()
                items.append({'orden': i, 'doc_version_id': doc_id, 'titulo': titulo})
            if not items:
                continue

            nombre_archivo = f"Expediente - {cat} - {self.licitacion.numero_proceso}.pdf"
            out_path = os.path.join(carpeta, nombre_archivo)
            meta = {
                'titulo_expediente': f"Expediente - {cat} - {self.licitacion.numero_proceso}",
                'creado_por': os.getenv("USERNAME") or os.getenv("USER") or "Usuario",
                'qr_text': f"{self.licitacion.numero_proceso} | {self.licitacion.institucion} | {cat}"
            }

            try:
                generar_expediente_pdf(self.db, self.licitacion, items, out_path, meta)
                generados.append(out_path)
            except Exception as e:
                logging.exception("Error generando expediente por categoría")
                messagebox.showerror("Error", f"No se pudo generar el PDF de {cat}:\n{e}", parent=self)

        if not generados:
            messagebox.showwarning("Sin salida", "No se generó ningún PDF (revisa selección/categorías).", parent=self)
            return

        msg = "✅ Se generaron los siguientes archivos:\n\n" + "\n".join(f"- {os.path.basename(p)}" for p in generados)
        messagebox.showinfo("Listo", msg, parent=self)
        try:
            previsualizar_expediente(generados[-1])
        except Exception:
            pass

    def _ui_generar_expediente_zip(self):
        """
        1) Abre la ventana para confirmar y reordenar por categoría.
        2) PERSISTE el orden (orden_pliego) en BD y en memoria.
        3) Pide carpeta destino.
        4) Genera un ZIP por cada categoría (Legal, Financiera, Técnica, Sobre B) marcada.
        """
        # 1) Confirmar y reordenar
        docs_obj = list(getattr(self.licitacion, "documentos_solicitados", []) or [])
        if not docs_obj:
            messagebox.showwarning("Sin documentos", "Esta licitación no tiene documentos cargados.", parent=self)
            return

        dlg = DialogoOrdenExpediente(self, documentos_obj=docs_obj)
        self.wait_window(dlg)
        if dlg.result_incluir is None:
            return  # canceló

        incluir = dlg.result_incluir       # dict {cat: bool}
        orden_por_cat = dlg.result_orden   # dict {cat: [objs Documento]}

        # 2) === PERSISTE EL ORDEN EN BD (y en memoria) ===
        pares_docid_orden = []
        orden_global = 1
        for cat in CATS_ORDEN_EXPD:
            if cat not in orden_por_cat:
                continue
            for d in orden_por_cat[cat]:
                # En memoria:
                setattr(d, "orden_pliego", orden_global)
                # A BD:
                doc_id = getattr(d, "id", None)
                if doc_id is not None:
                    pares_docid_orden.append((doc_id, orden_global))
                orden_global += 1

        try:
            ok = self.db.guardar_orden_documentos(self.licitacion.id, pares_docid_orden)
            if not ok:
                print("[WARN] No se pudo guardar el orden en BD.")
        except Exception as e:
            print("[WARN] Excepción al guardar orden en BD:", e)

        # Refrescar vistas si aplica
        try:
            self.actualizar_info_docs()  # o self.actualizar_listas_docs()
        except Exception:
            pass

        # 3) Carpeta destino
        carpeta = filedialog.askdirectory(parent=self, title="Selecciona la carpeta de salida de los ZIP")
        if not carpeta:
            return
        os.makedirs(carpeta, exist_ok=True)

        # 4) Generar un ZIP por categoría
        try:
            generados = generar_expediente_zip_por_categoria(
                self.db, self.licitacion, carpeta, orden_por_cat, incluir
            )
        except Exception as e:
            logging.exception("Error generando ZIP por categoría")
            messagebox.showerror("Error", f"No se pudieron generar los ZIP:\n{e}", parent=self)
            return

        if not generados:
            messagebox.showwarning("Sin salida", "No se generó ningún ZIP. Revisa selección/categorías.", parent=self)
            return

        # Aviso final + opción de abrir carpeta
        msg = "✅ Se generaron los siguientes ZIP:\n\n" + "\n".join(f"- {os.path.basename(p)}" for p in generados)
        messagebox.showinfo("Listo", msg, parent=self)

        try:
            if messagebox.askyesno("Abrir carpeta", "¿Deseas abrir la carpeta de salida?", parent=self):
                if platform.system() == "Windows":
                    os.startfile(carpeta)
                elif platform.system() == "Darwin":
                    subprocess.call(["open", carpeta])
                else:
                    subprocess.call(["xdg-open", carpeta])
        except Exception:
            pass


    def validar_faltantes(self):
        try:
            docs = list(self.licitacion.documentos_solicitados)
        except Exception:
            docs = []

        total = len(docs)
        con_archivo = 0
        faltan_list = []

        for d in docs:
            ruta_ok = isinstance(getattr(d, 'ruta_archivo', ''), str) and getattr(d, 'ruta_archivo', '').strip() != ""
            if ruta_ok:
                con_archivo += 1
            else:
                faltan_list.append(f"- [{getattr(d, 'codigo', '—')}] {getattr(d, 'nombre', 'Documento sin nombre')}")

        if total == 0:
            messagebox.showwarning(
                "Validación",
                "No hay documentos cargados en la lista.\n\nVe a “Gestionar Documentos...” para agregarlos.",
                parent=self
            )
            return

        if con_archivo == 0:
            messagebox.showwarning(
                "Validación",
                "Ningún documento tiene archivo asociado.\n\nAdjunta los archivos en “Gestionar Documentos...”.",
                parent=self
            )
            return

        if con_archivo < total:
            faltan = total - con_archivo
            faltantes_str = "\n".join(faltan_list[:15])
            extra = "" if len(faltan_list) <= 15 else f"\n... y {len(faltan_list)-15} más."
            messagebox.showwarning(
                "Validación",
                f"Faltan {faltan} documento(s) por adjuntar archivo:\n\n{faltantes_str}{extra}",
                parent=self
            )
            return

        messagebox.showinfo(
            "Validación",
            "Todos los documentos tienen archivo asociado.",
            parent=self
        )


class DialogoSeleccionarLicitacion(simpledialog.Dialog):
    # ... (sin cambios)
    def __init__(self, parent, title, todas_las_licitaciones, licitacion_actual_id):
        self.todas_las_licitaciones = todas_las_licitaciones
        self.licitacion_actual_id = licitacion_actual_id
        self.result = None
        super().__init__(parent, title)

    def body(self, master):
        self.geometry("850x500") 
        
        ttk.Label(master, text="Seleccione una licitación de la cual importar la lista de documentos:", wraplength=450).pack(padx=10, pady=10)
        
        tree_frame = ttk.Frame(master)
        tree_frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

        cols = ("proceso", "nombre", "institucion")
        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings", height=15)
        self.tree.heading("proceso", text="Código Proceso")
        self.tree.heading("nombre", text="Nombre del Proceso")
        self.tree.heading("institucion", text="Institución")
        
        self.tree.column("proceso", width=150, anchor=tk.W)
        self.tree.column("nombre", width=400, anchor=tk.W)
        self.tree.column("institucion", width=200, anchor=tk.W)

        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        licitaciones_ordenadas = sorted(self.todas_las_licitaciones, key=lambda l: (l.institucion, l.nombre_proceso))
        for lic in licitaciones_ordenadas:
            if lic.numero_proceso == self.licitacion_actual_id:
                continue
            
            self.tree.insert("", tk.END, iid=lic.numero_proceso, values=(
                lic.numero_proceso,
                lic.nombre_proceso,
                lic.institucion
            ))
            
        return self.tree

    def apply(self):
        try:
            self.result = self.tree.selection()[0]
        except IndexError:
            self.result = None


# Pega estas dos nuevas clases en gestor_licitaciones_db.py

# En gestor_licitaciones_db.py, reemplaza esta clase por completo

# Pega esta NUEVA clase en tu archivo

# REEMPLAZA esta clase por completo

class DialogoSeleccionarNuestrasEmpresas(simpledialog.Dialog):
    """Un diálogo moderno con checkboxes para seleccionar múltiples empresas nuestras."""
    def __init__(self, parent, todas_las_empresas, seleccion_actual):
        self.todas_las_empresas = sorted(todas_las_empresas, key=lambda x: x['nombre'])
        self.nombres_seleccionados_inicial = {str(e) for e in seleccion_actual}
        self.selection_status = {emp['nombre']: (emp['nombre'] in self.nombres_seleccionados_inicial) for emp in self.todas_las_empresas}
        super().__init__(parent, "Seleccionar Empresas Participantes")

    def body(self, master):
        self.geometry("600x450")
        
        # --- Búsqueda ---
        search_frame = ttk.Frame(master)
        search_frame.pack(fill=tk.X, padx=10, pady=5)
        ttk.Label(search_frame, text="🔍 Buscar:").pack(side=tk.LEFT, padx=(0, 5))
        self.search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=self.search_var)
        search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.search_var.trace_add("write", lambda *args: self._populate_treeview())

        # --- Treeview ---
        tree_frame = ttk.Frame(master)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        self.tree = ttk.Treeview(tree_frame, columns=('nombre',), show='tree')
        self.tree.column("#0", width=40, anchor=tk.CENTER)
        self.tree.heading("#0", text="Sel.")
        self.tree.heading('nombre', text='Nombre de la Empresa')
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.tree.bind("<Button-1>", self._toggle_selection)
        self._populate_treeview()
        
        return search_entry

    def _populate_treeview(self):
        """Filtra y llena el treeview con los checkboxes."""
        self.tree.delete(*self.tree.get_children())
        search_term = self.search_var.get().lower()
        
        for empresa in self.todas_las_empresas:
            nombre = empresa['nombre']
            if search_term in nombre.lower():
                estado_check = '☑' if self.selection_status.get(nombre) else '☐'
                self.tree.insert('', tk.END, text=estado_check, values=(nombre,), iid=nombre)

    def _toggle_selection(self, event):
        """Cambia el estado del checkbox al hacer clic."""
        row_id = self.tree.identify_row(event.y)
        if not row_id: return
        
        # Invertir el estado de selección
        self.selection_status[row_id] = not self.selection_status.get(row_id, False)
        self._populate_treeview() # Redibujar para reflejar el cambio
        
    def apply(self):
        # Devolver la lista de nombres de las empresas marcadas con ☑
        self.result = [nombre for nombre, seleccionado in self.selection_status.items() if seleccionado]



class DialogoSeleccionarCompetidores(simpledialog.Dialog):
    """Un diálogo para seleccionar múltiples competidores de una lista maestra, con búsqueda."""
    def __init__(self, parent, competidores_maestros, competidores_actuales):
        self.todos_competidores = competidores_maestros
        self.nombres_actuales = {c.nombre for c in competidores_actuales}
        self.competidores_disponibles = sorted(
            [c for c in self.todos_competidores if c['nombre'] not in self.nombres_actuales],
            key=lambda x: x['nombre']
        )
        self.competidores_filtrados = self.competidores_disponibles[:]
        self.seleccion = {}
        super().__init__(parent, "Seleccionar Competidores desde Catálogo")

    def body(self, master):
        self.geometry("600x450")
        master.pack(fill=tk.BOTH, expand=True)

        # --- Cuadro de Búsqueda ---
        search_frame = ttk.Frame(master, padding=(10, 5))
        search_frame.pack(fill=tk.X)
        ttk.Label(search_frame, text="🔍 Buscar:").pack(side=tk.LEFT, padx=(0, 5))
        self.search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=self.search_var, width=50)
        search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.search_var.trace_add("write", self._filtrar_lista)

        # --- Treeview con Resultados ---
        tree_frame = ttk.Frame(master)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        cols = ('sel', 'nombre', 'rnc')
        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings")
        self.tree.heading('sel', text='Sel.')
        self.tree.heading('nombre', text='Nombre del Competidor')
        self.tree.heading('rnc', text='RNC')
        self.tree.column('sel', width=40, anchor=tk.CENTER)
        self.tree.column('nombre', width=350)
        self.tree.column('rnc', width=120)

        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.tree.bind("<Button-1>", self.on_toggle_selection)
        
        self._poblar_treeview() # Llenar la tabla por primera vez
        
        return search_entry # Foco inicial en la búsqueda

    def _filtrar_lista(self, *args):
        """Filtra la lista de competidores según el texto en el cuadro de búsqueda."""
        termino = self.search_var.get().lower()
        if not termino:
            self.competidores_filtrados = self.competidores_disponibles[:]
        else:
            self.competidores_filtrados = [
                c for c in self.competidores_disponibles
                if termino in c.get('nombre', '').lower() or termino in c.get('rnc', '').lower()
            ]
        self._poblar_treeview()

    def _poblar_treeview(self):
        """Limpia y vuelve a llenar el treeview con la lista filtrada."""
        self.tree.delete(*self.tree.get_children())
        for comp in self.competidores_filtrados:
            nombre = comp['nombre']
            # Asegurarse de que cada competidor tenga una entrada en el diccionario de selección
            if nombre not in self.seleccion:
                self.seleccion[nombre] = False
            
            estado_actual = '☑' if self.seleccion[nombre] else '☐'
            self.tree.insert('', tk.END, iid=nombre, values=(estado_actual, nombre, comp.get('rnc', '')))

    def on_toggle_selection(self, event):
        """Maneja el clic para marcar/desmarcar la selección."""
        row_id = self.tree.identify_row(event.y)
        if not row_id: return
        
        # Invertir el estado de selección
        self.seleccion[row_id] = not self.seleccion.get(row_id, False)
        
        self._poblar_treeview() # Redibujar para reflejar el cambio
        
    def apply(self):
        self.result = [comp for comp in self.todos_competidores if self.seleccion.get(comp['nombre'])]

class VentanaMaestroKits(tk.Toplevel):
    """Ventana para gestionar los Kits de Requisitos por Institución."""
    def __init__(self, parent):
        super().__init__(parent)
        self.parent_app = parent
        self.db = parent.db
        self.title("Gestor de Kits de Requisitos por Institución")
        self.geometry("950x600")
        self.grab_set()

        # --- Variables de Estado ---
        self.institucion_actual = tk.StringVar()
        self.kit_actual_id = None
        self.kits_en_memoria = []

        # --- Crear Widgets ---
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Panel de selección de Institución
        filtro_frame = ttk.LabelFrame(main_frame, text="1. Seleccione una Institución", padding=10)
        filtro_frame.pack(fill=tk.X)
        
        nombres_instituciones = sorted([i['nombre'] for i in self.parent_app.instituciones_registradas])
        self.institucion_combo = ttk.Combobox(filtro_frame, textvariable=self.institucion_actual, values=nombres_instituciones, state="readonly")
        self.institucion_combo.pack(fill=tk.X, expand=True)
        self.institucion_combo.bind("<<ComboboxSelected>>", self.cargar_kits_por_institucion)

        # Paneles para Kits y sus Documentos
        paned_window = ttk.PanedWindow(main_frame, orient=tk.HORIZONTAL)
        paned_window.pack(fill=tk.BOTH, expand=True, pady=10)

        # Panel Izquierdo: Lista de Kits
        kits_frame = ttk.LabelFrame(paned_window, text="2. Kits de la Institución", padding=10)
        paned_window.add(kits_frame, weight=1)
        
        self.tree_kits = ttk.Treeview(kits_frame, columns=('nombre',), show='headings', selectmode="browse")
        self.tree_kits.heading('nombre', text='Nombre del Kit')
        self.tree_kits.pack(fill=tk.BOTH, expand=True)
        self.tree_kits.bind('<<TreeviewSelect>>', self.cargar_documentos_del_kit)
        
# En la clase VentanaMaestroKits, dentro del __init__
        # Reemplaza el bloque de botones de los kits por este

        btn_kits_frame = ttk.Frame(kits_frame)
        btn_kits_frame.pack(fill=tk.X, pady=(5,0))
        # Creamos una cuadrícula para organizar mejor los botones
        btn_kits_frame.columnconfigure(tuple(range(2)), weight=1)

        # Fila 1 de botones
        ttk.Button(btn_kits_frame, text="Agregar Kit Manual", command=self.agregar_kit).grid(row=0, column=0, sticky=tk.EW, padx=2, pady=1)
        ttk.Button(btn_kits_frame, text="Importar desde Licitación...", command=self.importar_kit_desde_licitacion).grid(row=0, column=1, sticky=tk.EW, padx=2, pady=1)
        
        # Fila 2 de botones
        self.btn_editar_kit = ttk.Button(btn_kits_frame, text="Editar/Renombrar Kit", command=self.editar_kit, state="disabled")
        self.btn_editar_kit.grid(row=1, column=0, sticky=tk.EW, padx=2, pady=1)
        self.btn_clonar_kit = ttk.Button(btn_kits_frame, text="Clonar Kit Seleccionado", command=self.clonar_kit_seleccionado, state="disabled")
        self.btn_clonar_kit.grid(row=1, column=1, sticky=tk.EW, padx=2, pady=1)

        # Fila 3 para el botón de eliminar
        self.btn_eliminar_kit = ttk.Button(
            btn_kits_frame, text="Eliminar Kit",
            command=self.eliminar_kit, state="disabled"
        )
        self.btn_eliminar_kit.grid(row=2, column=0, columnspan=2, sticky=tk.EW, padx=2, pady=1)

        # Panel Derecho: Documentos del Kit seleccionado
        kit_docs_frame = ttk.LabelFrame(paned_window, text="3. Documentos en el Kit", padding=10)
        paned_window.add(kit_docs_frame, weight=2)

        # --- Treeview con scrollbar vertical ---
        cols = ('nombre', 'codigo')
        self.tree_docs = ttk.Treeview(kit_docs_frame, columns=cols, show='tree headings')

        # Encabezados
        self.tree_docs.heading('#0', text='Categoría')
        self.tree_docs.heading('nombre', text='Nombre del Documento')
        self.tree_docs.heading('codigo', text='Código')

        # Anchos/alineación
        self.tree_docs.column('#0', width=160, anchor=tk.W, stretch=True)
        self.tree_docs.column('nombre', width=300, anchor=tk.W, stretch=True)
        self.tree_docs.column('codigo', width=120, anchor=tk.W, stretch=False)

        # Scrollbar vertical
        scroll_y = ttk.Scrollbar(kit_docs_frame, orient='vertical', command=self.tree_docs.yview)
        self.tree_docs.configure(yscrollcommand=scroll_y.set)

        # Layout
        self.tree_docs.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll_y.pack(side=tk.RIGHT, fill=tk.Y)

        # --- Botonera debajo del Treeview de documentos del kit ---
        btn_docs_frame = ttk.Frame(kit_docs_frame)  # <- antes: docs_frame
        btn_docs_frame.pack(fill=tk.X, pady=(5, 0))

        self.btn_agregar_doc = ttk.Button(
            btn_docs_frame, text="Añadir Documento",
            command=self.agregar_documento_a_kit, state="disabled"
        )
        self.btn_agregar_doc.pack(side=tk.LEFT, expand=True, padx=2)

        self.btn_quitar_doc = ttk.Button(
            btn_docs_frame, text="Quitar Documento",
            command=self.quitar_documento_de_kit, state="disabled"
        )
        self.btn_quitar_doc.pack(side=tk.LEFT, expand=True, padx=2)

        # Habilita/Deshabilita "Quitar Doc" según selección
        def _toggle_quitar_doc(_event=None):
            self.btn_quitar_doc.config(
                state=("normal" if self.tree_docs.selection() else "disabled")
            )

        self.tree_docs.bind('<<TreeviewSelect>>', _toggle_quitar_doc)

    # Pega estos nuevos métodos dentro de la clase VentanaMaestroKits

    def clonar_kit_seleccionado(self):
        if not self.kit_actual_id: return
        
        nombre_kit_original = self.tree_kits.item(self.kit_actual_id, 'values')[0]
        institucion = self.institucion_actual.get()

        nuevo_nombre = simpledialog.askstring("Clonar Kit", "Nombre para el nuevo kit clonado:",
                                              initialvalue=f"Copia de {nombre_kit_original}", parent=self)
        if not (nuevo_nombre and nuevo_nombre.strip()):
            return

        try:
            # 1. Crear el nuevo kit en la base de datos
            cursor = self.db.cursor
            cursor.execute("INSERT INTO kits_de_requisitos (nombre_kit, institucion_nombre) VALUES (?, ?)",
                           (nuevo_nombre.strip(), institucion))
            nuevo_kit_id = cursor.lastrowid

            # 2. Copiar todos los items del kit original al nuevo
            cursor.execute("SELECT documento_maestro_id FROM kit_items WHERE kit_id = ?", (self.kit_actual_id,))
            documentos_a_clonar = cursor.fetchall()

            if documentos_a_clonar:
                items_para_insertar = [(nuevo_kit_id, doc_id[0]) for doc_id in documentos_a_clonar]
                cursor.executemany("INSERT INTO kit_items (kit_id, documento_maestro_id) VALUES (?, ?)", items_para_insertar)
            
            self.db.conn.commit()
            self.cargar_kits_por_institucion()
            messagebox.showinfo("Éxito", f"El kit '{nombre_kit_original}' ha sido clonado como '{nuevo_nombre}'.", parent=self)

        except self.db.conn.IntegrityError:
            messagebox.showerror("Error", "Ya existe un kit con ese nombre para esta institución.", parent=self)
            self.db.conn.rollback()

# En la clase VentanaMaestroKits, reemplaza este método completo

# En la clase VentanaMaestroKits, reemplaza este método por completo

    def importar_kit_desde_licitacion(self):
        institucion = self.institucion_actual.get()
        if not institucion:
            messagebox.showwarning("Sin Selección", "Primero debe seleccionar una institución.", parent=self)
            return

        # 1) Filtrar licitaciones por institución
        licitaciones_filtradas = [l for l in self.parent_app.gestor_licitaciones if l.institucion == institucion]
        if not licitaciones_filtradas:
            messagebox.showinfo("Información", f"No se encontraron licitaciones para la institución '{institucion}'.", parent=self)
            return

        dialogo = DialogoSeleccionarLicitacion(self, "Importar Requisitos desde Licitación", licitaciones_filtradas, None)
        if not dialogo.result:
            return

        lic_origen = next((l for l in licitaciones_filtradas if l.numero_proceso == dialogo.result), None)
        if not (lic_origen and lic_origen.documentos_solicitados):
            messagebox.showinfo("Información", "La licitación seleccionada no tiene documentos para importar.", parent=self)
            return

        # 2) Nombre del nuevo kit
        nombre_kit = simpledialog.askstring(
            "Nuevo Kit",
            "Nombre para el kit importado:",
            initialvalue=f"Kit de {lic_origen.numero_proceso}",
            parent=self
        )
        if not (nombre_kit and nombre_kit.strip()):
            return

        try:
            cursor = self.db.cursor

            # 3) Crear el kit
            cursor.execute(
                "INSERT INTO kits_de_requisitos (nombre_kit, institucion_nombre) VALUES (?, ?)",
                (nombre_kit.strip(), institucion)
            )
            nuevo_kit_id = cursor.lastrowid

            # 4) Códigos únicos desde la licitación
            codigos = sorted({doc.codigo for doc in lic_origen.documentos_solicitados if doc.codigo})
            if not codigos:
                self.db.conn.commit()
                self.cargar_kits_por_institucion()
                messagebox.showinfo("Información", "La licitación no contiene códigos válidos para importar.", parent=self)
                return

            # 5) Tomar **un solo id** de documento maestro por cada código
            placeholders = ",".join("?" * len(codigos))
            cursor.execute(f"""
                SELECT MIN(id) AS id
                FROM documentos_maestros
                WHERE codigo IN ({placeholders})
                GROUP BY codigo
            """, codigos)
            ids_unicos = [row[0] for row in cursor.fetchall()]

            # 6) Insertar vínculos evitando duplicados
            items = [(nuevo_kit_id, doc_id) for doc_id in ids_unicos]
            if items:
                cursor.executemany(
                    "INSERT OR IGNORE INTO kit_items (kit_id, documento_maestro_id) VALUES (?, ?)",
                    items
                )

            self.db.conn.commit()
            self.cargar_kits_por_institucion()
            messagebox.showinfo(
                "Éxito",
                f"Se creó el kit '{nombre_kit.strip()}' con {len(items)} documentos únicos.",
                parent=self
            )

        except self.db.conn.IntegrityError:
            self.db.conn.rollback()
            messagebox.showerror("Error", "Ya existe un kit con ese nombre para esta institución.", parent=self)
        except Exception as e:
            self.db.conn.rollback()
            messagebox.showerror("Error", f"Ocurrió un error al importar el kit:\n{e}", parent=self)

    def cargar_documentos_del_kit(self, event=None):
        self.tree_docs.delete(*self.tree_docs.get_children())
        
        es_seleccion_valida = self.tree_kits.selection()
        estado_botones = "normal" if es_seleccion_valida else "disabled"

        self.btn_editar_kit.config(state=estado_botones)
        self.btn_eliminar_kit.config(state=estado_botones)
        self.btn_agregar_doc.config(state=estado_botones)
        self.btn_clonar_kit.config(state=estado_botones)
        self.btn_quitar_doc.config(state="disabled")

        if not es_seleccion_valida:
            self.kit_actual_id = None
            return

        self.kit_actual_id = self.tree_kits.selection()[0]
        
        query = """
            SELECT dm.categoria, dm.nombre, dm.codigo
            FROM kit_items ki
            JOIN documentos_maestros dm ON ki.documento_maestro_id = dm.id
            WHERE ki.kit_id = ?
            ORDER BY dm.categoria, dm.nombre
        """
        self.db.cursor.execute(query, (self.kit_actual_id,))
        
        docs_por_categoria = {}
        for cat, nombre, cod in self.db.cursor.fetchall():
            if cat not in docs_por_categoria:
                docs_por_categoria[cat] = []
            docs_por_categoria[cat].append({'nombre': nombre, 'codigo': cod})

        # Insertar en el Treeview agrupado
        for categoria, documentos in sorted(docs_por_categoria.items()):
            parent_id = self.tree_docs.insert('', tk.END, text=categoria, open=True)
            for doc in documentos:
                self.tree_docs.insert(parent_id, tk.END, values=(doc['nombre'], doc['codigo']))



# En la clase VentanaMaestroKits, reemplaza este método

# En la clase VentanaMaestroKits, pega estos dos métodos reemplazando los antiguos

    def cargar_kits_por_institucion(self, event=None):
        institucion = self.institucion_actual.get()
        # Limpia la lista de kits
        self.tree_kits.delete(*self.tree_kits.get_children())
        # Limpia la lista de documentos del kit (usando el nombre correcto: tree_docs)
        self.tree_docs.delete(*self.tree_docs.get_children()) 
        
        # Deshabilita los botones que dependen de una selección
        self.btn_editar_kit.config(state="disabled")
        self.btn_eliminar_kit.config(state="disabled")
        self.btn_clonar_kit.config(state="disabled")
        self.btn_agregar_doc.config(state="disabled")
        self.btn_quitar_doc.config(state="disabled")

        if not institucion:
            return
        
        # Carga los kits de la base de datos
        self.db.cursor.execute("SELECT id, nombre_kit FROM kits_de_requisitos WHERE institucion_nombre = ?", (institucion,))
        self.kits_en_memoria = self.db.cursor.fetchall()
        for kit_id, nombre_kit in self.kits_en_memoria:
            self.tree_kits.insert('', tk.END, iid=kit_id, values=(nombre_kit,))

    def cargar_documentos_del_kit(self, event=None):
        # Limpia la vista de documentos antes de cargar los nuevos
        self.tree_docs.delete(*self.tree_docs.get_children())
        
        es_seleccion_valida = self.tree_kits.selection()
        estado_botones = "normal" if es_seleccion_valida else "disabled"

        # Actualiza el estado de TODOS los botones
        self.btn_editar_kit.config(state=estado_botones)
        self.btn_eliminar_kit.config(state=estado_botones)
        self.btn_agregar_doc.config(state=estado_botones)
        self.btn_clonar_kit.config(state=estado_botones)
        self.btn_quitar_doc.config(state="disabled") # Este depende de la selección de un documento

        if not es_seleccion_valida:
            self.kit_actual_id = None
            return

        self.kit_actual_id = self.tree_kits.selection()[0]
        
        # Carga y muestra los documentos agrupados por categoría
        query = """
            SELECT dm.categoria, dm.nombre, dm.codigo
            FROM kit_items ki
            JOIN documentos_maestros dm ON ki.documento_maestro_id = dm.id
            WHERE ki.kit_id = ?
            ORDER BY dm.categoria, dm.nombre
        """
        self.db.cursor.execute(query, (self.kit_actual_id,))
        
        docs_por_categoria = {}
        for cat, nombre, cod in self.db.cursor.fetchall():
            if cat not in docs_por_categoria:
                docs_por_categoria[cat] = []
            docs_por_categoria[cat].append({'nombre': nombre, 'codigo': cod})

        # Inserta en el Treeview agrupado
        for categoria, documentos in sorted(docs_por_categoria.items()):
            parent_id = self.tree_docs.insert('', tk.END, text=categoria, open=True)
            for doc in documentos:
                self.tree_docs.insert(parent_id, tk.END, values=(doc['nombre'], doc['codigo']))

    def agregar_kit(self):
        institucion = self.institucion_actual.get()
        if not institucion:
            messagebox.showwarning("Sin Selección", "Primero debe seleccionar una institución.", parent=self)
            return
        
        nombre_kit = simpledialog.askstring("Nuevo Kit", "Nombre para el nuevo kit:", parent=self)
        if nombre_kit and nombre_kit.strip():
            try:
                self.db.cursor.execute("INSERT INTO kits_de_requisitos (nombre_kit, institucion_nombre) VALUES (?, ?)", (nombre_kit.strip(), institucion))
                self.db.conn.commit()
                self.cargar_kits_por_institucion()
            except self.db.conn.IntegrityError:
                messagebox.showerror("Error", "Ya existe un kit con este nombre para esta institución.", parent=self)

    def editar_kit(self):
        if not self.kit_actual_id: return
        nombre_actual = self.tree_kits.item(self.kit_actual_id, 'values')[0]
        
        nuevo_nombre = simpledialog.askstring("Editar Kit", "Nuevo nombre para el kit:", initialvalue=nombre_actual, parent=self)
        if nuevo_nombre and nuevo_nombre.strip() and nuevo_nombre != nombre_actual:
            try:
                self.db.cursor.execute("UPDATE kits_de_requisitos SET nombre_kit = ? WHERE id = ?", (nuevo_nombre.strip(), self.kit_actual_id))
                self.db.conn.commit()
                self.cargar_kits_por_institucion()
            except self.db.conn.IntegrityError:
                messagebox.showerror("Error", "Ya existe un kit con este nombre para esta institución.", parent=self)
                
    def eliminar_kit(self):
        if not self.kit_actual_id: return
        nombre_kit = self.tree_kits.item(self.kit_actual_id, 'values')[0]
        if messagebox.askyesno("Confirmar", f"¿Está seguro de que desea eliminar el kit '{nombre_kit}'?\nEsta acción es permanente.", parent=self, icon='warning'):
            self.db.cursor.execute("DELETE FROM kits_de_requisitos WHERE id = ?", (self.kit_actual_id,))
            self.db.conn.commit()
            self.cargar_kits_por_institucion()
            
# En la clase VentanaMaestroKits, REEMPLAZA este método:

# En la clase VentanaMaestroKits, REEMPLAZA este método:

    def agregar_documento_a_kit(self):
        if not self.kit_actual_id:
            return

        # Obtenemos los documentos que ya están en el kit para no mostrarlos
        self.db.cursor.execute("SELECT documento_maestro_id FROM kit_items WHERE kit_id = ?", (self.kit_actual_id,))
        ids_actuales = {row[0] for row in self.db.cursor.fetchall()}
        
        documentos_disponibles_maestros = [doc for doc in self.parent_app.documentos_maestros if doc.id not in ids_actuales]

        # Usamos el nuevo diálogo de selección para una experiencia consistente
        dialogo = DialogoSeleccionarDocumento(
            self,
            "Seleccionar Documentos para el Kit",
            documentos_disponibles_maestros,
            documentos_actuales=[] # No aplica filtro de código aquí, ya lo hicimos con IDs
        )
        if not dialogo.result:
            return

        items_para_insertar = []
        for doc_maestro in dialogo.result:
            items_para_insertar.append((self.kit_actual_id, doc_maestro.id))

        try:
            if items_para_insertar:
                self.db.cursor.executemany(
                    "INSERT OR IGNORE INTO kit_items (kit_id, documento_maestro_id) VALUES (?, ?)",
                    items_para_insertar
                )
                self.db.conn.commit()
                self.cargar_documentos_del_kit() # Refrescar la vista
        except Exception as e:
            self.db.conn.rollback()
            messagebox.showerror("Error", f"No se pudo añadir el/los documento(s):\n{e}", parent=self)



    def quitar_documento_de_kit(self):
        if not (self.kit_actual_id and self.tree_docs.selection()): return
        
        item_seleccionado = self.tree_docs.selection()[0]
        # Solo actuamos si se selecciona un documento (hijo), no una categoría (padre)
        if not self.tree_docs.parent(item_seleccionado):
            return

        valores = self.tree_docs.item(item_seleccionado, 'values')
        codigo_doc = valores[1]
        
        doc_maestro_obj = next((d for d in self.parent_app.documentos_maestros if d.codigo == codigo_doc), None)
        if not doc_maestro_obj:
            messagebox.showerror("Error", "No se pudo encontrar el documento maestro.", parent=self)
            return

        self.db.cursor.execute("DELETE FROM kit_items WHERE kit_id = ? AND documento_maestro_id = ?", (self.kit_actual_id, doc_maestro_obj.id))
        self.db.conn.commit()
        self.cargar_documentos_del_kit()


class DialogoElegirCategoria(simpledialog.Dialog):
    """Un simple diálogo para seleccionar una categoría de una lista."""
    def __init__(self, parent, title, categorias, categoria_sugerida):
        self.categorias = categorias
        self.categoria_sugerida = categoria_sugerida
        super().__init__(parent, title)

    def body(self, master):
        ttk.Label(master, text="Seleccione la categoría para este documento:").pack(padx=10, pady=(10,5))
        
        self.categoria_var = tk.StringVar(value=self.categoria_sugerida)
        combo = ttk.Combobox(master, textvariable=self.categoria_var, values=self.categorias, state="readonly", width=30)
        combo.pack(padx=10, pady=5)
        
        # Preseleccionar la categoría sugerida si existe en la lista
        if self.categoria_sugerida in self.categorias:
            combo.set(self.categoria_sugerida)
        else:
            combo.current(0)
            
        return combo

    def apply(self):
        self.result = self.categoria_var.get()


class VentanaMaestroResponsables(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent_app = parent
        self.responsables_copia = [dict(r) for r in self.parent_app.responsables_maestros]
        self.title("Catálogo de Responsables"); self.geometry("500x450"); self.grab_set()
        main_frame = ttk.Frame(self, padding="15"); main_frame.pack(fill=tk.BOTH, expand=True)
        tree_frame = ttk.LabelFrame(main_frame, text="Responsables Registrados", padding=10); tree_frame.pack(fill=tk.BOTH, expand=True)
        self.tree = ttk.Treeview(tree_frame, columns=('nombre',), show="headings")
        self.tree.heading('nombre', text='Nombre del Responsable o Departamento')
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview); self.tree.configure(yscrollcommand=scrollbar.set); scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        btn_frame = ttk.Frame(main_frame); btn_frame.pack(fill=tk.X, pady=10)
        ttk.Button(btn_frame, text="Agregar", command=self.agregar).pack(side=tk.LEFT, expand=True, padx=5)
        ttk.Button(btn_frame, text="Editar", command=self.editar).pack(side=tk.LEFT, expand=True, padx=5)
        ttk.Button(btn_frame, text="Eliminar", command=self.eliminar).pack(side=tk.LEFT, expand=True, padx=5)
        ttk.Button(main_frame, text="✅ Guardar y Cerrar", command=self.cerrar_y_guardar).pack(pady=(10,0), ipady=4)
        self.protocol("WM_DELETE_WINDOW", self.destroy)
        self.actualizar_lista()

    def actualizar_lista(self):
        self.tree.delete(*self.tree.get_children())
        for resp in sorted(self.responsables_copia, key=lambda x: x['nombre']):
            self.tree.insert('', tk.END, iid=resp['nombre'], values=(resp['nombre'],))

    def agregar(self):
        nombre = simpledialog.askstring("Agregar Responsable", "Nombre:", parent=self)
        if nombre and nombre.strip():
            nombre = nombre.strip()
            if any(r['nombre'].lower() == nombre.lower() for r in self.responsables_copia):
                messagebox.showerror("Error", "Ya existe un responsable con ese nombre.", parent=self); return
            self.responsables_copia.append({'nombre': nombre}); self.actualizar_lista()

    def editar(self):
        sel = self.tree.selection()
        if not sel: return
        nombre_actual = sel[0]
        responsable_actual = next((r for r in self.responsables_copia if r['nombre'] == nombre_actual), None)
        if not responsable_actual: return
        nuevo_nombre = simpledialog.askstring("Editar Responsable", "Nuevo nombre:", initialvalue=nombre_actual, parent=self)
        if nuevo_nombre and nuevo_nombre.strip():
            responsable_actual['nombre'] = nuevo_nombre.strip(); self.actualizar_lista()

    def eliminar(self):
        sel = self.tree.selection()
        if not sel: return
        nombre_a_eliminar = sel[0]
        if messagebox.askyesno("Confirmar", f"¿Eliminar a '{nombre_a_eliminar}' del catálogo?", parent=self):
            self.responsables_copia = [r for r in self.responsables_copia if r['nombre'] != nombre_a_eliminar]
            self.actualizar_lista()

    def cerrar_y_guardar(self):
        self.parent_app.responsables_maestros[:] = self.responsables_copia

        self.parent_app.db.save_master_lists(
            empresas=self.parent_app.empresas_registradas,
            instituciones=self.parent_app.instituciones_registradas,
            documentos_maestros=self.parent_app.documentos_maestros,
            competidores_maestros=self.parent_app.competidores_maestros,
            responsables_maestros=self.parent_app.responsables_maestros,
            replace_tables={'responsables_maestros'}
        )

        self.destroy()

class VentanaResultadosTests(tk.Toplevel):
    """Muestra los resultados de las pruebas de integridad y permite exportarlos."""
    def __init__(self, parent, results_log):
        super().__init__(parent)
        self.results_log = results_log
        self.title("Resultados de las Pruebas de Integridad (Smoke Tests)")
        self.geometry("800x500")
        self.grab_set()

        main_frame = ttk.Frame(self, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # --- Panel de Resultados ---
        log_frame = ttk.LabelFrame(main_frame, text="Registro de Pruebas", padding=10)
        log_frame.pack(fill=tk.BOTH, expand=True)
        
        self.text_widget = tk.Text(log_frame, wrap=tk.WORD, font=("Consolas", 10))
        self.text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.text_widget.yview)
        self.text_widget.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.text_widget.insert(tk.END, self.results_log)
        self.text_widget.config(state="disabled")

        # --- Botones de Acción (CORREGIDO) ---
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(10, 0))
        # Unificamos en un solo botón que llama a la función correcta
        ttk.Button(btn_frame, text="💾 Exportar Resultados a .txt", command=self._exportar_resultados).pack(side=tk.LEFT)
        ttk.Button(btn_frame, text="Cerrar", command=self.destroy).pack(side=tk.RIGHT)

    def _exportar_resultados(self):
        """Guarda el contenido del log de pruebas en un archivo de texto."""
        file_path = filedialog.asksaveasfilename(
            parent=self,
            title="Exportar Log de Pruebas",
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
        )
        if file_path:
            try:
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(self.results_log)
                messagebox.showinfo("Éxito", "El registro ha sido exportado correctamente.", parent=self)
            except IOError as e:
                messagebox.showerror("Error al Guardar", f"No se pudo guardar el archivo:\n{e}", parent=self)

    def _exportar(self):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("PDF", "*.pdf")]
        )
        if file_path:
            self.reporter.generate_bid_results_report(self.licitacion, file_path)
            messagebox.showinfo("Éxito", f"Reporte exportado en:\n{file_path}")
class VentanaMaestroCompetidores(tk.Toplevel):
# En gestor_licitaciones_db.py, clase VentanaMaestroCompetidores

# En gestor_licitaciones_db.py, clase VentanaMaestroCompetidores, reemplaza __init__

    def __init__(self, parent):
        super().__init__(parent)
        self.parent_app = parent
        self.competidores_copia = [dict(c) for c in self.parent_app.competidores_maestros]
        self.competidores_filtrados = self.competidores_copia

        self.title("Catálogo de Competidores")
        self.geometry("900x550") # Un poco más ancha para las nuevas columnas
        self.grab_set()

        main_frame = ttk.Frame(self, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Campo de búsqueda
        search_frame = ttk.Frame(main_frame)
        search_frame.pack(fill=tk.X, pady=(0, 10))
        ttk.Label(search_frame, text="🔍 Buscar:").pack(side=tk.LEFT, padx=(0, 5))
        self.search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=self.search_var, width=50)
        search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.search_var.trace_add("write", self.filtrar_lista)
        
        # Treeview con nuevas columnas
        tree_frame = ttk.LabelFrame(main_frame, text="Competidores Registrados", padding=10)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        cols = ('nombre', 'rnc', 'rpe', 'representante') # <-- COLUMNAS ACTUALIZADAS
        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings")
        self.tree.heading('nombre', text='Nombre')
        self.tree.heading('rnc', text='RNC')
        self.tree.heading('rpe', text='No. RPE')
        self.tree.heading('representante', text='Representante')
        
        self.tree.column('nombre', width=250)
        self.tree.column('rnc', width=120)
        self.tree.column('rpe', width=120)
        self.tree.column('representante', width=200)

        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.bind("<<TreeviewSelect>>", self._on_selection_change)

        # Botones (sin cambios en su creación)
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=10)
        btn_frame.columnconfigure(tuple(range(4)), weight=1)
        ttk.Button(btn_frame, text="Agregar", command=self.agregar).grid(row=0, column=0, sticky=tk.EW, padx=5)
        ttk.Button(btn_frame, text="Editar", command=self.editar).grid(row=0, column=1, sticky=tk.EW, padx=5)
        self.btn_ver_perfil = ttk.Button(btn_frame, text="📈 Ver Perfil", command=self.ver_perfil, state="disabled")
        self.btn_ver_perfil.grid(row=0, column=2, sticky=tk.EW, padx=5)
        ttk.Button(btn_frame, text="Eliminar", command=self.eliminar).grid(row=0, column=3, sticky=tk.EW, padx=5)
        ttk.Button(main_frame, text="✅ Guardar Cambios y Cerrar", command=self.cerrar_y_guardar).pack(pady=(10,0), ipady=4)
        
        self.protocol("WM_DELETE_WINDOW", self.destroy)
        self.actualizar_lista()
# In gestor_licitaciones_db.py, add this new method to VentanaMaestroCompetidores

    def filtrar_lista(self, *args):
        termino = self.search_var.get().lower()
        if not termino:
            # If the search box is empty, show all competitors
            self.competidores_filtrados = self.competidores_copia
        else:
            # Otherwise, filter the list by name or RNC
            self.competidores_filtrados = [
                c for c in self.competidores_copia
                if termino in c.get('nombre', '').lower() or termino in c.get('rnc', '').lower()
            ]
        self.actualizar_lista()

# Esta es la única versión de actualizar_lista que debe existir en VentanaMaestroCompetidores

# En gestor_licitaciones_db.py, clase VentanaMaestroCompetidores, reemplaza actualizar_lista

    def actualizar_lista(self):
        self.tree.delete(*self.tree.get_children())
        for comp in sorted(self.competidores_filtrados, key=lambda x: x['nombre']):
            # Añadir los nuevos valores a la tabla
            values = (
                comp.get('nombre', ''), 
                comp.get('rnc', ''),
                comp.get('rpe', ''),
                comp.get('representante', '')
            )
            self.tree.insert('', tk.END, iid=comp['nombre'], values=values)
        self._on_selection_change()
        
    def _on_selection_change(self, event=None):
        """Habilita o deshabilita el botón 'Ver Perfil' según la selección."""
        if self.tree.selection():
            self.btn_ver_perfil.config(state="normal")
        else:
            self.btn_ver_perfil.config(state="disabled")

    def ver_perfil(self):
        """Abre la ventana del perfil del competidor seleccionado."""
        if not self.tree.selection(): return
        
        nombre_competidor = self.tree.item(self.tree.selection()[0], 'values')[0]
        VentanaPerfilCompetidor(self, nombre_competidor, self.parent_app.gestor_licitaciones)




# In VentanaMaestroCompetidores, replace the 'agregar', 'editar', and 'eliminar' methods

    def agregar(self):
        data = DialogoGestionarEntidad(self, "Agregar Competidor", "competidor").result
        if data and data.get('nombre'):
            if any(c['nombre'].lower() == data['nombre'].lower() for c in self.competidores_copia):
                messagebox.showerror("Error", "Ya existe un competidor con ese nombre.", parent=self)
                return
            
            # Check for duplicate RNC only if it's provided
            if data.get('rnc') and any(c.get('rnc', '').lower() == data['rnc'].lower() and c.get('rnc', '') for c in self.competidores_copia):
                messagebox.showerror("Error", "Ya existe un competidor con ese RNC.", parent=self)
                return

            self.competidores_copia.append(data)
            self.filtrar_lista() # Re-apply filter to show the new item

    def editar(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Sin selección", "Por favor, selecciona un competidor para editar.", parent=self)
            return
            
        nombre_actual = self.tree.item(sel[0])['values'][0]
        competidor_actual = next((c for c in self.competidores_copia if c['nombre'] == nombre_actual), None)
        if not competidor_actual:
            return

        data = DialogoGestionarEntidad(self, "Editar Competidor", "competidor", initial_data=competidor_actual).result
        if data and data.get('nombre'):
            competidor_actual.update(data)
            self.filtrar_lista() # Re-apply filter

    def eliminar(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Sin selección", "Por favor, selecciona un competidor para eliminar.", parent=self)
            return

        nombre_a_eliminar = self.tree.item(sel[0])['values'][0]
        if messagebox.askyesno("Confirmar", f"¿Estás seguro de que quieres eliminar a '{nombre_a_eliminar}' del catálogo maestro?", parent=self):
            self.competidores_copia = [c for c in self.competidores_copia if c['nombre'] != nombre_a_eliminar]
            self.filtrar_lista() # Re-apply filter

    def cerrar_y_guardar(self):
        # Actualiza la lista maestra en sitio (no cambies la referencia)
        self.parent_app.competidores_maestros[:] = self.competidores_copia

        # Guardar SOLO el catálogo de competidores como reemplazo total
        self.parent_app.db.save_master_lists(
            empresas=self.parent_app.empresas_registradas,
            instituciones=self.parent_app.instituciones_registradas,
            documentos_maestros=self.parent_app.documentos_maestros,
            competidores_maestros=self.parent_app.competidores_maestros,
            responsables_maestros=self.parent_app.responsables_maestros,
            replace_tables={'competidores_maestros'}
        )

        self.destroy()


# En gestor_licitaciones_db_2.py, junto a las otras clases de ventana

# En gestor_licitaciones_db.py, pega esta clase actualizada
class VentanaPerfilCompetidor(tk.Toplevel):
    """Muestra un dashboard con estadísticas y el historial de un competidor."""
    def __init__(self, parent, competidor_nombre, todas_las_licitaciones):
        super().__init__(parent)
        self.competidor_nombre = competidor_nombre
        self.todas_las_licitaciones = todas_las_licitaciones

        # Intentar encontrar la DB en distintos niveles del padre
        self.db = (
            getattr(parent, "db", None)
            or getattr(getattr(parent, "parent_app", None), "db", None)
        )

        self.title(f"Perfil de Competidor: {self.competidor_nombre}")
        self.geometry("1100x700")
        self.grab_set()

        # 1) Procesar datos
        self.historial, self.kpis = self._procesar_historial()

        # 2) UI
        main_frame = ttk.Frame(self, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)

        kpi_frame = ttk.Frame(main_frame)
        kpi_frame.pack(fill=tk.X, pady=10)
        self.crear_widgets_kpi(kpi_frame)

        historial_frame = ttk.LabelFrame(
            main_frame,
            text="Historial de Participación (Doble clic para ver detalles)",
            padding=10
        )
        historial_frame.pack(fill=tk.BOTH, expand=True)
        self.crear_tabla_historial(historial_frame)

    # ------------------- Lógica -------------------
    def _procesar_historial(self):
        """
        Construye el historial del competidor y KPIs, detectando GANADORES por LOTE.
        Regla de detección por cada lote de la licitación:
        1) Si el lote tiene 'ganador_nombre' y coincide con el competidor => ganó ese lote.
        2) Si no hay 'ganador_nombre', pero en la oferta del competidor para ese lote existe o['ganador'] == True => ganó ese lote.
        """
        historial = []
        participaciones_por_institucion = {}

        total_participaciones = 0               # cuántas licitaciones participó este competidor (aparece en oferentes_participantes)
        total_licitaciones_ganadas = 0         # en cuántas licitaciones ganó al menos 1 lote
        total_lotes_ganados = 0                # suma de lotes ganados en todas las licitaciones
        monto_ofertado_total = 0

        estados_finalizados = ["Adjudicada", "Descalificado Fase A", "Descalificado Fase B", "Desierta", "Cancelada"]

        for lic in self.todas_las_licitaciones:
            # localizar al competidor en esta licitación
            oferente_obj = next((o for o in lic.oferentes_participantes if o.nombre == self.competidor_nombre), None)
            if not oferente_obj:
                continue  # este competidor no participó en esta licitación

            total_participaciones += 1

            # monto ofertado (tu método actual)
            try:
                monto_ofertado = oferente_obj.get_monto_total_ofertado(solo_habilitados=True)
            except Exception:
                monto_ofertado = sum((o.get('monto', 0) or 0) for o in getattr(oferente_obj, 'ofertas_por_lote', []))
            monto_ofertado_total += monto_ofertado

            # institución (para KPI de "institución favorita")
            institucion = lic.institucion
            participaciones_por_institucion[institucion] = participaciones_por_institucion.get(institucion, 0) + 1

            # ======== conteo de lotes ganados por esta licitación ========
            lotes_ganados_en_esta_lic = 0
            # 1) si en el objeto lote ya guardas ganador_nombre / ganado_por_nosotros:
            for lote in getattr(lic, 'lotes', []):
                ganador_nombre = getattr(lote, 'ganador_nombre', '') or ''
                if ganador_nombre:
                    if ganador_nombre == self.competidor_nombre:
                        lotes_ganados_en_esta_lic += 1
                    continue  # ya resuelto este lote

                # 2) si no hay ganador_nombre, revisa la oferta marcada como ganador en oferente_obj
                for o in getattr(oferente_obj, 'ofertas_por_lote', []):
                    if str(o.get('lote_numero')) == str(getattr(lote, 'numero', '')) and o.get('ganador'):
                        lotes_ganados_en_esta_lic += 1
                        break

            # Construir "resultado" amigable
            if lic.estado in estados_finalizados:
                if lotes_ganados_en_esta_lic > 0:
                    total_licitaciones_ganadas += 1
                    total_lotes_ganados += lotes_ganados_en_esta_lic
                    resultado = f"🏆 Ganador de {lotes_ganados_en_esta_lic} lote{'s' if lotes_ganados_en_esta_lic != 1 else ''}"
                else:
                    resultado = "Perdedor"
            else:
                # Si está en proceso igual queremos mostrar si ya hay radios marcados (opcional)
                if lotes_ganados_en_esta_lic > 0:
                    total_lotes_ganados += lotes_ganados_en_esta_lic
                    resultado = f"En Proceso (marcado {lotes_ganados_en_esta_lic} lote{'s' if lotes_ganados_en_esta_lic != 1 else ''})"
                else:
                    resultado = "En Proceso"

            historial.append({
                'proceso': lic.numero_proceso,
                'nombre': lic.nombre_proceso,
                'institucion': lic.institucion,
                'monto_ofertado': monto_ofertado,
                'resultado': resultado
            })

        # para tasa de éxito usamos solo licitaciones finalizadas en las que participó
        participaciones_finalizadas = sum(1 for item in historial if item['resultado'].startswith("🏆") or item['resultado'] == "Perdedor")

        kpis = {
            'participaciones': total_participaciones,
            'ganadas': total_licitaciones_ganadas,                     # cuántas lic ganó (>=1 lote)
            'lotes_ganados': total_lotes_ganados,                      # total de lotes ganados
            'tasa_exito': (total_licitaciones_ganadas / participaciones_finalizadas * 100) if participaciones_finalizadas > 0 else 0,
            'monto_promedio': (monto_ofertado_total / total_participaciones) if total_participaciones > 0 else 0,
            'top_institucion': max(participaciones_por_institucion, key=participaciones_por_institucion.get) if participaciones_por_institucion else "N/A"
        }

        return historial, kpis


    # ------------------- Widgets -------------------
    def crear_widgets_kpi(self, parent_frame):
        parent_frame.columnconfigure(tuple(range(7)), weight=1)

        kpi_widgets = [
            ("Participaciones Totales", f"{self.kpis['participaciones']}"),
            ("Licitaciones Ganadas", f"{self.kpis['ganadas']}"),
            ("Lotes Ganados", f"{self.kpis['lotes_ganados']}"),
            ("Tasa de Éxito", f"{self.kpis['tasa_exito']:.1f}%"),
            ("Monto Ofertado Promedio", f"RD$ {self.kpis['monto_promedio']:,.2f}"),
            ("Institución Favorita", self.kpis['top_institucion'])
        ]

        for i, (titulo, valor) in enumerate(kpi_widgets):
            frame = ttk.LabelFrame(parent_frame, text=titulo)
            frame.grid(row=0, column=i, sticky="ew", padx=5, pady=5)
            lbl = ttk.Label(frame, text=valor, font=("Helvetica", 14, "bold"), anchor="center")
            lbl.pack(pady=10, padx=10)

    def crear_tabla_historial(self, parent_frame):
        cols = ('proceso', 'institucion', 'nombre', 'monto', 'resultado')
        self.tree = ttk.Treeview(parent_frame, columns=cols, show="headings")

        self.tree.heading('proceso', text="Proceso");       self.tree.column('proceso', width=150)
        self.tree.heading('institucion', text="Institución"); self.tree.column('institucion', width=220)
        self.tree.heading('nombre', text="Nombre Licitación"); self.tree.column('nombre', width=380)
        self.tree.heading('monto', text="Monto Ofertado");     self.tree.column('monto', width=160, anchor=tk.E)
        self.tree.heading('resultado', text="Resultado");      self.tree.column('resultado', width=110, anchor=tk.CENTER)

        self.tree.tag_configure('ganador', background='#d4edda')

        # Doble clic: abrir reporte de licitación
        self.tree.bind("<Double-1>", self._abrir_reporte_licitacion)

        for item in self.historial:
            tags = ('ganador',) if item['resultado'] == "🏆 Ganador" else ()
            self.tree.insert(
                '', tk.END,
                values=(item['proceso'], item['institucion'], item['nombre'],
                        f"RD$ {item['monto_ofertado']:,.2f}", item['resultado']),
                tags=tags
            )
        self.tree.pack(fill=tk.BOTH, expand=True)

    def _abrir_reporte_licitacion(self, event=None):
        if not self.tree.selection():
            return
        sel = self.tree.selection()[0]
        numero_proceso = self.tree.item(sel, 'values')[0]
        lic = next((l for l in self.todas_las_licitaciones if l.numero_proceso == numero_proceso), None)
        if lic:
            VentanaReporte(self, lic)
        else:
            messagebox.showwarning("No Encontrado",
                                   "No se pudo encontrar el registro completo de la licitación seleccionada.",
                                   parent=self)


# Pega esta nueva clase completa en tu archivo gestor_licitaciones_db_4.0-gemini.py

class VentanaPerfilEmpresaNuestra(tk.Toplevel):
    """Muestra un dashboard con estadísticas y el historial de una de nuestras empresas."""
    def __init__(self, parent, empresa_nombre, todas_las_licitaciones):
        super().__init__(parent)
        self.empresa_nombre = empresa_nombre
        self.todas_las_licitaciones = todas_las_licitaciones
        self.parent_app = parent # Guardamos referencia a la app principal

        self.title(f"Perfil de Empresa: {self.empresa_nombre}")
        self.geometry("1100x700")
        self.grab_set()

        # 1) Procesar datos
        self.historial, self.kpis = self._procesar_historial()

        # 2) UI
        main_frame = ttk.Frame(self, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)

        kpi_frame = ttk.Frame(main_frame)
        kpi_frame.pack(fill=tk.X, pady=10)
        self.crear_widgets_kpi(kpi_frame)

        historial_frame = ttk.LabelFrame(
            main_frame,
            text="Historial de Participación (Doble clic para ver detalles de la licitación)",
            padding=10
        )
        historial_frame.pack(fill=tk.BOTH, expand=True)
        self.crear_tabla_historial(historial_frame)

    def _procesar_historial(self):
        """
        Construye el historial de la empresa y sus KPIs.
        """
        historial = []
        participaciones_por_institucion = {}
        total_participaciones = 0
        total_licitaciones_ganadas = 0
        total_lotes_ganados = 0
        monto_adjudicado_total = 0.0

        estados_finalizados = ["Adjudicada", "Descalificado Fase A", "Descalificado Fase B", "Desierta", "Cancelada"]

        for lic in self.todas_las_licitaciones:
            # Verificamos si nuestra empresa participó en esta licitación
            nombres_empresas_participantes = {str(e) for e in lic.empresas_nuestras}
            if self.empresa_nombre not in nombres_empresas_participantes:
                continue

            total_participaciones += 1
            institucion = lic.institucion
            participaciones_por_institucion[institucion] = participaciones_por_institucion.get(institucion, 0) + 1

            lotes_ganados_por_esta_empresa = 0
            monto_adjudicado_en_esta_lic = 0.0
            es_ganadora_de_la_lic = False

            if lic.estado == "Adjudicada":
                for lote in lic.lotes:
                    # Comprobamos si el lote fue ganado por nosotros Y por ESTA empresa en particular
                    if lote.ganado_por_nosotros and lote.empresa_nuestra == self.empresa_nombre:
                        lotes_ganados_por_esta_empresa += 1
                        monto_adjudicado_en_esta_lic += lote.monto_ofertado

            if lotes_ganados_por_esta_empresa > 0:
                es_ganadora_de_la_lic = True
                total_licitaciones_ganadas += 1
                total_lotes_ganados += lotes_ganados_por_esta_empresa
                monto_adjudicado_total += monto_adjudicado_en_esta_lic

            # Determinar el resultado para la tabla
            resultado = "En Proceso"
            if lic.estado in estados_finalizados:
                if es_ganadora_de_la_lic:
                    resultado = f"🏆 Ganador ({lotes_ganados_por_esta_empresa} lote{'s' if lotes_ganados_por_esta_empresa != 1 else ''})"
                else:
                    resultado = "Perdedor"

            historial.append({
                'proceso': lic.numero_proceso,
                'nombre': lic.nombre_proceso,
                'institucion': lic.institucion,
                'monto_adjudicado': monto_adjudicado_en_esta_lic,
                'resultado': resultado
            })

        participaciones_finalizadas = sum(1 for item in historial if item['resultado'].startswith("🏆") or item['resultado'] == "Perdedor")

        kpis = {
            'participaciones': total_participaciones,
            'licitaciones_ganadas': total_licitaciones_ganadas,
            'lotes_ganados': total_lotes_ganados,
            'tasa_exito': (total_licitaciones_ganadas / participaciones_finalizadas * 100) if participaciones_finalizadas > 0 else 0,
            'monto_adjudicado_total': monto_adjudicado_total,
            'top_institucion': max(participaciones_por_institucion, key=participaciones_por_institucion.get) if participaciones_por_institucion else "N/A"
        }
        return historial, kpis

    def crear_widgets_kpi(self, parent_frame):
        parent_frame.columnconfigure(tuple(range(6)), weight=1)
        kpi_widgets = [
            ("Participaciones", f"{self.kpis['participaciones']}"),
            ("Licitaciones Ganadas", f"{self.kpis['licitaciones_ganadas']}"),
            ("Lotes Ganados", f"{self.kpis['lotes_ganados']}"),
            ("Tasa de Éxito", f"{self.kpis['tasa_exito']:.1f}%"),
            ("Monto Total Adjudicado", f"RD$ {self.kpis['monto_adjudicado_total']:,.2f}"),
            ("Institución Frecuente", self.kpis['top_institucion'])
        ]
        for i, (titulo, valor) in enumerate(kpi_widgets):
            frame = ttk.LabelFrame(parent_frame, text=titulo)
            frame.grid(row=0, column=i, sticky="ew", padx=5, pady=5)
            lbl = ttk.Label(frame, text=valor, font=("Helvetica", 14, "bold"), anchor="center")
            lbl.pack(pady=10, padx=10)

    def crear_tabla_historial(self, parent_frame):
        cols = ('proceso', 'institucion', 'nombre', 'monto', 'resultado')
        self.tree = ttk.Treeview(parent_frame, columns=cols, show="headings")
        self.tree.heading('proceso', text="Proceso")
        self.tree.heading('institucion', text="Institución")
        self.tree.heading('nombre', text="Nombre Licitación")
        self.tree.heading('monto', text="Monto Adjudicado")
        self.tree.heading('resultado', text="Resultado")
        self.tree.column('proceso', width=150)
        self.tree.column('institucion', width=220)
        self.tree.column('nombre', width=380)
        self.tree.column('monto', width=160, anchor=tk.E)
        self.tree.column('resultado', width=110, anchor=tk.CENTER)
        self.tree.tag_configure('ganador', background='#d4edda')
        self.tree.bind("<Double-1>", self._abrir_detalles_licitacion)

        for item in self.historial:
            tags = ('ganador',) if item['resultado'].startswith("🏆") else ()
            self.tree.insert('', tk.END,
                values=(item['proceso'], item['institucion'], item['nombre'],
                        f"RD$ {item['monto_adjudicado']:,.2f}", item['resultado']),
                tags=tags)
        self.tree.pack(fill=tk.BOTH, expand=True)

    def _abrir_detalles_licitacion(self, event=None):
        if not self.tree.selection(): return
        
        sel = self.tree.selection()[0]
        numero_proceso = self.tree.item(sel, 'values')[0]
        lic = next((l for l in self.todas_las_licitaciones if l.numero_proceso == numero_proceso), None)
        
        if lic:
            # Llamamos a la función de la app principal para abrir la ventana de detalles
            self.parent_app.abrir_ventana_detalles_desde_objeto(lic)

class VentanaAnalisisPaquetes(tk.Toplevel):
    """
    Una ventana dedicada a mostrar la tabla pivote de Lote x Oferente y
    el análisis de los mejores paquetes de ofertas.
    """
    def __init__(self, parent, licitacion):
        super().__init__(parent)
        self.licitacion = licitacion
        # CORRECCIÓN: Aseguramos la referencia a la app principal
        self.parent_app = parent if isinstance(parent, AppLicitacionesGUI) else parent.parent_app
        self.title(f"Análisis de Paquetes: {licitacion.numero_proceso}")
        self.geometry("1000x650") # Un poco más de alto para el nuevo botón
        self.grab_set()

        main_frame = ttk.Frame(self, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Frame para la tabla pivote
        tabla_frame = ttk.LabelFrame(main_frame, text="Tabla Pivote de Ofertas (Lotes vs. Competidores)", padding=10)
        tabla_frame.pack(fill=tk.BOTH, expand=True)

        # Frame para el resumen del análisis
        resumen_frame = ttk.LabelFrame(main_frame, text="Resultados del Análisis", padding=15)
        resumen_frame.pack(fill=tk.X, pady=(15, 0))

        # --- Frame para botones inferiores ---
        bottom_frame = ttk.Frame(main_frame)
        bottom_frame.pack(fill=tk.X, pady=(10, 0), padx=10)
        
        ttk.Button(bottom_frame, text="Exportar Reporte...", command=self._exportar_analisis).pack(side=tk.LEFT)
        ttk.Button(bottom_frame, text="Cerrar", command=self.destroy).pack(side=tk.RIGHT)
        # --- Fin del frame de botones ---

        # Realizar los cálculos y construir la UI
        self._crear_tabla_pivote(tabla_frame)
        self._mostrar_resumen(resumen_frame)

    def _exportar_analisis(self):
        """Pide una ruta y llama al generador de reportes para el análisis de paquetes."""
        file_path = filedialog.asksaveasfilename(
            parent=self,
            title="Exportar Análisis de Paquetes",
            initialfile=f"Analisis_Paquetes_{self.licitacion.numero_proceso}",
            filetypes=[("Archivos de Excel", "*.xlsx"), ("Archivos PDF", "*.pdf")],
            defaultextension=".xlsx"
        )
        if not file_path:
            return

        try:
            # Usamos el generador de reportes de la app principal
            self.parent_app.reporter.generate_package_analysis_report(self.licitacion, file_path)
            messagebox.showinfo("Éxito", f"El reporte ha sido guardado exitosamente en:\n{file_path}", parent=self)
        except Exception as e:
            messagebox.showerror("Error de Exportación", f"No se pudo generar el reporte:\n{e}", parent=self)

    def _crear_tabla_pivote(self, parent_frame):
        # (Este método se mantiene exactamente igual que antes)
        matriz_ofertas = self.licitacion.get_matriz_ofertas()
        if not matriz_ofertas:
            ttk.Label(parent_frame, text="No hay ofertas habilitadas para analizar.").pack(pady=20)
            return

        todos_los_oferentes = sorted(list(set(
            oferente for ofertas_lote in matriz_ofertas.values() for oferente in ofertas_lote
        )))

        columnas = ['lote'] + todos_los_oferentes
        tree = ttk.Treeview(parent_frame, columns=columnas, show="headings")
        
        tree.heading('lote', text='Lote')
        tree.column('lote', width=250)
        for oferente in todos_los_oferentes:
            tree.heading(oferente, text=oferente)
            tree.column(oferente, width=120, anchor=tk.E)

        tree.tag_configure('oferta_minima', background='#d4edda', font=('Helvetica', 9, 'bold'))

        for lote_num, ofertas in sorted(matriz_ofertas.items()):
            monto_minimo = min((d['monto'] for d in ofertas.values() if isinstance(d.get('monto'), (int, float))), default=float('inf'))
            lote_obj = next((l for l in self.licitacion.lotes if str(l.numero) == lote_num), None)
            nombre_lote = lote_obj.nombre if lote_obj else 'N/D'
            
            valores_fila = [f"Lote {lote_num}: {nombre_lote}"]
            for oferente in todos_los_oferentes:
                oferta = ofertas.get(oferente)
                if oferta and isinstance(oferta.get('monto'), (int, float)):
                    valores_fila.append(f"RD$ {oferta['monto']:,.2f}")
                else:
                    valores_fila.append("---")

            item_id = tree.insert("", tk.END, values=valores_fila)
            
            for i, oferente in enumerate(todos_los_oferentes):
                oferta = ofertas.get(oferente)
                if oferta and oferta.get('monto') == monto_minimo:
                    tree.item(item_id, tags=('oferta_minima',))

        tree.pack(fill=tk.BOTH, expand=True)

    def _mostrar_resumen(self, parent_frame):
        # (Este método se mantiene exactamente igual que antes)
        paquete_individual = self.licitacion.calcular_mejor_paquete_individual()
        paquete_unico = self.licitacion.calcular_mejor_paquete_por_oferente()

        parent_frame.columnconfigure(1, weight=1)

        ttk.Label(parent_frame, text="Opción 1: Mejor Oferta por Lote Individual", font=('Helvetica', 10, 'bold')).grid(row=0, column=0, sticky="w")
        monto_ind = paquete_individual['monto_total']
        ttk.Label(parent_frame, text=f"RD$ {monto_ind:,.2f}", font=('Helvetica', 12, 'bold')).grid(row=0, column=1, sticky="e", padx=10)
        
        ttk.Label(parent_frame, text="Opción 2: Mejor Paquete de Oferente Único", font=('Helvetica', 10, 'bold')).grid(row=1, column=0, sticky="w", pady=(5,0))
        if paquete_unico:
            monto_uni = paquete_unico['monto_total']
            ttk.Label(parent_frame, text=f"RD$ {monto_uni:,.2f} ({paquete_unico['oferente']})", font=('Helvetica', 12, 'bold')).grid(row=1, column=1, sticky="e", padx=10)
        else:
            monto_uni = float('inf')
            ttk.Label(parent_frame, text="N/A (Ningún oferente cotizó todos los lotes)", font=('Helvetica', 10, 'italic')).grid(row=1, column=1, sticky="e", padx=10)

        ttk.Separator(parent_frame, orient='horizontal').grid(row=2, column=0, columnspan=2, sticky='ew', pady=10)

        if monto_ind > 0 and monto_ind <= monto_uni:
            conclusion_texto = f"🏆 La Opción 1 (Lotes Individuales) es la más económica."
            color = "green"
        elif monto_uni != float('inf'):
            conclusion_texto = f"🏆 La Opción 2 (Paquete de {paquete_unico['oferente']}) es la más económica."
            color = "blue"
        else:
            conclusion_texto = "Análisis completado. No hay paquetes de oferente único disponibles."
            color = "black"

        ttk.Label(parent_frame, text=conclusion_texto, font=('Helvetica', 12, 'bold'), foreground=color).grid(row=3, column=0, columnspan=2)        
class VentanaComparadorOfertas(tk.Toplevel):
    """Muestra una tabla comparativa de ofertas para un lote específico."""
    def __init__(self, parent, licitacion, num_lote):
        super().__init__(parent)
        self.parent = parent
        self.licitacion = licitacion
        # 🔴 CLAVE: guardar el parámetro en self para usarlo en toda la clase
        self.num_lote = str(num_lote)

        self.title(f"Comparador de Ofertas – Lote {self.num_lote}")
        self.geometry("800x500")
        try:
            self.grab_set()
        except Exception:
            pass

        main = ttk.Frame(self, padding=10)
        main.pack(fill=tk.BOTH, expand=True)

        # Tabla
        self.tree = ttk.Treeview(
            main,
            columns=("participante", "tipo", "monto", "pct", "ganador"),
            show="headings",
            height=14
        )
        self.tree.heading("participante", text="Participante")
        self.tree.heading("tipo", text="Tipo")
        self.tree.heading("monto", text="Monto Ofertado")
        self.tree.heading("pct", text="% vs Base")
        self.tree.heading("ganador", text="Ganador")

        self.tree.column("participante", width=260, anchor=tk.W)
        self.tree.column("tipo", width=100, anchor=tk.CENTER)
        self.tree.column("monto", width=140, anchor=tk.E)
        self.tree.column("pct", width=100, anchor=tk.E)
        self.tree.column("ganador", width=80, anchor=tk.CENTER)
        self.tree.pack(fill=tk.BOTH, expand=True)

        # Resaltado para ganador
        self.tree.tag_configure("ganador", background="#d4edda", font=("Helvetica", 9, "bold"))

        # Cargar datos desde el método que ya agregamos antes
        ofertas = self._recopilar_ofertas()  # ← EXISTE por el bloque que pegaste en el paso anterior

        for o in ofertas:
            pct = o.get('pct_vs_base', None)
            pct_txt = f"{pct:.2f}%" if isinstance(pct, (int, float)) else "N/D"
            tags = ("ganador",) if o.get('ganador') else ()
            self.tree.insert(
                "",
                tk.END,
                values=(
                    o.get('participante', ''),
                    o.get('tipo', ''),
                    f"RD$ {float(o.get('monto', 0) or 0):,.2f}",
                    pct_txt,
                    "Sí" if o.get('ganador') else "No"
                ),
                tags=tags
            )



    def _recopilar_ofertas(self):
        """
        Devuelve una lista de dicts con las ofertas del lote self.num_lote:
        [
        {
            'participante': str,
            'tipo': 'Nosotros' | 'Competidor',
            'monto': float,
            'ganador': bool,
            'pct_vs_base': float | None,
            'lote_num': str,
            'nombre_lote': str
        }, ...
        ]
        """
        lista = []
        lic = self.licitacion
        num = str(self.num_lote)

        # Buscar el lote
        lote = next((l for l in getattr(lic, "lotes", []) if str(getattr(l, "numero", "")) == num), None)
        if not lote:
            return lista

        nombre_lote = getattr(lote, "nombre", "") or ""
        base = float(getattr(lote, "monto_base_personal", 0) or getattr(lote, "monto_base", 0) or 0)

        # NUESTRA OFERTA (si existe)
        monto_nuestro = float(getattr(lote, "monto_ofertado", 0) or 0)
        if monto_nuestro > 0:
            nom_nuestra = (getattr(lote, "empresa_nuestra", None) or "NOSOTROS").strip() or "NOSOTROS"
            lista.append({
                'participante': nom_nuestra,
                'tipo': 'Nosotros',
                'monto': monto_nuestro,
                'ganador': bool(getattr(lote, "ganado_por_nosotros", False)),
                'pct_vs_base': ((monto_nuestro - base) / base * 100.0) if base > 0 else None,
                'lote_num': num,
                'nombre_lote': nombre_lote
            })

        # COMPETIDORES
        for comp in getattr(lic, "oferentes_participantes", []) or []:
            if isinstance(comp, dict):
                nombre_comp = (comp.get("nombre") or "").strip()
                ofertas = comp.get("ofertas_por_lote", []) or []
            else:
                nombre_comp = (getattr(comp, "nombre", "") or "").strip()
                ofertas = getattr(comp, "ofertas_por_lote", []) or []

            if not nombre_comp:
                continue

            for o in ofertas:
                if str(o.get("lote_numero")) == num:
                    monto = float(o.get("monto", 0) or 0)
                    ganador = bool(o.get("ganador", False))
                    lista.append({
                        'participante': nombre_comp,
                        'tipo': 'Competidor',
                        'monto': monto,
                        'ganador': ganador,
                        'pct_vs_base': ((monto - base) / base * 100.0) if base > 0 and monto > 0 else None,
                        'lote_num': num,
                        'nombre_lote': nombre_lote
                    })

        # Orden: primero los que tienen monto, de menor a mayor
        lista.sort(key=lambda d: (0 if d['monto'] > 0 else 1, d['monto']))
        return lista



    def _crear_tabla_comparativa(self, parent, ofertas):
        if not ofertas:
            ttk.Label(parent, text="No hay ofertas habilitadas para comparar en este lote.").pack()
            return

        oferentes = list(ofertas.keys())
        criterios = ["Monto Ofertado", "Plazo (días)", "Garantía (meses)"]

        tree = ttk.Treeview(parent, columns=['criterio'] + oferentes, show="headings")
        tree.heading('criterio', text='Criterio')
        tree.column('criterio', width=150, anchor=tk.W)
        for oferente in oferentes:
            tree.heading(oferente, text=oferente)
            tree.column(oferente, width=150, anchor=tk.E)

        # Encontrar los mejores valores para resaltar
        valores_monto = [d['monto'] for d in ofertas.values() if d['monto'] > 0]
        valores_plazo = [d['plazo'] for d in ofertas.values() if d['plazo'] > 0]
        valores_garantia = [d['garantia'] for d in ofertas.values() if d['garantia'] > 0]

        mejor_monto = min(valores_monto) if valores_monto else None
        mejor_plazo = min(valores_plazo) if valores_plazo else None
        mejor_garantia = max(valores_garantia) if valores_garantia else None

        # Poblar filas
        # Fila de Monto
        row_monto = [criterios[0]]
        for oferente in oferentes:
            monto = ofertas[oferente]['monto']
            display = f"RD$ {monto:,.2f}"
            if monto == mejor_monto: display += " ⭐" # Mejor oferta
            row_monto.append(display)
        tree.insert("", tk.END, values=row_monto)

        # Fila de Plazo
        row_plazo = [criterios[1]]
        for oferente in oferentes:
            plazo = ofertas[oferente]['plazo']
            display = f"{plazo} días"
            if plazo == mejor_plazo: display += " ⭐"
            row_plazo.append(display)
        tree.insert("", tk.END, values=row_plazo)

        # Fila de Garantía
        row_garantia = [criterios[2]]
        for oferente in oferentes:
            garantia = ofertas[oferente]['garantia']
            display = f"{garantia} meses"
            if garantia == mejor_garantia: display += " ⭐"
            row_garantia.append(display)
        tree.insert("", tk.END, values=row_garantia)

        tree.pack(fill=tk.BOTH, expand=True)


class DialogoSeleccionarDocumento(simpledialog.Dialog):
    """Un diálogo moderno con checkboxes, búsqueda y filtro por categoría para seleccionar múltiples documentos de plantilla."""
    def __init__(self, parent, title, documentos_maestros, documentos_actuales):
        codigos_actuales = {doc.codigo for doc in documentos_actuales}
        self.documentos_disponibles = [
            doc for doc in documentos_maestros if doc.codigo not in codigos_actuales
        ]
        
        # --- CAMBIO 1: Lógica para obtener categorías ---
        categorias_unicas = sorted(list(set(doc.categoria for doc in self.documentos_disponibles if doc.categoria)))
        self.categorias_filtro = ["Todas"] + categorias_unicas
        # --- FIN DEL CAMBIO ---
        
        self.selection_status = {doc.id: tk.BooleanVar(value=False) for doc in self.documentos_disponibles}
        super().__init__(parent, title)

# En la clase DialogoSeleccionarDocumento, REEMPLAZA este método:

    def body(self, master):
        self.geometry("800x400")
        
        filter_frame = ttk.Frame(master)
        filter_frame.pack(fill=tk.X, padx=10, pady=5)
        filter_frame.columnconfigure(1, weight=1)

        ttk.Label(filter_frame, text="🔍 Buscar:").grid(row=0, column=0, padx=(0, 5), pady=5)
        self.search_var = tk.StringVar()
        search_entry = ttk.Entry(filter_frame, textvariable=self.search_var)
        search_entry.grid(row=0, column=1, sticky="ew", pady=5)
        
        ttk.Label(filter_frame, text="Categoría:").grid(row=0, column=2, padx=(10, 5), pady=5)
        self.categoria_var = tk.StringVar(value="Todas")
        categoria_combo = ttk.Combobox(filter_frame, textvariable=self.categoria_var, values=self.categorias_filtro, state="readonly", width=15)
        categoria_combo.grid(row=0, column=3, pady=5)

        tree_frame = ttk.Frame(master)
        # --- LÍNEA CORREGIDA ---
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(5, 10))
        
        cols = ('codigo', 'nombre', 'categoria')
        self.tree = ttk.Treeview(tree_frame, columns=cols, show='tree headings')
        self.tree.column("#0", width=40, anchor=tk.CENTER, stretch=False)
        self.tree.heading("#0", text="Sel.")
        self.tree.heading('codigo', text='Código')
        self.tree.heading('nombre', text='Nombre del Documento')
        self.tree.heading('categoria', text='Categoría')
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.search_var.trace_add("write", lambda *args: self._populate_treeview())
        categoria_combo.bind("<<ComboboxSelected>>", lambda *args: self._populate_treeview())
        
        self.tree.bind("<Button-1>", self._toggle_selection)
        self._populate_treeview()
        
        return search_entry
    
    def _populate_treeview(self):
        self.tree.delete(*self.tree.get_children())
        search_term = self.search_var.get().lower()
        categoria_seleccionada = self.categoria_var.get()
        
        for doc in self.documentos_disponibles:
            nombre = doc.nombre or ""
            codigo = doc.codigo or ""
            categoria = doc.categoria or ""
            
            # --- CAMBIO 4: Lógica de filtrado combinada ---
            filtro_busqueda_pasa = (search_term in nombre.lower() or search_term in codigo.lower())
            filtro_categoria_pasa = (categoria_seleccionada == "Todas" or categoria_seleccionada == categoria)
            # --- FIN DEL CAMBIO ---

            if filtro_busqueda_pasa and filtro_categoria_pasa:
                check_char = '☑' if self.selection_status[doc.id].get() else '☐'
                self.tree.insert('', tk.END, text=check_char, values=(codigo, nombre, categoria), iid=doc.id)

    def _toggle_selection(self, event):
        row_id_str = self.tree.identify_row(event.y)
        if not row_id_str: return
        
        row_id = int(row_id_str)
        self.selection_status[row_id].set(not self.selection_status[row_id].get())
        
        check_char = '☑' if self.selection_status[row_id].get() else '☐'
        self.tree.item(row_id, text=check_char)
        
    def apply(self):
        self.result = [doc for doc in self.documentos_disponibles if self.selection_status[doc.id].get()]


# REEMPLAZA esta clase por completo en tu archivo:

# REEMPLAZA esta clase por completo en tu archivo:

class VentanaAnalisisFaseA(tk.Toplevel):
    """Ventana para registrar y analizar las causas de descalificación en Fase A."""
    def __init__(self, parent, licitacion, db_manager):
        super().__init__(parent)
        self.parent_app = parent.parent_app
        self.licitacion = licitacion
        self.db = db_manager # Lo mantenemos por si se usa en otro lado
        
        self.title(f"Análisis de Fallas Fase A - {self.licitacion.numero_proceso}")
        self.geometry("1200x700")
        self.grab_set()

        self.documentos_seleccionados = {doc.id: tk.BooleanVar(value=False) for doc in self.licitacion.documentos_solicitados}
        
        main_frame = ttk.Frame(self, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        main_pane = ttk.PanedWindow(main_frame, orient=tk.HORIZONTAL)
        main_pane.pack(fill=tk.BOTH, expand=True)
        left_pane = ttk.PanedWindow(main_pane, orient=tk.VERTICAL)
        main_pane.add(left_pane, weight=2)
        
        self._crear_panel_participantes(left_pane)
        self._crear_panel_documentos(left_pane)
        right_frame = ttk.Frame(main_pane)
        main_pane.add(right_frame, weight=3)
        self._crear_panel_registro(right_frame)

        # Cambiamos el botón para que solo cierre la ventana
        ttk.Button(main_frame, text="✅ Aceptar y Cerrar", command=self.destroy).pack(pady=(10,0), ipady=5)

        # Cargamos la vista con las fallas que ya tiene la licitación en memoria
        self._refrescar_vista_fallas()

    def _crear_panel_participantes(self, parent):
        frame = ttk.LabelFrame(parent, text="1. Seleccione Participante(s)", padding=10)
        parent.add(frame, weight=1)
        
        search_frame = ttk.Frame(frame)
        search_frame.pack(fill=tk.X, pady=(0, 5))
        ttk.Label(search_frame, text="🔍 Buscar:").pack(side=tk.LEFT, padx=(0, 5))
        self.participante_search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=self.participante_search_var)
        search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        tree_container = ttk.Frame(frame)
        tree_container.pack(fill=tk.BOTH, expand=True)

        self.tree_participantes = ttk.Treeview(tree_container, columns=('nombre', 'tipo'), show='headings', selectmode="extended")
        self.tree_participantes.heading('nombre', text='Nombre'); self.tree_participantes.heading('tipo', text='Tipo')
        self.tree_participantes.column('tipo', width=100, anchor='center')
        
        scrollbar = ttk.Scrollbar(tree_container, orient="vertical", command=self.tree_participantes.yview)
        self.tree_participantes.configure(yscrollcommand=scrollbar.set)
        self.tree_participantes.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.participante_search_var.trace_add("write", lambda *args: self._poblar_participantes())
        self._poblar_participantes()

    def _crear_panel_documentos(self, parent):
        frame = ttk.LabelFrame(parent, text="2. Marque Documento(s) con Fallas", padding=10)
        parent.add(frame, weight=2)

        self.tree_docs = ttk.Treeview(frame, columns=('nombre', 'codigo'), show='tree headings')
        self.tree_docs.column("#0", width=40, anchor=tk.CENTER, stretch=False)
        self.tree_docs.heading("#0", text="Sel.")
        self.tree_docs.heading('nombre', text='Nombre del Documento')
        self.tree_docs.heading('codigo', text='Código')
        
        scrollbar_docs = ttk.Scrollbar(frame, orient="vertical", command=self.tree_docs.yview)
        self.tree_docs.configure(yscrollcommand=scrollbar_docs.set)
        
        self.tree_docs.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar_docs.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.tree_docs.bind("<Button-1>", self._toggle_doc_selection)
        for doc in sorted(self.licitacion.documentos_solicitados, key=lambda d: d.codigo or ''):
            if doc.id:
                self.tree_docs.insert('', tk.END, text='☐', values=(doc.nombre, doc.codigo), iid=doc.id)

    def _crear_panel_registro(self, parent):
        frame = ttk.Frame(parent, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="3. Comentario (Opcional) y Añadir a la Lista", font=("Segoe UI", 11, "bold")).pack(anchor="w", pady=(0, 10))
        self.comentario_text = tk.Text(frame, height=4)
        self.comentario_text.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Button(frame, text="⬇️ Añadir Falla(s) a la Lista", command=self.anadir_fallas_a_memoria).pack(fill=tk.X, ipady=5, pady=(0, 20))
        
        resultados_frame = ttk.LabelFrame(frame, text="Fallas a Registrar (Lista Temporal)", padding=10)
        resultados_frame.pack(fill=tk.BOTH, expand=True)
        
        self.tree_fallas = ttk.Treeview(resultados_frame, columns=('participante', 'documento', 'comentario'), show='headings')
        self.tree_fallas.heading('participante', text='Participante')
        self.tree_fallas.heading('documento', text='Documento Fallido')
        self.tree_fallas.heading('comentario', text='Comentario')
        self.tree_fallas.pack(fill=tk.BOTH, expand=True)

    def _poblar_participantes(self):
        self.tree_participantes.delete(*self.tree_participantes.get_children())
        search_term = self.participante_search_var.get().lower()
        nuestras_empresas = self._nuestras_empresas_de_licitacion()
        
        todos_los_participantes = []
        for emp in nuestras_empresas: todos_los_participantes.append({'nombre': emp, 'tipo': "Nuestra"})
        for oferente in self.licitacion.oferentes_participantes:
            if oferente.nombre not in nuestras_empresas:
                todos_los_participantes.append({'nombre': oferente.nombre, 'tipo': "Competidor"})
        
        todos_los_participantes.sort(key=lambda p: p['nombre'])
        
        for p in todos_los_participantes:
            if search_term in p['nombre'].lower():
                self.tree_participantes.insert('', tk.END, values=(p['nombre'], p['tipo']), iid=p['nombre'])

    def _toggle_doc_selection(self, event):
        """Maneja el clic en una fila para marcar/desmarcar el checkbox."""
        iid = self.tree_docs.identify_row(event.y)
        if not iid: return
        
        try:
            doc_id = int(iid)
            if doc_id in self.documentos_seleccionados:
                var = self.documentos_seleccionados[doc_id]
                var.set(not var.get())
                check_char = '☑' if var.get() else '☐'
                self.tree_docs.item(iid, text=check_char)
        except (ValueError, KeyError):
            # Ignorar clics en áreas sin ID válido
            pass

    def _nuestras_empresas_de_licitacion(self):
        return {str(e) for e in self.licitacion.empresas_nuestras}


    def cargar_fallas_existentes(self):
        """Este método ya no es necesario, los datos vienen en self.licitacion."""
        self._refrescar_vista_fallas()

# En la clase VentanaAnalisisFaseA

    def _refrescar_vista_fallas(self):
        self.tree_fallas.delete(*self.tree_fallas.get_children())
        mapa_docs = {doc.id: doc.nombre for doc in self.licitacion.documentos_solicitados}
        
        # La fuente de datos ahora es la lista dentro del objeto licitacion
        for falla in self.licitacion.fallas_fase_a:
            # Accedemos a los datos usando las claves del diccionario
            participante = falla.get('participante_nombre')
            doc_id = falla.get('documento_id')
            comentario = falla.get('comentario')
            
            doc_nombre = mapa_docs.get(doc_id, "Documento no encontrado")
            self.tree_fallas.insert('', tk.END, values=(participante, doc_nombre, comentario))
        
    def anadir_fallas_a_memoria(self):
        """Añade las selecciones actuales a la lista de fallas del objeto Licitacion."""
        participantes_sel = [self.tree_participantes.item(iid, 'values')[0] for iid in self.tree_participantes.selection()]
        documentos_sel_ids = [doc_id for doc_id, var in self.documentos_seleccionados.items() if var.get()]
        comentario = self.comentario_text.get("1.0", tk.END).strip()
        
        if not participantes_sel or not documentos_sel_ids:
            messagebox.showwarning("Datos Faltantes", "Debe seleccionar al menos un participante y un documento.", parent=self)
            return

        nuestras_empresas = self._nuestras_empresas_de_licitacion()
        nuevas_fallas_count = 0
        
        for participante in participantes_sel:
            es_nuestro = participante in nuestras_empresas
            for doc_id in documentos_sel_ids:
                # Comprobar si la falla ya existe en la lista de la licitación
                if not any(f['participante_nombre'] == participante and f['documento_id'] == doc_id for f in self.licitacion.fallas_fase_a):
                    nueva_falla = {
                        "licitacion_id": self.licitacion.id,
                        "participante_nombre": participante,
                        "documento_id": doc_id,
                        "comentario": comentario,
                        "es_nuestro": es_nuestro
                    }
                    self.licitacion.fallas_fase_a.append(nueva_falla)
                    nuevas_fallas_count += 1
        
        if nuevas_fallas_count > 0:
            self._refrescar_vista_fallas()
            self.comentario_text.delete("1.0", tk.END)
            for var in self.documentos_seleccionados.values(): var.set(False)
            for iid in self.tree_docs.get_children(): self.tree_docs.item(iid, text='☐')
        else:
            messagebox.showinfo("Información", "Las fallas seleccionadas ya estaban en la lista.", parent=self)

    def _guardar_y_cerrar(self):
        """Este método ya no necesita guardar. La ventana de Detalles lo hará."""
        self.destroy() # Simplemente cerramos la ventana


# Pega esta NUEVA clase en tu archivo

class DialogoConfirmarImportacion(simpledialog.Dialog):
    """
    Una ventana avanzada para revisar documentos seleccionados,
    permitiendo cambiar su categoría de forma masiva o individual antes de importar.
    """
    def __init__(self, parent, documentos_seleccionados, categorias_disponibles):
        self.documentos = documentos_seleccionados
        self.categorias = categorias_disponibles
        super().__init__(parent, "Confirmar y Categorizar Documentos")

    def body(self, master):
        self.geometry("800x500")

        # --- Panel de Acción Masiva ---
        bulk_frame = ttk.Frame(master)
        bulk_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(bulk_frame, text="Aplicar esta categoría a TODOS:").pack(side=tk.LEFT, padx=(0, 5))
        self.bulk_categoria_var = tk.StringVar()
        self.bulk_combo = ttk.Combobox(bulk_frame, textvariable=self.bulk_categoria_var, values=self.categorias, state="readonly", width=20)
        self.bulk_combo.pack(side=tk.LEFT)
        ttk.Button(bulk_frame, text="Aplicar a Todos", command=self._aplicar_a_todos).pack(side=tk.LEFT, padx=5)

        # --- Treeview para edición individual ---
        tree_frame = ttk.Frame(master)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        cols = ('codigo', 'nombre', 'categoria')
        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings")
        self.tree.heading('codigo', text='Código'); self.tree.column('codigo', width=150)
        self.tree.heading('nombre', text='Nombre del Documento'); self.tree.column('nombre', width=400)
        self.tree.heading('categoria', text='Categoría (Doble Clic para Editar)'); self.tree.column('categoria', width=150)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        for doc in self.documentos:
            self.tree.insert('', tk.END, iid=doc.id, values=(doc.codigo, doc.nombre, doc.categoria))
            
        self.tree.bind("<Double-1>", self._editar_celda)
        return self.tree

    def _aplicar_a_todos(self):
        """Aplica la categoría seleccionada en el combo a todos los items del treeview."""
        nueva_categoria = self.bulk_categoria_var.get()
        if not nueva_categoria: return
        
        for iid in self.tree.get_children():
            self.tree.set(iid, 'categoria', nueva_categoria)

    def _editar_celda(self, event):
        """Crea un combobox de edición in-place sobre la celda de categoría."""
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell": return

        iid = self.tree.identify_row(event.y)
        column = self.tree.identify_column(event.x)
        if column != "#3": # Columna de categoría
            return

        x, y, width, height = self.tree.bbox(iid, column)
        
        valor_actual = self.tree.set(iid, column)
        combo = ttk.Combobox(self.tree, values=self.categorias, state="readonly")
        combo.place(x=x, y=y, width=width, height=height)
        combo.set(valor_actual)
        combo.focus_set()

        def on_selection_change(event=None):
            self.tree.set(iid, column, combo.get())
            combo.destroy()
        
        combo.bind("<<ComboboxSelected>>", on_selection_change)
        combo.bind("<FocusOut>", lambda e: combo.destroy())

    def apply(self):
        self.result = []
        for iid in self.tree.get_children():
            valores = self.tree.item(iid, 'values')
            self.result.append({
                'codigo': valores[0],
                'nombre': valores[1],
                'categoria': valores[2],
                'id_maestro': iid # Pasamos el ID del documento maestro original
            })



class VentanaGestionDocumentos(tk.Toplevel):
    def __init__(
        self,
        parent,
        licitacion,
        callback=None,
        documentos_maestros=None,
        categorias=None,
        todas_las_licitaciones=None,
        lista_responsables=None,
        on_docs_changed=None,
        *args, **kwargs
    ):
        super().__init__(parent)
        self.parent = parent
        self.parent_app = parent.parent_app 
        self.licitacion = licitacion

        self.callback_actualizar = callback
        self._on_docs_changed = on_docs_changed

        self.documentos_maestros = documentos_maestros or []
        self.categorias = categorias or ["Legal", "Financiera", "Técnica", "Sobre B", "Otros"]
        self.todas_las_licitaciones = todas_las_licitaciones or []

        lr = lista_responsables or []
        if isinstance(lr, dict):
            lr = [lr]
        self.lista_responsables = ["Sin Asignar"] + sorted(
            [r["nombre"] if isinstance(r, dict) else str(r) for r in lr]
        )

        self.title(f"Gestionar Documentos de {licitacion.nombre_proceso}")
        self.geometry("1200x700") # Un poco más ancha para las columnas
        self.grab_set()

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(pady=10, padx=10, fill="both", expand=True)
        
        # --- LÓGICA DE INTERFAZ ACTUALIZADA ---
        self.trees = {} # Usaremos un diccionario para los Treeviews
        for categoria in self.categorias:
            frame = ttk.Frame(self.notebook, padding="10")
            self.notebook.add(frame, text=categoria)

            # Creamos el Treeview con sus columnas
            cols = ('estado', 'rev', 'adj', 'codigo', 'nombre', 'condicion', 'responsable')
            tree = ttk.Treeview(frame, columns=cols, show="headings", selectmode=tk.EXTENDED)
            
            tree.heading('estado', text='✓')
            tree.heading('rev', text='👁️')
            tree.heading('adj', text='📎')
            tree.heading('codigo', text='Código')
            tree.heading('nombre', text='Nombre del Documento')
            tree.heading('condicion', text='Condición')
            tree.heading('responsable', text='Responsable')

            tree.column('estado', width=30, anchor=tk.CENTER, stretch=False)
            tree.column('rev', width=30, anchor=tk.CENTER, stretch=False)
            tree.column('adj', width=30, anchor=tk.CENTER, stretch=False)
            tree.column('codigo', width=120)
            tree.column('nombre', width=450)
            tree.column('condicion', width=100, anchor=tk.CENTER)
            tree.column('responsable', width=150)
            
            tree.pack(side=tk.LEFT, fill="both", expand=True)
            scrollbar = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
            tree.configure(yscrollcommand=scrollbar.set)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            self.trees[categoria] = tree
            tree.bind("<<TreeviewSelect>>", self.on_doc_select)
        # --- FIN DE LA LÓGICA DE INTERFAZ ---

        # Barra de acciones rápidas (sin cambios)
        action_frame = ttk.Frame(self)
        action_frame.pack(pady=5, fill="x", padx=10)
        ttk.Label(action_frame, text="Asignar Responsable:").pack(side=tk.LEFT, padx=(0, 5))
        self.responsable_var = tk.StringVar()
        self.responsable_combo = ttk.Combobox(action_frame, textvariable=self.responsable_var, values=self.lista_responsables, state="disabled", width=20)
        self.responsable_combo.pack(side=tk.LEFT, padx=5)
        self.responsable_combo.bind("<<ComboboxSelected>>", self._guardar_responsable_multiple)
        ttk.Separator(action_frame, orient="vertical").pack(side=tk.LEFT, padx=15, fill="y")
        self.revisado_button = ttk.Button(action_frame, text="👁️ Marcar como Revisado/No Revisado", command=self._toggle_estado_revisado, state="disabled")
        self.revisado_button.pack(side=tk.LEFT, padx=5)
        self.subsanable_button = ttk.Button(action_frame, text="⚖️ Cambiar Condición (Subsanable)", command=self.cambiar_estado_subsanable, state="disabled")
        self.subsanable_button.pack(side=tk.LEFT, padx=5)

        # Botones de gestión (sin cambios)
        btn_frame = ttk.Frame(self)
        btn_frame.pack(pady=5, fill="x", padx=10)
        btn_frame.columnconfigure(tuple(range(4)), weight=1)
        acciones = {
            "➕ Agregar Manual": self.agregar_manual, "📥 Importar de otra Licitación": self.importar_desde_licitacion,
            "✨ Agregar desde Plantilla": self.agregar_desde_plantilla, "✏️ Editar": self.editar_documento,
            "🟢/❌ Cambiar Estado": self.cambiar_estado_documento, "🗑️ Eliminar": self.eliminar_documento,
            "📎 Adjuntar Archivo": self.adjuntar_archivo, "📂 Ver Archivo": self.ver_archivo, "❌ Quitar Adjunto": self.quitar_adjunto,
        }
        self.buttons = {}
        for i, (text, cmd) in enumerate(acciones.items()):
            btn = ttk.Button(btn_frame, text=text, command=cmd)
            btn.grid(row=i // 4, column=i % 4, sticky=tk.EW, padx=5, pady=2)
            self.buttons[text] = btn

        # Render inicial
        self.actualizar_listas_docs()
        self.on_doc_select(None)

    # --- MÉTODO _get_active_category_and_tree ACTUALIZADO ---
    def _get_active_category_and_tree(self):
        try:
            tab_id = self.notebook.select()
            cat_activa = self.notebook.tab(tab_id, "text")
            return cat_activa, self.trees[cat_activa]
        except (KeyError, tk.TclError):
            return None, None

    # --- MÉTODO _find_docs_from_selection ACTUALIZADO ---
    def _find_docs_from_selection(self):
        """Devuelve los objetos Documento seleccionados en el Treeview activo."""
        _, tree = self._get_active_category_and_tree()
        if not tree:
            return []

        selected_iids = set(tree.selection())
        if not selected_iids:
            return []

        # Mapeamos por el iid que realmente metimos en el Treeview (id real o tmp-...)
        result = []
        for doc in self.licitacion.documentos_solicitados:
            iid = self._iid_for_doc(doc)
            if iid in selected_iids:
                result.append(doc)
        return result


    @staticmethod
    def _orden_sort_val(d):
        """Normaliza orden_pliego a int; si está vacío/None/no convertible, lo manda al final."""
        val = getattr(d, "orden_pliego", None)
        try:
            return int(val)
        except (TypeError, ValueError):
            return 999_999
        

    def _iid_for_doc(self, doc):
        """IID estable para Treeview: usa id de DB si existe; si no, un temporal."""
        return f"doc-{doc.id}" if getattr(doc, "id", None) else f"tmp-{id(doc)}"


    # --- MÉTODO actualizar_listas_docs TOTALMENTE REDISEÑADO ---
    def actualizar_listas_docs(self):
        """Redibuja los Treeviews por categoría con datos en columnas."""
        # Agrupar por categoría
        docs_por_categoria = {cat: [] for cat in self.categorias}
        for doc in self.licitacion.documentos_solicitados:
            cat = getattr(doc, "categoria", None)
            if cat in docs_por_categoria:
                docs_por_categoria[cat].append(doc)

        # Limpiar y volver a poblar cada tree por categoría
        for categoria, tree in self.trees.items():
            # 1) limpiar
            tree.delete(*tree.get_children())

            # 2) log opcional: detectar documentos sin orden
            for d in docs_por_categoria.get(categoria, []):
                if getattr(d, "orden_pliego", None) in (None, "", " ", "N/A"):
                    print("[WARN] Doc sin orden_pliego:", getattr(d, "codigo", ""), getattr(d, "nombre", ""))

            # 3) ordenar de forma robusta
            documentos_ordenados = sorted(
                docs_por_categoria.get(categoria, []),
                key=lambda d: (self._orden_sort_val(d), d.codigo or "", d.nombre or "")
            )

            # 4) insertar filas (incluye docs sin id)
            for doc in documentos_ordenados:
                iid = self._iid_for_doc(doc)  # <- clave: iid estable

                estado = "✓" if getattr(doc, "presentado", False) else "❌"
                revisado = "✓" if getattr(doc, "revisado", False) else ""
                adjunto = "✓" if getattr(doc, "ruta_archivo", "") else ""
                condicion = getattr(doc, "subsanable", None) or "No Definido"
                responsable = getattr(doc, "responsable", "") or "Sin Asignar"

                tree.insert(
                    "",
                    tk.END,
                    iid=iid,
                    values=(estado, revisado, adjunto, doc.codigo, doc.nombre, condicion, responsable),
                )


        # Refrescar estado de botones/combos
        self.on_doc_select(None)

    def _notify_docs_changed(self):
        try:
            if callable(self._on_docs_changed): self._on_docs_changed()
            elif callable(self.callback_actualizar): self.callback_actualizar()
            elif hasattr(self.parent, "actualizar_info_docs"): self.parent.actualizar_info_docs()
        except Exception: pass

    def _next_orden(self):
        docs = getattr(self.licitacion, "documentos_solicitados", []) or []
        if not docs: return 1
        try: return max(int(getattr(d, "orden_pliego", 0) or 0) for d in docs) + 1
        except Exception: return len(docs) + 1

    def on_doc_select(self, event=None):
        docs = self._find_docs_from_selection()
        state_if_selection = "normal" if docs else "disabled"
        for key in ["🟢/❌ Cambiar Estado", "🗑️ Eliminar", "📎 Adjuntar Archivo"]:
            self.buttons[key].config(state=state_if_selection)
        self.responsable_combo.config(state="readonly" if docs else "disabled")
        self.revisado_button.config(state=state_if_selection)
        state_if_single = "normal" if len(docs) == 1 else "disabled"
        self.buttons["✏️ Editar"].config(state=state_if_single)
        self.subsanable_button.config(state=state_if_single)
        doc = docs[0] if len(docs) == 1 else None
        self.buttons["📂 Ver Archivo"].config(state="normal" if (doc and getattr(doc, "ruta_archivo", "")) else "disabled")
        any_has_file = any(getattr(d, "ruta_archivo", "") for d in docs)
        self.buttons["❌ Quitar Adjunto"].config(state="normal" if (docs and any_has_file) else "disabled")
        if doc: self.responsable_var.set(doc.responsable or "Sin Asignar")
        else: self.responsable_var.set("")

# En la clase VentanaGestionDocumentos, REEMPLAZA este método:

# En la clase VentanaGestionDocumentos, REEMPLAZA este método:

    def agregar_desde_plantilla(self):
        nombres_empresas_participantes = {str(e) for e in self.licitacion.empresas_nuestras}
        if not nombres_empresas_participantes:
            messagebox.showinfo("Sin Empresas", "Primero debe asignar al menos una de sus empresas a esta licitación.", parent=self)
            return

        plantillas_disponibles = [
            d for d in self.documentos_maestros 
            if getattr(d, "empresa_nombre", None) in nombres_empresas_participantes
        ]
        if not plantillas_disponibles:
            empresas_str = ", ".join(sorted(list(nombres_empresas_participantes)))
            messagebox.showinfo("Sin Plantillas", f"No hay plantillas de documentos asociadas a: {empresas_str}.", parent=self)
            return

        dialogo_seleccion = DialogoSeleccionarDocumento(self, "Seleccionar de Plantilla", plantillas_disponibles, self.licitacion.documentos_solicitados)
        
        if dialogo_seleccion.result:
            documentos_a_importar = dialogo_seleccion.result
            
            # --- NUEVO PASO: Abrir el diálogo de confirmación y categorización ---
            dialogo_confirmacion = DialogoConfirmarImportacion(self, documentos_a_importar, self.categorias)
            
            if dialogo_confirmacion.result:
                nuevos_agregados = 0
                for doc_data in dialogo_confirmacion.result:
                    # Buscamos la plantilla original para copiar otros datos
                    doc_maestro = next((d for d in self.documentos_maestros if str(d.id) == str(doc_data['id_maestro'])), None)
                    if doc_maestro:
                        nuevo_doc = Documento(
                            codigo=doc_maestro.codigo,
                            nombre=doc_maestro.nombre,
                            categoria=doc_data['categoria'], # Usamos la categoría final elegida por el usuario
                            comentario=doc_maestro.comentario,
                            subsanable=getattr(doc_maestro, "subsanable", "Subsanable"),
                            obligatorio=bool(getattr(doc_maestro, "obligatorio", False))
                        )
                        self.licitacion.documentos_solicitados.append(nuevo_doc)
                        nuevos_agregados += 1
                
                if nuevos_agregados > 0:
                    self.actualizar_listas_docs()
                    self._notify_docs_changed()
                    messagebox.showinfo("Éxito", f"Se agregaron {nuevos_agregados} documentos.", parent=self)
        
        self.grab_set()

    def importar_desde_licitacion(self):
        dialogo = DialogoSeleccionarLicitacion(self, "Importar Documentos", self.todas_las_licitaciones, self.licitacion.numero_proceso)
        if dialogo.result:
            origen = next((l for l in self.todas_las_licitaciones if l.numero_proceso == dialogo.result), None)
            if not origen: return
            codigos_existentes = {d.codigo for d in self.licitacion.documentos_solicitados}
            for d_o in getattr(origen, "documentos_solicitados", []):
                if d_o.codigo in codigos_existentes: continue
                nuevo = Documento(codigo=d_o.codigo, nombre=d_o.nombre, categoria=d_o.categoria, subsanable=getattr(d_o, "subsanable", "No Definido"), comentario=getattr(d_o, "comentario", ""), obligatorio=bool(getattr(d_o, "obligatorio", False)))
                nuevo.orden_pliego = self._next_orden(); self.licitacion.documentos_solicitados.append(nuevo)
            self.actualizar_listas_docs(); self._notify_docs_changed()
        self.grab_set()

    def agregar_manual(self):
        cat, _ = self._get_active_category_and_tree()
        if not cat:
            messagebox.showwarning("Selección Requerida", "Por favor, selecciona una categoría (pestaña) primero.", parent=self)
            return
        empresa_principal = str(self.licitacion.empresas_nuestras[0]) if self.licitacion.empresas_nuestras else None
        dlg = DialogoAgregarDocumento(self, f"Nuevo Documento - {cat}", categorias=self.categorias, empresa_actual=empresa_principal)
        if dlg.result:
            datos = dlg.result
            nuevo_doc = Documento(codigo=datos["codigo"], nombre=datos["nombre"], categoria=datos["categoria"], comentario=datos["comentario"])
            self.licitacion.documentos_solicitados.append(nuevo_doc)
            if datos["guardar_plantilla"] and empresa_principal:
                if any(d.codigo == datos["codigo"] and d.empresa_nombre == empresa_principal for d in self.documentos_maestros):
                    messagebox.showwarning("Plantilla Duplicada", f"Ya existe una plantilla con el código '{datos['codigo']}' para la empresa '{empresa_principal}'.\nEl documento se agregó a la licitación, pero no a la plantilla.", parent=self)
                else:
                    doc_plantilla = Documento(codigo=datos["codigo"], nombre=datos["nombre"], categoria=datos["categoria"], comentario=datos["comentario"], empresa_nombre=empresa_principal)
                    self.parent_app.documentos_maestros.append(doc_plantilla)
                    self.parent_app.db.save_master_lists(empresas=self.parent_app.empresas_registradas, instituciones=self.parent_app.instituciones_registradas, documentos_maestros=self.parent_app.documentos_maestros, competidores_maestros=self.parent_app.competidores_maestros, responsables_maestros=self.parent_app.responsables_maestros, replace_tables={'documentos_maestros'})
            self.actualizar_listas_docs(); self._notify_docs_changed()
        self.grab_set()

    def editar_documento(self):
        docs = self._find_docs_from_selection()
        if not docs or len(docs) > 1: return
        doc = docs[0]
        dlg = DialogoAgregarDocumento(self, "Editar Documento", initial_data=doc, categorias=self.categorias)
        if dlg.result:
            # Corrección: El resultado de DialogoAgregarDocumento es un diccionario
            datos = dlg.result
            doc.codigo = datos['codigo']
            doc.nombre = datos['nombre']
            doc.categoria = datos['categoria']
            doc.comentario = datos['comentario']
            self.actualizar_listas_docs()
            self._notify_docs_changed()
        self.grab_set()

    def cambiar_estado_documento(self):
        docs = self._find_docs_from_selection()
        if not docs: return
        nuevo = not all(d.presentado for d in docs)
        for d in docs: d.presentado = nuevo
        self.actualizar_listas_docs(); self._notify_docs_changed()

    def eliminar_documento(self):
        docs = self._find_docs_from_selection()
        if not docs: return
        if messagebox.askyesno("Confirmar", f"¿Eliminar {len(docs)} documento(s) seleccionado(s)?", parent=self):
            for d in docs:
                try: self.licitacion.documentos_solicitados.remove(d)
                except ValueError: pass
            self.actualizar_listas_docs(); self._notify_docs_changed()
            self.grab_set()

    def cambiar_estado_subsanable(self):
        docs = self._find_docs_from_selection()
        if not docs or len(docs) > 1: return
        doc = docs[0]
        estados = ["No Definido", "Subsanable", "No Subsanable"]
        try: idx = estados.index(doc.subsanable)
        except ValueError: idx = 0
        doc.subsanable = estados[(idx + 1) % len(estados)]
        self.actualizar_listas_docs(); self._notify_docs_changed()
        self.grab_set()

    def _guardar_responsable_multiple(self, event=None):
        docs = self._find_docs_from_selection()
        nuevo = self.responsable_var.get()
        if docs and nuevo:
            for d in docs: d.responsable = nuevo
            self.actualizar_listas_docs(); self._notify_docs_changed()

    def adjuntar_archivo(self):
        docs = self._find_docs_from_selection()
        if not docs: messagebox.showwarning("Sin selección", "Seleccione al menos un documento.", parent=self); return
        ruta = filedialog.askopenfilename(parent=self, title="Seleccionar Archivo")
        if not ruta: return
        for doc in docs:
            doc.ruta_archivo = ruta; doc.presentado = True
        self.actualizar_listas_docs(); self._notify_docs_changed()

    def ver_archivo(self):
        docs = self._find_docs_from_selection()
        if not docs or len(docs) > 1 or not getattr(docs[0], "ruta_archivo", ""): return
        try: os.startfile(docs[0].ruta_archivo)
        except Exception as e: messagebox.showerror("Error", f"No se pudo abrir el archivo: {e}", parent=self)

    def quitar_adjunto(self):
        docs = self._find_docs_from_selection()
        if not docs: return
        con_archivo = [d for d in docs if getattr(d, "ruta_archivo", "")]
        if not con_archivo: messagebox.showinfo("Sin adjuntos", "Ninguno de los documentos seleccionados tiene archivo adjunto.", parent=self); return
        plural = "s" if len(con_archivo) > 1 else ""
        if not messagebox.askyesno("Confirmar", f"¿Quitar el archivo adjunto de {len(con_archivo)} documento{plural}?", parent=self): return
        for d in con_archivo: d.ruta_archivo = ""; d.presentado = False
        self.actualizar_listas_docs(); self._notify_docs_changed()

    def _toggle_estado_revisado(self):
        docs = self._find_docs_from_selection()
        if not docs: return
        nuevo = not all(getattr(d, "revisado", False) for d in docs)
        for d in docs: d.revisado = nuevo
        self.actualizar_listas_docs(); self._notify_docs_changed()

    def on_guardar_y_continuar(self):
        self._notify_docs_changed(); self.destroy()

# REEMPLAZA esta clase por completo en tu archivo:

class VentanaMaestroDocumentos(tk.Toplevel):
    def __init__(self, parent, documentos_maestros, categorias, db_manager):
        super().__init__(parent)
        self.parent_app = parent
        self.documentos_maestros = documentos_maestros
        self.categorias_documentos = ["Todas"] + categorias
        self.db = db_manager
        
        self.title("Gestor de Plantillas de Documentos por Empresa")
        self.geometry("950x650") # Un poco más ancha para la nueva tabla
        self.grab_set()
        
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # --- Panel superior con todos los filtros ---
        filter_frame = ttk.Frame(main_frame, padding=5)
        filter_frame.pack(fill=tk.X, pady=5)
        filter_frame.columnconfigure(1, weight=1) # Hacemos que la búsqueda se expanda

        # Filtro de Empresa
        ttk.Label(filter_frame, text="Seleccionar Empresa:").grid(row=0, column=0, padx=(0,5), sticky="w")
        self.empresa_var = tk.StringVar()
        nombres_empresas = sorted([e['nombre'] for e in self.parent_app.empresas_registradas])
        self.empresa_combo = ttk.Combobox(filter_frame, textvariable=self.empresa_var, values=nombres_empresas, state="readonly")
        self.empresa_combo.grid(row=0, column=1, columnspan=3, sticky="ew")
        self.empresa_combo.bind("<<ComboboxSelected>>", self.actualizar_lista)

        # Filtro de Búsqueda y Categoría
        ttk.Label(filter_frame, text="Buscar:").grid(row=1, column=0, padx=(0,5), pady=5, sticky="w")
        self.search_var = tk.StringVar()
        search_entry = ttk.Entry(filter_frame, textvariable=self.search_var)
        search_entry.grid(row=1, column=1, sticky="ew", pady=5)
        self.search_var.trace_add("write", lambda *args: self.actualizar_lista())

        ttk.Label(filter_frame, text="Categoría:").grid(row=1, column=2, padx=(10,5), pady=5, sticky="w")
        self.categoria_var = tk.StringVar(value="Todas")
        categoria_combo = ttk.Combobox(filter_frame, textvariable=self.categoria_var, values=self.categorias_documentos, state="readonly")
        categoria_combo.grid(row=1, column=3, sticky="ew", pady=5)
        categoria_combo.bind("<<ComboboxSelected>>", lambda *args: self.actualizar_lista())

        # --- Reemplazamos Listbox por Treeview ---
        list_frame = ttk.LabelFrame(main_frame, text="Documentos de la Plantilla", padding=10)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        cols = ('adjunto', 'codigo', 'nombre', 'categoria')
        self.tree = ttk.Treeview(list_frame, columns=cols, show="headings")
        self.tree.heading('adjunto', text='📎')
        self.tree.heading('codigo', text='Código')
        self.tree.heading('nombre', text='Nombre del Documento')
        self.tree.heading('categoria', text='Categoría')
        
        self.tree.column('adjunto', width=30, anchor=tk.CENTER, stretch=False)
        self.tree.column('codigo', width=150)
        self.tree.column('nombre', width=400)
        self.tree.column('categoria', width=120)

        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.tree.bind('<<TreeviewSelect>>', self.on_doc_select)
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # --- Botones de acción (sin cambios en su creación) ---
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=10)
        btn_frame.columnconfigure(tuple(range(3)), weight=1)
        acciones = {"➕ Agregar Nuevo": self.agregar_doc, "✏️ Editar": self.editar_doc, "🗑️ Eliminar": self.eliminar_doc, "📎 Adjuntar Plantilla": self.adjuntar_archivo, "📂 Ver Plantilla": self.ver_archivo, "❌ Quitar Plantilla": self.quitar_adjunto}
        self.buttons = {}
        for i, (text, cmd) in enumerate(acciones.items()):
            btn = ttk.Button(btn_frame, text=text, command=cmd)
            btn.grid(row=i//3, column=i%3, sticky=tk.EW, padx=5, pady=2)
            self.buttons[text] = btn
        
        self.on_doc_select(None)
        self.protocol("WM_DELETE_WINDOW", self.cerrar_y_recargar)

    def cerrar_y_recargar(self):
        self.parent_app.cargar_datos_desde_db()
        self.destroy()

    def actualizar_lista(self, event=None):
        """Método rediseñado para filtrar y poblar el Treeview."""
        self.tree.delete(*self.tree.get_children())
        
        # Obtenemos los valores de todos los filtros
        empresa_sel = self.empresa_var.get()
        search_term = self.search_var.get().lower()
        categoria_sel = self.categoria_var.get()

        if not empresa_sel:
            return

        # Filtramos la lista de documentos maestros
        docs_filtrados = []
        for doc in self.documentos_maestros:
            if doc.empresa_nombre == empresa_sel:
                # Filtro por categoría
                if categoria_sel != "Todas" and doc.categoria != categoria_sel:
                    continue
                # Filtro por búsqueda de texto
                if search_term not in doc.nombre.lower() and search_term not in doc.codigo.lower():
                    continue
                docs_filtrados.append(doc)
        
        # Llenamos el Treeview con los resultados
        for doc in sorted(docs_filtrados, key=lambda d: (d.categoria, d.nombre)):
            adjunto_icono = "📎" if hasattr(doc, 'ruta_archivo') and doc.ruta_archivo else ""
            self.tree.insert('', tk.END, iid=doc.id, values=(
                adjunto_icono,
                doc.codigo,
                doc.nombre,
                doc.categoria
            ))
        self.on_doc_select(None)

    def on_doc_select(self, event=None):
        doc = self._get_selected_doc()
        state_if_doc = "normal" if doc else "disabled"
        for key in ["✏️ Editar", "🗑️ Eliminar", "📎 Adjuntar Plantilla"]:
            self.buttons[key].config(state=state_if_doc)
        
        if doc:
            state_if_path = "normal" if hasattr(doc, 'ruta_archivo') and doc.ruta_archivo else "disabled"
            self.buttons["📂 Ver Plantilla"].config(state=state_if_path)
            self.buttons["❌ Quitar Plantilla"].config(state=state_if_path)
        else:
            self.buttons["📂 Ver Plantilla"].config(state="disabled")
            self.buttons["❌ Quitar Plantilla"].config(state="disabled")

    def _get_selected_doc(self):
        """Método actualizado para obtener el documento desde la selección del Treeview."""
        selection = self.tree.selection()
        if not selection:
            return None
        doc_id = int(selection[0])
        return next((doc for doc in self.documentos_maestros if doc.id == doc_id), None)

    def adjuntar_archivo(self):
        if (doc := self._get_selected_doc()) and (ruta := filedialog.askopenfilename(parent=self, title="Seleccionar Archivo de Plantilla")):
            doc.ruta_archivo = ruta
            self._save_and_reload()

    def ver_archivo(self):
        if (doc := self._get_selected_doc()) and hasattr(doc, 'ruta_archivo') and doc.ruta_archivo:
            try:
                os.startfile(doc.ruta_archivo)
            except Exception:
                messagebox.showerror("Error", "No se pudo abrir el archivo.", parent=self)

    def quitar_adjunto(self):
        if (doc := self._get_selected_doc()):
            doc.ruta_archivo = ""
            self._save_and_reload()

    def agregar_doc(self):
        empresa = self.empresa_var.get()
        if not empresa:
            messagebox.showerror("Error", "Primero debe seleccionar una empresa.", parent=self)
            return
        
        dialogo = DialogoAgregarDocumento(self, "Nueva Plantilla", categorias=self.categorias_documentos, empresa_actual=empresa)
        if dialogo.result:
            datos = dialogo.result
            if any(d.codigo == datos["codigo"] and d.empresa_nombre == empresa for d in self.documentos_maestros):
                messagebox.showerror("Error", f"Ya existe un documento con el código '{datos['codigo']}' para esta empresa.", parent=self)
                return
            
            nuevo_doc = Documento(
                codigo=datos["codigo"], nombre=datos["nombre"], categoria=datos["categoria"],
                comentario=datos["comentario"], empresa_nombre=empresa
            )
            self.documentos_maestros.append(nuevo_doc)
            self._save_and_reload()

    def editar_doc(self):
        if not (doc := self._get_selected_doc()):
            return
        dialogo = DialogoAgregarDocumento(self, "Editar Plantilla", initial_data=doc, categorias=self.categorias_documentos, empresa_actual=doc.empresa_nombre)
        if dialogo.result:
            datos = dialogo.result
            doc.codigo, doc.nombre, doc.categoria, doc.comentario = datos["codigo"], datos["nombre"], datos["categoria"], datos["comentario"]
            self._save_and_reload()

    def eliminar_doc(self):
        if (doc := self._get_selected_doc()) and messagebox.askyesno("Confirmar", f"¿Está seguro de que desea eliminar la plantilla '{doc.nombre}'?", parent=self):
            self.documentos_maestros.remove(doc)
            self._save_and_reload()

    def _save_and_reload(self):
        self.parent_app.db.save_master_lists(
            empresas=self.parent_app.empresas_registradas,
            instituciones=self.parent_app.instituciones_registradas,
            documentos_maestros=self.documentos_maestros,
            competidores_maestros=self.parent_app.competidores_maestros,
            responsables_maestros=self.parent_app.responsables_maestros,
            replace_tables={'documentos_maestros'}
        )
        self.parent_app.cargar_datos_desde_db()
        self.actualizar_lista()


class VentanaSeleccionMaestro(tk.Toplevel):
    # ... (sin cambios)
    def __init__(self, parent):
        super().__init__(parent)
        self.parent_app = parent
        self.title("Gestión de Datos Maestros")
        self.geometry("450x250")
        self.resizable(False, False)
        self.grab_set()

        main_frame = ttk.Frame(self, padding=20)
        main_frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(main_frame, text="Seleccione el área que desea gestionar:", font=("Helvetica", 11)).pack(pady=(0, 15))

        style = ttk.Style(self)
        style.configure("Maestro.TButton", font=("Helvetica", 10, "bold"), padding=10)

        ttk.Button(main_frame, text="📚 Gestionar Plantillas de Documentos", style="Maestro.TButton", 
                   command=self.abrir_maestro_docs).pack(fill=tk.X, pady=5)
        ttk.Button(main_frame, text="🏢 Gestionar Empresas e Instituciones", style="Maestro.TButton", 
                   command=self.abrir_maestro_entidades).pack(fill=tk.X, pady=5)

    def abrir_maestro_docs(self):
        self.destroy()
        self.parent_app.abrir_ventana_maestro_docs()

    def abrir_maestro_entidades(self):
        self.destroy()
        self.parent_app.abrir_ventana_maestro_entidades()

# En gestor_licitaciones_db.py, reemplaza esta clase por completo:

# En gestor_licitaciones_db.py, reemplaza esta clase:
class VentanaMaestroEntidades(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent_app = parent

        # Copias de trabajo (no tocan las listas reales hasta "Guardar y Cerrar")
        self.entidades_copia = {
            'empresa':       [dict(e) for e in self.parent_app.empresas_registradas],
            'institucion':   [dict(i) for i in self.parent_app.instituciones_registradas],
        }

        self.title("Gestor de Empresas e Instituciones")
        self.geometry("950x550")
        self.grab_set()

        notebook = ttk.Notebook(self, padding="10")
        notebook.pack(fill=tk.BOTH, expand=True)

        # --- Pestaña Empresas ---
        self.widgets_empresas = self._crear_panel_entidad(notebook, "empresa")
        notebook.add(self.widgets_empresas['frame'], text="Empresas")

        # --- Pestaña Instituciones ---
        self.widgets_instituciones = self._crear_panel_entidad(notebook, "institucion")
        notebook.add(self.widgets_instituciones['frame'], text="Instituciones")

        ttk.Button(self, text="✅ Guardar y Cerrar", command=self.cerrar_y_guardar)\
            .pack(pady=10, ipady=4)

        self.protocol("WM_DELETE_WINDOW", self.destroy)
        self.cargar_listas()

    def _crear_panel_entidad(self, parent, entity_type):
        frame = ttk.Frame(parent, padding="10")
        # ... (el código de tree_frame y tree se mantiene igual) ...
        tree_frame = ttk.LabelFrame(frame, text=f"Listado de {entity_type.capitalize()}s")
        tree_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        cols = ("nombre", "rnc", "telefono", "correo")
        tree = ttk.Treeview(tree_frame, columns=cols, show="headings", selectmode="browse")
        for col in cols:
            tree.heading(col, text=col.capitalize())
        tree.column("nombre",   width=250)
        tree.column("rnc",      width=120)
        tree.column("telefono", width=120)
        tree.column("correo",   width=250)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill=tk.X, pady=(10, 0))
        # --- INICIO DEL CAMBIO ---
        num_columns = 4 if entity_type == 'empresa' else 3
        btn_frame.columnconfigure(tuple(range(num_columns)), weight=1)

        btn_add    = ttk.Button(btn_frame, text="➕ Agregar",  command=lambda: self._agregar(entity_type))
        btn_edit   = ttk.Button(btn_frame, text="✏️ Editar",   state="disabled", command=lambda: self._editar(entity_type))
        btn_delete = ttk.Button(btn_frame, text="🗑️ Eliminar", state="disabled", command=lambda: self._eliminar(entity_type))

        btn_add.grid(row=0, column=0, sticky=tk.EW, padx=5)
        btn_edit.grid(row=0, column=1, sticky=tk.EW, padx=5)
        
        # Agregamos el botón de perfil solo para empresas
        if entity_type == 'empresa':
            btn_perfil = ttk.Button(btn_frame, text="📈 Ver Perfil", state="disabled", command=self._ver_perfil_empresa)
            btn_perfil.grid(row=0, column=2, sticky=tk.EW, padx=5)
            btn_delete.grid(row=0, column=3, sticky=tk.EW, padx=5)
        else:
            btn_delete.grid(row=0, column=2, sticky=tk.EW, padx=5)
        
        status_var = tk.StringVar()
        ttk.Label(frame, textvariable=status_var, anchor=tk.W, relief=tk.SUNKEN).pack(fill=tk.X, side=tk.BOTTOM, pady=(10,0), ipady=2)
        
        # Pasamos el nuevo botón de perfil a la función que lo habilita/deshabilita
        widgets_a_controlar = (btn_edit, btn_delete, btn_perfil) if entity_type == 'empresa' else (btn_edit, btn_delete)
        tree.bind("<<TreeviewSelect>>", lambda e: self._on_selection_change(tree, *widgets_a_controlar))
        # --- FIN DEL CAMBIO ---
        
        tree.bind("<Double-1>", lambda e: self._editar(entity_type))

        return {"frame": frame, "tree": tree, "status_var": status_var}

    def _on_selection_change(self, tree, *buttons_to_toggle):
        state = "normal" if tree.selection() else "disabled"
        for btn in buttons_to_toggle:
            if btn: # Comprobamos que el botón existe
                btn.config(state=state)

    def _get_entity_list_and_tree(self, entity_type):
        if entity_type == 'empresa':
            return self.entidades_copia['empresa'], self.widgets_empresas['tree']
        else:
            return self.entidades_copia['institucion'], self.widgets_instituciones['tree']
        
    def _ver_perfil_empresa(self):
        entity_list, tree = self._get_entity_list_and_tree('empresa')
        if not tree.selection(): return
        
        nombre_empresa = tree.item(tree.selection()[0], 'values')[0]
        VentanaPerfilEmpresaNuestra(self, nombre_empresa, self.parent_app.gestor_licitaciones)


    # ----------------- Carga / CRUD -----------------
# En la clase VentanaMaestroEntidades, REEMPLAZA este método:

    def cargar_listas(self):
        for entity_type, widgets in [("empresa", self.widgets_empresas), ("institucion", self.widgets_instituciones)]:
            tree, entity_list = widgets["tree"], self.entidades_copia[entity_type]
            tree.delete(*tree.get_children())
            
            # --- LÓGICA ANTI-DUPLICADOS ---
            nombres_vistos = set()
            entidades_unicas = []
            for e in entity_list:
                nombre = e.get('nombre', '').strip()
                if nombre and nombre.lower() not in nombres_vistos:
                    nombres_vistos.add(nombre.lower())
                    entidades_unicas.append(e)
            # --- FIN DE LA LÓGICA ---

            # Usar la lista ya filtrada y ordenada
            for e in sorted(entidades_unicas, key=lambda x: x.get('nombre', '')):
                values = (e.get('nombre', ''), e.get('rnc', ''), e.get('telefono', ''), e.get('correo', ''))
                if e.get('nombre'):
                    tree.insert("", tk.END, values=values, iid=e['nombre'])
            widgets["status_var"].set(f"Total: {len(entidades_unicas)} {entity_type}s")
# En la clase VentanaMaestroEntidades, REEMPLAZA este método:

    def _agregar(self, entity_type):
        dialogo = DialogoGestionarEntidad(self, f"Agregar {entity_type.capitalize()}", entity_type)
        if dialogo.result and dialogo.result.get('nombre'):
            nueva_entidad_data = dialogo.result
            nombre_nuevo = nueva_entidad_data['nombre'].strip()

            if not nombre_nuevo:
                messagebox.showwarning("Dato requerido", "El nombre no puede estar vacío.", parent=self)
                return

            entity_list, _ = self._get_entity_list_and_tree(entity_type)

            # --- CORRECCIÓN CLAVE ---
            # Verificar si ya existe, ignorando mayúsculas/minúsculas
            if any(e.get('nombre', '').strip().lower() == nombre_nuevo.lower() for e in entity_list):
                messagebox.showerror("Error", f"Ya existe un(a) {entity_type} con el nombre '{nombre_nuevo}'.", parent=self)
                return
            
            entity_list.append(nueva_entidad_data)
            self.cargar_listas()

    def _editar(self, entity_type):
        entity_list, tree = self._get_entity_list_and_tree(entity_type)
        if not tree.selection():
            return
        selected_name = tree.selection()[0]
        entidad_actual = next((e for e in entity_list if e.get('nombre') == selected_name), None)
        if not entidad_actual:
            return
        dialogo = DialogoGestionarEntidad(self, f"Editar {entity_type.capitalize()}", entity_type, initial_data=entidad_actual)
        if dialogo.result and dialogo.result.get('nombre'):
            entidad_actual.update(dialogo.result)
            self.cargar_listas()

    def _eliminar(self, entity_type):
        """
        Evita eliminar:
         - una INSTITUCIÓN usada por alguna licitación (lic.institucion == nombre)
         - una EMPRESA usada por alguna licitación en su lista multi-empresa (nombre ∈ [e.nombre for e in lic.empresas_nuestras])
        """
        entity_list, tree = self._get_entity_list_and_tree(entity_type)
        if not tree.selection():
            return
        selected_name = tree.selection()[0]

        # ¿Está en uso?
        en_uso = False
        for lic in getattr(self.parent_app, "gestor_licitaciones", []):
            if entity_type == 'institucion':
                if str(getattr(lic, "institucion", "")) == selected_name:
                    en_uso = True
                    break
            else:  # empresa
                # lic.empresas_nuestras es una lista de objetos Empresa
                empresas_lic = {str(e) for e in getattr(lic, "empresas_nuestras", [])}
                if selected_name in empresas_lic:
                    en_uso = True
                    break

        if en_uso:
            messagebox.showerror("Error",
                                 f"'{selected_name}' está en uso en una o más licitaciones y no puede ser eliminado.",
                                 parent=self)
            return

        if messagebox.askyesno("Confirmar", f"¿Eliminar a '{selected_name}' del catálogo?"):
            entity_list[:] = [e for e in entity_list if e.get('nombre') != selected_name]
            self.cargar_listas()

    def cerrar_y_guardar(self):
        # 1) Pasar lo editado (las copias) a las listas maestras reales
        self.parent_app.empresas_registradas[:]     = self.entidades_copia['empresa']
        self.parent_app.instituciones_registradas[:] = self.entidades_copia['institucion']

        # 2) Guardar en BD. replace_tables asegura reflejar altas/bajas
        self.parent_app.db.save_master_lists(
            empresas=self.parent_app.empresas_registradas,
            instituciones=self.parent_app.instituciones_registradas,
            documentos_maestros=self.parent_app.documentos_maestros,
            competidores_maestros=self.parent_app.competidores_maestros,
            responsables_maestros=self.parent_app.responsables_maestros,
            replace_tables={'empresas_maestras', 'instituciones_maestras'}
        )

        # 3) Recargar desde BD y cerrar
        self.parent_app.cargar_datos_desde_db()
        self.destroy()




class VentanaReporte(tk.Toplevel):
    # ... (sin cambios)
    def __init__(self, parent, licitacion):
        super().__init__(parent); self.licitacion = licitacion
        self.title(f"Reporte: {self.licitacion.nombre_proceso}"); self.geometry("950x800"); self.grab_set()

        main_frame = ttk.Frame(self, padding="15"); main_frame.pack(fill=tk.BOTH, expand=True)

        header_frame = ttk.Frame(main_frame); header_frame.pack(fill=tk.X, pady=(0, 10))
        kpis = {"Estado Actual": self.licitacion.estado, "Progreso Docs": f"{self.licitacion.get_porcentaje_completado():.1f}%", 
                "Días Restantes": self.licitacion.get_dias_restantes(), "Diferencia Oferta": f"{self.licitacion.get_diferencia_porcentual():.2f}%"}
        for i, (titulo, valor) in enumerate(kpis.items()): self.crear_kpi(header_frame, titulo, valor, i)

        middle_frame = ttk.Frame(main_frame); middle_frame.pack(fill=tk.X, pady=10)
        self.crear_seccion_cronograma(middle_frame); self.crear_seccion_financiera(middle_frame)

        notebook = ttk.Notebook(main_frame); notebook.pack(fill=tk.BOTH, expand=True, pady=10)
        tabs = {"Checklist de Documentos": self.crear_checklist_documentos, "Competidores y Resultados": self.crear_seccion_competidores}
        for text, func in tabs.items(): tab = ttk.Frame(notebook, padding=(0,10)); notebook.add(tab, text=text); func(tab)
        
        export_frame = ttk.Frame(main_frame); export_frame.pack(fill=tk.X, pady=(10,0), side=tk.BOTTOM)
        ttk.Label(export_frame, text="Exportar este reporte:").pack(side=tk.LEFT, padx=(0,10))
        if REPORTLAB_AVAILABLE: ttk.Button(export_frame, text="📄 Exportar a PDF", command=lambda: self.exportar_reporte('pdf')).pack(side=tk.LEFT, padx=5)
        if OPENPYXL_AVAILABLE: ttk.Button(export_frame, text="📈 Exportar a Excel", command=lambda: self.exportar_reporte('excel')).pack(side=tk.LEFT, padx=5)

    def exportar_reporte(self, formato):
        ext = '.pdf' if formato == 'pdf' else '.xlsx'
        filetypes = [('PDF', '*.pdf')] if formato == 'pdf' else [('Excel', '*.xlsx')]
        default_filename = f"Reporte_{self.licitacion.numero_proceso.replace(' ', '_')}{ext}"
        if file_path := filedialog.asksaveasfilename(parent=self, title=f"Guardar como {formato.upper()}", initialfile=default_filename, filetypes=filetypes, defaultextension=ext):
            ReportGenerator().generate_bid_results_report(self.licitacion, file_path)
            messagebox.showinfo("Éxito", f"Reporte guardado en:\n{file_path}", parent=self)

    def crear_kpi(self, parent, titulo, valor, columna):
        frame = ttk.LabelFrame(parent, text=titulo, padding=10); frame.grid(row=0, column=columna, padx=5, sticky="ew")
        parent.grid_columnconfigure(columna, weight=1)
        ttk.Label(frame, text=valor, font=("Helvetica", 14, "bold")).pack()
        if "%" in valor: ttk.Progressbar(frame, orient="horizontal", length=150, mode="determinate", value=float(valor.replace('%',''))).pack(pady=(5,0))

    def crear_seccion_cronograma(self, parent):
        frame = ttk.LabelFrame(parent, text="Cronograma", padding=10); frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        tree = ttk.Treeview(frame, columns=("evento", "fecha", "estado"), show="headings", height=7)
        tree.heading("evento", text="Hito"); tree.heading("fecha", text="Fecha Límite"); tree.heading("estado", text="Estado")
        tree.column("evento", width=250); tree.column("fecha", anchor=tk.CENTER); tree.column("estado", anchor=tk.CENTER)
        tree.tag_configure('cumplido', background='#d4edda'); tree.tag_configure('incumplido', background='#f8d7da')
        for evento, datos in sorted(self.licitacion.cronograma.items()):
            tree.insert("", tk.END, values=(evento, datos.get("fecha_limite", "N/D"), datos.get("estado", "Pendiente")), tags=(datos.get("estado", "Pendiente").lower(),))
        tree.pack(fill=tk.BOTH, expand=True)

# En la clase VentanaReporte, reemplaza este método

    def crear_seccion_financiera(self, parent):
        frame = ttk.LabelFrame(parent, text="Resumen Financiero (Solo Lotes Participados)", padding=10)
        frame.pack(side=tk.LEFT, fill=tk.BOTH)
        
        # Usar solo_participados=True para reflejar los montos relevantes
        base = self.licitacion.get_monto_base_total(solo_participados=True)
        ofertado = self.licitacion.get_oferta_total(solo_participados=True)
        diferencia_pct = self.licitacion.get_diferencia_porcentual(solo_participados=True)

        data = {
            "Monto Base (Presupuesto):": base, 
            "Monto de Nuestra Oferta:": ofertado, 
            "Diferencia Absoluta:": ofertado - base, 
            "Diferencia Porcentual:": diferencia_pct
        }

        for i, (label, value) in enumerate(data.items()):
            ttk.Label(frame, text=label).grid(row=i, column=0, sticky=tk.W, pady=3)
            is_pct = "Porcentual" in label
            format_str = "{:,.2f}%" if is_pct else "RD$ {:,.2f}"
            ttk.Label(frame, text=format_str.format(value), font=("Helvetica", 10, "bold")).grid(row=i, column=1, sticky=tk.E, padx=10)

    def crear_checklist_documentos(self, parent):
        tree = ttk.Treeview(parent, columns=("estado", "nombre", "categoria", "subsanable"), show="headings")
        tree.heading("estado", text="✓"); tree.heading("nombre", text="Documento"); tree.heading("categoria", text="Categoría"); tree.heading("subsanable", text="Condición")
        tree.column("estado", width=30, anchor=tk.CENTER); tree.column("nombre", width=400); tree.column("categoria", anchor=tk.CENTER); tree.column("subsanable", anchor=tk.CENTER)
        tree.tag_configure('no_subsanable_pendiente', background='#f8d7da')
        for doc in sorted(self.licitacion.documentos_solicitados, key=lambda d: (d.categoria, d.nombre)):
            tag = 'no_subsanable_pendiente' if doc.subsanable == "No Subsanable" and not doc.presentado else ""
            tree.insert("", tk.END, values=("✅" if doc.presentado else "❌", doc.nombre, doc.categoria, doc.subsanable), tags=(tag,))
        tree.pack(fill=tk.BOTH, expand=True)

    def crear_seccion_competidores(self, parent_frame):
        self.participants_frame = ttk.LabelFrame(parent_frame, text="Resultados Detallados"); self.participants_frame.pack(fill=tk.BOTH, expand=True)
        self._actualizar_vista_participantes()


    def _actualizar_vista_participantes(self, event=None):
        for widget in self.participants_frame.winfo_children():
            widget.destroy()

        tree = ttk.Treeview(
            self.participants_frame,
            columns=("participante", "monto_ofertado", "monto_habilitado",
                    "fase_a_general", "monto_base_lote", "dif_lote", "ganador"),
            show="headings"
        )

        tree.heading("participante", text="Participante / Lote Ofertado")
        tree.heading("monto_ofertado", text="Monto Ofertado")
        tree.heading("monto_habilitado", text="Monto Habilitado (Fase A)")
        tree.heading("fase_a_general", text="Estado Fase A")
        tree.heading("monto_base_lote", text="Monto Base Lote")
        tree.heading("dif_lote", text="% Diferencia")
        tree.heading("ganador", text="Ganador")

        tree.column("participante", width=350, anchor=tk.W)
        tree.column("monto_ofertado", width=130, anchor=tk.E)
        tree.column("monto_habilitado", width=150, anchor=tk.E)
        tree.column("fase_a_general", width=110, anchor=tk.CENTER)
        tree.column("monto_base_lote", width=130, anchor=tk.E)
        tree.column("dif_lote", width=90, anchor=tk.CENTER)
        tree.column("ganador", width=90, anchor=tk.CENTER)

        tree.tag_configure('competidor', font=('Helvetica', 10, 'bold'))
        tree.tag_configure('ganador_real', background='#d4edda', font=('Helvetica', 9, 'bold'))
        tree.tag_configure('nuestra_empresa', background='#cce5ff')
        tree.tag_configure('descalificado', foreground='red')
        tree.tag_configure('pendiente', foreground='orange')

        def _norm(s: str) -> str:
            s = (s or "").strip()
            s = s.replace("➡️", "").replace("(Nuestra Oferta)", "")
            while "  " in s:
                s = s.replace("  ", " ")
            return s.upper()

        # Fase A evaluada
        hito_eval = self.licitacion.cronograma.get("Informe de Evaluacion Tecnica", {})
        estado_hito_cumplido = hito_eval.get("estado") == "Cumplido"
        estados_que_implican_fase_A_evaluada = {"Adjudicada", "Descalificado Fase B", "Sobre B Entregado"}
        fase_A_evaluada = estado_hito_cumplido or (self.licitacion.estado in estados_que_implican_fase_A_evaluada)

        # Ganadores reales (por lote) + nuestras empresas
        ganadores_por_lote = {str(l.numero): (l.ganador_nombre or "").strip() for l in self.licitacion.lotes}
        nuestras_empresas = {_norm(str(e)) for e in getattr(self.licitacion, "empresas_nuestras", [])}

        # Participantes + nuestra fila
        participantes = [o.__dict__ for o in getattr(self.licitacion, "oferentes_participantes", [])]
        nuestras = ", ".join(str(e) for e in getattr(self.licitacion, "empresas_nuestras", [])) or "Nuestras Empresas"
        nuestras_ofertas = [
            {'lote_numero': l.numero, 'monto': l.monto_ofertado, 'paso_fase_A': l.fase_A_superada}
            for l in getattr(self.licitacion, "lotes", [])
            if l.participamos
        ]
        participantes.append({"nombre": f"➡️ {nuestras} (Nuestra Oferta)", "es_nuestra": True, "ofertas_por_lote": nuestras_ofertas})

        # Monto habilitado (si procede)
        for p in participantes:
            if fase_A_evaluada:
                p["monto_habilitado"] = sum(o.get('monto', 0) for o in p.get("ofertas_por_lote", []) if o.get('paso_fase_A', True))
            else:
                p["monto_habilitado"] = 0

        participantes_orden = sorted(
            participantes,
            key=lambda it: it["monto_habilitado"] if it["monto_habilitado"] > 0 else float('inf')
        )

        for p in participantes_orden:
            tags = ['competidor']
            if p.get("es_nuestra"):
                tags.append('nuestra_empresa')

            if fase_A_evaluada:
                habilitado = any(o.get('paso_fase_A', False) for o in p.get('ofertas_por_lote', [])) if p.get('ofertas_por_lote') else False
                estado_general = "Habilitado" if habilitado else "Descalificado"
                if not habilitado:
                    tags.append('descalificado')
                monto_habilitado_str = f"RD$ {p['monto_habilitado']:,.2f}"
            else:
                estado_general = "Pendiente"
                monto_habilitado_str = "N/D"
                tags.append('pendiente')

            parent_id = tree.insert(
                "", tk.END,
                values=(p['nombre'],
                        f"RD$ {sum(o.get('monto', 0) for o in p.get('ofertas_por_lote', [])):,.2f}",
                        monto_habilitado_str, estado_general, "", "", ""),
                tags=tuple(tags)
            )

            # === CLAVE: nombres del padre "desglosados" para matchear "BARNHOUSE SERVICES, ZOEC CIVIL" ===
            nombre_participante_norm = _norm(p['nombre'])
            nombres_en_padre = {x.strip() for x in nombre_participante_norm.split(",") if x.strip()}  # {"BARNHOUSE SERVICES", "ZOEC CIVIL"}

            gano_alguno = 0

            for oferta in sorted(p.get('ofertas_por_lote', []), key=lambda o: o.get('lote_numero', '')):
                lote_num = str(oferta.get('lote_numero'))
                lote_obj = next((l for l in getattr(self.licitacion, "lotes", []) if str(l.numero) == lote_num), None)
                lote_nombre = getattr(lote_obj, "nombre", "N/E")

                # Montos y %dif
                base_str, dif_pct_str = "N/D", "N/D"
                if lote_obj:
                    base = float(getattr(lote_obj, "monto_base", 0) or 0)
                    of_m = float(oferta.get('monto', 0) or 0)
                    base_str = f"RD$ {base:,.2f}"
                    if base > 0 and of_m > 0:
                        dif_pct_str = f"{((of_m - base)/base)*100:.2f}%"

                # Estado Fase A por lote
                if fase_A_evaluada:
                    paso_a = oferta.get('paso_fase_A', True) if p.get('es_nuestra') else oferta.get('paso_fase_A', False)
                    estado_a = "✅" if paso_a else "❌"
                    lote_tags = [] if paso_a else ['descalificado']
                else:
                    estado_a = "⏳"; lote_tags = ['pendiente']

                # === DECISIÓN DE GANADOR (mejorada) ===
                ganador_real = _norm(ganadores_por_lote.get(lote_num, ""))

                es_ganador_esta_fila = False
                if ganador_real:
                    # 1) Si es nuestra fila y el ganador está en nuestras empresas
                    if p.get('es_nuestra') and (ganador_real in nuestras_empresas):
                        es_ganador_esta_fila = True
                    # 2) Si el nombre del ganador aparece en el texto del padre (separado por comas)
                    elif ganador_real in nombres_en_padre:
                        es_ganador_esta_fila = True
                    # 3) fallback: si el texto del padre comienza exactamente con el ganador
                    elif nombre_participante_norm.startswith(ganador_real):
                        es_ganador_esta_fila = True

                ganador_txt = "Sí" if es_ganador_esta_fila else "No"
                if es_ganador_esta_fila:
                    lote_tags.append('ganador_real')
                    gano_alguno += 1

                tree.insert(
                    parent_id, tk.END,
                    values=(f"    ↳ Lote {lote_num}: {lote_nombre}",
                            f"RD$ {oferta.get('monto', 0):,.2f}",
                            "", estado_a, base_str, dif_pct_str, ganador_txt),
                    tags=tuple(lote_tags)
                )

            # Si ganó al menos un lote: pinto el padre y muestro conteo
            if gano_alguno > 0:
                current_tags = set(tree.item(parent_id, 'tags') or ())
                current_tags.add('ganador_real')
                tree.item(parent_id, tags=tuple(current_tags))
                vals = list(tree.item(parent_id, 'values'))
                vals[-1] = f"Sí ({gano_alguno})"
                tree.item(parent_id, values=tuple(vals))

        tree.pack(fill=tk.BOTH, expand=True)





class DialogoSeleccionarReporteGlobal(simpledialog.Dialog):
    def __init__(self, parent, title="Generar Reporte Global"): super().__init__(parent, title)
    def body(self, master):
        ttk.Label(master, text="Tipo de Reporte:").pack(pady=5)
        self.report_type_var = tk.StringVar(value="Historial por Institución")
        ttk.Combobox(master, state="readonly", textvariable=self.report_type_var, values=["Historial por Institución"]).pack(pady=5)
        ttk.Label(master, text="Formato de Salida:").pack(pady=5)
        self.format_var = tk.StringVar(value="Excel")
        ttk.Combobox(master, state="readonly", textvariable=self.format_var, values=["Excel", "PDF"]).pack(pady=5)
    def apply(self): self.result = (self.report_type_var.get(), self.format_var.get())


# En gestor_licitaciones_db_2.py
# Pega estas dos nuevas clases

class VentanaRestauracion(tk.Toplevel):
    """Muestra una lista de backups disponibles y permite restaurar uno."""
    def __init__(self, parent):
        super().__init__(parent)
        self.parent_app = parent
        self.db = parent.db
        self.title("Restaurar desde Copia de Seguridad")
        self.geometry("800x400")
        self.grab_set()

        main_frame = ttk.Frame(self, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(main_frame, text="Seleccione una copia de seguridad para restaurar. La aplicación se reiniciará.",
                  wraplength=750, justify=tk.LEFT).pack(pady=(0, 10))

        cols = ('fecha', 'comentario', 'ruta')
        self.tree = ttk.Treeview(main_frame, columns=cols, show='headings')
        self.tree.heading('fecha', text='Fecha de Creación')
        self.tree.heading('comentario', text='Comentario')
        self.tree.heading('ruta', text='Ruta del Archivo')
        self.tree.column('fecha', width=150)
        self.tree.column('comentario', width=300)
        self.tree.pack(fill=tk.BOTH, expand=True)
        self.tree.bind("<Double-1>", self._on_restore)

        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(10, 0))
        ttk.Button(btn_frame, text="Restaurar Selección", command=self._on_restore).pack(side=tk.LEFT)
        ttk.Button(btn_frame, text="Abrir Manualmente...", command=self._restore_manual).pack(side=tk.RIGHT)

        self._cargar_backups()

    def _cargar_backups(self):
        self.tree.delete(*self.tree.get_children())
        try:
            self.db.cursor.execute("SELECT timestamp, comentario, ruta_archivo FROM backups_log ORDER BY timestamp DESC")
            for ts, com, ruta in self.db.cursor.fetchall():
                # Solo mostrar backups que todavía existen en el disco
                if os.path.exists(ruta):
                    fecha = datetime.datetime.fromisoformat(ts).strftime('%Y-%m-%d %H:%M:%S')
                    self.tree.insert('', tk.END, values=(fecha, com, ruta))
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo cargar el historial de backups:\n{e}", parent=self)

    def _on_restore(self, event=None):
        if not self.tree.selection():
            messagebox.showwarning("Sin Selección", "Por favor, seleccione un respaldo de la lista.", parent=self)
            return

        ruta_backup = self.tree.item(self.tree.selection()[0], 'values')[2]
        self._proceder_restauracion(ruta_backup)

    def _restore_manual(self):
        ruta_backup = filedialog.askopenfilename(parent=self, title="Seleccionar Copia de Seguridad Manualmente",
                                                 filetypes=[("DB files", "*.db")])
        if ruta_backup:
            self._proceder_restauracion(ruta_backup)

    def _proceder_restauracion(self, ruta_backup):
        if messagebox.askyesno("¡ADVERTENCIA!",
                                 "Se reemplazarán TODOS los datos actuales con los del respaldo.\n\nEsta acción no se puede deshacer. ¿Desea continuar?",
                                 icon='warning', parent=self):
            try:
                self.parent_app.db.close()
                shutil.copyfile(ruta_backup, self.parent_app.db_path)
                messagebox.showinfo("Éxito", "Base de datos restaurada. La aplicación se reiniciará.", parent=self)
                self.parent_app._reiniciar_app()
            except Exception as e:
                messagebox.showerror("Error", f"Falló la restauración:\n{e}", parent=self)
                # Intentar reconectar a la BD original si la restauración falla
                self.parent_app._conectar_a_db(self.parent_app.db_path)


# En gestor_licitaciones_db_2.py
# Pega esta nueva clase completa

class VentanaSanityCheck(tk.Toplevel):
    """Una interfaz para ejecutar chequeos de integridad y reparar la base de datos."""
    def __init__(self, parent):
        super().__init__(parent)
        self.parent_app = parent
        self.db = parent.db
        self.issues_found = {}

        self.title("Diagnóstico y Reparación de Base de Datos")
        self.geometry("700x500")
        self.grab_set()

        main_frame = ttk.Frame(self, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # --- Controles Superiores ---
        controls_frame = ttk.Frame(main_frame)
        controls_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Button(controls_frame, text="🔎 Ejecutar Diagnóstico", command=self.run_checks).pack(side=tk.LEFT)
        self.btn_repair = ttk.Button(controls_frame, text="🛠️ Aplicar Correcciones", state="disabled", command=self.apply_fixes)
        self.btn_repair.pack(side=tk.LEFT, padx=10)

        # --- Ventana de Reporte ---
        report_frame = ttk.LabelFrame(main_frame, text="Reporte de Diagnóstico", padding=10)
        report_frame.pack(fill=tk.BOTH, expand=True)

        self.report_text = tk.Text(report_frame, wrap=tk.WORD, height=10, font=("Consolas", 10))
        self.report_text.pack(fill=tk.BOTH, expand=True)
        self.report_text.insert(tk.END, "Presione 'Ejecutar Diagnóstico' para comenzar...")
        self.report_text.config(state="disabled")

    def run_checks(self):
        self.report_text.config(state="normal")
        self.report_text.delete("1.0", tk.END)
        self.report_text.insert(tk.END, "Ejecutando diagnóstico, por favor espere...\n\n")
        self.update_idletasks() # Forzar actualización de la UI

        self.issues_found = self.db.run_sanity_checks()

        has_orphans = bool(self.issues_found.get('orphans'))
        has_missing_indexes = bool(self.issues_found.get('missing_indexes'))

        if not (has_orphans or has_missing_indexes):
            self.report_text.insert(tk.END, "✅ ¡Excelente! No se encontraron problemas de integridad en la base de datos.")
            self.btn_repair.config(state="disabled")
        else:
            self.report_text.insert(tk.END, "⚠️ Se encontraron los siguientes problemas:\n\n")
            if has_orphans:
                self.report_text.insert(tk.END, "--- Registros Huérfanos Encontrados ---\n")
                for table, ids in self.issues_found['orphans'].items():
                    self.report_text.insert(tk.END, f"  - Tabla '{table}': {len(ids)} registros sin padre.\n")

            if has_missing_indexes:
                self.report_text.insert(tk.END, "\n--- Índices de Rendimiento Faltantes ---\n")
                for index in self.issues_found['missing_indexes']:
                    self.report_text.insert(tk.END, f"  - Falta el índice '{index['name']}' en la tabla '{index['table']}'.\n")

            self.report_text.insert(tk.END, "\nSe recomienda aplicar las correcciones.")
            self.btn_repair.config(state="normal")

        self.report_text.config(state="disabled")

    def apply_fixes(self):
        if not self.issues_found:
            messagebox.showinfo("Información", "No hay problemas que corregir.", parent=self)
            return

        if messagebox.askyesno("Confirmar Reparación",
                                 "Se aplicarán las siguientes correcciones:\n"
                                 "  - Se eliminarán permanentemente los registros huérfanos.\n"
                                 "  - Se crearán los índices de rendimiento faltantes.\n\n"
                                 "¿Desea continuar?", icon='warning', parent=self):

            success, message = self.db.auto_repair(self.issues_found)
            if success:
                messagebox.showinfo("Éxito", message, parent=self)
                # Volver a ejecutar los chequeos para confirmar que todo está limpio
                self.run_checks()
            else:
                messagebox.showerror("Error", message, parent=self)

def seleccionar_o_crear_db_inicial():
    """
    Usa una ventana raíz temporal para manejar la selección de la base de datos
    antes de que la aplicación principal sea creada.
    """
    root_temp = tk.Tk()
    root_temp.withdraw()

    config_file = "config.json"
    db_path = None
    try:
        with open(config_file, 'r') as f: config = json.load(f)
        last_db = config.get("db_path")
        # --- MEJORA: Comprobar que la ruta no solo exista, sino que sea un archivo ---
        if last_db and os.path.isfile(last_db):
            if messagebox.askyesno("Reanudar Sesión", f"¿Desea abrir la última base de datos utilizada?\n\n{last_db}", parent=root_temp):
                db_path = last_db
    except (FileNotFoundError, json.JSONDecodeError):
        pass

    if not db_path:
        if messagebox.askyesno("Iniciar", "¿Desea abrir un archivo de base de datos existente?", parent=root_temp):
            db_path = filedialog.askopenfilename(parent=root_temp, title="Seleccionar Base de Datos", filetypes=[("Database files", "*.db")])
        else:
            db_path = filedialog.asksaveasfilename(parent=root_temp, title="Crear Nueva Base de Datos", filetypes=[("Database files", "*.db")], defaultextension=".db")

    root_temp.destroy()
    
    # Se retorna la ruta solo si no es una cadena vacía.
    return db_path if db_path else None



class AppLicitacionesGUI(ThemedTk):

    def __init__(self, db_path):
        super().__init__()
        self.set_theme("arc")

        self.__version__ = "10.0-Stable"
        self.db_path = db_path
        self.api_key = None # Atributo para guardar la clave en memoria
        self._cargar_configuracion() # Nueva función para leer el archivo config.json

        self.title(f"Gestor de Licitaciones v{self.__version__} - [{os.path.basename(db_path)}]")
        self.geometry("1400x750")

        self._conectar_a_db(db_path)
        self._guardar_configuracion(db_path=self.db_path) # Ya no guarda el usuario

        
        # Inicialización de atributos
        self.gestor_licitaciones = []
        self.empresas_registradas = []
        self.instituciones_registradas = []
        self.documentos_maestros = []
        self.competidores_maestros = LoggingList("competidores_maestros", self) 
        self.responsables_maestros = []
        self.categorias_documentos = ["Legal", "Financiera", "Técnica", "Sobre B", "Otros"]
        self.perfil_entorno = tk.StringVar()
        self.posibles_perfiles = ["Local (Rápido)", "Red / Dropbox (Seguro)"]
        self.debug_mode = False
        self.debug_viewer = None
        self.debug_mode_var = tk.BooleanVar(value=False)
        
        # Estilos
        style = ttk.Style(self)
        style.configure("Urgent.TLabel", background="red", foreground="white", font=('Helvetica', 10, 'bold'), padding=5)
        style.configure("Soon.TLabel", background="orange", foreground="black", font=('Helvetica', 10, 'bold'), padding=5)
        style.configure("Safe.TLabel", background="green", foreground="white", font=('Helvetica', 10, 'bold'), padding=5)
        style.configure("Done.TLabel", background="grey", foreground="white", font=('Helvetica', 10, 'bold'), padding=5)
        style.configure("Small.TButton", font=("Helvetica", 9), padding=4)
        
        # Creación de la interfaz
        self.crear_widgets()
        self.crear_menu_contextual()
        self._crear_menu_superior()
        self.protocol("WM_DELETE_WINDOW", self.al_cerrar)
        
        # Carga de datos
        self.cargar_datos_desde_db()
        self._realizar_backup_automatico()
        self.reporter = ReportGenerator()


# Pega esta NUEVA función DENTRO de la clase AppLicitacionesGUI

    def _cargar_configuracion(self):
        """Lee el archivo config.json y carga la clave API si existe."""
        try:
            with open("config.json", 'r') as f:
                config = json.load(f)
                self.api_key = config.get("api_key", None)
        except (FileNotFoundError, json.JSONDecodeError):
            # Si el archivo no existe o está vacío, no hacemos nada.
            self.api_key = None



    def abrir_ventana_detalles_desde_objeto(self, licitacion_obj):
        if licitacion_obj:
            # --- AÑADE self.instituciones_registradas AL FINAL ---
            VentanaDetalles(self, licitacion_obj, self.cargar_datos_desde_db, self.documentos_maestros, self.categorias_documentos, self.db, self.instituciones_registradas)

    def _actualizar_contadores_barra_estado(self):
        """Calcula y actualiza las etiquetas de la barra de estado inferior."""
        if not hasattr(self, 'status_label_total'): # Evita errores si aún no se ha creado la UI
            return

        total_ganadas = 0
        total_perdidas = 0
        total_lotes_ganados = 0
        estados_perdida_directa = ["Descalificado Fase A", "Descalificado Fase B", "Desierta", "Cancelada"]

        for lic in self.gestor_licitaciones:
            if lic.estado == "Adjudicada":
                lotes_ganados_en_esta_lic = sum(1 for lote in lic.lotes if getattr(lote, "ganado_por_nosotros", False))
                if lotes_ganados_en_esta_lic > 0:
                    total_ganadas += 1
                    total_lotes_ganados += lotes_ganados_en_esta_lic
                else:
                    total_perdidas += 1
            elif lic.estado in estados_perdida_directa:
                total_perdidas += 1

        total_activas = len(self.gestor_licitaciones) - total_ganadas - total_perdidas
        
        self.status_label_total.config(text=f"Datos Cargados. {len(self.gestor_licitaciones)} Licitaciones en Total")
        self.status_label_activas.config(text=f"Activas: {total_activas}")
        self.status_label_ganadas.config(text=f"Ganadas: {total_ganadas}")
        self.status_label_lotes_ganados.config(text=f"Lotes Ganados: {total_lotes_ganados}")
        self.status_label_perdidas.config(text=f"Perdidas: {total_perdidas}")

    def _realizar_backup_automatico(self):
        """Crea un backup si el último tiene más de 1 día de antigüedad."""
        try:
            self.db.cursor.execute("SELECT MAX(timestamp) FROM backups_log")
            last_backup_ts_str = self.db.cursor.fetchone()[0]

            if last_backup_ts_str:
                last_backup_date = datetime.datetime.fromisoformat(last_backup_ts_str).date()
                if (datetime.date.today() - last_backup_date).days < 1:
                    print("INFO: Backup automático omitido, ya se hizo uno hoy.")
                    return # Ya se hizo un backup hoy
            
            print("INFO: Realizando backup automático...")
            timestamp_str = datetime.datetime.now().strftime("%Y-%m-%d")
            backup_folder = os.path.join(os.path.dirname(self.db_path), "Backups")
            os.makedirs(backup_folder, exist_ok=True)
            
            base_filename = os.path.basename(self.db_path).replace('.db', '')
            backup_filename = f"{base_filename}_auto_{timestamp_str}.db"
            backup_path = os.path.join(backup_folder, backup_filename)

            # Pausar, copiar, reconectar
            self.db.close()
            shutil.copyfile(self.db_path, backup_path)
            self._conectar_a_db(self.db_path)

            self.db.cursor.execute(
                "INSERT INTO backups_log (timestamp, ruta_archivo, comentario) VALUES (?, ?, ?)",
                (datetime.datetime.now().isoformat(), backup_path, "Backup Automático")
            )
            self.db.conn.commit()
            print(f"INFO: Backup automático creado en {backup_path}")

        except Exception as e:
            print(f"ERROR: Falló el backup automático: {e}")
            # Asegurarse de reconectar si algo falla
            if not getattr(self.db, 'conn', None):
                self._conectar_a_db(self.db_path)

    def _get_tooltip_text(self):
        try:
            row_id = self.tree.identify_row(self.tree.winfo_pointery() - self.tree.winfo_rooty())
            column_id = self.tree.identify_column(self.tree.winfo_pointerx() - self.tree.winfo_rootx())
            if not row_id or row_id == "finalizadas_parent": return None
            
            # <-- CORRECCIÓN: El nombre del proceso ahora es la columna #3 y el índice 2
# Ahora 'nombre' es la columna #2 (índice 1)
            if column_id == '#2':
                return self.tree.item(row_id, 'values')[1]
        except Exception: 
            return None
        return None

    def _seleccionar_o_crear_db(self):
        config_file = "config.json"
        try:
            with open(config_file, 'r') as f: config = json.load(f)
            last_db = config.get("db_path")
            if last_db and os.path.exists(last_db):
                if messagebox.askyesno("Reanudar Sesión", f"¿Desea abrir la última base de datos utilizada?\n\n{last_db}"):
                    return last_db
        except (FileNotFoundError, json.JSONDecodeError): pass

        if messagebox.askyesno("Iniciar", "¿Desea abrir un archivo de base de datos existente?"):
            return filedialog.askopenfilename(title="Seleccionar Base de Datos", filetypes=[("Database files", "*.db")])
        else:
            return filedialog.asksaveasfilename(title="Crear Nueva Base de Datos", filetypes=[("Database files", "*.db")], defaultextension=".db")
            
    def _conectar_a_db(self, db_path):
        try:
            self.db = DatabaseManager(db_path)
        except Exception as e:
            messagebox.showerror("Error de Conexión", f"No se pudo abrir la base de datos:\n{e}")
            self.destroy()
            sys.exit()

# En gestor_licitaciones_db_2.py, dentro de la clase AppLicitacionesGUI

# En la clase AppLicitacionesGUI
# En la clase AppLicitacionesGUI, REEMPLAZA este método:

    def _guardar_configuracion(self, db_path=None):
        config = {}
        config_file = "config.json"
        try:
            with open(config_file, 'r') as f:
                config = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            pass

        if db_path and isinstance(db_path, str) and db_path.strip():
            config["db_path"] = db_path

        # --- LÍNEA NUEVA ---
        # Añade la clave API al diccionario de configuración si la tenemos
        if self.api_key:
            config["api_key"] = self.api_key
        # --- FIN LÍNEA NUEVA ---

        try:
            with open(config_file, 'w') as f:
                json.dump(config, f, indent=4)
        except IOError as e:
            print(f"Advertencia: No se pudo escribir en {config_file}: {e}")

    def _reiniciar_app(self): self.destroy(); main()

    def analizar_competidores(self, licitaciones_filtradas): 
        """
        Analiza participación de competidores en las licitaciones filtradas,
        excluyendo nuestras empresas.
        """
        competidores = {}

        for lic in licitaciones_filtradas:
            # Lista de nuestras empresas en esta licitación
            nombres_empresas_nuestras = {str(e) for e in lic.empresas_nuestras}

            for oferente in lic.oferentes_participantes:
                # Saltar nuestras propias empresas
                if oferente.nombre in nombres_empresas_nuestras:
                    continue

                stats = competidores.setdefault(oferente.nombre, {
                    "participaciones": 0,
                    "monto_total_habilitado": 0.0,
                    "conteo_monto": 0
                })

                stats["participaciones"] += 1

                monto_hab = oferente.get_monto_total_ofertado(solo_habilitados=True)
                if monto_hab > 0:
                    stats["monto_total_habilitado"] += monto_hab
                    stats["conteo_monto"] += 1
        
        # Convertir a lista y calcular promedios
        resultado = []
        for nombre, datos in competidores.items():
            promedio = (datos["monto_total_habilitado"] / datos["conteo_monto"]) if datos["conteo_monto"] > 0 else 0
            resultado.append({
                "nombre": nombre,
                "participaciones": datos["participaciones"],
                "monto_promedio": promedio
            })

        # Ordenar por participaciones
        return sorted(resultado, key=lambda x: x["participaciones"], reverse=True)

    def abrir_ventana_maestro_responsables(self):
        VentanaMaestroResponsables(self)

# En la clase AppLicitacionesGUI, reemplaza este método por completo

    def _crear_menu_superior(self):
        self.menu_bar = tk.Menu(self)
        self.winfo_toplevel().config(menu=self.menu_bar)

        menus_data = {
            "Archivo": [
                ("Seleccionar/Crear Base de Datos...", self._reiniciar_app), None,
                ("Crear Copia de Seguridad...", self._crear_copia_seguridad),
                ("Restaurar desde Copia...", self._restaurar_desde_copia), None,
                ("Salir", self.al_cerrar)
            ],
            "Editar": [
                ("Agregar Licitación...", self.abrir_ventana_agregar),
                ("Gestionar Empresas e Instituciones", self.abrir_ventana_maestro_entidades),
                ("Gestionar Plantillas de Documentos", self.abrir_ventana_maestro_docs),
                ("Gestionar Catálogo de Competidores", self.abrir_ventana_maestro_competidores),
                ("Gestionar Catálogo de Responsables", self.abrir_ventana_maestro_responsables),
                ("Gestionar Kits de Requisitos", self.abrir_ventana_maestro_kits), None,
            ],
            "Ver": [
                ("Dashboard Global", self.abrir_dashboard_global)
            ],
            "Reportes": [
                ("Reporte de Licitación Seleccionada", self.abrir_ventana_reporte),
                ("Reportes Globales", self.abrir_ventana_reportes_globales)
            ],
            "Herramientas": [
                # --- AÑADE ESTA LÍNEA TEMPORALMENTE ---
                ("Diagnóstico y Reparación de BD...", self.abrir_ventana_sanity_check),
                ("Ejecutar Pruebas de Integridad...", self.ejecutar_smoke_tests),
                {"type": "checkbutton", "label": "Activar Modo Diagnóstico",
                 "variable": self.debug_mode_var, "command": self._toggle_debug_mode},
                None,
                {"type": "submenu", "label": "Perfil de Entorno", "options": self.posibles_perfiles, "variable": self.perfil_entorno}
            ],            
            "Ayuda": [
                ("Acerca de...", self._mostrar_acerca_de)
            ]
        }

        for menu_label, command_list in menus_data.items():
            menu = tk.Menu(self.menu_bar, tearoff=0)
            self.menu_bar.add_cascade(label=menu_label, menu=menu)

            for item in command_list:
                if item is None:
                    menu.add_separator()

                elif isinstance(item, dict) and item.get("type") == "submenu":
                    submenu = tk.Menu(menu, tearoff=0)
                    menu.add_cascade(label=item["label"], menu=submenu)
                    for option in item["options"]:
                        submenu.add_radiobutton(label=option, variable=item["variable"], command=self._guardar_perfil_entorno)

                elif isinstance(item, dict) and item.get("type") == "checkbutton":
                    menu.add_checkbutton(
                        label=item["label"],
                        variable=item["variable"],
                        command=item["command"]
                    )

                else:
                    menu.add_command(label=item[0], command=item[1])

    def _on_tree_select(self, event):
        selection = self.tree.selection()
        if not selection or selection[0] == "finalizadas_parent":
            self.status_display_label.config(text="-- Selecciona una Fila --", style="Done.TLabel")
            return
        
        # <-- CORRECCIÓN: Índice cambiado a [1] para obtener el código del proceso
        numero_proceso_sel = self.tree.item(selection[0], 'values')[0]

        if licitacion := next((lic for lic in self.gestor_licitaciones if lic.numero_proceso == numero_proceso_sel), None):
            self._update_status_display(licitacion)

    def _update_status_display(self, licitacion):
        self.status_display_label.config(text=licitacion.get_dias_restantes())
        style_to_use = "Done.TLabel"
        hoy = datetime.date.today()
        eventos_futuros = [datetime.datetime.strptime(d["fecha_limite"], '%Y-%m-%d').date() for d in licitacion.cronograma.values() if d.get("estado") == "Pendiente" and d.get("fecha_limite")]
        if eventos_futuros:
            diferencia = (min(eventos_futuros) - hoy).days
            if 0 <= diferencia <= 7: style_to_use = "Urgent.TLabel"
            elif 8 <= diferencia <= 30: style_to_use = "Soon.TLabel"
            elif diferencia > 30: style_to_use = "Safe.TLabel"
        self.status_display_label.config(style=style_to_use)

    def _mostrar_acerca_de(self):
        messagebox.showinfo("Acerca de", f"Gestor de Licitaciones\nVersión {self.__version__}\n\nDesarrollado por ZOEC CIVIL.")

# En la clase AppLicitacionesGUI, reemplaza este método

    def _crear_copia_seguridad(self):
        if not self.db_path: return

        # Pedir un comentario opcional al usuario
        comentario = simpledialog.askstring("Crear Copia de Seguridad",
                                            "Añada un comentario para este respaldo (ej: 'Antes de importar masivamente'):",
                                            parent=self)
        if comentario is None: # Si el usuario presiona "Cancelar"
            return

        try:
            timestamp_str = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            backup_folder = os.path.join(os.path.dirname(self.db_path), "Backups")
            os.makedirs(backup_folder, exist_ok=True)

            base_filename = os.path.basename(self.db_path).replace('.db', '')
            backup_filename = f"{base_filename}_backup_{timestamp_str}.db"
            backup_path = os.path.join(backup_folder, backup_filename)

            # Pausamos la conexión a la BD actual para copiar el archivo de forma segura
            self.db.close()
            shutil.copyfile(self.db_path, backup_path)

            # Nos reconectamos y guardamos el registro del backup
            self._conectar_a_db(self.db_path)
            self.db.cursor.execute(
                "INSERT INTO backups_log (timestamp, ruta_archivo, comentario) VALUES (?, ?, ?)",
                (datetime.datetime.now().isoformat(), backup_path, comentario)
            )
            self.db.conn.commit()

            messagebox.showinfo("Éxito", f"Copia de seguridad creada con éxito en:\n{backup_path}", parent=self)

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo crear la copia de seguridad:\n{e}", parent=self)
            # Intentar reconectar incluso si falla
            if not self.db.conn:
                self._conectar_a_db(self.db_path)


    def _restaurar_desde_copia(self):
        """Abre la ventana de gestión de restauración."""
        VentanaRestauracion(self)

    def aplicar_filtros(self):
        criterios = { 'estado': self.filtro_estado_var.get(), 'empresa': self.filtro_empresa_var.get(),
                      'busqueda': self.filtro_busqueda_var.get().lower(), 'lote': self.filtro_lote_var.get().lower() }
        lista_filtrada = self.gestor_licitaciones[:]
        if criterios['estado']: lista_filtrada = [l for l in lista_filtrada if l.estado == criterios['estado']]
        if criterios['empresa']: lista_filtrada = [l for l in lista_filtrada if str(l.empresa_nuestra) == criterios['empresa']]
        if criterios['busqueda']: lista_filtrada = [l for l in lista_filtrada if criterios['busqueda'] in f"{l.nombre_proceso} {l.numero_proceso} {l.institucion}".lower()]
        if criterios['lote']: lista_filtrada = [l for l in lista_filtrada if any(criterios['lote'] in f"{lt.numero} {lt.nombre}".lower() for lt in l.lotes)]
        self.actualizar_tabla_gui(lista_filtrada)

    def limpiar_filtros(self):
        for var in [self.filtro_estado_var, self.filtro_empresa_var, self.filtro_busqueda_var, self.filtro_lote_var]: var.set('')
        self.aplicar_filtros()

    def actualizar_combos_filtros(self):
        self.filtro_estado_combo['values'] = [""] + sorted({l.estado for l in self.gestor_licitaciones})
        # Reunir todas las empresas de todas las licitaciones
        todas_empresas = set()
        for l in self.gestor_licitaciones:
            for e in l.empresas_nuestras:
                todas_empresas.add(str(e))

        self.filtro_empresa_combo['values'] = [""] + sorted(todas_empresas)

# EN LA CLASE AppLicitacionesGUI, DENTRO DE gestor_licitaciones_db.py

    def cargar_datos_desde_db(self):
        lic_data, emp_data, inst_data, docs_data, comp_maestros, resp_maestros = self.db.get_all_data()
        self.gestor_licitaciones = [Licitacion(**data) for data in lic_data]
        for lic in self.gestor_licitaciones:
            if not getattr(lic, 'id', None):
                continue
            # --- Inyectar ganadores por lote desde la BD (ESQUEMA NUEVO) ---
            try:
                filas = self.db.get_ganadores_por_lote(lic.id)  # devuelve dicts con ganador_nombre y empresa_nuestra (si aplica)
                by_lote = {str(r.get("lote_numero")): r for r in filas}

                # Conjunto con nombres de nuestras empresas para comparar rápido
                nuestras_empresas = {str(e).strip() for e in getattr(lic, "empresas_nuestras", [])}

                for lote in lic.lotes:
                    key = str(getattr(lote, "numero", ""))
                    info = by_lote.get(key)

                    if not info:
                        # limpiar por si venía algo colgado
                        lote.ganador_nombre = ""
                        lote.ganado_por_nosotros = False
                        continue

                    ganador = (info.get("ganador_nombre") or "").strip()
                    emp_nuestra_row = (info.get("empresa_nuestra") or "").strip()
                    emp_nuestra_lote = (getattr(lote, "empresa_nuestra", "") or "").strip()

                    lote.ganador_nombre = ganador

                    # Regla: es nuestro si (a) la fila tiene empresa_nuestra, o
                    # (b) la empresa_nuestra del lote coincide con el ganador, o
                    # (c) el ganador es alguna de nuestras empresas cargadas en la licitación.
                    es_nuestro = bool(emp_nuestra_row) \
                                or (emp_nuestra_lote and ganador and emp_nuestra_lote == ganador) \
                                or (ganador in nuestras_empresas)

                    lote.ganado_por_nosotros = bool(es_nuestro)

                    # (Opcional) marca 'ganador' en las ofertas de competidores para este lote
                    if hasattr(lic, "oferentes_participantes") and lic.oferentes_participantes:
                        for comp in lic.oferentes_participantes:
                            for o in comp.ofertas_por_lote:
                                if str(o.get("lote_numero")) == key:
                                    o["ganador"] = (comp.nombre.strip() == ganador)

            except Exception as e:
                print("[WARN] No se pudo inyectar ganadores por lote al cargar:", e)
            # --- fin bloque de inyección ---

        self.empresas_registradas = emp_data
        self.instituciones_registradas = inst_data
        self.documentos_maestros = [Documento(**data) for data in docs_data]
        self.competidores_maestros.clear()
        self.competidores_maestros.extend(comp_maestros)
        self.responsables_maestros = resp_maestros
        perfil_guardado = self.db.get_setting('env_profile', self.posibles_perfiles[0])
        self.perfil_entorno.set(perfil_guardado)
        print(f"Perfil de Entorno cargado: '{perfil_guardado}'")
        debug_state = self.db.get_setting('debug', 'false').lower() == 'true'
        self.debug_mode_var.set(debug_state)
        self._toggle_debug_mode(inicializando=True)
        self.actualizar_tabla_gui()
        self.actualizar_combos_filtros()

        # --- ✅ INICIO DE LA LÓGICA CORREGIDA ---
# Al final de la función cargar_datos_desde_db, reemplaza el bloque del contador

        self.actualizar_tabla_gui()
        self.actualizar_combos_filtros()
        
        # Llama a la nueva función centralizada para actualizar los contadores
        self._actualizar_contadores_barra_estado()
    def al_cerrar(self):
        # (opcional) prints de diagnóstico
        try:
            print("\n=== [APP] Cierre: guardando listas maestras (no destructivo) ===")
            print(f" - empresas: {len(self.empresas_registradas)}")
            print(f" - instituciones: {len(self.instituciones_registradas)}")
            print(f" - documentos_maestros: {len(self.documentos_maestros)}")
            print(f" - competidores_maestros: {len(self.competidores_maestros)}")
            print(f" - responsables_maestros: {len(self.responsables_maestros)}")
        except Exception as e:
            print(f"[APP] (warning) No se pudo contar alguna lista: {e}")

        # Guardado NO destructivo: no pasar replace_tables aquí
        try:
            self.db.save_master_lists(
                empresas=self.empresas_registradas,
                instituciones=self.instituciones_registradas,
                documentos_maestros=self.documentos_maestros,
                competidores_maestros=self.competidores_maestros,
                responsables_maestros=self.responsables_maestros,
                replace_tables=None
            )
        except Exception as e:
            print(f"[APP] Error guardando listas maestras al cerrar: {e}")
        finally:
            try:
                self.db.close()
            except Exception:
                pass
            self.destroy()

    def crear_menu_contextual(self):
        self.context_menu = tk.Menu(self, tearoff=0)
        self.context_menu.add_command(label="📝 Editar Detalles", command=self.abrir_ventana_detalles)
        self.context_menu.add_command(label="🐑 Duplicar Licitación", command=self.duplicar_licitacion)
        self.context_menu.add_separator(); self.context_menu.add_command(label="🗑️ Eliminar Licitación", command=self.eliminar_licitacion)
        self.tree.bind("<Button-3>", self.mostrar_menu_contextual)
        Tooltip(self.tree, text_func=self._get_tooltip_text)

    def mostrar_menu_contextual(self, event):
        if iid := self.tree.identify_row(event.y):
            self.tree.selection_set(iid); self.tree.focus(iid); self.context_menu.post(event.x_root, event.y_root)

    def crear_widgets(self):
        main_frame = ttk.Frame(self, padding="10"); main_frame.pack(fill=tk.BOTH, expand=True)
        filter_frame = ttk.LabelFrame(main_frame, text="Filtros y Búsqueda", padding="10"); filter_frame.pack(fill=tk.X, pady=(0, 10))
        self.filtro_busqueda_var = tk.StringVar(); self.filtro_busqueda_var.trace_add('write', lambda *a: self.aplicar_filtros())
        self.filtro_lote_var = tk.StringVar(); self.filtro_lote_var.trace_add('write', lambda *a: self.aplicar_filtros())
        self.filtro_estado_var = tk.StringVar(); self.filtro_empresa_var = tk.StringVar()
        ttk.Label(filter_frame, text="🔍 Buscar Proceso:").grid(row=0, column=0, padx=(0,5), sticky="w")
        ttk.Entry(filter_frame, textvariable=self.filtro_busqueda_var, width=30).grid(row=0, column=1, padx=5, pady=5)
        ttk.Label(filter_frame, text="📦 Contiene Lote:").grid(row=0, column=2, padx=(10,5), sticky="w")
        ttk.Entry(filter_frame, textvariable=self.filtro_lote_var, width=30).grid(row=0, column=3, padx=5, pady=5)
        ttk.Label(filter_frame, text="Estado:").grid(row=1, column=0, padx=(0,5), sticky="w")
        self.filtro_estado_combo = ttk.Combobox(filter_frame, textvariable=self.filtro_estado_var, state="readonly", width=28); self.filtro_estado_combo.grid(row=1, column=1, padx=5, pady=5)
        self.filtro_estado_combo.bind("<<ComboboxSelected>>", lambda e: self.aplicar_filtros())
        ttk.Label(filter_frame, text="Empresa:").grid(row=1, column=2, padx=(10,5), sticky="w")
        self.filtro_empresa_combo = ttk.Combobox(filter_frame, textvariable=self.filtro_empresa_var, state="readonly", width=28); self.filtro_empresa_combo.grid(row=1, column=3, padx=5, pady=5)
        self.filtro_empresa_combo.bind("<<ComboboxSelected>>", lambda e: self.aplicar_filtros())
        ttk.Button(filter_frame, text="🧹 Limpiar Filtros", command=self.limpiar_filtros).grid(row=0, column=4, rowspan=2, padx=(20,0), ipady=5)
        ttk.Separator(filter_frame, orient=tk.HORIZONTAL).grid(row=2, column=0, columnspan=5, sticky="ew", pady=10)
        status_display_frame = ttk.LabelFrame(filter_frame, text="Próximo Vencimiento")
        status_display_frame.grid(row=0, column=5, rowspan=2, padx=(20, 5), pady=5, sticky="nsew")
        self.status_display_label = ttk.Label(status_display_frame, text="-- Selecciona una Fila --", anchor="center", style="Done.TLabel")
        self.status_display_label.pack(fill=tk.BOTH, expand=True)
        filter_frame.columnconfigure(5, weight=1)
        table_frame = ttk.Frame(main_frame); table_frame.pack(fill=tk.BOTH, expand=True)
        cols = ('codigo', 'nombre', 'empresa', 'dias_restantes', 'porcentaje_docs', 'diferencia', 'monto_ofertado', 'estatus')
        self.tree = ttk.Treeview(table_frame, columns=cols, show='headings')
        # Enlazar selección y doble clic
        self.tree.bind('<<TreeviewSelect>>', self._on_tree_select)
        self.tree.bind("<Double-1>", self.abrir_vista_detallada_lotes)


        headings = {'codigo':'Código', 'nombre':'Nombre Proceso', 'empresa':'Empresa', 
                    'dias_restantes':'Restan', 'porcentaje_docs':'% Docs', 'diferencia':'% Dif.', 
                    'monto_ofertado':'Monto Ofertado', 'estatus':'Estatus'}
        for col, txt in headings.items():
            self.tree.heading(col, text=txt)


        widths = {'codigo': 140, 'nombre': 450, 'empresa': 150, 'dias_restantes': 120,
                'porcentaje_docs': 75, 'diferencia': 75, 'monto_ofertado': 140, 'estatus': 100}

        for col, width in widths.items():
            anchor = tk.W
            if col not in ['codigo', 'nombre', 'empresa', 'monto_ofertado', 'dias_restantes']:
                anchor = tk.CENTER
            elif col == 'monto_ofertado':
                anchor = tk.E
            self.tree.column(col, width=width, anchor=anchor, stretch=True)

        # 🔽 que el Treeview recalcule anchos al cambiar de tamaño
        self.tree.bind('<Configure>', self._ajustar_ancho_columnas)
        # === COLORES / ESTILOS PARA FILAS ===
        # Ganada: verde suave
        self.tree.tag_configure('ganada', background='#E6F4EA', foreground='#1E7D32')  # verde Google-like
        # Perdida: rojo suave
        self.tree.tag_configure('perdida', background='#FDECEA', foreground='#B71C1C')
        # En proceso (opcional): amarillo suave
        self.tree.tag_configure('proceso', background='#FFF8E1', foreground='#8D6E00')
        # Encabezado de finalizadas (si usas fila de sección)
        self.tree.tag_configure('finalizadas_header', background='#F3F4F6', foreground='#374151', font=('Helvetica', 9, 'bold'))
                # (opcional) Próximo a vencer: solo cambia el texto; el fondo lo pone 'en_proceso'
        self.tree.tag_configure('proximo', foreground='#D35400', font=('Helvetica', 9, 'bold'))

        # (opcional) que la selección no tape el color de fondo
        style = ttk.Style(self)
        style.map('Treeview',
                background=[('selected', '#C7F0D8')],
                foreground=[('selected', '#0B6B32')])

        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
# Scrollbar vertical
        scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Scrollbar horizontal
        hscroll = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(xscrollcommand=hscroll.set)
        hscroll.pack(side=tk.BOTTOM, fill=tk.X)
      
# En la función crear_widgets, reemplaza el bloque del status_frame por este:

        status_frame = ttk.Frame(main_frame, relief="sunken", padding=(5,2))
        status_frame.pack(fill=tk.X, side=tk.BOTTOM, pady=(10, 0))
        
        self.status_label_total = ttk.Label(status_frame, font=("Helvetica", 9))
        self.status_label_total.pack(side=tk.LEFT, padx=(0, 5))
        
        ttk.Separator(status_frame, orient='vertical').pack(side=tk.LEFT, padx=5, fill='y')
        
        self.status_label_activas = ttk.Label(status_frame, font=("Helvetica", 9, "bold"), foreground="#007bff")
        self.status_label_activas.pack(side=tk.LEFT, padx=5)
        
        self.status_label_ganadas = ttk.Label(status_frame, font=("Helvetica", 9, "bold"), foreground="green")
        self.status_label_ganadas.pack(side=tk.LEFT, padx=5)
        
        self.status_label_lotes_ganados = ttk.Label(status_frame, font=("Helvetica", 9, "bold"), foreground="#2E7D32")
        self.status_label_lotes_ganados.pack(side=tk.LEFT, padx=5)

        self.status_label_perdidas = ttk.Label(status_frame, font=("Helvetica", 9, "bold"), foreground="red")
        self.status_label_perdidas.pack(side=tk.LEFT, padx=5)
        style = ttk.Style(self); style.configure("Accion.TButton", font=("Helvetica", 10, "bold"))
        botones_frame = ttk.Frame(self); botones_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=(5, 10))
        ttk.Button(botones_frame, text="➕ Agregar", style="Accion.TButton", command=self.abrir_ventana_agregar).pack(side=tk.LEFT, padx=5, ipady=4)
        ttk.Button(botones_frame, text="📝 Ver/Editar", style="Accion.TButton", command=self.abrir_ventana_detalles).pack(side=tk.LEFT, padx=5, ipady=4)
        Tooltip(self.tree, text_func=self._get_tooltip_text)


    def _ajustar_ancho_columnas(self, event=None):
        # Porcentajes (pesos) de cada columna — deben sumar 1.0
        pesos = {
            'codigo':           0.12,
            'nombre':           0.38,
            'empresa':          0.10,
            'dias_restantes':   0.12,
            'porcentaje_docs':  0.08,
            'diferencia':       0.08,
            'monto_ofertado':   0.07,
            'estatus':          0.05,
        }
        total = max(self.tree.winfo_width(), 1)
        for col, p in pesos.items():
            # Mínimos simpáticos para que no colapsen
            minimo = 70 if col != 'nombre' else 150
            self.tree.column(col, width=max(int(total * p), minimo))

# Pega estos dos nuevos métodos DENTRO de la clase AppLicitacionesGUI

# En la clase AppLicitacionesGUI, REEMPLAZA este método:

    def _nuestras_empresas_de(self, lic):
        """Devuelve un set con los nombres de nuestras empresas que participan en una licitación."""
        empresas = set()
        # Primero, revisamos las empresas asignadas a nivel de lote
        for lote in getattr(lic, "lotes", []):
            nombre_empresa_lote = (getattr(lote, "empresa_nuestra", None) or "").strip()
            # --- LÓGICA MEJORADA: Ignoramos si está vacío o es '(Sin Asignar)' ---
            if nombre_empresa_lote and nombre_empresa_lote != "(Sin Asignar)":
                empresas.add(nombre_empresa_lote)
        
        # Si después de revisar los lotes no encontramos ninguna empresa,
        # usamos la lista general de la licitación como respaldo.
        if not empresas:
            for item in getattr(lic, "empresas_nuestras", []):
                nombre = None
                if hasattr(item, 'nombre'):
                    nombre = item.nombre
                elif isinstance(item, dict) and 'nombre' in item:
                    nombre = item.get('nombre')
                
                if nombre and isinstance(nombre, str) and nombre.strip() and nombre.lower() != 'none':
                    empresas.add(nombre.strip())
        return empresas
    
# En la clase AppLicitacionesGUI, REEMPLAZA este método:

# En la clase AppLicitacionesGUI, REEMPLAZA este método:

    def _display_empresas_de(self, lic):
        """Devuelve un string amigable para mostrar las empresas de una licitación."""
        # La función _nuestras_empresas_de ahora solo devuelve nombres de empresas reales
        emps = sorted(self._nuestras_empresas_de(lic))
        
        # Si la lista de empresas reales no está vacía, las mostramos.
        # Si está vacía, significa que no hay ninguna asignada, y mostramos "(Sin Asignar)".
        return ", ".join(emps) if emps else "(Sin Asignar)"
    # En la clase AppLicitacionesGUI, REEMPLAZA tu método actual con este:

    def actualizar_tabla_gui(self, lista_a_mostrar=None):
        lista_para_usar = lista_a_mostrar if lista_a_mostrar is not None else self.gestor_licitaciones
        self.tree.delete(*self.tree.get_children())

        estados_finalizados = ["Adjudicada", "Descalificado Fase A", "Descalificado Fase B", "Desierta", "Cancelada"]

        licitaciones_activas = [l for l in lista_para_usar if l.estado not in estados_finalizados]
        licitaciones_finalizadas = [l for l in lista_para_usar if l.estado in estados_finalizados]

        def obtener_clave_ordenamiento(licitacion):
            hoy = datetime.date.today()
            fechas = []
            for d in licitacion.cronograma.values():
                if d.get("estado") == "Pendiente" and d.get("fecha_limite"):
                    try:
                        f = datetime.datetime.strptime(d["fecha_limite"], "%Y-%m-%d").date()
                        if f >= hoy:
                            fechas.append(f)
                    except Exception:
                        pass
            return min(fechas) if fechas else datetime.date.max

        activas_ordenadas = sorted(licitaciones_activas, key=obtener_clave_ordenamiento)

        for lic in activas_ordenadas:
            tags = ['proceso']
            dias_restantes_str = lic.get_dias_restantes()
            if "días" in dias_restantes_str:
                try:
                    dias = int(dias_restantes_str.split()[1])
                    if dias <= 7:
                        tags.append('proximo')
                except (ValueError, IndexError):
                    pass

            # --- Lógica de visualización mejorada ---
            monto_ofertado = lic.get_oferta_total(solo_participados=True)
            monto_ofertado_str = f"RD$ {monto_ofertado:,.2f}" if monto_ofertado > 0 else "N/D"
            dif_str = f"{lic.get_diferencia_porcentual(solo_participados=True, usar_base_personal=False):.2f}%" if monto_ofertado > 0 else "N/D"
            # --- Fin de la lógica ---

            values = (
                lic.numero_proceso,
                lic.nombre_proceso,
                self._display_empresas_de(lic),
                dias_restantes_str,
                f"{lic.get_porcentaje_completado():.1f}%",
                dif_str,
                monto_ofertado_str,
                lic.estado
            )
            self.tree.insert('', tk.END, values=values, tags=tuple(tags))

        if licitaciones_finalizadas:
            tv = self.tree  # <— usa SIEMPRE el Treeview real

            parent_id = "finalizadas_parent"
            tv.insert(
                '', tk.END, iid=parent_id,
                values=("", f"▶ Licitaciones Finalizadas ({len(licitaciones_finalizadas)})", "", "", "", "", "", ""),
                tags=('finalizadas_header',)
            )

            finalizadas_ordenadas = sorted(licitaciones_finalizadas, key=lambda l: l.nombre_proceso, reverse=True)

            for lic in finalizadas_ordenadas:
                # 1) Calcula los valores de la fila
                monto_ofertado = lic.get_oferta_total(solo_participados=True)
                monto_ofertado_str = f"RD$ {monto_ofertado:,.2f}" if monto_ofertado > 0 else "N/D"
                dif_str = (
                    f"{lic.get_diferencia_porcentual(solo_participados=True, usar_base_personal=False):.2f}%"
                    if monto_ofertado > 0 else "N/D"
                )

                values = (
                    lic.numero_proceso,
                    lic.nombre_proceso,
                    self._display_empresas_de(lic),
                    lic.get_dias_restantes(),
                    f"{lic.get_porcentaje_completado():.1f}%",
                    dif_str,
                    monto_ofertado_str,
                    lic.estado
                )

                # 2) Construye los tags (color de fila)
                tags = []
                estado = getattr(lic, "estado", "Iniciada") or "Iniciada"
                if estado == "Adjudicada":
                    hay_lote_nuestro = any(getattr(l, "ganado_por_nosotros", False) for l in getattr(lic, "lotes", []))
                    tags.append("ganada" if hay_lote_nuestro else "perdida")
                else:
                    tags.append("en_proceso")

                # 3) Inserta la fila usando el Treeview correcto
                tv.insert(parent_id, tk.END, values=values, tags=tuple(tags))

        self._actualizar_contadores_barra_estado()

    def agregar_licitacion_callback(self, nueva_licitacion):
        self.db.save_licitacion(nueva_licitacion); self.cargar_datos_desde_db()




    def duplicar_licitacion(self):
        # 1) Verificación de selección
        iid = self.tree.focus()
        if not iid:
            messagebox.showwarning("Sin Selección", "Selecciona una licitación.")
            return
        if iid == "finalizadas_parent":
            return

        # 2) Obtener la licitación original desde el Treeview
        #    Tomamos el número de proceso (columna 0 de values)
        numero_original = self.tree.item(iid, 'values')[0]
        original = next((l for l in self.gestor_licitaciones if l.numero_proceso == numero_original), None)
        if not original:
            messagebox.showerror("Error", "No se encontró la licitación original.")
            return

        # 3) Pedir la nueva empresa
        empresas_nombres = [e['nombre'] for e in self.empresas_registradas]
        dlg = DialogoSeleccionarEmpresa(self, "Duplicar para...", empresas_nombres)
        nueva_empresa_nombre = dlg.result
        if not nueva_empresa_nombre:
            return  # usuario canceló

        # 4) Clonar usando dict plano para "romper" ids y referencias
        #    (esto evita que se conserve 'id' y que el guardado sea tratado como UPDATE)
        data = original.to_dict()

        # 4.1) Limpiar identificadores y campos que forzarían UPDATE
        data['id'] = None
        data['last_modified'] = None

        # 4.2) Asignar la nueva empresa y ajustar nombre del proceso (opcional)
        data['empresa_nuestra'] = {'nombre': nueva_empresa_nombre}
        data['nombre_proceso'] = f"{original.nombre_proceso} ({nueva_empresa_nombre})"

        # 4.3) Generar un número de proceso NUEVO y único
        base_code = original.numero_proceso
        sufijo = ("".join(filter(str.isalnum, nueva_empresa_nombre))[:10]).upper()
        propuesto = f"{base_code}-{sufijo}"

        existentes = {l.numero_proceso for l in self.gestor_licitaciones}
        nuevo_codigo = propuesto
        contador = 2
        while nuevo_codigo in existentes:
            nuevo_codigo = f"{propuesto}-{contador}"
            contador += 1
        data['numero_proceso'] = nuevo_codigo  # <- clave para que sea un registro distinto

        # 4.4) Resetear documentos (ids/estado/archivos)
        for d in data.get('documentos_solicitados', []):
            d['id'] = None
            d['empresa_nombre'] = nueva_empresa_nombre
            d['presentado'] = False
            d['revisado'] = False
            d['ruta_archivo'] = ""

        # 4.5) Resetear lotes (ids/montos/ganadores)
        for l in data.get('lotes', []):
            l['id'] = None
            l['monto_ofertado'] = 0.0
            l['ganado_por_nosotros'] = False
            l['ganador_nombre'] = ""

        # 4.6) Resetear marcas de ganador en ofertas de competidores
        for comp in data.get('oferentes_participantes', []):
            if 'id' in comp:
                comp['id'] = None
            for o in comp.get('ofertas_por_lote', []):
                o['ganador'] = False  # limpiar cualquier marca previa

        # 4.7) Dejar la licitación en estado inicial (ajusta a tu flujo si lo prefieres)
        data['adjudicada'] = False
        data['adjudicada_a'] = ""
        data['estado'] = "Iniciada"

        # 5) Construir el objeto y guardar (se insertará porque id=None)
        copia = Licitacion(**data)
        try:
            self.db.save_licitacion(copia)
            self.cargar_datos_desde_db()
            messagebox.showinfo(
                "Éxito",
                f"Licitación duplicada para '{nueva_empresa_nombre}' con código '{nuevo_codigo}'.",
                parent=self
            )
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo duplicar la licitación:\n{e}", parent=self)


    def abrir_vista_detallada_lotes(self, event=None):
        if not (iid := self.tree.focus()): return
        if iid == "finalizadas_parent": return
        try:
            # <-- CORRECCIÓN: Índice cambiado a [1] para obtener el código del proceso
            numero_proceso_sel = self.tree.item(iid, 'values')[0]
            licitacion = next((l for l in self.gestor_licitaciones if l.numero_proceso == numero_proceso_sel), None)
            if licitacion:
                VentanaVistaLotes(self, licitacion)
        except (IndexError, StopIteration):
            pass

    def abrir_ventana_agregar(self): VentanaAgregarLicitacion(self, self.empresas_registradas, self.instituciones_registradas, self.agregar_licitacion_callback)
    
    def abrir_ventana_detalles(self):
        if not (iid := self.tree.focus()): messagebox.showwarning("Sin Selección", "Selecciona una licitación."); return
        if iid == "finalizadas_parent": return
        if licitacion := next((l for l in self.gestor_licitaciones if l.numero_proceso == self.tree.item(iid, 'values')[0]), None):
            # --- AÑADE self.instituciones_registradas AL FINAL ---
            VentanaDetalles(self, licitacion, self.cargar_datos_desde_db, self.documentos_maestros, self.categorias_documentos, self.db, self.instituciones_registradas)
    
    
    
    def abrir_ventana_reporte(self):
        if not (iid := self.tree.focus()): messagebox.showwarning("Sin Selección", "Selecciona una licitación."); return
        if iid == "finalizadas_parent": return
        # <-- CORRECCIÓN: Índice cambiado a [1] para obtener el código del proceso
        if licitacion := next((l for l in self.gestor_licitaciones if l.numero_proceso == self.tree.item(iid, 'values')[0]), None):
            VentanaReporte(self, licitacion)
            
    def abrir_ventana_reportes_globales(self):
        if not (dialogo_result := DialogoSeleccionarReporteGlobal(self).result): return
        _, formato = dialogo_result; ext = '.xlsx' if formato == 'Excel' else '.pdf'
        if file_path := filedialog.asksaveasfilename(parent=self, title="Guardar Reporte Histórico", initialfile=f"Reporte_Historial_{datetime.date.today()}{ext}", filetypes=[(formato, f"*{ext}")], defaultextension=ext):
            ReportGenerator().generate_institution_history_report(self.gestor_licitaciones, file_path)
            messagebox.showinfo("Éxito", f"Reporte guardado en:\n{file_path}", parent=self)
            
    def abrir_dashboard_global(self):
        if not self.gestor_licitaciones: messagebox.showinfo("Sin Datos", "No hay licitaciones para generar un dashboard."); return
        VentanaDashboardGlobal(self, self.gestor_licitaciones)
        
    def abrir_ventana_maestro_docs(self): VentanaMaestroDocumentos(self, self.documentos_maestros, self.categorias_documentos, self.db)
    
    def abrir_ventana_maestros(self): VentanaSeleccionMaestro(self)
    
    def abrir_ventana_maestro_entidades(self): VentanaMaestroEntidades(self)

    def abrir_ventana_maestro_competidores(self):
        VentanaMaestroCompetidores(self)


    def abrir_ventana_maestro_kits(self):
        """Abre la ventana para gestionar los kits de requisitos."""
        VentanaMaestroKits(self)

    # En la clase AppLicitacionesGUI, pega este nuevo método

    def abrir_ventana_sanity_check(self):
        """Abre la ventana de diagnóstico de la base de datos."""
        VentanaSanityCheck(self)

    # En la clase AppLicitacionesGUI, pega este nuevo método

    def _guardar_perfil_entorno(self, *args):
        """Se activa cuando el perfil de entorno cambia y lo guarda en la BD."""
        nuevo_perfil = self.perfil_entorno.get()
        self.db.set_setting('env_profile', nuevo_perfil)
        messagebox.showinfo("Perfil Actualizado",
                            f"Se ha cambiado el perfil a '{nuevo_perfil}'.\n\n"
                            "Se recomienda reiniciar la aplicación para que todos los ajustes surtan efecto.",
                            parent=self)

    # Pega estos 5 nuevos métodos dentro de la clase AppLicitacionesGUI

    def debug_log(self, evento, payload=None):
        """
        Registra un evento en el visor de diagnóstico si el modo está activo.
        """
        if not self.debug_mode:
            return

        # Si el modo está activo pero la ventana no existe, la creamos.
        if not self.debug_viewer or not self.debug_viewer.winfo_exists():
            self.debug_viewer = VentanaVisorDebug(self)

        timestamp = datetime.datetime.now().strftime("%H:%M:%S.%f")[:-3]
        payload_str = ""
        if payload is not None:
            if isinstance(payload, (dict, list)):
                # Usamos json para una visualización bonita de diccionarios o listas
                try:
                    payload_str = json.dumps(payload, indent=2, ensure_ascii=False)
                except TypeError:
                    payload_str = str(payload)
            else:
                payload_str = str(payload)

        mensaje = f"[{timestamp}] -- {evento}"
        if payload_str:
            mensaje += f":\n{payload_str}"

        self.debug_viewer.add_log_entry(mensaje)

    def _toggle_debug_mode(self, inicializando=False):
        """
        Activa o desactiva el modo de diagnóstico y lo guarda en la BD.
        """
        self.debug_mode = self.debug_mode_var.get()
        self.db.set_setting('debug', str(self.debug_mode))

        if self.debug_mode:
            # Si se activa, nos aseguramos de que la ventana exista
            if not self.debug_viewer or not self.debug_viewer.winfo_exists():
                self.debug_viewer = VentanaVisorDebug(self)
            self.debug_viewer.deiconify() # La trae al frente si estaba minimizada
        else:
            # Si se desactiva, cerramos la ventana si existe
            if self.debug_viewer and self.debug_viewer.winfo_exists():
                self.debug_viewer.destroy()
                self.debug_viewer = None

        # No mostramos el mensaje si solo estamos cargando el estado al inicio
        if not inicializando:
            self.debug_log("Modo Diagnóstico", f"Estado: {'Activado' if self.debug_mode else 'Desactivado'}")

    def on_debug_viewer_closed(self):
        """
        Llamado por la ventana de debug cuando el usuario la cierra.
        Actualiza el estado para que el checkbutton del menú se desmarque.
        """
        self.debug_viewer = None
        if self.debug_mode: # Si el modo seguía activo
            self.debug_mode_var.set(False)
            self._toggle_debug_mode()

    def abrir_ventana_maestro_competidores(self):
        self.debug_log("Acción de UI", "Abriendo ventana 'Maestro de Competidores'")
        VentanaMaestroCompetidores(self)

    def eliminar_licitacion(self):
        if not (iid := self.tree.focus()): return
        if iid == "finalizadas_parent": return

        numero_proceso_sel = self.tree.item(iid, 'values')[0]
        licitacion = next((l for l in self.gestor_licitaciones if l.numero_proceso == numero_proceso_sel), None)

        if licitacion:
            self.debug_log("Eliminación Licitación (Inicio)", licitacion.to_summary_dict())
            if messagebox.askyesno("Confirmar", f"¿Eliminar '{licitacion.nombre_proceso}'?"):
                if self.db.delete_licitacion(numero_proceso_sel):
                    self.debug_log("Eliminación Licitación (Éxito)", {"numero_proceso": numero_proceso_sel})
                else:
                    self.debug_log("Eliminación Licitación (Fallo)", {"numero_proceso": numero_proceso_sel})
                self.cargar_datos_desde_db()


    def ejecutar_smoke_tests(self):
            """Orquesta la ejecución de todas las pruebas y muestra los resultados."""
            log = [f"--- INICIO DE PRUEBAS DE INTEGRIDAD (SMOKE TESTS) v{self.__version__} ---"]
            log.append(f"Fecha y Hora: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            log.append(f"Base de Datos: {self.db_path}\n")

            self.config(cursor="wait")
            self.update_idletasks()

            try:
                self.db.begin_transaction()
                log.append("INFO: Transacción iniciada. Todos los cambios serán revertidos.")

                # --- Ejecutar Casos de Prueba ---
                self._test_crud_licitacion(log)
                self._test_crud_maestros(log)
                # --- NUEVAS PRUEBAS AVANZADAS ---
                self._test_ganadores_por_lote(log)
                self._test_fallas_fase_a(log)
                # Nota: La prueba de duplicación se omite en el modo transaccional
                # porque requiere que los IDs se confirmen, lo cual no ocurre
                # hasta el commit, pero la lógica está disponible si se necesita.
                # self._test_duplicacion(log)

                log.append("\n--- PRUEBAS COMPLETADAS ---")
                if any("[FAIL]" in line for line in log):
                    log.append("RESULTADO: ❌ FALLO. Se encontraron uno o más errores.")
                else:
                    log.append("RESULTADO: ✅ ÉXITO. Todas las pruebas avanzadas pasaron.")

            except Exception as e:
                log.append(f"\n[FAIL] ERROR CATASTRÓFICO DURANTE LAS PRUEBAS: {e}")
                log.append(traceback.format_exc())
            finally:
                self.db.rollback_transaction()
                log.append("INFO: Transacción revertida. La base de datos no ha sido modificada.")
                self.config(cursor="")

            VentanaResultadosTests(self, "\n".join(log))

    def _test_crud_licitacion(self, log):
        """Prueba la creación, actualización y eliminación de una licitación."""
        log.append("\n--- Módulo: Licitaciones ---")
        test_id = f"SMOKETEST-{int(datetime.datetime.now().timestamp())}"
        lic_obj = None
        try:
            # 1. CREATE
            datos_lic = {
                "nombre_proceso": "Licitación de Prueba", "numero_proceso": test_id,
                "institucion": "Institucion Maestra", "empresa_nuestra": {"nombre": "Empresa Maestra"},
                "lotes": [{"numero": 1, "nombre": "Lote de Prueba", "monto_base": 1000}]
            }
            lic_obj = Licitacion(**datos_lic)
            self.db.save_licitacion(lic_obj)
            log.append(f"[PASS] CREATE: Se creó la licitación de prueba '{test_id}'.")
        except Exception as e:
            log.append(f"[FAIL] CREATE: No se pudo crear la licitación de prueba. Error: {e}")
            return # Si la creación falla, no podemos continuar

        try:
            # 2. READ (implícito en el update y delete)
            # 3. UPDATE
            lic_obj.estado = "En Proceso"
            self.db.save_licitacion(lic_obj)
            log.append(f"[PASS] UPDATE: Se actualizó el estado de la licitación '{test_id}'.")
        except Exception as e:
            log.append(f"[FAIL] UPDATE: No se pudo actualizar la licitación '{test_id}'. Error: {e}")

        try:
            # 4. DELETE
            if self.db.delete_licitacion(test_id):
                log.append(f"[PASS] DELETE: Se eliminó la licitación de prueba '{test_id}'.")
            else:
                log.append(f"[FAIL] DELETE: El borrado de '{test_id}' no reportó filas afectadas.")
        except Exception as e:
            log.append(f"[FAIL] DELETE: No se pudo eliminar la licitación '{test_id}'. Error: {e}")

    def _test_crud_maestros(self, log):
            """Prueba la escritura y eliminación en tablas maestras (ej: competidores)."""
            log.append("\n--- Módulo: Datos Maestros (Competidores) ---")
            test_name = "Competidor de Prueba Auto"

            try:
                # 1. CREATE
                # Creamos una copia para no modificar la lista en memoria durante la prueba
                temp_competidores = list(self.competidores_maestros)
                temp_competidores.append({'nombre': test_name, 'rnc': '000-0000000-0'})
                
                self.db.save_master_lists(
                    empresas=self.empresas_registradas,
                    instituciones=self.instituciones_registradas,
                    documentos_maestros=[d.to_dict() for d in self.documentos_maestros],
                    competidores_maestros=temp_competidores, # Usamos la lista temporal
                    responsables_maestros=self.responsables_maestros,
                    replace_tables={'competidores_maestros'}
                )
                log.append(f"[PASS] CREATE: Se guardó un nuevo competidor maestro '{test_name}'.")

            except Exception as e:
                log.append(f"[FAIL] CREATE: No se pudo guardar el competidor maestro. Error: {e}")
                log.append(traceback.format_exc()) # Añadimos más detalle al log
                return

            try:
                # 2. DELETE
                temp_competidores_del = [c for c in temp_competidores if c['nombre'] != test_name]
                self.db.save_master_lists(
                    empresas=self.empresas_registradas,
                    instituciones=self.instituciones_registradas,
                    documentos_maestros=[d.to_dict() for d in self.documentos_maestros],
                    competidores_maestros=temp_competidores_del, # Usamos la lista sin el competidor de prueba
                    responsables_maestros=self.responsables_maestros,
                    replace_tables={'competidores_maestros'}
                )
                log.append(f"[PASS] DELETE: Se eliminó el competidor maestro de prueba.")
            except Exception as e:
                log.append(f"[FAIL] DELETE: No se pudo eliminar el competidor maestro. Error: {e}")
                log.append(traceback.format_exc()) # Añadimos más detalle al log

    def _test_ganadores_por_lote(self, log):
            """Prueba la asignación y eliminación de ganadores por lote."""
            log.append("\n--- Módulo: Ganadores por Lote ---")
            test_id = f"SMOKETEST-GANADOR-{int(datetime.datetime.now().timestamp())}"
            
            try:
                # Setup: Crear una licitación con un lote y un competidor
                datos_lic = {
                    "nombre_proceso": "Prueba Ganadores", "numero_proceso": test_id,
                    "lotes": [{"numero": "101", "nombre": "Lote Ganador"}],
                    "oferentes_participantes": [{"nombre": "Competidor Ganador", "ofertas_por_lote": [{"lote_numero": "101", "monto": 500}]}]
                }
                lic_obj = Licitacion(**datos_lic)
                self.db.save_licitacion(lic_obj)
                lic_id_db = lic_obj.id

                # 1. ASIGNAR GANADOR
                self.db.marcar_ganador_lote(lic_id_db, "101", "Competidor Ganador", None)
                ganadores = self.db.get_ganadores_por_lote(lic_id_db)
                if ganadores and ganadores[0]['ganador_nombre'] == "Competidor Ganador":
                    log.append("[PASS] ASIGNAR: Se asignó correctamente un ganador al lote.")
                else:
                    log.append(f"[FAIL] ASIGNAR: No se pudo verificar la asignación del ganador. Se obtuvo: {ganadores}")

                # 2. ELIMINAR GANADOR
                self.db.borrar_ganador_lote(lic_id_db, "101")
                ganadores_despues = self.db.get_ganadores_por_lote(lic_id_db)
                if not ganadores_despues:
                    log.append("[PASS] ELIMINAR: Se eliminó correctamente la asignación del ganador.")
                else:
                    log.append(f"[FAIL] ELIMINAR: La asignación del ganador no se eliminó. Se obtuvo: {ganadores_despues}")

            except Exception as e:
                log.append(f"[FAIL] PRUEBA DE GANADORES: La prueba falló con una excepción. Error: {e}")
                log.append(traceback.format_exc())
            finally:
                # Limpieza
                try:
                    self.db.delete_licitacion(test_id)
                except Exception:
                    pass

    def _test_fallas_fase_a(self, log):
        """Prueba el registro de fallas de Fase A."""
        log.append("\n--- Módulo: Fallas Fase A ---")
        test_id = f"SMOKETEST-FALLAS-{int(datetime.datetime.now().timestamp())}"
        
        try:
            # Setup: Crear licitación con documento y competidor
            datos_lic = {
                "nombre_proceso": "Prueba Fallas", "numero_proceso": test_id,
                "documentos_solicitados": [{"codigo": "F-01", "nombre": "Doc de Falla"}],
                "oferentes_participantes": [{"nombre": "Competidor con Falla"}]
            }
            lic_obj = Licitacion(**datos_lic)
            self.db.save_licitacion(lic_obj)
            lic_id_db = lic_obj.id
            doc_id_db = lic_obj.documentos_solicitados[0].id

            # 1. REGISTRAR FALLA
            fallas_a_registrar = [{
                "licitacion_id": lic_id_db,
                "participante_nombre": "Competidor con Falla",
                "documento_id": doc_id_db,
                "comentario": "Falla de prueba",
                "es_nuestro": False
            }]
            # Usamos el método genérico _save_related_data que ya existe
            self.db._save_related_data('descalificaciones_fase_a', lic_id_db, fallas_a_registrar, 
                                    ['licitacion_id', 'participante_nombre', 'documento_id', 'comentario', 'es_nuestro'])
            
            # 2. VERIFICAR FALLA
            # Recargamos la licitación completa para verificar
            lic_data, _, _, _, _, _ = self.db.get_all_data()
            lic_recargada_data = next((l for l in lic_data if l.get('id') == lic_id_db), None)
            
            if lic_recargada_data and lic_recargada_data.get('fallas_fase_a'):
                falla_registrada = lic_recargada_data['fallas_fase_a'][0]
                if falla_registrada['documento_id'] == doc_id_db and falla_registrada['participante_nombre'] == "Competidor con Falla":
                    log.append("[PASS] REGISTRAR: Se registró y verificó correctamente una falla de Fase A.")
                else:
                    log.append(f"[FAIL] VERIFICAR: La falla registrada no coincide con los datos esperados. Se obtuvo: {falla_registrada}")
            else:
                log.append("[FAIL] VERIFICAR: No se encontraron fallas registradas para la licitación de prueba.")

        except Exception as e:
            log.append(f"[FAIL] PRUEBA DE FALLAS: La prueba falló con una excepción. Error: {e}")
            log.append(traceback.format_exc())
        finally:
            # Limpieza
            try:
                self.db.delete_licitacion(test_id)
            except Exception:
                pass


def _render_portada_pdf_bytes(titulo_expediente, lic_data, qr_text=None):
    """
    Genera un PDF (en memoria) con portada: título, info de licitación y QR.
    Devuelve bytes.
    """
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    W, H = letter

    # Título
    c.setFont("Helvetica-Bold", 20)
    c.drawString(1.0*inch, H - 1.5*inch, titulo_expediente)

    # Info básica de la licitación
    c.setFont("Helvetica", 11)
    y = H - 2.0*inch
    info_lines = [
        f"Proceso: {lic_data.get('numero_proceso','')}",
        f"Nombre:  {lic_data.get('nombre_proceso','')}",
        f"Institución: {lic_data.get('institucion','')}",
        f"Empresa: {lic_data.get('empresa_nuestra','')}",
        f"Generado: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}",
    ]
    for line in info_lines:
        c.drawString(1.0*inch, y, line); y -= 0.25*inch

    # QR (opcional)
    if qr_text and QR_AVAILABLE:
        qr_img = qrcode.make(qr_text)
        qr_buf = io.BytesIO()
        qr_img.save(qr_buf, format="PNG")
        qr_buf.seek(0)
        from reportlab.lib.utils import ImageReader
        img = ImageReader(qr_buf)
        c.drawImage(img, W - 2.2*inch, H - 2.7*inch, width=1.8*inch, height=1.8*inch)

    c.showPage()
    c.save()
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


def _render_indice_pdf_bytes(items_con_paginas):
    """
    Genera un PDF (en memoria) con un índice paginado.
    items_con_paginas: lista de dicts [{ 'titulo':..., 'pagina_inicio': int }, ...]
    Devuelve bytes.
    """
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    W, H = letter

    c.setFont("Helvetica-Bold", 16)
    c.drawString(1.0*inch, H - 1.0*inch, "ÍNDICE DEL EXPEDIENTE")
    c.setFont("Helvetica", 10)
    y = H - 1.5*inch

    for it in items_con_paginas:
        titulo = it['titulo']
        pag = it['pagina_inicio']
        line = f"{titulo}"
        c.drawString(1.0*inch, y, line)
        c.drawRightString(W - 1.0*inch, y, f"Pág. {pag}")
        y -= 0.28*inch
        if y < 1.0*inch:
            c.showPage()
            c.setFont("Helvetica", 10)
            y = H - 1.0*inch

    c.showPage()
    c.save()
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


def generar_expediente_pdf(db: DatabaseManager, licitacion, items, out_path, meta):
    """
    Une PDFs en un solo expediente con:
      - Portada (con QR)
      - Índice paginado
      - Marcadores (si PyPDF2 lo soporta)
    Parámetros:
      db: DatabaseManager
      licitacion: objeto Licitacion (ya cargado en UI)
      items: lista de dicts [{ 'orden':1, 'doc_version_id': <documentos.id>, 'titulo':'...' }, ...]
      out_path: ruta de salida PDF
      meta: dict con campos como {'titulo_expediente': '...', 'qr_text': '...'}
    """
    # 1) Guardar cabecera + items en BD (para trazabilidad)
    exp_id = db.crear_expediente(licitacion.id, meta.get('titulo_expediente','Expediente'), meta.get('creado_por','Usuario'))
    db.agregar_items_expediente(exp_id, items)

    # 2) Construir lista de PDFs existentes y títulos
    #    (Portada y índice se calculan con PDF en memoria)
    merger = PdfMerger()

    # --- Portada ---
    portada_bytes = _render_portada_pdf_bytes(
        meta.get('titulo_expediente', 'Expediente'),
        {
            'numero_proceso': licitacion.numero_proceso,
            'nombre_proceso': licitacion.nombre_proceso,
            'institucion': licitacion.institucion,
            'empresa_nuestra': str(licitacion.empresa_nuestra),
        },
        qr_text=meta.get('qr_text')
    )
    merger.append(PdfReader(io.BytesIO(portada_bytes)))

    # Para calcular págs de inicio, iremos añadiendo y contando.
    pagina_actual = len(merger.pages)  # tras portada
    indice_tmp = []

    # --- Recorrer items (documentos) ---
    for it in sorted(items, key=lambda x: x['orden']):
        doc_id = it['doc_version_id']
        titulo = it['titulo']
        # Obtenemos ruta y abrimos
        db.cursor.execute("SELECT ruta_archivo FROM documentos WHERE id=?", (doc_id,))
        row = db.cursor.fetchone()
        ruta = (row[0] or "") if row else ""
        if not ruta or not os.path.isfile(ruta):
            # Si falta, insertamos una página en blanco con aviso
            aviso = _render_indice_pdf_bytes([{'titulo': f"[FALTANTE] {titulo}", 'pagina_inicio': 0}])
            merger.append(PdfReader(io.BytesIO(aviso)))
            indice_tmp.append({'titulo': f"[FALTANTE] {titulo}", 'pagina_inicio': pagina_actual + 1})
            pagina_actual = len(merger.pages)
            continue

        reader = PdfReader(ruta)
        num_pages = len(reader.pages)

        # Añadimos marcador (si es posible) después de anexar
        merger.append(reader)
        try:
            # Algunos PyPDF2 exponen .add_outline_item, otros usan .addBookmark en writer.
            # PdfMerger permite "add_outline_item(title, page_number)" en versiones recientes.
            merger.add_outline_item(titulo, pagina_actual)
        except Exception:
            pass

        indice_tmp.append({'titulo': titulo, 'pagina_inicio': pagina_actual + 1})
        pagina_actual += num_pages

    # --- Índice (después de portada) ---
    indice_bytes = _render_indice_pdf_bytes(indice_tmp)
    # Truco: insertar el índice en la segunda posición (pos=1)
    merger.merge(1, PdfReader(io.BytesIO(indice_bytes)))

    # --- Guardar PDF final ---
    with open(out_path, "wb") as f:
        merger.write(f)
    merger.close()

    return exp_id

def previsualizar_expediente(ruta_pdf):
    if not os.path.isfile(ruta_pdf):
        messagebox.showerror("Archivo no encontrado", "No existe el PDF generado.", parent=None)
        return
    try:
        if platform.system() == "Windows":
            os.startfile(ruta_pdf)
        elif platform.system() == "Darwin":
            subprocess.call(["open", ruta_pdf])
        else:
            subprocess.call(["xdg-open", ruta_pdf])
    except Exception as e:
        messagebox.showwarning("Aviso", f"No se pudo abrir el PDF automáticamente.\nRuta: {ruta_pdf}\n\n{e}")

# ===================== Confirmador de Orden por Categoría =====================

# ===== Categorías fijas del expediente (exactamente 4) =====
CATS_ORDEN_EXPD = ["Legal", "Financiera", "Técnica", "Sobre B"]

def _cat_norm_exp(s: str) -> str:
    """
    Normaliza la categoría a una de las 4 fijas:
    Legal, Financiera, Técnica, Sobre B.
    Si no reconoce, la envía a 'Técnica' como default para no perderla.
    """
    if not isinstance(s, str):
        return "Técnica"
    s0 = s.strip().lower()

    # quitar acentos
    import unicodedata, re
    s1 = "".join(c for c in unicodedata.normalize("NFD", s0) if unicodedata.category(c) != "Mn")

    if "sobre" in s1 and "b" in s1:
        return "Sobre B"
    if "finan" in s1:
        return "Financiera"
    if "legal" in s1 or "jurid" in s1:
        return "Legal"
    if "tec" in s1 or "tecnic" in s1:
        return "Técnica"

    # Default (para no perder documentos por nombre raro)
    return "Técnica"


# REEMPLAZA esta clase por completo en tu archivo:

# REEMPLAZA esta clase por completo en tu archivo:

class DialogoOrdenExpediente(tk.Toplevel):
    """
    Revisa y reordena documentos agrupados en las 4 categorías fijas del expediente,
    utilizando una interfaz moderna de tablas.
    """
# En DialogoOrdenExpediente, reemplaza el método __init__

    def __init__(self, parent, documentos_obj, cats_prioridad=CATS_ORDEN_EXPD):
        super().__init__(parent)
        self.title("Confirmar orden del expediente")
        self.geometry("950x550")
        self.transient(parent)
        self.grab_set()

        grupos = {cat: [] for cat in cats_prioridad}
        for d in documentos_obj:
            cat = _cat_norm_exp(getattr(d, "categoria", ""))
            grupos[cat].append(d)

        self._data = {}
        self._trees = {} # La variable correcta es _trees
        self._incluir = {}
        self._tabs = ttk.Notebook(self)
        self._tabs.pack(fill="both", expand=True, padx=10, pady=10)

        for cat in cats_prioridad:
            docs = grupos.get(cat, [])
            
            def get_sort_key(documento):
                orden = getattr(documento, "orden_pliego", None)
                return 999999 if orden is None else int(orden)
            
            docs = sorted(docs, key=get_sort_key)
            
            f = ttk.Frame(self._tabs, padding=10)
            self._tabs.add(f, text=cat)
            self._data[cat] = list(docs)
            self._incluir[cat] = tk.BooleanVar(value=True)

            top = ttk.Frame(f)
            top.pack(fill="x")
            ttk.Checkbutton(top, text=f"Incluir {cat} en el expediente", variable=self._incluir[cat]).pack(side="left")

            mid = ttk.Frame(f)
            mid.pack(fill="both", expand=True, pady=6)

            cols = ('presentado', 'codigo', 'nombre')
            tree = ttk.Treeview(mid, columns=cols, show="headings", selectmode=tk.EXTENDED)
            tree.heading('presentado', text='✓'); tree.heading('codigo', text='Código'); tree.heading('nombre', text='Nombre del Documento')
            tree.column('presentado', width=30, anchor=tk.CENTER, stretch=False); tree.column('codigo', width=150); tree.column('nombre', width=500)
            
            tree.pack(side="left", fill="both", expand=True)
            scrollbar = ttk.Scrollbar(mid, orient="vertical", command=tree.yview)
            tree.configure(yscrollcommand=scrollbar.set)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            self._trees[cat] = tree

            btns = ttk.Frame(mid)
            btns.pack(side="left", fill="y", padx=8)
            ttk.Button(btns, text="⬆️ Subir", command=lambda c=cat: self._mover(c, -1)).pack(fill="x", pady=3)
            ttk.Button(btns, text="⬇️ Bajar", command=lambda c=cat: self._mover(c, 1)).pack(fill="x", pady=3)
            ttk.Button(btns, text="⤒ Arriba", command=lambda c=cat: self._to_edge(c, top=True)).pack(fill="x", pady=3)
            ttk.Button(btns, text="⤓ Abajo", command=lambda c=cat: self._to_edge(c, top=False)).pack(fill="x", pady=3)
            ttk.Button(btns, text="↺ Reset (orden guardado)", command=lambda c=cat: self._reset(c)).pack(fill="x", pady=12)

            self._render(cat)

        bottom = ttk.Frame(self)
        bottom.pack(fill="x", padx=10, pady=(0, 10))
        ttk.Button(bottom, text="Cancelar", command=self._cancel).pack(side="right", padx=6)
        ttk.Button(bottom, text="Guardar y continuar", command=self._accept).pack(side="right")

        self.result_incluir = None
        self.result_orden = None


    def _render(self, cat):
        tree = self._trees[cat] # CORRECCIÓN: Usar self._trees
        tree.delete(*tree.get_children())
        for i, d in enumerate(self._data[cat]):
            presentado_icono = "✓" if getattr(d, "presentado", False) else "❌"
            codigo = getattr(d, 'codigo', '')
            nombre = getattr(d, 'nombre', '')
            tree.insert('', tk.END, iid=str(i), values=(presentado_icono, codigo, nombre))

# En DialogoOrdenExpediente, reemplaza el método _mover

    def _mover(self, cat, delta):
        tree = self._trees[cat] # CORRECCIÓN: Usar self._trees
        sel_iids = tree.selection()
        if not sel_iids: return

        sel_indices = sorted([int(iid) for iid in sel_iids])
        
        items = self._data[cat]
        if delta < 0:
            for i in sel_indices:
                if i > 0 and str(i - 1) not in sel_iids:
                    items[i], items[i-1] = items[i-1], items[i]
        else:
            for i in reversed(sel_indices):
                if i < len(items) - 1 and str(i + 1) not in sel_iids:
                    items[i], items[i+1] = items[i+1], items[i]
        
        self._render(cat)
        
        nuevos_iids_a_seleccionar = []
        for i in sel_indices:
            try:
                # Recalculamos la nueva posición del objeto en la lista
                obj = self._data[cat][i+delta]
                nuevo_indice = self._data[cat].index(obj)
                nuevos_iids_a_seleccionar.append(str(nuevo_indice))
            except (ValueError, IndexError):
                pass
        
        tree.selection_set(nuevos_iids_a_seleccionar)

# En DialogoOrdenExpediente, reemplaza el método _to_edge

    def _to_edge(self, cat, top=True):
        tree = self._trees[cat] # CORRECCIÓN: Usar self._trees
        sel_iids = tree.selection()
        if not sel_iids: return
        
        sel_indices = {int(iid) for iid in sel_iids}
        items = self._data[cat]
        
        picked = [item for i, item in enumerate(items) if i in sel_indices]
        rest = [item for i, item in enumerate(items) if i not in sel_indices]
        
        self._data[cat] = picked + rest if top else rest + picked
        
        self._render(cat)
        
        if top:
            nuevos_iids = [str(i) for i in range(len(sel_indices))]
        else:
            nuevos_iids = [str(i) for i in range(len(rest), len(items))]
        tree.selection_set(nuevos_iids)

    def _reset(self, cat):
        def get_sort_key(d):
            orden = getattr(d, "orden_pliego", None)
            return 999999 if orden is None else int(orden)
        self._data[cat].sort(key=get_sort_key)
        self._render(cat)

    def _accept(self):
        self.result_incluir = {cat: var.get() for cat, var in self._incluir.items()}
        self.result_orden = {cat: list(self._data[cat]) for cat in self._data}
        self.destroy()

    def _cancel(self):
        self.result_incluir = None
        self.result_orden = None
        self.destroy()

    def _render(self, cat):
        tree = self._trees[cat]
        tree.delete(*tree.get_children())
        for i, d in enumerate(self._data[cat]):
            presentado_icono = "✓" if getattr(d, "presentado", False) else "❌"
            codigo = getattr(d, 'codigo', '')
            nombre = getattr(d, 'nombre', '')
            tree.insert('', tk.END, iid=str(i), values=(presentado_icono, codigo, nombre))

    def _mover(self, cat, delta):
        tree = self._trees[cat]
        sel_iids = tree.selection()
        if not sel_iids: return
        sel_indices = sorted([int(iid) for iid in sel_iids])
        items = self._data[cat]
        if delta < 0:
            for i in sel_indices:
                if i > 0 and str(i - 1) not in sel_iids:
                    items[i], items[i-1] = items[i-1], items[i]
        else:
            for i in reversed(sel_indices):
                if i < len(items) - 1 and str(i + 1) not in sel_iids:
                    items[i], items[i+1] = items[i+1], items[i]
        self._render(cat)
        nuevos_iids = []
        for i in sel_indices:
            j = i + delta
            if 0 <= j < len(self._data[cat]):
                nuevos_iids.append(str(j))
        tree.selection_set(nuevos_iids)

    def _to_edge(self, cat, top=True):
        tree = self._trees[cat]
        sel_iids = tree.selection()
        if not sel_iids: return
        sel_indices = {int(iid) for iid in sel_iids}
        items = self._data[cat]
        picked = [item for i, item in enumerate(items) if i in sel_indices]
        rest = [item for i, item in enumerate(items) if i not in sel_indices]
        self._data[cat] = picked + rest if top else rest + picked
        self._render(cat)
        if top:
            nuevos_iids = [str(i) for i in range(len(picked))]
        else:
            nuevos_iids = [str(i) for i in range(len(self._data[cat]) - len(picked), len(self._data[cat]))]
        tree.selection_set(nuevos_iids)

    def _reset(self, cat):
        def get_sort_key(d):
            orden = getattr(d, "orden_pliego", None)
            return 999999 if orden is None else int(orden)
        self._data[cat].sort(key=get_sort_key)
        self._render(cat)

    def _accept(self):
        self.result_incluir = {cat: var.get() for cat, var in self._incluir.items()}
        self.result_orden = {cat: list(self._data[cat]) for cat in self._data}
        self.destroy()

    def _cancel(self):
        self.result_incluir = None
        self.result_orden = None
        self.destroy()


def generar_expediente_zip_por_categoria(db, licitacion, carpeta_salida, orden_por_cat, incluir):
    """
    Crea un ZIP por cada categoría marcada en 'incluir', respetando el orden manual
    confirmado en 'orden_por_cat' (que contiene listas de OBJETOS Documento).
    - CATS: ["Legal", "Financiera", "Técnica", "Sobre B"]
    - Dentro del ZIP:
        * index.csv -> [orden, codigo, nombre, categoria, archivo]
        * Archivos en orden (prefijo 001-, 002-, ...). Si falta, crea FALTANTE_###.txt
    Devuelve: lista de rutas zip generadas.
    """
    import os, io
    from zipfile import ZipFile, ZIP_DEFLATED
    from csv import writer

    os.makedirs(carpeta_salida, exist_ok=True)
    generados = []

    for cat in CATS_ORDEN_EXPD:
        if not incluir.get(cat, False):
            continue
        docs_obj = list(orden_por_cat.get(cat, []) or [])
        if not docs_obj:
            continue

        nombre_zip = f"Expediente - {cat} - {licitacion.numero_proceso}.zip"
        out_zip_path = os.path.join(carpeta_salida, nombre_zip)

        try:
            with ZipFile(out_zip_path, "w", compression=ZIP_DEFLATED) as zf:
                # 1) index.csv con el orden
                buf = io.StringIO()
                w = writer(buf)
                w.writerow(["orden", "codigo", "nombre", "categoria", "archivo"])
                for i, d in enumerate(docs_obj, start=1):
                    codigo = getattr(d, "codigo", "") or ""
                    nombre = getattr(d, "nombre", "") or ""
                    categoria = getattr(d, "categoria", "") or ""
                    ruta = getattr(d, "ruta_archivo", "") or ""
                    archivo = os.path.basename(ruta) if ruta else f"FALTANTE_{i:03d}.txt"
                    w.writerow([i, codigo, nombre, categoria, archivo])
                zf.writestr("index.csv", buf.getvalue())

                # 2) Archivos en orden (prefijo para mantener orden y evitar duplicados)
                for i, d in enumerate(docs_obj, start=1):
                    ruta = getattr(d, "ruta_archivo", "") or ""
                    if ruta and os.path.isfile(ruta):
                        base = os.path.basename(ruta)
                        arcname = f"{i:03d} - {base}"
                        zf.write(ruta, arcname=arcname)
                    else:
                        zf.writestr(f"FALTANTE_{i:03d}.txt", "Documento no encontrado o sin archivo adjunto.")

            generados.append(out_zip_path)

        except Exception as e:
            try:
                messagebox.showerror("Error ZIP", f"No se pudo crear el ZIP de {cat}:\n{e}")
            except Exception:
                pass

    return generados


def generar_expediente_zip(db: DatabaseManager, licitacion, out_zip_path, items=None):
    """
    Crea un ZIP con:
      - Archivos PDF/DOCX existentes de los documentos seleccionados (o todos)
      - index.csv con orden, título, nombre de archivo
    """
    docs = db.obtener_documentos_de_licitacion(licitacion.id)
    if items:
        ids = {it['doc_version_id'] for it in items}
        docs = [d for d in docs if d['id'] in ids]

    if not docs:
        messagebox.showwarning("Sin documentos", "No hay documentos para incluir.", parent=None)
        return False

        # Orden: por items (si se pasa), si no: categoria+codigo
    if items:
        # Si te pasaron items, ese orden manda.
        orden_map = {it['doc_version_id']: it['orden'] for it in items}
        docs.sort(key=lambda d: orden_map.get(d['id'], 999999))
    else:
        # No hay items -> intentamos orden del pliego.
        # Construimos un mapa (codigo+nombre normalizados -> orden_pliego) desde los OBJETOS en memoria.
# ... else:
        try:
            objs = list(getattr(licitacion, "documentos_solicitados", []) or [])
        except Exception:
            objs = []

        orden_map_pliego = {getattr(o, "id", -1): getattr(o, "orden_pliego", 999999) for o in objs}
        docs.sort(key=lambda d: (orden_map_pliego.get(d.get("id", -1), 999999), d.get("categoria") or "", d.get("codigo") or ""))

        def _norm(s):
            s = str(s or "").strip().lower()
            import unicodedata, re
            s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
            return re.sub(r"\s+", " ", s)

        orden_map_pliego = {}
        # Si tus objetos Documento tienen 'id' de BD, podrías usar un map por id;
        # este enfoque por (codigo,nombre) funciona aunque no tengas id sincronizado.
        for obj in objs:
            key = (_norm(getattr(obj, "codigo", "")), _norm(getattr(obj, "nombre", "")))
            orden_map_pliego[key] = getattr(obj, "orden_pliego", 999999)

        def _zip_key(d):
            key = (_norm(d.get("codigo", "")), _norm(d.get("nombre", "")))
            # Orden principal: orden_pliego (si no está, 999999 para dejarlos al final)
            # Orden secundario estable: categoria+codigo
            return (orden_map_pliego.get(key, 999999), d.get("categoria") or "", d.get("codigo") or "")

        docs.sort(key=_zip_key)


    try:
        with ZipFile(out_zip_path, "w", compression=ZIP_DEFLATED) as zf:
            # índice CSV
            from csv import writer
            buf = io.StringIO()
            w = writer(buf)
            w.writerow(["orden", "titulo", "archivo"])
            for i, d in enumerate(docs, start=1):
                titulo = f"[{d.get('codigo') or ''}] {d.get('nombre') or ''}".strip()
                ruta = d.get('ruta_archivo') or ''
                nombre_archivo = os.path.basename(ruta) if ruta else f"FALTANTE_{i}.txt"
                w.writerow([i, titulo, nombre_archivo])
            zf.writestr("index.csv", buf.getvalue())

            # ficheros
            for i, d in enumerate(docs, start=1):
                ruta = d.get('ruta_archivo') or ''
                if os.path.isfile(ruta):
                    arcname = os.path.basename(ruta)
                    zf.write(ruta, arcname=arcname)
                else:
                    # marcador “faltante”
                    zf.writestr(f"FALTANTE_{i}.txt", "Documento no encontrado.")
        return True
    except Exception as e:
        messagebox.showerror("Error ZIP", f"No se pudo crear el ZIP:\n{e}")
        return False




def main():
    """Función principal para lanzar la aplicación de forma robusta."""
    setup_logging()
    
    db_path = seleccionar_o_crear_db_inicial()
    
    if not db_path:
        print("No se seleccionó ninguna base de datos. Cerrando aplicación.")
        return
        
    app = AppLicitacionesGUI(db_path=db_path)
    app.mainloop()

if __name__ == "__main__":
    main()